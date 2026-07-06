# -*- coding: utf-8 -*-
from __future__ import annotations

from typing import Dict, List
import csv
import math
import os
import re
import html as _html

from qgis.PyQt import QtCore, QtWidgets, QtGui
from qgis.PyQt.QtGui import QGuiApplication
from .i18n import apply_i18n, current_language, install_runtime_i18n_patches, translate_html, tr_text as _tr
from .ui_core.responsive import fit_to_screen, configure_table
from qgis.core import QgsFeatureRequest, QgsVectorLayer

try:
    from .noise_core.noise_common import OCTAVE_BANDS, A_WEIGHTING
except Exception:
    from noise_core.noise_common import OCTAVE_BANDS, A_WEIGHTING

try:
    from openpyxl import Workbook
except Exception:
    Workbook = None


# Client-facing receiver table/export schema.  Detailed MDT/path diagnostics are
# still kept internally in the result payload, but the default dialog and exports
# should stay readable for consultancy workflows.
NOISE_I18N_NATIVE_BUILD = "2026-07-06-v7-consultancy-report-clean"

CONSULTANCY_RECEIVER_COLUMNS = [
    ("rec_id", "ID receptor"),
    ("rec_type", "tipo"),
    ("noise_dba", "nivel total dB(A)"),
    ("limit_dba", "límite dB(A)"),
    ("margin_db", "margen respecto al límite dB"),
    ("state", "estado"),
    ("exceeds", "supera el límite"),
    ("n_src", "nº aerogeneradores"),
    ("near_m", "dist. aerogenerador cercano (m)"),
    ("dom_model", "modelo dominante"),
    ("dom_group", "grupo fuente dom."),
    ("dom_park", "parque dom."),
    ("src_lwa", "LwA fuente dom. dB(A)"),
    ("adiv_db", "Adiv dB"),
    ("aatm_db", "Aatm dB"),
    ("aground_db", "Agr/Aground dB"),
    ("abar_max_db", "Abar máx. dB"),
    ("ground_g", "G suelo"),
    ("ground_md", "modo suelo"),
    ("rec_h_m", "h receptor m"),
    ("rec_z_m", "z terreno receptor m"),
    ("rec_ac_z_m", "z acústica receptor m"),
    ("dom_src_lyr", "capa fuente dominante"),
]

CONSULTANCY_RECEIVER_KEYS = [key for key, _label in CONSULTANCY_RECEIVER_COLUMNS]
CONSULTANCY_RECEIVER_HEADERS = [label for _key, label in CONSULTANCY_RECEIVER_COLUMNS]



def _cleanup_german_noise_html(html: str) -> str:
    """Post-process German HTML generated from the French report template.

    The runtime translator is intentionally conservative and non-cascading to
    avoid corrupting short UI labels.  The noise report is long HTML assembled
    from older French text blocks, so a final DE-only cleanup makes the visible
    report much less mixed without affecting other languages.
    """
    repl = [

        ("Récepteur critique (niveau sonore le plus élevé)", "Kritischer Rezeptor (höchster Schallpegel)"),
        ("ID récepteur", "Rezeptor-ID"),
        ("Niveau total", "Gesamtpegel"),
        ("Limite applicable", "Anwendbarer Grenzwert"),
        ("Marge", "Abstand zum Grenzwert"),
        ("Modèle dominant", "Dominantes Modell"),
        ("Groupe source", "Quellgruppe"),
        ("Éoliennes contributrices dans le rayon", "Beitragende Windturbinen im Radius"),
        ("Décomposition des atténuations", "Aufschlüsselung der Dämpfungen"),
        ("Les valeurs affichées ci-dessous sont les amplitudes d’atténuation utilisées par le modèle. Dans l’équation principale, ces termes sont soustraits au niveau de source.", "Die unten angezeigten Werte sind die vom Modell verwendeten Dämpfungsbeträge. In der Hauptgleichung werden diese Terme vom Quellpegel abgezogen."),
        ("Puissance acoustique de l’éolienne", "Schallleistung der Windturbine"),
        ("Dispersion géométrique", "Geometrische Ausbreitungsdämpfung"),
        ("Absorption dans l’air", "Luftabsorption"),
        ("Effet du sol", "Bodeneffekt"),
        ("Diffraction topographique", "Topografische Abschirmung"),
        ("Abar maximal des contributeurs", "Maximaler Abar der Beitragenden"),
        ("Abar maximal parmi toutes les éoliennes qui contribuent au récepteur", "Maximaler Abar-Wert unter allen Windturbinen, die zum Rezeptor beitragen"),
        ("Abar pondéré par énergie", "Energiegewichteter Abar"),
        ("Moyenne pondérée par la contribution acoustique de chaque éolienne", "Mittelwert, gewichtet nach dem akustischen Beitrag jeder Windturbine"),
        ("Trajets écrantés", "Abgeschirmte Pfade"),
        ("Nombre d’éoliennes contributrices avec Abar &gt; 0 dB", "Anzahl beitragender Windturbinen mit Abar &gt; 0 dB"),
        ("Note : le niveau résultant inclut la sommation énergétique multi-source et multi-bande ; ce n’est pas une soustraction directe depuis une seule éolienne.", "Hinweis: Der Ergebnispegel enthält die energetische Summierung über mehrere Quellen und Frequenzbänder; er ist keine direkte Subtraktion von einer einzelnen Windturbine."),
        ("Note : le niveau résultant inclut", "Hinweis: Der Ergebnispegel enthält"),
        ("sommation énergétique multi-source et multi-bande", "energetische Summierung über mehrere Quellen und Frequenzbänder"),
        ("ce n’est pas une soustraction directe depuis une seule éolienne", "dies ist keine direkte Subtraktion von einer einzelnen Windturbine"),
        ("Bande dominante", "Dominantes Band"),
        ("Origine du spectre", "Spektrumquelle"),
        ("Glossaire des symboles", "Symbolglossar"),
        ("Définition compacte des symboles qui apparaissent dans les formules et tableaux de ce rapport.", "Kompakte Definition der Symbole, die in Formeln und Tabellen dieses Berichts erscheinen."),
        ("Symbole", "Symbol"),
        ("Signification", "Bedeutung"),
        ("Portée de ce rapport — à lire avant d’utiliser les résultats", "Geltungsbereich dieses Berichts — vor Nutzung der Ergebnisse lesen"),
        ("Ce que c’est", "Was es ist"),
        ("Ce que ce n’est pas", "Was es nicht ist"),
        ("Simplifications appliquées dans ce mode", "In diesem Modus angewendete Vereinfachungen"),
        ("Recommandation", "Empfehlung"),
        ("Statistiques des atténuations", "Dämpfungsstatistik"),
        ("Statistiques des Dämpfungen", "Dämpfungsstatistik"),
        ("Estadísticos de Dämpfungen", "Dämpfungsstatistik"),
        ("Estadísticos de", "Statistik der"),
        ("Dämpfungen (Abgedeckte Rezeptoren)", "Dämpfungen (abgedeckte Rezeptoren)"),
        ("Abgedeckte Rezeptoren", "abgedeckte Rezeptoren"),
        ("Abar Maximum parmi les Turbinen contributrices", "Maximaler Abar-Wert unter den beitragenden Windturbinen"),
        ("Abar Maximum", "Maximaler Abar"),
        ("Abar moyen", "Mittlerer Abar"),
        ("au Rezeptor", "zum Rezeptor"),
        ("bandes", "Frequenzbänder"),
        ("bandes.", "Frequenzbänder."),
        ("Turbinen contributrices", "beitragende Windturbinen"),
        ("Turbinen contribuyentes", "beitragende Windturbinen"),
        ("con Turbinen", "mit Windturbinen"),
        ("con Abar", "mit Abar"),
        ("Numero de", "Anzahl"),
        ("Número de", "Anzahl"),
        ("numéro de", "Anzahl"),
        ("d’éoliennes", "Windturbinen"),
        ("de Windturbinen", "Windturbinen"),
        ("eólica", "Windenergie"),
        ("Turbine", "Windturbine"),
        ("Turbinen", "Windturbinen"),
        ("Rezeptor ne", "Rezeptor bedeutet nicht"),
        (" ne signifie pas que ", " bedeutet nicht, dass "),
        (" désactivé", " deaktiviert"),
        ("désactivé", "deaktiviert"),
        ("pour le dominanten Pfad", "für den dominanten Pfad"),
        ("für den dominanten Pfad konnte kein gültiges DGM-Profil extrahiert werden", "für den dominanten Pfad kein gültiges DGM-Profil extrahiert werden konnte"),
        ("Der Wert Abar dominanter Pfad", "Der Abar-Wert des dominanten Pfads"),
        ("la Wert", "Der Wert"),
        ("est obtenu par", "wird ermittelt durch"),
        ("par sommation énergétique", "durch energetische Summierung"),
        ("de toutes", "aller"),
        ("und Frequenzbänder.", "und Frequenzbänder."),
        ("Im ISO-orientierten Rechenkern, das DGM ändert nicht", "Im ISO-orientierten Rechenkern ändert das DGM nicht"),
        ("Im schnellen Rechenkern, das DGM", "Im schnellen Rechenkern das DGM"),
        ("In dieser Berechnung,", "In dieser Berechnung"),
        ("l’atmosphärische Absorption", "die atmosphärische Absorption"),
        ("l’Bodeneffekt", "der Bodeneffekt"),
        ("l’Landnutzung", "die Landnutzung"),
        ("l’Windturbine", "die Windturbine"),
        ("l’Emission", "die Emission"),
        ("l’équation", "die Gleichung"),
        ("l’itération", "die Iteration"),
        ("d’Dämpfung", "Dämpfung"),
        ("d’topografische", "topografische"),
        ("d’Landnutzung", "Landnutzung"),
        ("mit ein <b>einziger", "mit einem <b>einzigen"),
        ("ein <b>einziger manueller G-Wert</b>", "einem <b>einzigen manuellen G-Wert</b>"),
        ("ein <b>einziger G-Wert je Pfad</b>", "ein <b>einziger G-Wert je Pfad</b>"),
        ("with ein", "mit einem"),
        ("with eine", "mit einer"),
        ("with un", "mit einem"),
        ("with une", "mit einer"),
        ("with ", "mit "),
        ("Adiv représente la geometrische Divergenz", "Adiv steht für die geometrische Divergenz"),
        ("Aatm wird berechnet je Band", "Aatm wird je Band berechnet"),
        ("und dépend von", "und hängt ab von"),
        ("dépend von", "hängt ab von"),
        ("und des Drucks", "und vom Druck"),
        ("mit einer formulation vereinfacht", "mit einer vereinfachten Formulierung"),
        ("Agr ist appliqué comme terme", "Agr wird als Term angewendet"),
        ("terme de Boden/Gelände", "Boden-/Geländeterm"),
        ("topografische Abschirmung de base", "grundlegende topografische Abschirmung"),
        ("lorsqu’un MDT ist disponible", "wenn ein DGM verfügbar ist"),
        ("Table synthétique pour la Beratung", "Übersichtstabelle für die Beratung"),
        ("résultats acoustiques par Rezeptor", "akustische Ergebnisse je Rezeptor"),
        ("source dominante", "dominante Quelle"),
        ("atténuations principales", "wichtigste Dämpfungen"),
        ("Les diagnostics internes MDT par paire", "Die internen DGM-Paardiagnosen"),
        ("sind conservés en mémoire", "werden im Speicher gehalten"),
        ("mais ne sind pas affichés por defecto", "werden aber standardmäßig nicht angezeigt"),
        ("n’introduit pas de terme explicite", "führt keinen expliziten Term ein"),
        ("Même si une Layer de relief existe dans le projet", "Auch wenn im Projekt ein Gelände-Layer vorhanden ist"),
        ("ce mode ne calcule pas", "berechnet dieser Modus nicht"),
        ("n’extrait pas de ligne de visée", "extrahiert keine Sichtlinie"),
        ("n’applique pas de diffraction", "wendet keine Beugung an"),
        ("la Physik se base donc uniquement sur", "die Physik basiert daher nur auf"),
        ("la correction empirique de Gelände", "der empirischen Geländekorrektur"),
        ("Les Werte affichées ci-dessous", "Die unten angezeigten Werte"),
        ("sind les Beträge", "sind die Beträge"),
        ("verwendet par le Modell", "die vom Modell verwendet werden"),
        ("ces termes sind soustraits", "diese Terme werden abgezogen"),
        ("au niveau de source", "vom Quellpegel"),
        ("Entrées tatsächlich verwendet dans ce Berechnung", "In dieser Berechnung tatsächlich verwendete Eingaben"),
        ("Ce moteur travaille en", "Dieser Rechenkern arbeitet mit"),
        ("Les bandes ne sind pas un résultat", "Die Bänder sind kein Ergebnis"),
        ("mais la <b>grille fréquentielle de la Methode</b>", "sondern das <b>Frequenzraster der Methode</b>"),
        ("le Berechnung a besoin", "die Berechnung benötigt"),
        ("d’une <b>entrée acoustique je Band</b>", "einen <b>akustischen Eingang je Band</b>"),
        ("Cette entrée kann provenir", "Dieser Eingang kann stammen"),
        ("d’un Spektrum mesuré/importé", "aus einem gemessenen/importierten Spektrum"),
        ("ou d’un gabarit/fallback ajusté", "oder aus einer angepassten Vorlage/einem Fallback"),
        ("au niveau global opérationnel", "an den globalen Betriebspegel"),
        ("Der Term de Boden se décompose en", "Der Bodenterm wird aufgeteilt in"),
        ("trois paramètres de Boden indépendants", "drei unabhängige Bodenparameter"),
        ("werden nicht verwendet ;", "werden nicht verwendet;"),
        ("ist verwendet", "wird verwendet"),
        ("Mathématiquement, le plugin applique", "Mathematisch wendet das Plugin an"),
        ("Lecture correcte d’Abar", "Korrekte Interpretation von Abar"),
        ("correspond uniquement à", "bezieht sich nur auf"),
        ("qui contribue le plus", "die am stärksten beiträgt"),
        ("à sa bande dominante", "und auf ihr dominantes Band"),
        ("niveen total", "Gesamtpegel"),
        ("est obtenu par sommation énergétique", "wird durch energetische Summierung ermittelt"),
        ("Limites und recommandations", "Grenzen und Empfehlungen"),
        ("Engine rapide", "Schneller Rechenkern"),
        ("Engine ISO-orientiert", "ISO-orientierter Rechenkern"),
        ("Adapté au Screening préliminaire", "Geeignet für vorläufiges Screening"),
        ("aux cartes agiles", "und schnelle Karten"),
        ("Adapté aux études techniques préliminaires", "Geeignet für vorläufige technische Studien"),
        ("aux comparaisons", "Vergleiche"),
        ("à la conception", "für die Auslegung"),
        ("Simplifications connues", "Bekannte Vereinfachungen"),
        ("Modèles multiples", "Mehrere Modelle"),
        ("pris en charge", "unterstützt"),
        ("n’est pas activé", "ist nicht aktiviert"),
        ("peut être coûteux", "kann rechenintensiv sein"),
        ("sur de grandes cartes", "auf großen Karten"),
        ("Pour les études réglementaires critiques", "Für kritische regulatorische Studien"),
        ("valider avec des mesures", "mit Messungen validieren"),
        ("logiciel commercial certifié", "zertifizierte kommerzielle Software"),
        ("calculée", "berechnet"),
        ("calculé", "berechnet"),
        ("appliqué", "angewendet"),
        ("appliquée", "angewendet"),
        ("simplificada", "vereinfacht"),
        ("simplifié", "vereinfacht"),
        ("simplifiée", "vereinfacht"),
        ("terme de", "Term für"),
        ("de sol", "des Bodens"),
        ("de terrain", "des Geländes"),
        (" pour ", " für "),
        ("Acoustic sources", "Akustische Quellen"),
        ("Noise · Sources", "Schall · Quellen"),
        ("Sources", "Quellen"),
        ("Receptores", "Rezeptoren"),
        ("Rezeptor cubiertos", "abgedeckte Rezeptoren"),
        ("Rezeptoren cubiertos", "abgedeckte Rezeptoren"),
        ("Trayectorias apantalladas", "Abgeschirmte Pfade"),
        ("Número de Turbinen contribuyentes con Abar &gt; 0 dB", "Anzahl beitragender Windturbinen mit Abar &gt; 0 dB"),
        ("Número de Turbinen contribuyentes con Abar > 0 dB", "Anzahl beitragender Windturbinen mit Abar > 0 dB"),
        ("NIVEL RESULTANTE", "ERGEBNISPEGEL"),
        ("NIVEAU RÉSULTANT", "ERGEBNISPEGEL"),
        ("Nivel resultante", "Ergebnispegel"),
        ("Banda dominante", "Dominantes Band"),
        ("Bande dominante", "Dominantes Band"),
        ("Origen Spektrum", "Spektrumquelle"),
        ("Origine du spectre", "Spektrumquelle"),
        ("Estadísticos de Dämpfungen (Abgedeckte Rezeptoren)", "Dämpfungsstatistik (abgedeckte Rezeptoren)"),
        ("Estadísticos de Dämpfungen", "Dämpfungsstatistik"),
        ("Statistiques des Dämpfungen", "Dämpfungsstatistik"),
        ("DGM-Lesart :", "DGM-Hinweis:"),
        ("au kritischer Rezeptor", "am kritischen Rezeptor"),
        ("ne signifie pas que das DGM ist désactivé", "bedeutet nicht, dass das DGM deaktiviert ist"),
        ("es bedeutet, dass für den dominanten Pfad konnte kein gültiges DGM-Profil extrahiert werden", "es bedeutet, dass für den dominanten Pfad kein gültiges DGM-Profil extrahiert werden konnte"),
        ("Korrekte Interpretation von Abar :", "Korrekte Interpretation von Abar:"),
        ("la Wert Abar dominanter Pfad", "Der Wert Abar des dominanten Pfads"),
        ("bezieht sich nur auf die Windturbine die am stärksten beiträgt", "bezieht sich nur auf die Windturbine, die am stärksten zum Rezeptor beiträgt"),
        ("und auf ihr dominantes Band", "und auf ihr dominantes Band"),
        ("Le Gesamtpegel des Rezeptors", "Der Gesamtpegel des Rezeptors"),
        ("de toutes die Windturbinen und bandes", "aller Windturbinen und Frequenzbänder"),
        ("toutes die Windturbinen und bandes", "aller Windturbinen und Frequenzbänder"),
        ("Abar Maximal entre les Turbinen contributrices", "Maximaler Abar-Wert unter den beitragenden Windturbinen"),
        ("Abar maximal parmi les Turbinen contributrices", "Maximaler Abar-Wert unter den beitragenden Windturbinen"),
        ("Abar moyen", "Mittlerer Abar"),
        ("nach akustischem Beitrag gewichtetes Abar", "Nach akustischem Beitrag gewichteter Abar"),
        ("trajets écrantés", "abgeschirmte Pfade"),
        ("Höhen des dominanten Pfads :", "Höhen des dominanten Pfads:"),
        ("Gelände Windturbine", "Gelände Turbine"),
        ("hub=", "Nabenhöhe="),
        ("akustische Turbinenhöhe", "akustische Turbinenhöhe"),
        ("Gelände Rezeptor", "Gelände Rezeptor"),
        ("h Rezeptor", "Rezeptorhöhe"),
        ("akustische Rezeptorhöhe", "akustische Rezeptorhöhe"),
        ("Nota: el nivel resultante incluye la suma energética multi-Quelle y multi-banda; no es una resta directa de una única Turbine.", "Hinweis: Der Ergebnispegel enthält die energetische Summierung über mehrere Quellen und Frequenzbänder; er ist keine direkte Subtraktion von einer einzelnen Turbine."),
        ("Nota: el nivel resultante incluye", "Hinweis: Der Ergebnispegel enthält"),
        ("la suma energética multi-Quelle y multi-banda", "die energetische Summierung über mehrere Quellen und Frequenzbänder"),
        ("no es una resta directa de una única Turbine", "er ist keine direkte Subtraktion von einer einzelnen Turbine"),
        ("Aatm (atmospheric)", "Aatm (Atmosphäre)"),
        ("Agr (sol)", "Agr (Boden)"),
        ("Abar trayectoria dominante", "Abar dominanter Pfad"),
        ("Maximum contributors Abar", "Maximaler Abar-Wert der beitragenden Turbinen"),
        ("Energy-weighted Abar", "Energiegewichteter Abar"),
        ("Average weighted by the acoustic contribution of each turbine", "Mittelwert, gewichtet nach dem akustischen Beitrag jeder Turbine"),
        ("Number de Windturbinen contributrices avec Abar &gt; 0 dB", "Anzahl beitragender Windturbinen mit Abar &gt; 0 dB"),
        ("Number de Windturbinen contributrices avec Abar > 0 dB", "Anzahl beitragender Windturbinen mit Abar > 0 dB"),
        ("Description", "Beschreibung"),
        ("Valeur [dB]", "Wert [dB]"),
        ("Terme", "Term"),
        ("Moyenne [dB]", "Mittelwert [dB]"),
        ("Maximum [dB]", "Maximum [dB]"),
        ("Promedio ponderado por la contribución acústica de cada turbina", "Mittelwert, gewichtet nach dem akustischen Beitrag jeder Turbine"),
        ("Número de turbinas contribuyentes", "Anzahl beitragender Windturbinen"),
        ("contributrices", "beitragend"),
        ("contribuyentes", "beitragend"),
        (" avec ", " mit "),
    ]

    repl.extend([
        # Critical receptor table / visible summary leftovers
        ("Dämpfung due zur Bodeneffekt", "Dämpfung durch Bodeneffekt"),
        ("Dämpfung due zur Bodenefekt", "Dämpfung durch Bodeneffekt"),
        ("Dämpfung due zur Bodeneffekt", "Dämpfung durch Bodeneffekt"),
        ("Máximo Abar zwischen todas las turbinas que contribuyen al receptor", "Maximaler Abar-Wert unter allen Windturbinen, die zum Rezeptor beitragen"),
        ("Máximo Abar entre todas las turbinas que contribuyen al receptor", "Maximaler Abar-Wert unter allen Windturbinen, die zum Rezeptor beitragen"),
        ("Anzahl beitragenWindturbinen mit Abar &gt; 0 dB", "Anzahl beitragender Windturbinen mit Abar &gt; 0 dB"),
        ("Anzahl beitragenWindturbinen mit Abar > 0 dB", "Anzahl beitragender Windturbinen mit Abar > 0 dB"),
        ("Anzahl beitragenWindturbinen", "Anzahl beitragender Windturbinen"),
        ("Número de Turbinen contribuyentes con Abar &gt; 0 dB", "Anzahl beitragender Windturbinen mit Abar &gt; 0 dB"),
        ("Número de Turbinen contribuyentes con Abar > 0 dB", "Anzahl beitragender Windturbinen mit Abar > 0 dB"),
        ("Nota: el nivel resultante incluye la suma energética multi-Quelle y multi-banda; no es una resta directa de una única Turbine.", "Hinweis: Der Ergebnispegel enthält die energetische Summierung über mehrere Quellen und Frequenzbänder; er ist keine direkte Subtraktion von einer einzelnen Windturbine."),
        ("Nota: el nivel resultante incluye la suma energética multi-Quelle y multi-banda; no es una resta directa de una única Windturbine.", "Hinweis: Der Ergebnispegel enthält die energetische Summierung über mehrere Quellen und Frequenzbänder; er ist keine direkte Subtraktion von einer einzelnen Windturbine."),
        ("Banda dominante:", "Dominantes Band:"),
        ("Origen Spectrum:", "Spektrumquelle:"),
        ("Origen Spektrum:", "Spektrumquelle:"),
        ("Die unten angezeigten Werte sind die Beträge Dämpfung die vom Modell verwendet werden. In die Gleichung principale, diese Term werden abgezogen au niveau der Quelle.", "Die unten angezeigten Werte sind die vom Modell verwendeten Dämpfungsbeträge. In der Hauptgleichung werden diese Terme vom Quellpegel abgezogen."),
        ("Beträge Dämpfung", "Dämpfungsbeträge"),
        ("In die Gleichung principale", "In der Hauptgleichung"),
        ("diese Term werden abgezogen au niveau der Quelle", "diese Terme werden vom Quellpegel abgezogen"),
        ("Dämpfung die vom Modell verwendet werden", "Dämpfungsbeträge, die vom Modell verwendet werden"),

        # Topographic-screening section leftovers
        ("Écran topographique mit MDT", "Topografische Abschirmung mit DGM"),
        ("Écran topographique avec MDT", "Topografische Abschirmung mit DGM"),
        ("Das DGM ändert nicht die Emission der Windturbine noch die atmosphärische Absorption. Seine Funktion ist es, zu beschreiben la géométrie real des Pfads und d'alimenter der Term Abar,b.", "Das DGM ändert weder die Schallemission der Windturbine noch die atmosphärische Absorption. Seine Funktion besteht darin, die reale Geometrie des Quelle-Rezeptor-Pfads zu beschreiben und den topografischen Abschirmungsterm Abar,b zu speisen."),
        ("Das DGM ändert nicht die Emission der Windturbine", "Das DGM ändert die Schallemission der Windturbine nicht"),
        ("noch die atmosphärische Absorption", "und nicht die atmosphärische Absorption"),
        ("Seine Funktion ist es, zu beschreiben la géométrie real des Pfads und d'alimenter der Term Abar,b", "Seine Funktion besteht darin, die reale Geometrie des Quelle-Rezeptor-Pfads zu beschreiben und den Term Abar,b zu speisen"),
        ("la géométrie real des Pfads", "die reale Geometrie des Pfads"),
        ("d'alimenter der Term", "den Term zu speisen"),
        ("Perfil del terreno", "Geländeprofil"),
        ("le profil Quelle–Rezeptor ist extrait du MDT mit einem adaptive Abtastung", "das Quelle-Rezeptor-Profil wird aus dem DGM mit adaptiver Abtastung extrahiert"),
        ("le profil Quelle-Rezeptor ist extrait du MDT mit einem adaptive Abtastung", "das Quelle-Rezeptor-Profil wird aus dem DGM mit adaptiver Abtastung extrahiert"),
        ("Línea de visión", "Sichtlinie"),
        ("la droite zwischen la Höhe efectiva der Quelle und la Rezeptorhöhe ist construite", "die Gerade zwischen der effektiven Quellhöhe und der Rezeptorhöhe wird konstruiert"),
        ("Si le Gelände bleibt toujours en dessous, alors", "Wenn das Gelände stets darunter bleibt, gilt"),
        ("Obstáculo dominante", "Dominantes Hindernis"),
        ("si une colline ou une crête dépasse", "wenn ein Hügel oder Grat die Sichtlinie überschreitet"),
        ("la Höhe au-dessus de la ligne de visée wird berechnet", "wird die Höhe über der Sichtlinie berechnet"),
        ("le relief coupe la vision directe", "das Gelände schneidet die direkte Sichtlinie"),
        ("eine Dämpfung supplémentaire par diffraction kann apparaître", "eine zusätzliche Dämpfung durch Beugung kann auftreten"),
        ("Géométrie real de l'obstacle", "Reale Geometrie des Hindernisses"),
        ("le plugin utilise la position réelle de l'obstacle dominant", "das Plugin verwendet die reale Position des dominanten Hindernisses"),
        ("et calcule", "und berechnet"),
        ("Activation conservative", "Konservative Aktivierung"),
        ("Abar ist nicht aktiviert für de petites irrégularités du MDT", "Abar wird nicht für kleine DGM-Unregelmäßigkeiten aktiviert"),
        ("un seuil minimal lié à la résolution du raster ist exigé", "es wird ein Mindestschwellwert in Bezug auf die Rasterauflösung verlangt"),
        ("Diffraction de tipo Fresnel", "Fresnel-artige Beugung"),
        ("Diffraction de type Fresnel", "Fresnel-artige Beugung"),
        ("mit cette géométrie", "mit dieser Geometrie"),
        ("une différence de chemins und un nombre de Fresnel sind estimés", "werden eine Weglängendifferenz und eine Fresnel-Zahl geschätzt"),
        ("Ce nombre est ensuite transformé en une Dämpfung", "Diese Zahl wird anschließend in eine Dämpfung umgewandelt"),
        ("dépendante de la fréquence au moyen de l’approximation actuelle du plugin", "die mit der aktuellen Plugin-Näherung frequenzabhängig ist"),
        ("En l’implémentation actuelle", "In der aktuellen Implementierung"),
        ("Abar ist également limité à des Werte raisonnables", "Abar wird außerdem auf plausible Werte begrenzt"),
        ("plafonnement supérieur", "obere Begrenzung"),
        ("afin d’éviter des suratténuations parasites", "um unerwünschte Überdämpfungen zu vermeiden"),
        ("En l’absence de MDT ou d’obstacle pertinent", "Ohne DGM oder relevantes Hindernis"),
        ("alors Abar,b = 0", "gilt Abar,b = 0"),
    ])


    repl.extend([
        ("Abar dominanter Pfad", "Abar des dominanten Pfads"),
        ("Maximaler Abar beitragend", "Maximaler Abar-Wert der Beitragenden"),
        ("Abar maximal contrib.", "Maximaler Abar-Wert der Beitragenden"),
        ("Maximaler Abar zwischen todas las turbinas que contribuyen al receptor", "Maximaler Abar-Wert unter allen Windturbinen, die zum Rezeptor beitragen"),
        ("Máximo Abar zwischen todas las turbinas que contribuyen al receptor", "Maximaler Abar-Wert unter allen Windturbinen, die zum Rezeptor beitragen"),
        ("Máximo Abar entre toutes les éoliennes qui contribuent au récepteur", "Maximaler Abar-Wert unter allen Windturbinen, die zum Rezeptor beitragen"),
        ("Mittelwert, gewichtet nach dem akustischen Beitrag jeder Turbine", "Mittelwert, gewichtet nach dem akustischen Beitrag jeder Windturbine"),
        ("Anzahl beitragenWindturbinen", "Anzahl beitragender Windturbinen"),
        ("Anzahl beitragendWindturbinen", "Anzahl beitragender Windturbinen"),
        ("Anzahl beitragen Turbinen", "Anzahl beitragender Windturbinen"),
        ("Anzahl de Windturbinen", "Anzahl der Windturbinen"),
        ("Número de Turbinen contribuyentes con Abar &gt; 0 dB", "Anzahl beitragender Windturbinen mit Abar &gt; 0 dB"),
        ("Número de Turbinen contribuyentes con Abar > 0 dB", "Anzahl beitragender Windturbinen mit Abar > 0 dB"),
        ("ERGENISPEGEL", "ERGEBNISPEGEL"),
        ("NIVEAU RÉSULTANT", "ERGEBNISPEGEL"),
        ("NIVEL RESULTANTE", "ERGEBNISPEGEL"),
        ("Nota : le niveau résultant inclut", "Hinweis: Der Ergebnispegel enthält"),
        ("Nota: el nivel resultante incluye", "Hinweis: Der Ergebnispegel enthält"),
        ("la suma energética multi-Quelle y multi-banda", "die energetische Summierung über mehrere Quellen und Frequenzbänder"),
        ("la sommation énergétique multi-Quelle y multi-banda", "die energetische Summierung über mehrere Quellen und Frequenzbänder"),
        ("multi-Quelle y multi-banda", "über mehrere Quellen und Frequenzbänder"),
        ("no es una resta directa de una única Windturbine", "dies ist keine direkte Subtraktion von einer einzelnen Windturbine"),
        ("no es una resta directa de una única Turbine", "dies ist keine direkte Subtraktion von einer einzelnen Windturbine"),
        ("ce n’est pas une resta directa de una única Windturbine", "dies ist keine direkte Subtraktion von einer einzelnen Windturbine"),
        ("Dominantes Band:", "Dominantes Frequenzband:"),
        ("Banda dominante", "Dominantes Frequenzband"),
        ("Origen Spectrum", "Spektrumquelle"),
        ("Origen Spektrum", "Spektrumquelle"),
        ("Spektrumquelle:", "Spektrumquelle:"),
        ("Estadísticos de Dämpfungen", "Dämpfungsstatistik"),
        ("Statistiques de Dämpfungen", "Dämpfungsstatistik"),
        ("(Abgedeckte Rezeptoren)", "(abgedeckte Rezeptoren)"),
        ("Dämpfung due zur Bodeneffekt", "Dämpfung durch Bodeneffekt"),
        ("Dämpfung due zur Bodenefekt", "Dämpfung durch Bodeneffekt"),
        ("Dämpfung due à l’effet de sol", "Dämpfung durch Bodeneffekt"),
        ("DEM auf dem dominanten Pfad", "DGM auf dem dominanten Pfad"),
        ("Abar trayectoria dominante", "Abar des dominanten Pfads"),
        ("Trayectorias apantalladas", "Abgeschirmte Pfade"),
        ("Trajectoires écrantées", "Abgeschirmte Pfade"),
        ("con Abar", "mit Abar"),
        ("contribuyen al receptor", "zum Rezeptor beitragen"),
        ("entre todas las turbinas", "unter allen Windturbinen"),
        ("zwischen todas las turbinas", "unter allen Windturbinen"),
        ("todas las turbinas", "allen Windturbinen"),
        ("éoliennes qui contribuent au récepteur", "Windturbinen, die zum Rezeptor beitragen"),
        # Topographic-screening report section
        ("Écran topographique mit MDT", "Topografische Abschirmung mit DGM"),
        ("Écran topographique avec MDT", "Topografische Abschirmung mit DGM"),
        ("Das DGM ändert nicht die Emission der Windturbine noch die atmosphärische Absorption", "Das DGM ändert weder die Schallemission der Windturbine noch die atmosphärische Absorption"),
        ("Das DGM ändert die Schallemission der Windturbine nicht und nicht die atmosphärische Absorption", "Das DGM ändert weder die Schallemission der Windturbine noch die atmosphärische Absorption"),
        ("Seine Funktion ist es, zu beschreiben la géométrie real des Pfads und d'alimenter der Term Abar,b", "Seine Funktion besteht darin, die reale Geometrie des Quelle-Rezeptor-Pfads zu beschreiben und den Term Abar,b zu speisen"),
        ("zu beschreiben la géométrie real des Pfads", "die reale Geometrie des Quelle-Rezeptor-Pfads zu beschreiben"),
        ("d'alimenter der Term Abar,b", "den Term Abar,b zu speisen"),
        ("Perfil del terreno", "Geländeprofil"),
        ("le profil Quelle–Rezeptor ist extrait du MDT mit einem adaptive Abtastung", "das Quelle-Rezeptor-Profil wird aus dem DGM mit adaptiver Abtastung extrahiert"),
        ("le profil Quelle-Rezeptor ist extrait du MDT mit einem adaptive Abtastung", "das Quelle-Rezeptor-Profil wird aus dem DGM mit adaptiver Abtastung extrahiert"),
        ("Línea de visión", "Sichtlinie"),
        ("la droite zwischen la Höhe efectiva der Quelle und la Rezeptorhöhe ist construite", "die Gerade zwischen der effektiven Quellhöhe und der Rezeptorhöhe wird konstruiert"),
        ("Si le Gelände bleibt toujours en dessous, alors", "Wenn das Gelände stets darunter bleibt, gilt"),
        ("Obstáculo dominante", "Dominantes Hindernis"),
        ("si une colline ou une crête dépasse", "wenn ein Hügel oder Grat die Sichtlinie überschreitet"),
        ("la Höhe au-dessus de la ligne de visée wird berechnet", "wird die Höhe über der Sichtlinie berechnet"),
        ("le relief coupe la vision directe", "das Gelände schneidet die direkte Sichtlinie"),
        ("eine Dämpfung supplémentaire par diffraction kann apparaître", "eine zusätzliche Dämpfung durch Beugung kann auftreten"),
        ("Géométrie real de l'obstacle", "Reale Geometrie des Hindernisses"),
        ("le plugin utilise la position réelle de l'obstacle dominant", "das Plugin verwendet die reale Position des dominanten Hindernisses"),
        ("et calcule", "und berechnet"),
        ("Activation conservative", "Konservative Aktivierung"),
        ("Abar ist nicht activé für de petites irrégularités du MDT", "Abar wird nicht für kleine DGM-Unregelmäßigkeiten aktiviert"),
        ("Abar ist nicht aktiviert für de petites irrégularités du MDT", "Abar wird nicht für kleine DGM-Unregelmäßigkeiten aktiviert"),
        ("un seuil minimal lié à la résolution du raster ist exigé", "es wird ein Mindestschwellwert in Bezug auf die Rasterauflösung verlangt"),
        ("Diffraction de tipo Fresnel", "Fresnel-artige Beugung"),
        ("Diffraction de type Fresnel", "Fresnel-artige Beugung"),
        ("mit cette géométrie", "mit dieser Geometrie"),
        ("une différence de chemins und un nombre de Fresnel sind estimés", "werden eine Weglängendifferenz und eine Fresnel-Zahl geschätzt"),
        ("Ce nombre est ensuite transformé en une Dämpfung", "Diese Zahl wird anschließend in eine Dämpfung umgewandelt"),
        ("dépendante de la fréquence au moyen de l’approximation actuelle du plugin", "die mit der aktuellen Plugin-Näherung frequenzabhängig ist"),
        ("En l’implémentation actuelle", "In der aktuellen Implementierung"),
        ("Abar ist également limité à des Werte raisonnables", "Abar wird außerdem auf plausible Werte begrenzt"),
        ("plafonnement supérieur", "obere Begrenzung"),
        ("afin d’éviter des suratténuations parasites", "um unerwünschte Überdämpfungen zu vermeiden"),
        ("En l’absence de MDT ou d’obstacle pertinent", "Ohne DGM oder relevantes Hindernis"),
        ("alors Abar,b = 0", "gilt Abar,b = 0"),
    ])

    repl.extend([
        ("Dämpfung due zur Bodeneffekt", "Dämpfung durch Bodeneffekt"),
        ("Dämpfung due zur Bodenwirkung", "Dämpfung durch Bodeneffekt"),
        ("Atenuación por MDT en la trayectoria dominante", "Dämpfung durch DGM auf dem dominanten Pfad"),
        ("Atenuación por apantallamiento topográfico", "Dämpfung durch topografische Abschirmung"),
        ("Trayectorias apantalladas", "Abgeschirmte Pfade"),
        ("Trajets écrantés", "Abgeschirmte Pfade"),
        ("NIVEL RESULTANTE", "ERGEBNISPEGEL"),
        ("NIVEAU RÉSULTANT", "ERGEBNISPEGEL"),
        ("Banda dominante", "Dominantes Frequenzband"),
        ("Origen Spectrum", "Spektrumquelle"),
        ("Origen Spektrum", "Spektrumquelle"),
        ("Nota: el nivel resultante incluye la suma energética multi-Quelle y multi-banda; no es una resta directa de una única Windturbine.", "Hinweis: Der Ergebnispegel enthält die energetische Summierung über mehrere Quellen und Frequenzbänder; dies ist keine direkte Subtraktion von einer einzelnen Windturbine."),
        ("Nota: el nivel resultante incluye la suma energética multi-fuente y multi-banda; no es una resta directa de una única turbina.", "Hinweis: Der Ergebnispegel enthält die energetische Summierung über mehrere Quellen und Frequenzbänder; dies ist keine direkte Subtraktion von einer einzelnen Windturbine."),
        ("Máximo Abar zwischen todas las turbinas que contribuyen al receptor", "Maximaler Abar-Wert unter allen Windturbinen, die zum Rezeptor beitragen"),
        ("Maximo Abar zwischen todas las turbinas que contribuyen al receptor", "Maximaler Abar-Wert unter allen Windturbinen, die zum Rezeptor beitragen"),
        ("Número de Turbinen contribuyentes con Abar", "Anzahl beitragender Windturbinen mit Abar"),
        ("Número de Windturbinen contribuyentes con Abar", "Anzahl beitragender Windturbinen mit Abar"),
        ("Anzahl beitragenWindturbinen", "Anzahl beitragender Windturbinen"),
        ("Anzahl beitragen Windturbinen", "Anzahl beitragender Windturbinen"),
        ("Estadísticos de Dämpfungen", "Dämpfungsstatistik"),
        ("Estadísticos de Dämpfungen (Abgedeckte Rezeptoren)", "Dämpfungsstatistik (abgedeckte Rezeptoren)"),
        ("Écran topographique mit MDT", "Topografische Abschirmung mit DGM"),
        ("mit MDT", "mit DGM"),
        ("del receptor", "des Rezeptors"),
        ("al receptor", "zum Rezeptor"),
        ("de una única", "einer einzelnen"),
        ("turbina", "Windturbine"),
    ])

    repl.extend([
        # Final grammar smoothing after broad fragment replacements
        ("Anzahl beitragende Windturbinen", "Anzahl beitragender Windturbinen"),
        ("Anzahl beitragend Windturbinen", "Anzahl beitragender Windturbinen"),
        ("Anzahl beitragende Turbinen", "Anzahl beitragender Turbinen"),
        ("Anzahl beitragend Turbinen", "Anzahl beitragender Turbinen"),
        ("Dämpfung Dämpfung", "Dämpfung"),
        ("Dämpfungsbeträge Dämpfung", "Dämpfungsbeträge"),
        ("in der Hauptgleichung, diese Terme", "In der Hauptgleichung werden diese Terme"),
        ("In der Hauptgleichung, diese Terme", "In der Hauptgleichung werden diese Terme"),
        ("diese Terme werden abgezogen vom Quellpegel", "diese Terme werden vom Quellpegel abgezogen"),
        ("diese Terme werden vom Quellpegel abgezogen vom Quellpegel", "diese Terme werden vom Quellpegel abgezogen"),
    ])

    for a, b in repl:
        html = html.replace(a, b)
    # Final conservative regex pass for common mixed-language connectors.
    html = re.sub(r"Número de\s+(?:Turbinen|Windturbinen)\s+contribuyentes\s+con\s+Abar\s*(&gt;|>)\s*0\s*dB", r"Anzahl beitragender Windturbinen mit Abar \1 0 dB", html)
    html = re.sub(r"Máximo Abar.*?(?:contribuyen al receptor|contribuent au récepteur)", "Maximaler Abar-Wert unter allen Windturbinen, die zum Rezeptor beitragen", html)
    html = re.sub(r"Nota\s*:\s*el nivel resultante.*?(?:Turbine|Windturbine)\.", "Hinweis: Der Ergebnispegel enthält die energetische Summierung über mehrere Quellen und Frequenzbänder; dies ist keine direkte Subtraktion von einer einzelnen Windturbine.", html)
    html = re.sub(r"Banda dominante\s*:", "Dominantes Frequenzband:", html)
    html = re.sub(r"Origen\s+(?:Spectrum|Spektrum)\s*:", "Spektrumquelle:", html)
    return html



def _cleanup_spanish_noise_html(html: str) -> str:
    """Post-process the noise HTML report when Spanish is active.

    The i18n pivot of VelantisWind is Spanish, but this dialog still contains an
    older French report template.  Until the full report is rewritten with
    Spanish source strings, this cleanup prevents Spanish sessions from showing
    a French/mixed-language technical summary.
    """
    repl = [
        # Window/report titles and main sections
        ("📊 RAPPORT TECHNIQUE D’IMPACT ACOUSTIQUE", "📊 INFORME TÉCNICO DE IMPACTO ACÚSTICO"),
        ("Évaluation du bruit généré par les éoliennes", "Evaluación del ruido generado por aerogeneradores"),
        ("1. RÉSUMÉ EXÉCUTIF", "1. RESUMEN EJECUTIVO"),
        ("2. COMMENT LE RÉSULTAT A ÉTÉ GÉNÉRÉ", "2. CÓMO SE HA GENERADO EL RESULTADO"),
        ("3. RÉCEPTEUR CRITIQUE", "3. RECEPTOR CRÍTICO"),
        ("4. CONFIGURATION ET PARAMÈTRES", "4. CONFIGURACIÓN Y PARÁMETROS"),
        ("5. PHYSIQUE DÉTAILLÉE ET TRAÇABILITÉ DU CALCUL", "5. FÍSICA DETALLADA Y TRAZABILIDAD DEL CÁLCULO"),
        ("6. GROUPES SOURCE ACOUSTIQUES", "6. GRUPOS FUENTE ACÚSTICOS"),
        ("7. DISTRIBUTION PAR TYPE DE RÉCEPTEUR", "7. DISTRIBUCIÓN POR TIPO DE RECEPTOR"),
        ("Portée de ce rapport — à lire avant d’utiliser les résultats", "Alcance de este informe — leer antes de usar los resultados"),
        ("Ce que c’est :", "Qué es:"),
        ("Ce que ce n’est pas :", "Qué no es:"),
        ("Simplifications appliquées dans ce mode :", "Simplificaciones aplicadas en este modo:"),
        ("Recommandation :", "Recomendación:"),
        ("Glossaire des symboles", "Glosario de símbolos"),
        ("Définition compacte des symboles qui apparaissent dans les formules et tableaux de ce rapport.", "Definición compacta de los símbolos que aparecen en las fórmulas y tablas de este informe."),
        ("Symbole", "Símbolo"),
        ("Signification", "Significado"),
        # Executive summary cards
        ("Éoliennes", "Aerogeneradores"),
        ("Récepteurs évalués", "Receptores evaluados"),
        ("Niveau maximal", "Nivel máximo"),
        ("Couverture de l’analyse", "Cobertura del análisis"),
        ("récepteurs</strong> dans le rayon", "receptores</strong> dentro del radio"),
        ("de couverture", "de cobertura"),
        ("récepteurs hors rayon", "receptores fuera del radio"),
        ("Conformité réglementaire", "Cumplimiento regulatorio"),
        ("récepteurs</strong> dépassent les limites", "receptores</strong> superan los límites"),
        ("de conformité sur les récepteurs couverts", "de cumplimiento sobre los receptores cubiertos"),
        ("Limite :", "Límite:"),
        ("Méthodologie de calcul", "Metodología de cálculo"),
        ("Moteur utilisé :", "Motor utilizado:"),
        ("Groupes source acoustiques :", "Grupos fuente acústicos:"),
        ("modèle(s) d’éolienne", "modelo(s) de aerogenerador"),
        ("Méthode :", "Método:"),
        ("Carte raster :", "Mapa raster:"),
        ("Propagation par bandes d’octave selon la méthodologie ISO-aligned", "Propagación por bandas de octava según la metodología ISO-aligned"),
        ("Calcul acoustique simplifié pour le criblage", "Cálculo acústico simplificado para screening"),
        # Critical receiver
        ("Récepteur critique (niveau sonore le plus élevé)", "Receptor crítico (mayor nivel sonoro)"),
        ("Récepteur critique non disponible.", "Receptor crítico no disponible."),
        ("ID récepteur :", "ID receptor:"),
        ("Niveau total :", "Nivel total:"),
        ("Limite applicable :", "Límite aplicable:"),
        ("Marge :", "Margen:"),
        ("Modèle dominant :", "Modelo dominante:"),
        ("Groupe source :", "Grupo fuente:"),
        ("Éoliennes contributrices dans le rayon :", "Aerogeneradores contribuyentes dentro del radio:"),
        ("Distance :", "Distancia:"),
        ("Décomposition des atténuations", "Desglose de atenuaciones"),
        ("Les valeurs affichées ci-dessous sont les amplitudes d’atténuation utilisées par le modèle. Dans l’équation principale, ces termes sont soustraits au niveau de source.", "Los valores mostrados abajo son las magnitudes de atenuación usadas por el modelo. En la ecuación principal, estos términos se restan al nivel de fuente."),
        ("Terme", "Término"),
        ("Valeur [dB]", "Valor [dB]"),
        ("Puissance acoustique de l’éolienne", "Potencia acústica del aerogenerador"),
        ("Dispersion géométrique", "Dispersión geométrica"),
        ("Absorption dans l’air", "Absorción en el aire"),
        ("Effet du sol", "Efecto del suelo"),
        ("Atténuation due à l’effet de sol", "Atenuación debida al efecto del suelo"),
        ("Atténuation due au MDT sur le trajet dominant", "Atenuación debida al MDT en la trayectoria dominante"),
        ("Abar trajet dominant", "Abar trayectoria dominante"),
        ("Abar maximal des contributeurs", "Abar máximo de los contribuyentes"),
        ("Abar maximal parmi toutes les éoliennes qui contribuent au récepteur", "Abar máximo entre todos los aerogeneradores que contribuyen al receptor"),
        ("Abar pondéré par énergie", "Abar ponderado por energía"),
        ("Moyenne pondérée par la contribution acoustique de chaque éolienne", "Media ponderada por la contribución acústica de cada aerogenerador"),
        ("Trajets écrantés", "Trayectorias apantalladas"),
        ("Nombre d’éoliennes contributrices avec Abar &gt; 0 dB", "Número de aerogeneradores contribuyentes con Abar &gt; 0 dB"),
        ("Nombre d’éoliennes contributrices avec Abar > 0 dB", "Número de aerogeneradores contribuyentes con Abar > 0 dB"),
        ("NIVEAU RÉSULTANT", "NIVEL RESULTANTE"),
        ("Note : le niveau résultant inclut la sommation énergétique multi-source et multi-bande ; ce n’est pas une soustraction directe depuis une seule éolienne.", "Nota: el nivel resultante incluye la suma energética multi-fuente y multi-banda; no es una resta directa desde un único aerogenerador."),
        ("Bande dominante :", "Banda dominante:"),
        ("Origine du spectre :", "Origen del espectro:"),
        ("CONFORME", "CUMPLE"),
        ("DÉPASSE", "EXCEDE"),
        # Abar/DEM notes
        ("Lecture MDT :", "Lectura MDT:"),
        ("Abar=0 au récepteur critique ne signifie pas que le MDT est désactivé ; cela signifie que", "Abar=0 en el receptor crítico no significa que el MDT esté desactivado; significa que"),
        ("obstacle dominant estimé", "obstáculo dominante estimado"),
        ("Seuil d’activation:", "Umbral de activación:"),
        ("D’autres récepteurs présentent bien un écran", "Otros receptores sí presentan apantallamiento"),
        ("un relief a été détecté, mais sous le seuil conservateur d’activation", "se ha detectado relieve, pero por debajo del umbral conservador de activación"),
        ("aucun profil MDT valide n’a pu être extrait pour le trajet dominant", "no se ha podido extraer un perfil MDT válido para la trayectoria dominante"),
        ("aucun MDT n’était disponible sur ce trajet", "no había MDT disponible en esa trayectoria"),
        ("aucun obstacle topographique pertinent n’a été détecté sur le trajet dominant", "no se ha detectado ningún obstáculo topográfico relevante en la trayectoria dominante"),
        ("état=", "estado="),
        ("actif", "activo"),
        ("non", "no"),
        ("oui", "sí"),
        ("sans nom", "sin nombre"),
        # Attenuation stats and config
        ("Statistiques des atténuations (récepteurs couverts)", "Estadísticos de atenuaciones (receptores cubiertos)"),
        ("Les amplitudes brutes d’atténuation sont affichées (et non le signe algébrique dans l’équation). Pour Abar, le maximum parmi les éoliennes contributrices de chaque récepteur est utilisé, pas uniquement le trajet dominant.", "Se muestran las magnitudes brutas de atenuación, no el signo algebraico dentro de la ecuación. Para Abar se usa el máximo entre los aerogeneradores contribuyentes de cada receptor, no solo la trayectoria dominante."),
        ("Moyenne [dB]", "Media [dB]"),
        ("Maximum [dB]", "Máximo [dB]"),
        ("divergence géométrique", "divergencia geométrica"),
        ("absorption atmosphérique", "absorción atmosférica"),
        ("effet de sol", "efecto del suelo"),
        ("maximum parmi les contributeurs", "máximo entre los contribuyentes"),
        ("Équation utilisée", "Ecuación utilizada"),
        ("Paramètres du calcul", "Parámetros del cálculo"),
        ("Termes actifs", "Términos activos"),
        ("Trajets avec G différent du global", "Trayectorias con G distinto del global"),
        ("Moteur :", "Motor:"),
        ("Hauteur du récepteur :", "Altura del receptor:"),
        ("Rayon maximal :", "Radio máximo:"),
        ("Mode sol :", "Modo suelo:"),
        ("G global de secours:", "G global de respaldo:"),
        ("G_eff moyen utilisé:", "G_eff medio utilizado:"),
        ("G_eff du récepteur critique utilisé:", "G_eff del receptor crítico utilizado:"),
        ("G utilisé:", "G utilizado:"),
        ("G_eff moyen:", "G_eff medio:"),
        ("G_eff du récepteur critique:", "G_eff del receptor crítico:"),
        ("Occupation du sol:", "Ocupación del suelo:"),
        ("Scénario acoustique :", "Escenario acústico:"),
        ("Température :", "Temperatura:"),
        ("Humidité relative :", "Humedad relativa:"),
        ("Pression :", "Presión:"),
        ("α atmosphérique :", "α atmosférica:"),
        ("depuis couche", "desde capa"),
        ("non généré", "no generado"),
        ("résolution demandée", "resolución solicitada"),
        ("effective", "efectiva"),
        ("auto-ajustée", "autoajustada"),
        ("diurne", "diurno"),
        ("nocturne", "nocturno"),
        ("personnalisé", "personalizado"),
        ("Limites appliquées :", "Límites aplicados:"),
        ("depuis les champs des récepteurs", "desde los campos de los receptores"),
        ("valeur unique", "valor único"),
        ("plage", "rango"),
        ("Limite de référence :", "Límite de referencia:"),
        # Scope banner
        ("une évaluation acoustique préliminaire alignée sur la méthodologie ISO 9613-2, destinée à la conception, à la comparaison d’alternatives et au criblage des récepteurs sensibles.", "una evaluación acústica preliminar alineada con la metodología ISO 9613-2, orientada al diseño, la comparación de alternativas y el cribado de receptores sensibles."),
        ("ce n’est pas un rapport acoustique certifié et ne remplace pas une étude réglementaire définitive réalisée avec un logiciel commercial validé.", "no es un informe acústico certificado y no sustituye a un estudio regulatorio definitivo realizado con software comercial validado."),
        ("une estimation rapide de criblage pour des cartes agiles et la comparaison d’alternatives d’implantation.", "una estimación rápida de screening para mapas ágiles y comparación de alternativas de implantación."),
        ("ce n’est ni un calcul spectral détaillé ni un rapport réglementaire ; pour les récepteurs proches de la limite, il convient de recalculer en mode ISO-aligned.", "no es un cálculo espectral detallado ni un informe regulatorio; para receptores cercanos al límite conviene recalcular en modo ISO-aligned."),
        ("Absorption atmosphérique Aatm via une table de référence avec corrections simplifiées de température, humidité et pression, et non la formulation analytique complète de l’ISO 9613-1.", "Absorción atmosférica Aatm mediante tabla de referencia con correcciones simplificadas de temperatura, humedad y presión, no mediante la formulación analítica completa de la ISO 9613-1."),
        ("Sans correction météorologique de long terme Cmet.", "Sin corrección meteorológica de largo plazo Cmet."),
        ("Diffraction topographique d’un obstacle dominant unique : sans diffraction latérale ni écrans multiples.", "Difracción topográfica de un único obstáculo dominante: sin difracción lateral ni pantallas múltiples."),
        ("Résolution spectrale en 8 bandes d’octave de 63 à 8000 Hz, pas en tiers d’octave.", "Resolución espectral en 8 bandas de octava de 63 a 8000 Hz, no en tercios de octava."),
        ("Directivité de source Dc supposée égale à 0 dB.", "Directividad de fuente Dc asumida igual a 0 dB."),
        ("Sans propagation par bandes d’octave.", "Sin propagación por bandas de octava."),
        ("Absorption atmosphérique résumée par un seul coefficient alpha constant.", "Absorción atmosférica resumida mediante un único coeficiente alfa constante."),
        ("Sans écran topographique Abar depuis le MDT.", "Sin apantallamiento topográfico Abar desde el MDT."),
        ("Effet de sol via une correction empirique simplifiée.", "Efecto de suelo mediante una corrección empírica simplificada."),
        ("Pour les décisions réglementaires critiques, validez les résultats avec des mesures de terrain ou un logiciel commercial certifié.", "Para decisiones regulatorias críticas, valida los resultados con mediciones de campo o con software comercial certificado."),
        # Glossary
        ("Niveau de puissance acoustique pondéré A de la source, en dB(A).", "Nivel de potencia acústica ponderado A de la fuente, en dB(A)."),
        ("Puissance acoustique de la source par bande d’octave, en dB.", "Potencia acústica de la fuente por banda de octava, en dB."),
        ("Forme spectrale de référence par bande utilisée comme gabarit, en dB.", "Forma espectral de referencia por banda utilizada como plantilla, en dB."),
        ("Pondération A appliquée à chaque bande d’octave, en dB.", "Ponderación A aplicada a cada banda de octava, en dB."),
        ("Décalage global appliqué au gabarit spectral pour reproduire le LwA cible, en dB.", "Desplazamiento global aplicado a la plantilla espectral para reproducir el LwA objetivo, en dB."),
        ("Niveau de pression acoustique pondéré A résultant au récepteur, en dB(A).", "Nivel de presión acústica ponderado A resultante en el receptor, en dB(A)."),
        ("Atténuation par divergence géométrique avec la distance, en dB.", "Atenuación por divergencia geométrica con la distancia, en dB."),
        ("Atténuation due à l’absorption atmosphérique de l’air, en dB.", "Atenuación debida a la absorción atmosférica del aire, en dB."),
        ("Atténuation due à l’effet de sol, en dB.", "Atenuación debida al efecto del suelo, en dB."),
        ("Atténuation due à l’écran topographique, uniquement en mode ISO avec MDT, en dB.", "Atenuación debida al apantallamiento topográfico, solo en modo ISO con MDT, en dB."),
        ("Distance tridimensionnelle entre source et récepteur, en mètres.", "Distancia tridimensional entre fuente y receptor, en metros."),
        ("Facteur de sol de 0 (dur) à 1 (meuble) et sa valeur effective par trajet.", "Factor de suelo de 0 (duro) a 1 (blando) y su valor efectivo por trayectoria."),
        ("Correction météorologique de long terme, non appliquée dans ce plugin.", "Corrección meteorológica de largo plazo, no aplicada en este plugin."),
        ("Correction de directivité de la source, supposée égale à 0 dB.", "Corrección de directividad de la fuente, asumida igual a 0 dB."),
        # Physics detail - generic
        ("Développement physique détaillé du moteur ISO-aligned", "Desarrollo físico detallado del motor ISO-aligned"),
        ("Développement physique détaillé du moteur rapide", "Desarrollo físico detallado del motor rápido"),
        ("Entrées réellement utilisées dans ce calcul", "Entradas realmente utilizadas en este cálculo"),
        ("Origine de chaque terme de l’équation", "Origen de cada término de la ecuación"),
        ("Comment il est obtenu dans ce plugin", "Cómo se obtiene en este plugin"),
        ("Divergence géométrique", "Divergencia geométrica"),
        ("Absorption atmosphérique simplifiée", "Absorción atmosférica simplificada"),
        ("Effet de sol simplifié", "Efecto de suelo simplificado"),
        ("Effet de sol simplifié avec occupation du sol", "Efecto de suelo simplificado con ocupación del suelo"),
        ("MDT / topographie", "MDT / topografía"),
        ("Ce que ce mode ne fait pas", "Qué no hace este modo"),
        ("Source acoustique :", "Fuente acústica:"),
        ("Niveau opérationnel global:", "Nivel operativo global:"),
        ("Géométrie :", "Geometría:"),
        ("Atmosphère:", "Atmósfera:"),
        ("Sol:", "Suelo:"),
        ("Topographie :", "Topografía:"),
        ("Entrée globale de la source.", "Entrada global de la fuente."),
        ("Calculé à partir de la distance 3D source–récepteur.", "Calculado a partir de la distancia 3D fuente–receptor."),
        ("Calculé avec un coefficient constant unique", "Calculado con un único coeficiente constante"),
        ("multiplié par la distance", "multiplicado por la distancia"),
        ("Correction empirique de l’effet de sol.", "Corrección empírica del efecto del suelo."),
        ("Représente la dispersion géométrique de l’onde sonore avec la distance 3D source–récepteur.", "Representa la dispersión geométrica de la onda sonora con la distancia 3D fuente–receptor."),
        ("Le moteur rapide travaille avec un seul niveau global", "El motor rápido trabaja con un único nivel global"),
        ("par groupe source", "por grupo fuente"),
        ("Il est conçu pour le criblage, les cartes rapides et les comparaisons rapides", "Está diseñado para screening, mapas rápidos y comparaciones rápidas"),
        ("en sacrifiant le détail spectral au profit de la vitesse", "sacrificando detalle espectral en favor de la velocidad"),
        ("Dans ce mode, il n’y a <b>pas de propagation par bandes</b> ni de terme explicite d’écran topographique.", "En este modo no hay <b>propagación por bandas</b> ni término explícito de apantallamiento topográfico."),
        ("Scénario opérationnel de ce calcul:", "Escenario operativo de este cálculo:"),
        ("provient d’un <b>LwA fixe</b> ou d’une <b>courbe acoustique LwA(ws)</b>", "procede de un <b>LwA fijo</b> o de una <b>curva acústica LwA(ws)</b>"),
        ("pour la vitesse ou le cas le plus défavorable sélectionnés", "para la velocidad o el peor caso seleccionados"),
        ("coordonnées de source et de récepteur", "coordenadas de fuente y receptor"),
        ("hauteur du récepteur", "altura del receptor"),
        ("hauteur effective de source", "altura efectiva de fuente"),
        ("distance 3D", "distancia 3D"),
        ("dans ce mode, T/HR/P ne sont pas utilisés", "en este modo, T/HR/P no se utilizan"),
        ("l’absorption est résumée par un coefficient unique", "la absorción se resume mediante un coeficiente único"),
        ("un <b>G global manuel</b> ou un <b>G_eff</b> dérivé depuis la couche d’occupation du sol", "un <b>G global manual</b> o un <b>G_eff</b> derivado desde la capa de ocupación del suelo"),
        ("le MDT n’entre pas comme écran explicite dans ce mode", "el MDT no entra como pantalla explícita en este modo"),
        ("Le terme", "El término"),
        ("est une correction empirique du terrain contrôlée par un seul paramètre manuel", "es una corrección empírica del terreno controlada por un único parámetro manual"),
        ("Dans ce calcul, aucun G_eff n’a été dérivé depuis une couche d’occupation du sol.", "En este cálculo no se ha derivado ningún G_eff desde una capa de ocupación del suelo."),
        ("Ici", "Aquí"),
        ("est la distance horizontale", "es la distancia horizontal"),
        ("la hauteur de source", "la altura de fuente"),
        ("la hauteur du récepteur", "la altura del receptor"),
        ("Dans le moteur rapide, le MDT n’introduit pas de terme explicite d’écran topographique.", "En el motor rápido, el MDT no introduce un término explícito de apantallamiento topográfico."),
        ("Même si une couche de relief existe dans le projet", "Aunque exista una capa de relieve en el proyecto"),
        ("ce mode ne calcule pas", "este modo no calcula"),
        ("n’extrait pas de ligne de visée", "no extrae línea de visión"),
        ("n’applique pas de diffraction", "no aplica difracción"),
        ("la physique se base donc uniquement sur", "por tanto, la física se basa únicamente en"),
        ("la correction empirique de terrain", "la corrección empírica de terreno"),
        ("Le moteur rapide ne travaille pas par bandes", "El motor rápido no trabaja por bandas"),
        ("ne calcule pas", "no calcula"),
        ("n’introduit pas", "no introduce"),
        ("Il est donc adapté au criblage et aux comparaisons rapides", "Por tanto, es adecuado para screening y comparaciones rápidas"),
        ("mais pas à l’analyse spectrale détaillée", "pero no para análisis espectral detallado"),
        # ISO detail / spectrum
        ("Équation générale par bande", "Ecuación general por banda"),
        ("Sommation finale pondérée A", "Suma final ponderada A"),
        ("Données d’entrée réellement utilisées dans cette exécution", "Datos de entrada realmente utilizados en esta ejecución"),
        ("Spectre utilisé par le groupe source", "Espectro utilizado por el grupo fuente"),
        ("Modèle :", "Modelo:"),
        ("Ce que représente chaque colonne :", "Qué representa cada columna:"),
        ("Bande [Hz]", "Banda [Hz]"),
        ("final", "final"),
        ("Δ appliqué :", "Δ aplicado:"),
        ("Ce décalage augmente ou réduit toute la forme spectrale afin que sa somme pondérée A reproduise le", "Este desplazamiento aumenta o reduce toda la forma espectral para que su suma ponderada A reproduzca el"),
        ("de la courbe acoustique ou du LwA fixe.", "de la curva acústica o del LwA fijo."),
        ("Interprétation :", "Interpretación:"),
        ("le spectre final", "el espectro final"),
        ("est celui qui entre réellement dans l’équation par bandes", "es el que entra realmente en la ecuación por bandas"),
        ("Si", "Si"),
        ("existe, il correspond à la forme de référence avant l’ajustement global", "existe, corresponde a la forma de referencia antes del ajuste global"),
        ("pour ce groupe, aucune forme interne visible n’a été utilisée", "para este grupo no se ha usado ninguna forma interna visible"),
        ("provient directement du spectre chargé/importé ou d’une bibliothèque externe", "procede directamente del espectro cargado/importado o de una biblioteca externa"),
        # Source groups and distributions
        ("LwA effectif par groupe", "LwA efectivo por grupo"),
        ("Non disponible", "No disponible"),
        ("Récepteurs par catégorie", "Receptores por categoría"),
        ("Conformité par catégorie", "Cumplimiento por categoría"),
        ("dépassent la limite", "superan el límite"),
        ("couverts", "cubiertos"),
        ("Limites et recommandations", "Límites y recomendaciones"),
        ("Moteur rapide :", "Motor rápido:"),
        ("Moteur ISO-aligned :", "Motor ISO-aligned:"),
        ("Adapté au criblage préliminaire et aux cartes agiles.", "Adecuado para screening preliminar y mapas ágiles."),
        ("Adapté aux études techniques préliminaires, aux comparaisons et à l’itération de conception.", "Adecuado para estudios técnicos preliminares, comparaciones e iteración de diseño."),
        ("Simplifications connues :", "Simplificaciones conocidas:"),
        ("Modèles multiples :", "Modelos múltiples:"),
        ("pris en charge au moyen de couches/groupes source indépendants", "soportados mediante capas/grupos fuente independientes"),
        ("Mélanger plusieurs modèles dans une seule couche via attributs n’est pas activé dans cette version expérimentale.", "Mezclar varios modelos en una sola capa mediante atributos no está activado en esta versión experimental."),
        ("Raster ISO + MDT :", "Raster ISO + MDT:"),
        ("utilise la même logique d’écran topographique que les récepteurs ponctuels", "usa la misma lógica de apantallamiento topográfico que los receptores puntuales"),
        ("mais peut être coûteux sur de grandes cartes", "pero puede ser costoso en mapas grandes"),
        ("Pour les études réglementaires critiques, valider avec des mesures ou un logiciel commercial certifié.", "Para estudios regulatorios críticos, validar con mediciones o con software comercial certificado."),
        # Acoustic scenario/source group fragments
        ("Courbes acoustiques LwA(ws) en cas le plus défavorable", "Curvas acústicas LwA(ws) en peor caso"),
        ("Courbes acoustiques LwA(ws) à", "Curvas acústicas LwA(ws) a"),
        ("Courbes acoustiques LwA(ws)", "Curvas acústicas LwA(ws)"),
        ("LwA fixe par groupe de source acoustique", "LwA fijo por grupo de fuente acústica"),
        ("Groupe", "Grupo"),
        ("modèle", "modelo"),
        ("parc", "parque"),
        ("spectre", "espectro"),
        ("sans valeur", "sin valor"),
    ]
    for a, b in repl:
        html = html.replace(a, b)
    # Regex cleanup for frequent remaining French connector fragments in dynamic text.
    html = re.sub(r"\brécepteur(s?)\b", lambda m: "receptor" + ("es" if m.group(1) else ""), html, flags=re.IGNORECASE)
    html = re.sub(r"\brécepteurs\b", "receptores", html, flags=re.IGNORECASE)
    html = re.sub(r"\béolienne(s?)\b", lambda m: "aerogenerador" + ("es" if m.group(1) else ""), html, flags=re.IGNORECASE)
    html = re.sub(r"\béoliennes\b", "aerogeneradores", html, flags=re.IGNORECASE)
    html = re.sub(r"\batténuation(s?)\b", lambda m: "atenuación" + ("es" if m.group(1) else ""), html, flags=re.IGNORECASE)
    html = re.sub(r"\bdépasse(nt)?\b", "supera", html, flags=re.IGNORECASE)
    html = re.sub(r"\blimites\b", "límites", html, flags=re.IGNORECASE)
    html = re.sub(r"\blimite\b", "límite", html, flags=re.IGNORECASE)

    # Final ES report hardening.  Older report blocks were authored in French;
    # the normal i18n layer is fragment-based and can leave long technical
    # paragraphs mixed.  Keep this cleanup ES-only and report-local.
    extra_repl = [
        ("Absorption atmosphérique Aatm via une table de référence avec corrections simplifiées de température, humidité et pression, et no la formulation analytique complète de l’ISO 9613-1.", "Absorción atmosférica Aatm mediante una tabla de referencia con correcciones simplificadas de temperatura, humedad y presión, no mediante la formulación analítica completa de la ISO 9613-1."),
        ("🧭 Comment le calcul ISO-aligned a été exécuté", "🧭 Cómo se ha ejecutado el cálculo ISO-aligned"),
        ("🧭 Comment le calcul Screening a été exécuté", "🧭 Cómo se ha ejecutado el cálculo Screening"),
        ("Cette section explique le flux réel suivi par le plugin afin que le résultat par receptor soit traçable. Le niveau final de chaque receptor ne provient pas d’une simple soustraction unique, mais du calcul de toutes les contributions source–receptor dans le rayon de calcul, puis de leur sommation énergétique.", "Esta sección explica el flujo real seguido por el plugin para que el resultado por receptor sea trazable. El nivel final de cada receptor no procede de una simple resta única, sino del cálculo de todas las contribuciones fuente–receptor dentro del radio de cálculo y de su suma energética."),
        ("Lecture des entrées SIG: les aerogeneradores/sources acoustiques, les receptores, la altura del receptor et le rayon maximal de calcul sont pris en compte", "Lectura de las entradas SIG: se tienen en cuenta los aerogeneradores/fuentes acústicas, los receptores, la altura del receptor y el radio máximo de cálculo"),
        ("la couche d’occupation du sol si elle existe et le MDT/DSM s’il est activo", "la capa de ocupación del suelo si existe y el MDT/DSM si está activo"),
        ("État acoustique de chaque groupe source : pour chaque modelo ou groupe d’aerogeneradores, un LwA opérationnel est obtenu à partir d’une valeur fixe ou d’une courbe LwA(ws). Dans ce calcul:", "Estado acústico de cada grupo fuente: para cada modelo o grupo de aerogeneradores se obtiene un LwA operativo a partir de un valor fijo o de una curva LwA(ws). En este cálculo:"),
        ("Conversion en bandes: le moteur ISO-aligned a besoin d’un espectro Lw,b en 8 bandes d’octave. S’il n’existe pas de espectro spécifique, le plugin en reconstruit un à partir d’un gabarit/fallback et l’ajuste pour reproduire le LwA opérationnel.", "Conversión a bandas: el motor ISO-aligned necesita un espectro Lw,b en 8 bandas de octava. Si no existe un espectro específico, el plugin lo reconstruye a partir de una plantilla/fallback y lo ajusta para reproducir el LwA operativo."),
        ("Sélection des contributeurs par receptor: pour chaque receptor, les aerogeneradores situées dans le rayon maximal sont recherchées. Les receptores sans sources dans ce rayon sont marqués comme hors rayon et ne produisent pas de niveau acoustique utile.", "Selección de contribuyentes por receptor: para cada receptor se buscan los aerogeneradores situados dentro del radio máximo. Los receptores sin fuentes dentro de ese radio se marcan como fuera de radio y no producen un nivel acústico útil."),
        ("Calcul par trajet source–receptor : pour chaque aerogenerador contributrice, la distancia 3D, les cotes acoustiques, G ou G_eff du sol et, si un MDT/DSM est disponible, l’éventuel écran topographique du trajet sont calculés.", "Cálculo por trayectoria fuente–receptor: para cada aerogenerador contribuyente se calculan la distancia 3D, las cotas acústicas, G o G_eff del suelo y, si hay un MDT/DSM disponible, el posible apantallamiento topográfico de la trayectoria."),
        ("Propagation par bande: dans chaque bande, on applique Lp,b = Lw,b - Adiv - Aatm,b - Agr,b - Abar,b. Adiv dépend de la distance, Aatm,b de la fréquence/de l’atmosphère, Agr,b du sol et Abar,b du MDT s’il existe un obstacle pertinent.", "Propagación por banda: en cada banda se aplica Lp,b = Lw,b - Adiv - Aatm,b - Agr,b - Abar,b. Adiv depende de la distancia, Aatm,b de la frecuencia y de la atmósfera, Agr,b del suelo y Abar,b del MDT si existe un obstáculo relevante."),
        ("Sommation par source : les 8 bandes sont pondérées A puis sommées énergétiquement pour obtenir le niveau pondéré A de cette aerogenerador au receptor.", "Suma por fuente: las 8 bandas se ponderan en A y se suman energéticamente para obtener el nivel ponderado A de ese aerogenerador en el receptor."),
        ("Sommation du receptor: toutes les aerogeneradores contributrices sont sommées énergétiquement pour obtenir le niveau total dB(A) du receptor.", "Suma del receptor: todos los aerogeneradores contribuyentes se suman energéticamente para obtener el nivel total en dB(A) del receptor."),
        ("Comparaison avec les límites: le niveau total est comparé à la límite attribuée au receptor ou à la límite de référence. La marge, l’état de conformité et le tableau des dépassements en découlent.", "Comparación con los límites: el nivel total se compara con el límite asignado al receptor o con el límite de referencia. De ahí se obtienen el margen, el estado de cumplimiento y la tabla de excedencias."),
        ("Lecture pratique : le receptor critique est celui qui présente le niveau total le plus élevé ou la marge la plus défavorable par rapport à la límite. La colonne « source dominante » identifie l’aerogenerador/le groupe qui contribue le plus, mais le résultat final du receptor inclut toutes les sources dans le rayon.", "Lectura práctica: el receptor crítico es el que presenta el nivel total más alto o el margen más desfavorable respecto al límite. La columna «fuente dominante» identifica el aerogenerador/grupo que más contribuye, pero el resultado final del receptor incluye todas las fuentes dentro del radio."),
        ("🔎 Ce qui distingue ce mode du mode Screening", "🔎 Qué diferencia este modo del modo Screening"),
        ("🔎 Ce qui distingue ce mode du mode ISO-aligned", "🔎 Qué diferencia este modo del modo ISO-aligned"),
        ("Le mode ISO-aligned est plus lourd mais plus traçable : il utilise les bandes d’octave, la pondération A finale, l’absorción atmosférica dépendante de la fréquence, le sol par régions et l’écran topographique Abar lorsqu’un MDT/DSM est disponible. C’est le mode recommandé pour les rapports techniques préliminaires et la revue des receptores sensibles.", "El modo ISO-aligned es más pesado, pero también más trazable: utiliza bandas de octava, ponderación A final, absorción atmosférica dependiente de la frecuencia, suelo por regiones y apantallamiento topográfico Abar cuando hay MDT/DSM disponible. Es el modo recomendado para informes técnicos preliminares y revisión de receptores sensibles."),
        ("Aerogeneradores contributrices dans le rayon", "Aerogeneradores contribuyentes dentro del radio"),
        ("LwA source dominante", "LwA fuente dominante"),
        ("Les amplitudes brutes d’atenuación sont affichées (et no le signe algébrique dans l’équation). Pour Abar, le maximum parmi les aerogeneradores contributrices de chaque receptor est utilisé, pas uniquement le trajet dominant.", "Se muestran las magnitudes brutas de atenuación, no el signo algebraico de la ecuación. Para Abar se usa el máximo entre los aerogeneradores contribuyentes de cada receptor, no solo la trayectoria dominante."),
        ("Adiv représente la divergencia geométrica. Aatm est calculé par bande et dépend de T, HR et de la pression, avec une formulation simplifiée. Agr est appliqué comme terme de sol/terrain et Abar comme écran topographique de base lorsqu’un MDT est disponible.", "Adiv representa la divergencia geométrica. Aatm se calcula por banda y depende de T, HR y presión mediante una formulación simplificada. Agr se aplica como término de suelo/terreno y Abar como apantallamiento topográfico básico cuando hay un MDT disponible."),
        ("Révision recommandée : la pression atmosphérique saisie", "Revisión recomendada: la presión atmosférica introducida"),
        ("est hors de la rango typique utilisée comme référence dans de nombreuses études préliminaires. Si ce n’est pas une mesure du site, vérifier si elle devrait être proche de 101,325 kPa ou ajustée à l’altitude.", "está fuera del rango típico usado como referencia en muchos estudios preliminares. Si no es una medición del emplazamiento, conviene comprobar si debería estar cerca de 101,325 kPa o ajustarse a la altitud."),
        ("G effectif depuis l’occupation du sol", "G efectivo desde ocupación del suelo"),
        ("atenuación par divergencia geométrica avec la distance", "atenuación por divergencia geométrica con la distancia"),
        ("atenuación due à l’absorción atmosférica de l’air", "atenuación debida a la absorción atmosférica del aire"),
        ("Facteur de sol de 0 (dur) à 1 (meuble) et sa valeur efectiva par trajet", "Factor de suelo de 0 (duro) a 1 (blando/poroso) y su valor efectivo por trayectoria"),
        ("Correction météorologique de long terme, no appliquée dans ce plugin", "Corrección meteorológica de largo plazo, no aplicada en este plugin"),
        ("Ce moteur travaille en 8 bandes d’octave (63–8000 Hz). Les bandes ne sont pas un résultat du calcul, mais la grille fréquentielle de la méthode. Pour appliquer la propagation par bandes, le calcul a besoin d’une entrée acoustique par bande de la source Lw,b. Cette entrée peut provenir d’un espectro mesuré/importé ou d’un gabarit/fallback ajusté au niveau global opérationnel.", "Este motor trabaja en 8 bandas de octava (63–8000 Hz). Las bandas no son un resultado del cálculo, sino la malla frecuencial de la metodología. Para aplicar la propagación por bandas, el cálculo necesita una entrada acústica por banda de la fuente Lw,b. Esta entrada puede proceder de un espectro medido/importado o de una plantilla/fallback ajustada al nivel global operativo."),
        ("Fuente acústica: Lw,b par bandes d’octave. S’il existe un espectro spécifique du groupe source, c’est l’entrée utilisée. Sino, le plugin utilise une bibliothèque/un gabarit/un fallback et l’ajuste au niveau global opérationnel.", "Fuente acústica: Lw,b por bandas de octava. Si existe un espectro específico del grupo fuente, esa es la entrada utilizada. Si no, el plugin usa una biblioteca/plantilla/fallback y la ajusta al nivel global operativo."),
        ("Nivel operativo global: il procede de un LwA fijo o de una curva acústica LwA(ws) selon le scénario sélectionné. Ce niveau global ne remplace pas les bandes : il fixe l’état opérationnel et le espectro fournit la répartition fréquentielle.", "Nivel operativo global: procede de un LwA fijo o de una curva acústica LwA(ws), según el escenario seleccionado. Este nivel global no sustituye a las bandas: fija el estado operativo y el espectro aporta la distribución frecuencial."),
        ("Geometría: coordenadas de fuente y receptor, altura del receptor, hauteur efectiva de source et distancia 3D.", "Geometría: coordenadas de fuente y receptor, altura del receptor, altura efectiva de fuente y distancia 3D."),
        ("Topographie: MDT/DSM optionnel. Il n’affecte que le calcul de Abar,b.", "Topografía: MDT/DSM opcional. Solo afecta al cálculo de Abar,b."),
        ("Entrée acoustique par bandes. Elle provient du espectro du groupe source (CSV, bibliothèque, gabarit ou fallback ajusté au niveau global). La courbe acoustique LwA(ws) ou le LwA fixe définit le niveau global opérationnel de l’aerogenerador, et le espectro par bandes répartit ce niveau entre les 8 bandes.", "Entrada acústica por bandas. Procede del espectro del grupo fuente (CSV, biblioteca, plantilla o fallback ajustado al nivel global). La curva acústica LwA(ws) o el LwA fijo define el nivel global operativo del aerogenerador, y el espectro por bandas reparte ese nivel entre las 8 bandas."),
        ("Calculé par bande avec une table de base d’absorption α_ref(f) et des corrections simplifiées de température, humidité relative et pression. L’implémentation actuelle utilise la formulation exacte du plugin : α = α_ref(f)·corr_T·corr_HR·corr_P.", "Calculado por banda con una tabla base de absorción α_ref(f) y correcciones simplificadas de temperatura, humedad relativa y presión. La implementación actual usa la formulación exacta del plugin: α = α_ref(f)·corr_T·corr_HR·corr_P."),
        ("Calculé comme efecto del suelo par régions. Le paramètre de sol utilisé est un G unique par trajet : manuel/global ou G_eff dérivé de la couche d’occupation du sol.", "Calculado como efecto del suelo por regiones. El parámetro de suelo utilizado es un G único por trayectoria: manual/global o G_eff derivado desde la capa de ocupación del suelo."),
        ("N’intervient que s’il existe un MDT/DSM et si un écran topographique est détecté. En l’absence de MDT ou d’obstacle pertinent, Abar,b = 0.", "Solo interviene si existe un MDT/DSM y se detecta apantallamiento topográfico. En ausencia de MDT o de obstáculo relevante, Abar,b = 0."),
        ("2. Entrée acoustique de la source et bandes", "2. Entrada acústica de la fuente y bandas"),
        ("Dans ce moteur, le terme Lw,b est une donnée d’entrée par bande. Les bandes d’octave (63–8000 Hz) ne sont pas un résultat ISO ni un tableau calculé par le plugin : ce sont la grille fréquentielle sur laquelle la propagation est résolue.", "En este motor, el término Lw,b es un dato de entrada por banda. Las bandas de octava (63–8000 Hz) no son un resultado ISO ni una tabla calculada por el plugin: son la malla frecuencial sobre la que se resuelve la propagación."),
        ("Le plugin combine deux éléments:", "El plugin combina dos elementos:"),
        ("Courbe acoustique globale LwA(ws): fixe le niveau opérationnel global de l’aerogenerador pour la vitesse de vent ou le cas le plus défavorable sélectionné.", "Curva acústica global LwA(ws): fija el nivel operativo global del aerogenerador para la velocidad de viento o el peor caso seleccionado."),
        ("Spectre par bandes Lw,b: répartit ce niveau global entre les 8 bandes et constitue l’entrée réelle utilisée dans l’équation par bandes.", "Espectro por bandas Lw,b: reparte este nivel global entre las 8 bandas y constituye la entrada real usada en la ecuación por bandas."),
        ("Ce espectro peut provenir d’un fichier spécifique du fabricant/de l’utilisateur ou d’un gabarit de référence. Si seule une courbe globale LwA(ws) est disponible, le plugin fixe d’abord le niveau global opérationnel LwA_cible, puis construit un espectro absolu par bandes à partir d’une forme spectrale de référence S_b^ref.", "Este espectro puede proceder de un archivo específico del fabricante/usuario o de una plantilla de referencia. Si solo hay una curva global LwA(ws), el plugin fija primero el nivel global operativo LwA_cible y después construye un espectro absoluto por bandas a partir de una forma espectral de referencia S_b^ref."),
        ("Reconstruction mathématique des bandes lorsqu’il n’existe que LwA(ws):", "Reconstrucción matemática de las bandas cuando solo existe LwA(ws):"),
        ("Autrement dit : la courbe acoustique fournit le niveau global opérationnel et le gabarit/la bibliothèque fournit la forme spectrale. Le décalage Δ est calculé de façon à ce que, après pondération A et sommation énergétique des 8 bandes, le espectro reconstruit reproduise exactement le LwA_cible de la courbe importée.", "Dicho de otra forma: la curva acústica proporciona el nivel global operativo y la plantilla/biblioteca proporciona la forma espectral. El desplazamiento Δ se calcula para que, después de aplicar la ponderación A y la suma energética de las 8 bandas, el espectro reconstruido reproduzca exactamente el LwA_cible de la curva importada."),
        ("S_b^ref est la forme spectrale de référence (si elle existe), A_weight,b la pondération A de chaque bande et Lw,b le niveau final en dB réellement utilisé par le calcul.", "S_b^ref es la forma espectral de referencia (si existe), A_weight,b la ponderación A de cada banda y Lw,b el nivel final en dB realmente utilizado por el cálculo."),
        ("Representa la dispersión geométrica de la onda sonora con la distancia 3D fuente–receptor. Aquí, d provient des coordonnées de l’aerogenerador et du receptor avec leurs hauteurs d’évaluation.", "Representa la dispersión geométrica de la onda sonora con la distancia 3D fuente–receptor. Aquí, d procede de las coordenadas del aerogenerador y del receptor con sus alturas de evaluación."),
        ("L’absorción atmosférica est calculée par bande à partir d’un coefficient de référence et de trois facteurs correcteurs. La dépendance physique à la température, l’humidité relative et la pression est bien représentée, mais au moyen d’une approximation simplifiée du plugin, et no de la formulation analytique complète de l’ISO 9613-1.", "La absorción atmosférica se calcula por banda a partir de un coeficiente de referencia y tres factores correctores. La dependencia física con temperatura, humedad relativa y presión se representa mediante una aproximación simplificada del plugin, no mediante la formulación analítica completa de la ISO 9613-1."),
        ("Interprétation des corrections: T est introduit en °C par rapport à une référence de 15 °C ; HR est comparée à une humidité optimale de référence de 50 % et la correction augmente lorsque l’on s’en éloigne ; P est introduite en kPa par rapport à une référence de 101,325 kPa avec une correction inverse. Ces facteurs ne modifient que le bloc atmosphérique Aatm,b : ils ne modifient ni l’émission de l’aerogenerador, ni l’efecto del suelo, ni le terme MDT/écran.", "Interpretación de las correcciones: T se introduce en °C respecto a una referencia de 15 °C; HR se compara con una humedad óptima de referencia del 50 % y la corrección aumenta al alejarse de ella; P se introduce en kPa respecto a una referencia de 101,325 kPa con una corrección inversa. Estos factores solo modifican el bloque atmosférico Aatm,b: no modifican ni la emisión del aerogenerador, ni el efecto del suelo, ni el término MDT/apantallamiento."),
        ("5. Effet de sol par régions", "5. Efecto del suelo por regiones"),
        ("El término de sol se décompose en As (région de source), Am (région intermédiaire) et Ar (région du receptor). Dans cette implémentation, trois paramètres de sol indépendants Gs/Gm/Gr ne sont pas utilisés ; un G unique par trajet est utilisé. Mathématiquement, le plugin applique :", "El término de suelo se descompone en As (región de fuente), Am (región intermedia) y Ar (región del receptor). En esta implementación no se usan tres parámetros de suelo independientes Gs/Gm/Gr; se usa un G único por trayectoria. Matemáticamente, el plugin aplica:"),
        ("où h_s est la hauteur caractéristique de la source, h_r celle du receptor, h_moy la hauteur moyenne du trajet et G_m≈0 dans l’approximation actuelle pour des conditions favorables de propagation. Cette valor único de sol peut être :", "donde h_s es la altura característica de la fuente, h_r la del receptor, h_moy la altura media de la trayectoria y G_m≈0 en la aproximación actual para condiciones favorables de propagación. Este valor único de suelo puede ser:"),
        ("G manuel/global, si l’utilisateur fixe une valor único.", "G manual/global, si el usuario fija un valor único."),
        ("G_eff, si une couche d’occupation du sol existe et si une moyenne pondérée par la longueur du trajet est calculée.", "G_eff, si existe una capa de ocupación del suelo y se calcula una media ponderada por la longitud de la trayectoria."),
        ("Significado physique de G: représente le caractère acoustique du terrain et contrôle l’influence du sol sur la propagation. G≈0 indique un sol dur (urbain, asphalte, roche), G≈1 un sol meuble/poreux (agricole, prairie, forestier) et les valeurs intermédiaires représentent un terrain mixte.", "Significado físico de G: representa el carácter acústico del terreno y controla la influencia del suelo sobre la propagación. G≈0 indica suelo duro (urbano, asfalto, roca), G≈1 suelo blando/poroso (agrícola, pradera, forestal) y los valores intermedios representan terreno mixto."),
        ("Ce que signifie « desde capa » : le plugin intersecte le trajet source–receptor avec la couche d’occupation du sol, attribue une valeur G_i à chaque polygone intercepté et calcule un G_eff unique pour ce trajet. C’est cette valeur qui entre réellement dans Agr,b ; le G global affiché dans le rapport reste uniquement une valeur de secours.", "Qué significa «desde capa»: el plugin intersecta la trayectoria fuente–receptor con la capa de ocupación del suelo, asigna un valor G_i a cada polígono interceptado y calcula un G_eff único para esa trayectoria. Ese es el valor que entra realmente en Agr,b; el G global mostrado en el informe queda solo como valor de respaldo."),
        ("Convention du rapport : Agr,b est affiché ici comme une amplitude positive d’atenuación. Dans l’équation principale, il est soustrait au niveau de source comme Adiv, Aatm et Abar.", "Convención del informe: Agr,b se muestra aquí como una magnitud positiva de atenuación. En la ecuación principal se resta al nivel de fuente, igual que Adiv, Aatm y Abar."),
        ("6. Écran topographique avec MDT", "6. Apantallamiento topográfico con MDT"),
        ("Le MDT ne modifie pas l’émission de l’aerogenerador ni l’absorción atmosférica. Sa fonction est de décrire la géométrie réelle du trajet et d’alimenter le terme Abar,b.", "El MDT no modifica la emisión del aerogenerador ni la absorción atmosférica. Su función es describir la geometría real de la trayectoria y alimentar el término Abar,b."),
        ("Profil du terrain : le profil source–receptor est extrait du MDT avec un échantillonnage adaptatif.", "Perfil del terreno: el perfil fuente–receptor se extrae del MDT con un muestreo adaptativo."),
        ("Ligne de visée: la droite entre la hauteur efectiva de source et la altura del receptor est construite. Si le terrain reste toujours en dessous, alors Abar,b = 0.", "Línea de visión: se construye la recta entre la altura efectiva de fuente y la altura del receptor. Si el terreno queda siempre por debajo, entonces Abar,b = 0."),
        ("Obstacle dominant: si une colline ou une crête supera, la hauteur au-dessus de la ligne de visée est calculée :", "Obstáculo dominante: si una colina o cresta sobresale, se calcula la altura por encima de la línea de visión:"),
        ("Lorsque h_obs > 0, le relief coupe la vision directe et une atenuación supplémentaire par diffraction peut apparaître.", "Cuando h_obs > 0, el relieve corta la visión directa y puede aparecer una atenuación adicional por difracción."),
        ("Géométrie réelle de l’obstacle : le plugin utilise la position réelle de l’obstacle dominant et calcule d1 (source → obstacle) et d2 (obstacle → receptor).", "Geometría real del obstáculo: el plugin usa la posición real del obstáculo dominante y calcula d1 (fuente → obstáculo) y d2 (obstáculo → receptor)."),
        ("Activation conservatrice: Abar n’est pas activé pour de petites irrégularités du MDT ; un seuil minimal lié à la résolution du raster est exigé.", "Activación conservadora: Abar no se activa para pequeñas irregularidades del MDT; se exige un umbral mínimo ligado a la resolución del raster."),
        ("Diffraction de type Fresnel: avec cette géométrie, une différence de chemins et un nombre de Fresnel sont estimés :", "Difracción tipo Fresnel: con esta geometría se estiman una diferencia de caminos y un número de Fresnel:"),
        ("Ce nombre est ensuite transformé en une atenuación Abar,b dépendante de la fréquence au moyen de l’approximation actuelle du plugin :", "Ese número se transforma después en una atenuación Abar,b dependiente de la frecuencia mediante la aproximación actual del plugin:"),
        ("Dans l’implémentation actuelle, Abar est également limité à des valeurs raisonnables (plafonnement supérieur) afin d’éviter des suratténuations parasites. En l’absence de MDT ou d’obstacle pertinent, alors Abar,b = 0.", "En la implementación actual, Abar también se limita a valores razonables (tope superior) para evitar sobreatenuaciones parásitas. En ausencia de MDT o de obstáculo relevante, Abar,b = 0."),
        ("Lecture du receptor critique: le tableau de la section du receptor critique affiche des amplitudes d’atenuación pour la traçabilité. Le niveau résultant ne doit pas être interprété comme une soustraction directe depuis une seule aerogenerador : il est obtenu par sommation énergétique par bandes et par sommation des sources contributrices dans le rayon de calcul.", "Lectura del receptor crítico: la tabla de la sección del receptor crítico muestra magnitudes de atenuación para trazabilidad. El nivel resultante no debe interpretarse como una resta directa desde un único aerogenerador: se obtiene mediante suma energética por bandas y suma de las fuentes contribuyentes dentro del radio de cálculo."),
        ("🌱 Physique de l’occupation du sol et calcul de G", "🌱 Física del uso del suelo y cálculo de G"),
        ("🌱 Physique de l’occupation du sol et calcul de G_eff", "🌱 Física del uso del suelo y cálculo de G_eff"),
        ("Dans ce calcul, l’efecto del suelo a été calculé avec un G manuel unique pour tout le trajet:", "En este cálculo, el efecto del suelo se ha calculado con un G manual único para toda la trayectoria:"),
        ("Cette valeur est appliquée dans le terme de sol du modelo. Aucun G_eff n’a été dérivé depuis une couche d’occupation du sol.", "Este valor se aplica en el término de suelo del modelo. No se ha derivado ningún G_eff desde una capa de ocupación del suelo."),
        ("🗺️ Physique du MDT et de l’écran topographique", "🗺️ Física del MDT y del apantallamiento topográfico"),
        ("Dans ce calcul, aucun MDT/DSM n’a été utilisé, donc le terme d’écran topographique est fixé à:", "En este cálculo no se ha usado ningún MDT/DSM, por lo que el término de apantallamiento topográfico se fija en:"),
        ("L’évaluation est réalisée sans introduire d’écrans topographiques. La géométrie du trajet est résolue sans profil de terrain et le calcul dépend de Lw,b, Adiv, Aatm,b et Agr,b.", "La evaluación se realiza sin introducir pantallas topográficas. La geometría de la trayectoria se resuelve sin perfil de terreno y el cálculo depende de Lw,b, Adiv, Aatm,b y Agr,b."),
        ("Simplificaciones conocidas: Aatm simplifié (tables + corrections) ; Agr et Abar avec approximations de base ; directivité Dc supposée égale à 0 dB ; Cmet/correction météorologique de long terme no appliquée.", "Simplificaciones conocidas: Aatm simplificada (tablas + correcciones); Agr y Abar con aproximaciones básicas; directividad Dc asumida igual a 0 dB; Cmet/corrección meteorológica de largo plazo no aplicada."),
    ]
    for a, b in extra_repl:
        html = html.replace(a, b)

    # Smaller fragments used when dynamic values split a sentence.
    frag_repl = [
        ("avec corrections simplifiées de température, humidité et pression", "con correcciones simplificadas de temperatura, humedad y presión"),
        ("via une table de référence", "mediante una tabla de referencia"),
        ("une table de référence", "una tabla de referencia"),
        ("formulation analytique complète", "formulación analítica completa"),
        ("flux réel suivi", "flujo real seguido"),
        ("résultat par receptor soit traçable", "resultado por receptor sea trazable"),
        ("niveau final", "nivel final"),
        ("ne provient pas", "no procede"),
        ("soustraction unique", "resta única"),
        ("contributions source–receptor", "contribuciones fuente–receptor"),
        ("dans le rayon de calcul", "dentro del radio de cálculo"),
        ("sommation énergétique", "suma energética"),
        ("entrées SIG", "entradas SIG"),
        ("sources acoustiques", "fuentes acústicas"),
        ("rayon maximal", "radio máximo"),
        ("sont pris en compte", "se tienen en cuenta"),
        ("couche d’occupation du sol", "capa de ocupación del suelo"),
        ("occupation du sol", "ocupación del suelo"),
        ("si elle existe", "si existe"),
        ("s’il est activo", "si está activo"),
        ("État acoustique", "Estado acústico"),
        ("groupe source", "grupo fuente"),
        ("groupe d’aerogeneradores", "grupo de aerogeneradores"),
        ("valeur fixe", "valor fijo"),
        ("courbe", "curva"),
        ("opérationnel", "operativo"),
        ("Conversion en bandes", "Conversión a bandas"),
        ("le moteur", "el motor"),
        ("a besoin d’un", "necesita un"),
        ("bandes d’octave", "bandas de octava"),
        ("n’existe pas", "no existe"),
        ("gabarit/fallback", "plantilla/fallback"),
        ("l’ajuste", "lo ajusta"),
        ("reproduire", "reproducir"),
        ("Sélection des contributeurs", "Selección de contribuyentes"),
        ("situées dans le rayon", "situados dentro del radio"),
        ("sont recherchées", "se buscan"),
        ("sans sources", "sin fuentes"),
        ("sont marqués", "se marcan"),
        ("hors rayon", "fuera de radio"),
        ("niveau acoustique utile", "nivel acústico útil"),
        ("trajet source–receptor", "trayectoria fuente–receptor"),
        ("contributrice", "contribuyente"),
        ("cotes acoustiques", "cotas acústicas"),
        ("éventuel écran topographique", "posible apantallamiento topográfico"),
        ("du trajet", "de la trayectoria"),
        ("sont calculés", "se calculan"),
        ("Propagation par bande", "Propagación por banda"),
        ("dans chaque bande", "en cada banda"),
        ("on applique", "se aplica"),
        ("dépend de", "depende de"),
        ("fréquence", "frecuencia"),
        ("atmosphère", "atmósfera"),
        ("obstacle pertinent", "obstáculo relevante"),
        ("Sommation par source", "Suma por fuente"),
        ("pondérées A", "ponderadas A"),
        ("sommées énergétiquement", "sumadas energéticamente"),
        ("obtenir", "obtener"),
        ("niveau pondéré A", "nivel ponderado A"),
        ("Sommation du receptor", "Suma del receptor"),
        ("toutes les", "todos los"),
        ("Comparaison avec", "Comparación con"),
        ("le niveau total", "el nivel total"),
        ("est comparé", "se compara"),
        ("attribuée", "asignado"),
        ("référence", "referencia"),
        ("La marge", "El margen"),
        ("état de conformité", "estado de cumplimiento"),
        ("dépassements", "excedencias"),
        ("Lecture pratique", "Lectura práctica"),
        ("présente", "presenta"),
        ("plus élevé", "más alto"),
        ("plus défavorable", "más desfavorable"),
        ("par rapport à", "respecto a"),
        ("colonne", "columna"),
        ("source dominante", "fuente dominante"),
        ("identifie", "identifica"),
        ("contribue le plus", "más contribuye"),
        ("résultat final", "resultado final"),
        ("inclut", "incluye"),
        ("plus lourd", "más pesado"),
        ("traçable", "trazable"),
        ("utilise", "usa"),
        ("pondération A", "ponderación A"),
        ("dépendante", "dependiente"),
        ("rapports techniques préliminaires", "informes técnicos preliminares"),
        ("revue des receptores sensibles", "revisión de receptores sensibles"),
        ("sont affichées", "se muestran"),
        ("signe algébrique", "signo algebraico"),
        ("équation", "ecuación"),
        ("maximum parmi", "máximo entre"),
        ("pas uniquement", "no únicamente"),
        ("trajet dominant", "trayectoria dominante"),
        ("Pression", "Presión"),
        ("pression", "presión"),
        ("plage typique", "rango típico"),
        ("études préliminaires", "estudios preliminares"),
        ("mesure du site", "medición del emplazamiento"),
        ("proche de", "cerca de"),
        ("ajustée à l’altitude", "ajustada a la altitud"),
        ("effectif depuis", "efectivo desde"),
        ("par bande", "por banda"),
        ("par bandes", "por bandas"),
        ("de l’air", "del aire"),
        ("de l’aerogenerador", "del aerogenerador"),
        ("du receptor", "del receptor"),
        ("au receptor", "al receptor"),
        ("des receptores", "de los receptores"),
        ("Facteur de sol", "Factor de suelo"),
        ("valeur efectiva", "valor efectivo"),
        ("par trajet", "por trayectoria"),
        ("no appliquée", "no aplicada"),
        ("Ce moteur travaille", "Este motor trabaja"),
        ("grille fréquentielle", "malla frecuencial"),
        ("méthode", "metodología"),
        ("entrée acoustique", "entrada acústica"),
        ("peut provenir", "puede proceder"),
        ("mesuré/importé", "medido/importado"),
        ("ajusté", "ajustado"),
        ("S’il existe", "Si existe"),
        ("spécifique", "específico"),
        ("c’est l’entrée utilisée", "esa es la entrada usada"),
        ("Sino", "Si no"),
        ("bibliothèque", "biblioteca"),
        ("il procède", "procede"),
        ("il procede", "procede"),
        ("selon le scénario sélectionné", "según el escenario seleccionado"),
        ("ne remplace pas", "no sustituye"),
        ("les bandes", "las bandas"),
        ("fixe l’état", "fija el estado"),
        ("fournit", "aporta"),
        ("répartition fréquentielle", "distribución frecuencial"),
        ("hauteur efectiva de source", "altura efectiva de fuente"),
        ("Topographie", "Topografía"),
        ("optionnel", "opcional"),
        ("n’affecte que", "solo afecta a"),
        ("Elle provient", "Procede"),
        ("La courbe", "La curva"),
        ("définit", "define"),
        ("répartit", "reparte"),
        ("Calculé", "Calculado"),
        ("calculé", "calculado"),
        ("table de base", "tabla base"),
        ("absorption", "absorción"),
        ("corrections simplifiées", "correcciones simplificadas"),
        ("température", "temperatura"),
        ("humidité relative", "humedad relativa"),
        ("humidité", "humedad"),
        ("L’implémentation actuelle", "La implementación actual"),
        ("formulation exacte", "formulación exacta"),
        ("comme efecto del suelo", "como efecto del suelo"),
        ("régions", "regiones"),
        ("paramètre de sol", "parámetro de suelo"),
        ("G unique", "G único"),
        ("manuel/global", "manual/global"),
        ("dérivé", "derivado"),
        ("N’intervient", "Solo interviene"),
        ("écran topographique", "apantallamiento topográfico"),
        ("est détecté", "se detecta"),
        ("En l’absence", "En ausencia"),
        ("Entrée acoustique", "Entrada acústica"),
        ("donnée d’entrée", "dato de entrada"),
        ("ce sont", "son"),
        ("résolue", "resuelta"),
        ("deux éléments", "dos elementos"),
        ("niveau opérationnel global", "nivel operativo global"),
        ("vitesse de vent", "velocidad de viento"),
        ("cas le plus défavorable", "peor caso"),
        ("sélectionné", "seleccionado"),
        ("entrée réelle", "entrada real"),
        ("fichier spécifique", "archivo específico"),
        ("fabricant", "fabricante"),
        ("utilisateur", "usuario"),
        ("forme spectrale", "forma espectral"),
        ("décalage", "desplazamiento"),
        ("de façon à ce que", "para que"),
        ("après", "después de"),
        ("reconstruit", "reconstruido"),
        ("exactement", "exactamente"),
        ("courbe importée", "curva importada"),
        ("forme spectrale de référence", "forma espectral de referencia"),
        ("niveau final", "nivel final"),
        ("réellement utilisé", "realmente utilizado"),
        ("provient des coordonnées", "procede de las coordenadas"),
        ("hauteurs d’évaluation", "alturas de evaluación"),
        ("dépendance physique", "dependencia física"),
        ("bien représentée", "representada"),
        ("au moyen d’une approximation", "mediante una aproximación"),
        ("facteurs correcteurs", "factores correctores"),
        ("ne modifient", "no modifican"),
        ("émission", "emisión"),
        ("le terme", "el término"),
        ("Effet de sol", "Efecto del suelo"),
        ("se décompose", "se descompone"),
        ("région de source", "región de fuente"),
        ("région intermédiaire", "región intermedia"),
        ("Dans cette implémentation", "En esta implementación"),
        ("ne sont pas utilisés", "no se usan"),
        ("Mathématiquement", "Matemáticamente"),
        ("où", "donde"),
        ("hauteur caractéristique", "altura característica"),
        ("hauteur moyenne", "altura media"),
        ("conditions favorables", "condiciones favorables"),
        ("Cette valor único", "Este valor único"),
        ("G manuel", "G manual"),
        ("l’utilisateur", "el usuario"),
        ("une moyenne pondérée", "una media ponderada"),
        ("longueur", "longitud"),
        ("physique", "físico"),
        ("représente", "representa"),
        ("caractère acoustique", "carácter acústico"),
        ("terrain", "terreno"),
        ("contrôle", "controla"),
        ("influence", "influencia"),
        ("urbain", "urbano"),
        ("asphalte", "asfalto"),
        ("meuble/poreux", "blando/poroso"),
        ("agricole", "agrícola"),
        ("prairie", "pradera"),
        ("forestier", "forestal"),
        ("valeurs intermédiaires", "valores intermedios"),
        ("terrain mixte", "terreno mixto"),
        ("Ce que signifie", "Qué significa"),
        ("intersecte", "intersecta"),
        ("attribue", "asigna"),
        ("polygone intercepté", "polígono interceptado"),
        ("entre réellement", "entra realmente"),
        ("affiché", "mostrado"),
        ("rapport", "informe"),
        ("valeur de secours", "valor de respaldo"),
        ("Convention du", "Convención del"),
        ("amplitude positive", "magnitud positiva"),
        ("il est soustrait", "se resta"),
        ("niveau de source", "nivel de fuente"),
        ("Écran topographique", "Apantallamiento topográfico"),
        ("ne modifie pas", "no modifica"),
        ("Sa fonction", "Su función"),
        ("décrire", "describir"),
        ("géométrie réelle", "geometría real"),
        ("Profil du terrain", "Perfil del terreno"),
        ("échantillonnage adaptatif", "muestreo adaptativo"),
        ("Ligne de visée", "Línea de visión"),
        ("droite", "recta"),
        ("est construite", "se construye"),
        ("reste toujours en dessous", "queda siempre por debajo"),
        ("colline", "colina"),
        ("crête", "cresta"),
        ("hauteur au-dessus", "altura por encima"),
        ("relief", "relieve"),
        ("vision directe", "visión directa"),
        ("supplémentaire", "adicional"),
        ("peut apparaître", "puede aparecer"),
        ("Géométrie réelle", "Geometría real"),
        ("position réelle", "posición real"),
        ("Activation conservatrice", "Activación conservadora"),
        ("n’est pas activé", "no se activa"),
        ("petites irrégularités", "pequeñas irregularidades"),
        ("seuil minimal", "umbral mínimo"),
        ("est exigé", "se exige"),
        ("Diffraction de type Fresnel", "Difracción tipo Fresnel"),
        ("différence de chemins", "diferencia de caminos"),
        ("sont estimés", "se estiman"),
        ("Ce nombre", "Ese número"),
        ("est ensuite transformé", "se transforma después"),
        ("moyen de", "mediante"),
        ("également limité", "también limitado"),
        ("valeurs raisonnables", "valores razonables"),
        ("plafonnement supérieur", "tope superior"),
        ("afin d’éviter", "para evitar"),
        ("suratténuations parasites", "sobreatenuaciones parásitas"),
        ("Physique de", "Física de"),
        ("aucun MDT/DSM", "ningún MDT/DSM"),
        ("n’a été utilisé", "se ha usado"),
        ("donc", "por tanto"),
        ("est fixé", "se fija"),
        ("L’évaluation", "La evaluación"),
        ("écrans topographiques", "pantallas topográficas"),
        ("profil de terrain", "perfil de terreno"),
        ("simplifié", "simplificada"),
        ("directivité", "directividad"),
        ("supposée égale", "asumida igual"),
        ("météorologique", "meteorológica"),
        ("long terme", "largo plazo"),
    ]
    for a, b in frag_repl:
        html = html.replace(a, b)

    # Last safety net for short French grammatical leftovers in Spanish reports.
    regex_repl = [
        (r"\bCette\b", "Esta"), (r"\bCe\b", "Este"), (r"\bce\b", "este"),
        (r"\bLe\b", "El"), (r"\ble\b", "el"), (r"\bLa\b", "La"), (r"\bla\b", "la"),
        (r"\bLes\b", "Los"), (r"\bles\b", "los"), (r"\bDes\b", "De los"), (r"\bdes\b", "de los"),
        (r"\bDans\b", "En"), (r"\bdans\b", "en"), (r"\bPour\b", "Para"), (r"\bpour\b", "para"),
        (r"\bAvec\b", "Con"), (r"\bavec\b", "con"), (r"\bSans\b", "Sin"), (r"\bsans\b", "sin"),
        (r"\bEt\b", "Y"), (r"\bet\b", "y"), (r"\bOu\b", "O"), (r"\bou\b", "o"),
        (r"\bIl\b", "Este"), (r"\bil\b", "este"), (r"\bElle\b", "Esta"), (r"\belle\b", "esta"),
        (r"\bEst\b", "Es"), (r"\best\b", "es"), (r"\bSont\b", "Son"), (r"\bsont\b", "son"),
        (r"\bUne\b", "Una"), (r"\bune\b", "una"), (r"\bUn\b", "Un"),
        (r"\bplus\b", "más"), (r"\bmoins\b", "menos"), (r"\bmais\b", "pero"),
        (r"\bdonc\b", "por tanto"), (r"\bpuis\b", "después"),
        (r"\bau\b", "al"), (r"\bdu\b", "del"), (r"\bde la\b", "de la"),
        (r"\blorsque\b", "cuando"), (r"\bLorsque\b", "Cuando"),
        (r"\bs’il\b", "si"), (r"\bS’il\b", "Si"), (r"\bsinon\b", "si no"), (r"\bSinon\b", "Si no"),
        (r"\bn’est\b", "no es"), (r"\bne\b", "no"),
    ]
    for pat, repl in regex_repl:
        html = re.sub(pat, repl, html)
    html = html.replace("l’", "").replace("L’", "").replace("d’", "de ").replace("D’", "De ").replace("n’", "no ").replace("N’", "No ")
    html = html.replace("7. Pondération A utilisée à la fin", "7. Ponderación A utilizada al final")
    html = html.replace("Description", "Descripción")
    html = html.replace("(divergence)", "(divergencia)")
    html = html.replace("(atmosphérique)", "(atmosférica)")
    html = html.replace("Agr (sol)", "Agr (suelo)")
    html = html.replace("fabricantee", "fabricante")
    html = html.replace("fabricantee", "fabricante")
    return html



def _cleanup_english_noise_html(html: str) -> str:
    """Post-process the noise HTML report when English is active.

    The report template is legacy French.  For English sessions we first
    convert that legacy template to the Spanish pivot and then translate the
    Spanish text to English.  This final pass catches long report paragraphs
    and dynamic fragments that the generic i18n layer may leave partially in
    Spanish or French.
    """
    repl = [
        # Main report titles and static labels
        ("Evaluación del ruido generado por aerogeneradores", "Assessment of wind turbine noise"),
        ("Alcance de este informe — leer antes de usar los resultados", "Scope of this report — read before using the results"),
        ("Qué es:", "What it is:"),
        ("Qué no es:", "What it is not:"),
        ("Simplificaciones aplicadas en este modo:", "Simplifications applied in this mode:"),
        ("Recomendación:", "Recommendation:"),
        ("RESUMEN EJECUTIVO", "EXECUTIVE SUMMARY"),
        ("Aerogeneradores", "Wind turbines"),
        ("Receptores evaluados", "Receivers evaluated"),
        ("Nivel máximo (dB(A))", "Maximum level (dB(A))"),
        ("Cobertura del análisis", "Analysis coverage"),
        ("receptores dentro del radio", "receivers within the radius"),
        ("de cobertura", "coverage"),
        ("receptores fuera del radio", "receivers outside the radius"),
        ("Cumplimiento regulatorio", "Regulatory compliance"),
        ("receptores superan los límites", "receivers exceed the limits"),
        ("de cumplimiento sobre los receptores cubiertos", "compliance over covered receivers"),
        ("Límite:", "Limit:"),
        ("Metodología de cálculo", "Calculation methodology"),
        ("Motor utilizado:", "Engine used:"),
        ("Grupos fuente acústicos:", "Acoustic source groups:"),
        ("modelo(s) de aerogenerador", "wind turbine model(s)"),
        ("Método:", "Method:"),
        ("Mapa raster:", "Raster map:"),
        ("sí · resolución solicitada", "yes · requested resolution"),
        ("efectiva", "effective"),
        ("no generado", "not generated"),
        ("autoajustada", "auto-adjusted"),
        ("CÓMO SE HA GENERADO EL RESULTADO", "HOW THE RESULT WAS GENERATED"),
        ("Cómo se ha ejecutado el cálculo ISO-aligned", "How the ISO-aligned calculation was executed"),
        ("Cómo se ha ejecutado el cálculo Screening", "How the Screening calculation was executed"),
        ("Qué diferencia este modo del modo Screening", "What distinguishes this mode from Screening mode"),
        ("RECEPTOR CRÍTICO", "CRITICAL RECEIVER"),
        ("Receptor crítico (mayor nivel sonoro)", "Critical receiver (highest sound level)"),
        ("ID receptor:", "Receiver ID:"),
        ("Nivel total:", "Total level:"),
        ("Límite aplicable:", "Applicable limit:"),
        ("Margen:", "Margin:"),
        ("Modelo dominante:", "Dominant model:"),
        ("Grupo fuente:", "Source group:"),
        ("Aerogeneradores contribuyentes dentro del radio", "Contributing wind turbines within the radius"),
        ("Distancia:", "Distance:"),
        ("Desglose de atenuaciones", "Attenuation breakdown"),
        ("Los valores mostrados abajo son las magnitudes de atenuación usadas por el modelo. En la ecuación principal, estos términos se restan al nivel de fuente.", "The values shown below are the attenuation magnitudes used by the model. In the main equation, these terms are subtracted from the source level."),
        ("Término", "Term"),
        ("Valor [dB]", "Value [dB]"),
        ("Descripción", "Description"),
        ("LwA fuente dominante", "Dominant-source LwA"),
        ("Potencia acústica del aerogenerador", "Wind turbine sound power"),
        ("Adiv (divergencia)", "Adiv (divergence)"),
        ("Dispersión geométrica", "Geometrical spreading"),
        ("Aatm (atmosférica)", "Aatm (atmospheric)"),
        ("Absorción en el aire", "Air absorption"),
        ("Agr (suelo)", "Agr (ground)"),
        ("Atenuación debida al efecto del suelo", "Attenuation due to ground effect"),
        ("Abar trayectoria dominante", "Dominant-path Abar"),
        ("Atenuación debida al MDT en la trayectoria dominante", "Attenuation due to the DTM along the dominant path"),
        ("Abar máximo de los contribuyentes", "Maximum Abar among contributors"),
        ("Abar máximo entre todos los aerogeneradores que contribuyen al receptor", "Maximum Abar among all wind turbines contributing to the receiver"),
        ("Abar ponderado por energía", "Energy-weighted Abar"),
        ("Media ponderada por la contribución acústica de cada aerogenerador", "Average weighted by the acoustic contribution of each wind turbine"),
        ("Trayectorias apantalladas", "Screened paths"),
        ("Número de aerogeneradores contribuyentes con Abar", "Number of contributing wind turbines with Abar"),
        ("NIVEL RESULTANTE", "RESULTING LEVEL"),
        ("Banda dominante:", "Dominant band:"),
        ("Origen del espectro:", "Spectrum source:"),
        ("Estadísticos de atenuaciones (receptores cubiertos)", "Attenuation statistics (covered receivers)"),
        ("Se muestran las magnitudes brutas de atenuación, no el signo algebraico de la ecuación. Para Abar se usa el máximo entre los aerogeneradores contribuyentes de cada receptor, no solo la trayectoria dominante.", "Raw attenuation magnitudes are shown, not the algebraic sign in the equation. For Abar, the maximum among the wind turbines contributing to each receiver is used, not only the dominant path."),
        ("CONFIGURACIÓN Y PARÁMETROS", "CONFIGURATION AND PARAMETERS"),
        ("Ecuación utilizada", "Equation used"),
        ("Parámetros del cálculo", "Calculation parameters"),
        ("Altura del receptor", "Receiver height"),
        ("Radio máximo", "Maximum radius"),
        ("Modo suelo", "Ground mode"),
        ("G utilizado", "G used"),
        ("G_eff medio", "Mean G_eff"),
        ("G_eff del receptor crítico", "Critical receiver G_eff"),
        ("Ocupación del suelo", "Land use"),
        ("Escenario acústico", "Acoustic scenario"),
        ("Temperatura", "Temperature"),
        ("Humedad relativa", "Relative humidity"),
        ("Presión", "Pressure"),
        ("Revisión recomendada", "Recommended review"),
        ("la presión atmosférica introducida", "the entered atmospheric pressure"),
        ("está fuera del rango típico usado como referencia en muchos estudios preliminares", "is outside the typical range used as a reference in many preliminary studies"),
        ("Si no es una medición del emplazamiento, conviene comprobar si debería estar cerca de 101,325 kPa o ajustarse a la altitud.", "If it is not a site measurement, check whether it should be close to 101.325 kPa or adjusted to altitude."),
        ("Trayectorias con G distinto del global", "Paths with G different from the global value"),
        ("Términos activos", "Active terms"),
        ("activo", "active"),
        ("inactivo", "inactive"),
        ("FÍSICA DETALLADA Y TRAZABILIDAD DEL CÁLCULO", "DETAILED PHYSICS AND CALCULATION TRACEABILITY"),
        ("Glosario de símbolos", "Glossary of symbols"),
        ("Definición compacta de los símbolos que aparecen en las fórmulas y tablas de este informe.", "Compact definition of the symbols appearing in this report's formulas and tables."),
        ("Símbolo", "Symbol"),
        ("Significado", "Meaning"),
        ("Desarrollo físico detallado del motor ISO-aligned", "Detailed physical development of the ISO-aligned engine"),
        ("Desarrollo físico detallado del motor rápido", "Detailed physical development of the fast engine"),
        ("Escenario operativo de este cálculo", "Operating scenario of this calculation"),
        ("Entradas realmente utilizadas en este cálculo", "Inputs actually used in this calculation"),
        ("Origen de cada término de la ecuación", "Origin of each term in the equation"),
        ("Cómo se obtiene en este plugin", "How it is obtained in this plugin"),
        ("Entrada acústica de la fuente y bandas", "Acoustic source input and bands"),
        ("Espectro utilizado por el grupo fuente", "Spectrum used by the source group"),
        ("Divergencia geométrica", "Geometrical divergence"),
        ("Absorción atmosférica simplificada", "Simplified atmospheric absorption"),
        ("Efecto del suelo por regiones", "Ground effect by regions"),
        ("Apantallamiento topográfico con MDT", "Topographic screening with DTM"),
        ("Ponderación A utilizada al final", "A-weighting used at the end"),
        ("Física del uso del suelo y cálculo de G", "Land-use physics and G calculation"),
        ("Física del MDT y del apantallamiento topográfico", "DTM and topographic-screening physics"),
        ("GRUPOS FUENTE ACÚSTICOS", "ACOUSTIC SOURCE GROUPS"),
        ("LwA efectivo por grupo", "Effective LwA by group"),
        ("DISTRIBUCIÓN POR TIPO DE RECEPTOR", "DISTRIBUTION BY RECEIVER TYPE"),
        ("Receptores por categoría", "Receivers by category"),
        ("Cumplimiento por categoría", "Compliance by category"),
        ("Límites y recomendaciones", "Limits and recommendations"),
        ("Motor rápido", "Fast engine"),
        ("Adecuado para screening preliminar y mapas ágiles.", "Suitable for preliminary screening and quick maps."),
        ("Motor ISO-aligned", "ISO-aligned engine"),
        ("Adecuado para estudios técnicos preliminares, comparaciones e iteración de diseño.", "Suitable for preliminary technical studies, comparisons and design iteration."),
        ("Simplificaciones conocidas", "Known simplifications"),
        ("Modelos múltiples", "Multiple models"),
        ("Raster ISO + MDT", "ISO raster + DTM"),
        # Common report prose
        ("una evaluación acústica preliminar alineada con la metodología ISO 9613-2, orientada al diseño, la comparación de alternativas y el cribado de receptores sensibles.", "a preliminary acoustic assessment aligned with ISO 9613-2 methodology, intended for design, alternative comparison and screening of sensitive receivers."),
        ("no es un informe acústico certificado y no sustituye a un estudio regulatorio definitivo realizado con software comercial validado.", "it is not a certified acoustic report and does not replace a definitive regulatory study prepared with validated commercial software."),
        ("Absorción atmosférica Aatm mediante una tabla de referencia con correcciones simplificadas de temperatura, humedad y presión, no mediante la formulación analítica completa de la ISO 9613-1.", "Atmospheric absorption Aatm using a reference table with simplified temperature, humidity and pressure corrections, not the complete analytical formulation of ISO 9613-1."),
        ("Sin corrección meteorológica de largo plazo Cmet.", "No long-term meteorological correction Cmet."),
        ("Difracción topográfica de un único obstáculo dominante: sin difracción lateral ni pantallas múltiples.", "Topographic diffraction for a single dominant obstacle: no lateral diffraction or multiple barriers."),
        ("Resolución espectral en 8 bandas de octava de 63 a 8000 Hz, no en tercios de octava.", "Spectral resolution in 8 octave bands from 63 to 8000 Hz, not in one-third octave bands."),
        ("Directividad de fuente Dc asumida igual a 0 dB.", "Source directivity Dc assumed equal to 0 dB."),
        ("Para decisiones regulatorias críticas, valida los resultados con mediciones de campo o con software comercial certificado.", "For critical regulatory decisions, validate the results with field measurements or certified commercial software."),
        ("el nivel resultante incluye la suma energética multi-fuente y multi-banda; no es una resta directa desde un único aerogenerador.", "the resulting level includes multi-source and multi-band energy summation; it is not a direct subtraction from a single wind turbine."),
        ("Propagación por bandas de octava según la metodología ISO-aligned", "Octave-band propagation according to the ISO-aligned methodology"),
        ("Cálculo acústico simplificado para el cribado", "Simplified acoustic calculation for screening"),
        # Technical vocabulary
        ("aerogenerador", "wind turbine"),
        ("aerogeneradores", "wind turbines"),
        ("receptor", "receiver"),
        ("receptores", "receivers"),
        ("fuente acústica", "acoustic source"),
        ("fuentes acústicas", "acoustic sources"),
        ("fuente", "source"),
        ("fuentes", "sources"),
        ("banda de octava", "octave band"),
        ("bandas de octava", "octave bands"),
        ("banda", "band"),
        ("bandas", "bands"),
        ("nivel", "level"),
        ("niveles", "levels"),
        ("límite", "limit"),
        ("límites", "limits"),
        ("margen", "margin"),
        ("cumple", "complies"),
        ("supera", "exceeds"),
        ("suelo", "ground"),
        ("sol/terreno", "ground/terrain"),
        ("distancia", "distance"),
        ("trayectoria", "path"),
        ("trayectorias", "paths"),
        ("apantallamiento", "screening"),
        ("topográfico", "topographic"),
        ("absorción atmosférica", "atmospheric absorption"),
        ("divergencia geométrica", "geometrical divergence"),
        ("atenuación", "attenuation"),
        ("atenuaciones", "attenuations"),
        ("ponderación", "weighting"),
        ("suma energética", "energy summation"),
        ("capa de ocupación del suelo", "land-use layer"),
        ("capa", "layer"),
        ("global", "global"),
    ]
    for a, b in repl:
        html = html.replace(a, b)
        html = html.replace(a.capitalize(), b.capitalize())

    # Direct French leftovers that should never appear in the English report.
    fr_repl = [
        ("RAPPORT TECHNIQUE D’IMPACT ACOUSTIQUE", "TECHNICAL NOISE IMPACT REPORT"),
        ("Évaluation du bruit généré par les éoliennes", "Assessment of wind turbine noise"),
        ("RÉSUMÉ EXÉCUTIF", "EXECUTIVE SUMMARY"),
        ("Éoliennes", "Wind turbines"),
        ("Récepteurs évalués", "Receivers evaluated"),
        ("Niveau maximal", "Maximum level"),
        ("Couverture de l’analyse", "Analysis coverage"),
        ("Conformité réglementaire", "Regulatory compliance"),
        ("Méthodologie de calcul", "Calculation methodology"),
        ("COMMENT LE RÉSULTAT A ÉTÉ GÉNÉRÉ", "HOW THE RESULT WAS GENERATED"),
        ("Comment le calcul ISO-aligned a été exécuté", "How the ISO-aligned calculation was executed"),
        ("Cette section explique", "This section explains"),
        ("chaque récepteur", "each receiver"),
        ("chaque receptor", "each receiver"),
        ("dans le rayon", "within the radius"),
        ("hors rayon", "outside the radius"),
        ("niveau total", "total level"),
        ("éolienne", "wind turbine"),
        ("éoliennes", "wind turbines"),
        ("récepteur", "receiver"),
        ("récepteurs", "receivers"),
        ("source dominante", "dominant source"),
        ("groupe source", "source group"),
        ("sommation énergétique", "energy summation"),
        ("bandes d’octave", "octave bands"),
        ("pondération A", "A-weighting"),
        ("absorption atmosphérique", "atmospheric absorption"),
        ("écran topographique", "topographic screening"),
        ("lorsqu’un", "when a"),
        ("disponible", "available"),
        ("Révision recommandée", "Recommended review"),
        ("pression atmosphérique saisie", "entered atmospheric pressure"),
        ("plage typique", "typical range"),
        ("PHYSIQUE DÉTAILLÉE ET TRAÇABILITÉ DU CALCUL", "DETAILED PHYSICS AND CALCULATION TRACEABILITY"),
        ("GROUPES SOURCE ACOUSTIQUES", "ACOUSTIC SOURCE GROUPS"),
        ("DISTRIBUTION PAR TYPE DE RÉCEPTEUR", "DISTRIBUTION BY RECEIVER TYPE"),
        ("Limites et recommandations", "Limits and recommendations"),
        ("Non disponible", "Not available"),
        ("oui", "yes"),
        ("non", "no"),
        ("actif", "active"),
        ("inactif", "inactive"),
    ]
    for a, b in fr_repl:
        html = html.replace(a, b)

    regex_repl = [
        (r"\bpar\b", "by"),
        (r"\bchaque\b", "each"),
        (r"\bcalcul\b", "calculation"),
        (r"\bCalcul\b", "Calculation"),
        (r"\bniveau\b", "level"),
        (r"\bNiveau\b", "Level"),
        (r"\bsource\b", "source"),
        (r"\bsources\b", "sources"),
        (r"\brayon\b", "radius"),
        (r"\bmode\b", "mode"),
        (r"\bmoteur\b", "engine"),
        (r"\btableau\b", "table"),
        (r"\bvaleur\b", "value"),
        (r"\bvaleurs\b", "values"),
        (r"\bdepuis\b", "from"),
        (r"\bavec\b", "with"),
        (r"\bpour\b", "for"),
        (r"\bselon\b", "according to"),
        (r"\bsi\b", "if"),
        (r"\bS’il\b", "If"),
        (r"\bS’este\b", "If"),
        (r"\bs’este\b", "if"),
        (r"\best\b", "is"),
        (r"\bsont\b", "are"),
        (r"\bno\b", "no"),
        (r"\bleur\b", "their"),
        (r"\bleurs\b", "their"),
        (r"\bqui\b", "that"),
        (r"\bcomme\b", "as"),
        (r"\bou\b", "or"),
        (r"\bet\b", "and"),
        (r"\bmais\b", "but"),
        (r"\baprès\b", "after"),
        (r"\bavant\b", "before"),
    ]
    for pat, repl_text in regex_repl:
        html = re.sub(pat, repl_text, html)

    # Clean frequent hybrids created by fragment translation.
    html = html.replace("HR", "RH")
    html = html.replace("T, HR, P", "T, RH, P")
    html = html.replace("corr_HR", "corr_RH")
    html = html.replace("G effective", "effective G")
    html = html.replace("source–receptor", "source–receiver")
    html = html.replace("receiver–receiver", "receiver")
    html = html.replace("wind turbines contribuyentes", "contributing wind turbines")
    html = html.replace("contributing wind turbinesturbines", "contributing wind turbines")
    html = html.replace("wind turbinees", "wind turbines")
    html = html.replace("receiveres", "receivers")
    html = html.replace("the source level", "the source level")

    # Extra cleanup for long mixed ES/FR/EN paragraphs generated by older
    # older fragment-level i18n misses.
    extra_repl = [
        ("Receivers evaluados", "Receivers evaluated"),
        ("compliance sobre los covered receivers", "compliance over covered receivers"),
        ("resolución solicitada", "requested resolution"),
        ("Height del receiver", "Receiver height"),
        ("G_eff del critical receiver", "Critical receiver G_eff"),
        ("Aatm: active (T, RH, P simplificada)", "Aatm: active (simplified T, RH, P)"),
        ("G efectivo desde ocupación del ground", "Effective G from land use"),
        ("Level de potencia acústica", "Sound power level"),
        ("Esta sección explica el flujo real seguido por el plugin para que el resultado por receiver sea trazable.", "This section explains the actual workflow followed by the plugin so that each receiver result is traceable."),
        ("Esta sección explica el flujo réelle seguido by el plugin afin que el resultado por receiver sea trazable.", "This section explains the actual workflow followed by the plugin so that each receiver result is traceable."),
        ("El level finale de cada receiver no procede de una simple resta única, sino del cálculo de todas las contribuciones source–receiver dentro del radio de cálculo y de su energy summation.", "The final level of each receiver does not come from a single subtraction, but from all source–receiver contributions within the calculation radius and their energy summation."),
        ("El level finale de each receiver no procede de una simple resta única, pero del calculation de todos los contribuciones source–receiver dentro del radio de cálculo, después de their energy summation.", "The final level of each receiver does not come from a single subtraction, but from all source–receiver contributions within the calculation radius and their energy summation."),
        ("Lecture de los entradas SIG", "GIS input reading"),
        ("los wind turbines/acoustic sources", "the wind turbines/acoustic sources"),
        ("los receivers", "the receivers"),
        ("la altura del receiver", "the receiver height"),
        ("el radio máximo de calculation", "the maximum calculation radius"),
        ("se tienen en cuenta", "are taken into account"),
        ("the layer de ocupación del ground", "the land-use layer"),
        ("si existe", "if it exists"),
        ("if existe", "if it exists"),
        ("si está active", "if it is active"),
        ("Estado acústico de each grupo source", "Acoustic state of each source group"),
        ("para each modelo o grupo de wind turbines", "for each wind turbine model or group"),
        ("un LwA operativo es obtenu à partir de una valor fijo o de una curva LwA(ws)", "an operational LwA is obtained from a fixed value or from an LwA(ws) curve"),
        ("En este calculation", "In this calculation"),
        ("LwA fijo por grupo de acoustic source", "fixed LwA per acoustic source group"),
        ("Band conversion", "Band conversion"),
        ("el motor ISO-aligned necesita un spectrum Lw,b en 8 octave bands", "the ISO-aligned engine needs an Lw,b spectrum in 8 octave bands"),
        ("If no existe de spectrum específico", "If no specific spectrum exists"),
        ("el plugin en reconstruido un à partir de un plantilla/fallback", "the plugin reconstructs one from a template/fallback"),
        ("lo ajusta para reproducir el LwA operativo", "and adjusts it to reproduce the operational LwA"),
        ("Selección de contribuyentes by receiver", "Selection of contributors by receiver"),
        ("para each receiver", "for each receiver"),
        ("los wind turbines situées en el radio máximo se buscan", "the wind turbines located within the maximum radius are searched"),
        ("Los receivers sin sources en este radius se marcan as outside radius", "Receivers with no sources within this radius are marked as outside the radius"),
        ("no produisent pas de level acústico útil", "do not produce a useful acoustic level"),
        ("Calculation by path source–receiver", "Calculation by source–receiver path"),
        ("para each wind turbine contribuyente", "for each contributing wind turbine"),
        ("la distance 3D", "the 3D distance"),
        ("los cotas acústicas", "the acoustic elevations"),
        ("posible screening topographic de la path se calculan", "the possible topographic screening of the path is calculated"),
        ("Propagación por band", "Band-wise propagation"),
        ("en cada band", "in each band"),
        ("se aplica", "is applied"),
        ("Adiv depende de la distance", "Adiv depends on distance"),
        ("Aatm,b de la frecuencia/de atmósfera", "Aatm,b on frequency/atmosphere"),
        ("Agr,b del sol", "Agr,b on ground"),
        ("Abar,b del MDT", "Abar,b on the DTM"),
        ("si existe un obstáculo relevante", "if a relevant obstacle exists"),
        ("Suma por source", "Source summation"),
        ("los 8 bandes son ponderadas A después sumadas energéticamente", "the 8 bands are A-weighted and then summed energetically"),
        ("para obtener el level ponderado A de cette wind turbine al receiver", "to obtain the A-weighted level of that wind turbine at the receiver"),
        ("Suma del receiver", "Receiver summation"),
        ("todos los wind turbines contribuyentes son sumadas energéticamente", "all contributing wind turbines are summed energetically"),
        ("para obtener el niveen total dB(A) del receiver", "to obtain the receiver total level in dB(A)"),
        ("Comparación con los limits", "Comparison with the limits"),
        ("el level total se compara à la limit asignado al receiver o à la limit de referencia", "the total level is compared with the limit assigned to the receiver or the reference limit"),
        ("El margin, status compliance y el table de los excedencias en découlent", "The margin, compliance status and exceedance table follow from this"),
        ("Este motor trabaja en 8 octave bands (63–8000 Hz). Los bandes no son pas un résultat del calculation, pero la malla frecuencial de la metodología.", "This engine works in 8 octave bands (63–8000 Hz). The bands are not a calculation result, but the frequency grid of the methodology."),
        ("Este motor trabaja en 8 octave bands", "This engine works in 8 octave bands"),
        ("Los bandes no son pas un résultat del calculation", "The bands are not a calculation result"),
        ("la malla frecuencial de la metodología", "the frequency grid of the methodology"),
        ("Para appliquer la propagation por bands", "To apply band-wise propagation"),
        ("el calculation necesita una entrada acústica por band de la source Lw,b", "the calculation needs a per-band acoustic input from the source Lw,b"),
        ("Esta entrée puede proceder de un spectrum medido/importado o de un plantilla/fallback ajustado al level global operativo", "This input may come from a measured/imported spectrum or from a template/fallback adjusted to the global operating level"),
        ("Acoustic source: Lw,b by octave bands. Si existe un spectrum específico del grupo source, esa es la entrada usada. Si no, el plugin usa una biblioteca/un gabarit/un fallback y lo ajusta al level global operativo.", "Acoustic source: Lw,b by octave bands. If a specific source-group spectrum exists, that is the input used. Otherwise, the plugin uses a library/template/fallback and adjusts it to the global operating level."),
        ("Este level global no sustituye las bands", "This global level does not replace the bands"),
        ("este fija el estado operativo y el spectrum aporta la distribución frecuencial", "it sets the operating state and the spectrum provides the frequency distribution"),
        ("MDT/DSM optional. Este solo afecta a el calculation de Abar,b", "optional DTM/DSM. This only affects the calculation of Abar,b"),
        ("Entrada acústica por bands", "Acoustic input by bands"),
        ("Procede del spectrum del grupo source", "It comes from the source-group spectrum"),
        ("La curva acústica LwA(ws) o el LwA fijo define el level global operativo del wind turbine", "The LwA(ws) acoustic curve or fixed LwA defines the wind turbine global operating level"),
        ("el spectrum por bands reparte ese level entre las 8 bands", "the band spectrum distributes that level across the 8 bands"),
        ("Calculado a partir de la 3D source-receiver distance", "Calculated from the 3D source–receiver distance"),
        ("Calculado por band con una tabla base de absorción", "Calculated by band with a base absorption table"),
        ("correcciones simplificadas de temperatura, humedad relativa y presión", "simplified temperature, relative humidity and pressure corrections"),
        ("La implementación actual usa la formulación exacta del plugin", "The current implementation uses the exact plugin formulation"),
        ("Calculado como efecto del ground by regiones", "Calculated as ground effect by regions"),
        ("El parámetro de ground utilisé es un G único por path", "The ground parameter used is a single G per path"),
        ("Only interviene que if existe un MDT/DSM", "Only applies if a DTM/DSM exists"),
        ("si un screening topographic se detecta", "if topographic screening is detected"),
        ("En ausencia de MDT o de obstáculo relevante", "In the absence of a DTM or relevant obstacle"),
        ("Los octave bands", "The octave bands"),
        ("no son pas un résultat ISO ni un table calculado by el plugin", "are not an ISO result or a table calculated by the plugin"),
        ("son la malla frecuencial sur laquelle la propagation es resuelta", "they are the frequency grid on which propagation is solved"),
        ("El plugin combina dos elementos", "The plugin combines two elements"),
        ("Global acoustic curve LwA(ws)", "Global acoustic curve LwA(ws)"),
        ("fixe el level operativo global del wind turbine", "sets the wind turbine global operating level"),
        ("Spectre por bands Lw,b", "Band spectrum Lw,b"),
        ("reparte este level global entre los 8 bandes", "distributes this global level across the 8 bands"),
        ("constituye entrada real utilisée en ecuación por bands", "constitutes the real input used in the band equation"),
        ("Este spectrum puede proceder de un fichier específico del fabricante/de usuario o de un gabarit de referencia", "This spectrum may come from a manufacturer/user file or from a reference template"),
        ("Autrement dit", "In other words"),
        ("la curva acoustique aporta el level global operativo y el gabarit/la biblioteca aporta la forma espectral", "the acoustic curve provides the global operating level and the template/library provides the spectral shape"),
        ("el spectrum reconstruido reproduise exactamente el LwA_cible de la curva importée", "the reconstructed spectrum exactly reproduces the LwA_cible of the imported curve"),
        ("Qué representa cada columna", "What each column represents"),
        ("Lw,b finale", "Final Lw,b"),
        ("Δ aplicado", "Applied Δ"),
        ("Este desplazamiento aumenta o reduce toda la forma espectral", "This shift increases or decreases the entire spectral shape"),
        ("para que su suma ponderada A reproduzca el LwA_cible", "so that its A-weighted sum reproduces the LwA_cible"),
        ("Interpretación", "Interpretation"),
        ("el espectro final Lw,b es el que entra realmente en la ecuación por bands", "the final Lw,b spectrum is what actually enters the band equation"),
        ("Aatm,b = α(f, T, RH, P) · d", "Aatm,b = α(f, T, RH, P) · d"),
        ("atmospheric absorption es calculadoe por band", "atmospheric absorption is calculated by band"),
        ("coefficient de referencia", "reference coefficient"),
        ("trois factores correctores", "three correction factors"),
        ("dependencia física", "physical dependence"),
        ("humedad relativa", "relative humidity"),
        ("presión", "pressure"),
        ("aproximación simplificadae del plugin", "simplified plugin approximation"),
        ("Interprétation de los corrections", "Interpretation of the corrections"),
        ("es introduit en °C respecto a una referencia", "is entered in °C relative to a reference"),
        ("se comparae à una humedad optimale", "is compared with an optimum humidity"),
        ("la correction augmente cuando on s’en éloigne", "the correction increases as it moves away from it"),
        ("es introduite en kPa respecto a una referencia", "is entered in kPa relative to a reference"),
        ("correction inverse", "inverse correction"),
        ("Ces factores no modifican que el bloc atmosphérique", "These factors modify only the atmospheric block"),
        ("ni emisión del wind turbine, ni efecto del ground, ni el término MDT/écran", "neither the wind turbine emission, nor the ground effect, nor the DTM/screening term"),
        ("se descompone", "is decomposed"),
        ("región de source", "source region"),
        ("región intermedia", "middle region"),
        ("région del receiver", "receiver region"),
        ("En esta implementación", "In this implementation"),
        ("trois paramètres de sol indépendants", "three independent ground parameters"),
        ("no se usan", "are not used"),
        ("un G único por path es utilisé", "a single G is used per path"),
        ("Matemáticamente", "Mathematically"),
        ("altura característica de la source", "characteristic source height"),
        ("celle del receiver", "that of the receiver"),
        ("altura media de la path", "mean path height"),
        ("condiciones favorables de propagation", "favourable propagation conditions"),
        ("Este single value de sol puede ser", "This single ground value may be"),
        ("G manual/global", "manual/global G"),
        ("si usuario fixe una single value", "if the user sets a single value"),
        ("media ponderada by la longitud de la path", "length-weighted average along the path"),
        ("Meaning físico de G", "Physical meaning of G"),
        ("representa el carácter acústico del terreno", "represents the acoustic character of the terrain"),
        ("controla influencia del sol sur la propagation", "controls the influence of ground on propagation"),
        ("indique un sol dur", "indicates hard ground"),
        ("un sol blando/poroso", "soft/porous ground"),
        ("terreno mixte", "mixed terrain"),
        ("Qué significa", "What it means"),
        ("desde capa", "from layer"),
        ("intersecta el path source–receiver", "intersects the source–receiver path"),
        ("asigna una value G_i à each polígono interceptado", "assigns a G_i value to each intercepted polygon"),
        ("calcule un G_eff unique para este trajet", "calculates a unique G_eff for this path"),
        ("C’es cette value that entra realmente en Agr,b", "This is the value that actually enters Agr,b"),
        ("el G global mostrado en el informe reste uniquement una valor de respaldo", "the global G shown in the report remains only a fallback value"),
        ("Convención del informe", "Report convention"),
        ("es mostrado ici as una magnitud positiva", "is shown here as a positive magnitude"),
        ("En ecuación principale", "In the main equation"),
        ("se resta al level de source", "it is subtracted from the source level"),
        ("El MDT no modifica emisión del wind turbine ni atmospheric absorption", "The DTM does not modify wind turbine emission or atmospheric absorption"),
        ("describir la geometría real de la path", "describe the real geometry of the path"),
        ("alimenter el término Abar,b", "feed the Abar,b term"),
        ("Profil del terreno", "Terrain profile"),
        ("es extrait del MDT", "is extracted from the DTM"),
        ("Línea de visión", "Line of sight"),
        ("Si el terreno queda siempre por debajo", "If the terrain always remains below it"),
        ("Obstacle dominant", "Dominant obstacle"),
        ("si una colina o una cresta exceeds", "if a hill or ridge protrudes above it"),
        ("Cuando h_obs > 0", "When h_obs > 0"),
        ("el relieve coupe la visión directa", "the terrain cuts the direct line of sight"),
        ("una attenuation adicional by diffraction puede aparecer", "additional attenuation by diffraction may appear"),
        ("Geometry réelle de obstacle", "Actual obstacle geometry"),
        ("posición réelle", "actual position"),
        ("Activación conservadora", "Conservative activation"),
        ("pequeñas irregularidades del MDT", "small DTM irregularities"),
        ("umbral mínimo", "minimum threshold"),
        ("resolución del raster", "raster resolution"),
        ("Difracción type Fresnel", "Fresnel-type diffraction"),
        ("diferencia de caminos", "path difference"),
        ("nombre de Fresnel", "Fresnel number"),
        ("Ese número se transforma después", "This number is then transformed"),
        ("dependiente de la frecuencia", "frequency-dependent"),
        ("En implémentation actuelle", "In the current implementation"),
        ("valores razonables", "reasonable values"),
        ("tope superior", "upper cap"),
        ("sobreattenuations parásitas", "spurious over-attenuation"),
        ("alors Abar,b = 0", "then Abar,b = 0"),
        ("A-weighting used at the ende", "A-weighting used at the end"),
        ("Lecture del receiver critique", "Reading the critical receiver"),
        ("no doit pas être interprété as una soustraction directe", "must not be interpreted as a direct subtraction"),
        ("una seule wind turbine", "a single wind turbine"),
        ("este es obtenu by energy summation por bands", "it is obtained by energy summation over bands"),
        ("by sommation de los sources contribuyentes dentro del radio de cálculo", "by summing the contributing sources within the calculation radius"),
        ("En este calculation", "In this calculation"),
        ("efecto del ground a été calculado", "the ground effect was calculated"),
        ("con un G manual unique para tout el trajet", "with a single manual G value for the whole path"),
        ("Este valor se aplica", "This value is applied"),
        ("No se ha derivado ningún G_eff", "No G_eff has been derived"),
        ("ningún MDT/DSM se ha usado", "no DTM/DSM was used"),
        ("por tanto el término de screening topographic se fija à", "therefore the topographic-screening term is set to"),
        ("La evaluación se realiza", "The assessment is performed"),
        ("sin introducir pantallas topográficas", "without introducing topographic barriers"),
        ("se resuelve sin perfil de terreno", "is solved without a terrain profile"),
        ("el cálculo depende de", "the calculation depends on"),
        ("Aatm simplificada", "simplified Aatm"),
        ("directividad Dc asumida igual à 0 dB", "directivity Dc assumed equal to 0 dB"),
        ("correction meteorológica de largo plazo", "long-term meteorological correction"),
        ("comparaciones e iteración de diseño", "comparisons and design iteration"),
        ("múltiples", "multiple"),
        ("soportados mediante", "supported through"),
        ("capas/grupos source independientes", "independent source layers/groups"),
        ("Mezclar varios modelos en una sola layer", "Mixing several models in a single layer"),
        ("mediante atributos no está activado", "through attributes is not enabled"),
        ("usa la misma lógica", "uses the same logic"),
        ("que los receivers puntuales", "as point receivers"),
        ("puede ser costoso en mapas grandes", "can be costly on large maps"),
        ("Para estudios regulatorios críticos", "For critical regulatory studies"),
        ("validar con mediciones o con software comercial certificado", "validate with measurements or certified commercial software"),
    ]
    for a, b in extra_repl:
        html = html.replace(a, b)

    extra_regex = [
        (r"\bpara\b", "for"),
        (r"\bpor\b", "by"),
        (r"\bdel\b", "of the"),
        (r"\bde los\b", "of the"),
        (r"\bde las\b", "of the"),
        (r"\bde la\b", "of the"),
        (r"\bel\b", "the"),
        (r"\bla\b", "the"),
        (r"\blos\b", "the"),
        (r"\blas\b", "the"),
        (r"\buna\b", "a"),
        (r"\bun\b", "a"),
        (r"\beste\b", "this"),
        (r"\besta\b", "this"),
        (r"\bEste\b", "This"),
        (r"\bEsta\b", "This"),
        (r"\bes\b", "is"),
        (r"\bson\b", "are"),
        (r"\bse\b", "is"),
        (r"\bsi\b", "if"),
        (r"\bexiste\b", "exists"),
        (r"\bestá\b", "is"),
        (r"\busado\b", "used"),
        (r"\busada\b", "used"),
        (r"\busados\b", "used"),
        (r"\busadas\b", "used"),
        (r"\bcalculadoe\b", "calculated"),
        (r"\bcalculado\b", "calculated"),
        (r"\bcalculada\b", "calculated"),
        (r"\bcalculan\b", "are calculated"),
        (r"\bcalculae\b", "calculates"),
        (r"\bcalcule\b", "calculates"),
        (r"\baplicado\b", "applied"),
        (r"\baplicada\b", "applied"),
        (r"\baplica\b", "applies"),
        (r"\butilizado\b", "used"),
        (r"\butilizada\b", "used"),
        (r"\butilisées\b", "used"),
        (r"\butilisé\b", "used"),
        (r"\bdisponible\b", "available"),
        (r"\bresultado\b", "result"),
        (r"\binforme\b", "report"),
        (r"\bmodelo\b", "model"),
        (r"\bgrupo\b", "group"),
        (r"\bvalor\b", "value"),
        (r"\bvalores\b", "values"),
        (r"\btabla\b", "table"),
        (r"\bcurva\b", "curve"),
        (r"\bplantilla\b", "template"),
        (r"\bbiblioteca\b", "library"),
        (r"\bforma espectral\b", "spectral shape"),
        (r"\bhumedad\b", "humidity"),
        (r"\btemperatura\b", "temperature"),
        (r"\bpresión\b", "pressure"),
        (r"\batmósfera\b", "atmosphere"),
        (r"\bterreno\b", "terrain"),
        (r"\baltura\b", "height"),
        (r"\bmedia\b", "mean"),
        (r"\bmáximo\b", "maximum"),
        (r"\bcrítico\b", "critical"),
        (r"\bcrítica\b", "critical"),
        (r"\bexactamente\b", "exactly"),
        (r"\breferencia\b", "reference"),
        (r"\bintermedia\b", "middle"),
        (r"\burbano\b", "urban"),
        (r"\basfalto\b", "asphalt"),
        (r"\bagricícola\b", "agricultural"),
        (r"\bpradera\b", "grassland"),
        (r"\bforestal\b", "forest"),
        (r"\bmixto\b", "mixed"),
        (r"\bmezcla\b", "mix"),
    ]
    for pat, repl_text in extra_regex:
        html = re.sub(pat, repl_text, html)

    html = html.replace("the the", "the")
    html = html.replace("of the the", "of the")
    html = html.replace("by el", "by the")
    html = html.replace("for el", "for the")
    html = html.replace("for the", "for the")
    html = html.replace("this calculation", "this calculation")
    html = html.replace("calculationado", "calculated")
    html = html.replace("calculationada", "calculated")
    html = html.replace("calculationan", "are calculated")
    html = html.replace("niveen", "level")
    html = html.replace("finale", "final")
    html = html.replace("critique", "critical")
    html = html.replace("réelle", "real")
    html = html.replace("résultat", "result")
    html = html.replace("gabarit", "template")
    html = html.replace("fichier", "file")
    html = html.replace("importée", "imported")
    html = html.replace("reproduise", "reproduces")
    html = html.replace("utilisée", "used")
    html = html.replace("utilisé", "used")
    html = html.replace("introduit", "entered")
    html = html.replace("introduite", "entered")
    html = html.replace("optimale", "optimum")
    html = html.replace("bloque atmosphérique", "atmospheric block")
    html = html.replace("écran", "screening")
    html = html.replace("visée", "sight")
    html = html.replace("ligne", "line")
    html = html.replace("liée", "linked")
    html = html.replace("lié", "linked")
    html = html.replace("résolution", "resolution")
    html = html.replace("traçabilité", "traceability")
    html = html.replace("soustraction", "subtraction")
    html = html.replace("depuis", "from")
    html = html.replace("jusqu", "until")
    html = html.replace("à ", "to ")
    html = html.replace("ISO-alined", "ISO-aligned")
    html = html.replace("alined with", "aligned with")

    # The two long explanatory sections were the main source of mixed language
    # in EN.  Until the whole template is authored natively in Spanish/English,
    # replace those long legacy blocks with stable English text.  Numeric result
    # sections, parameter tables and critical receiver values are preserved.
    clean_methodology = """
        <h2>2. HOW THE RESULT WAS GENERATED</h2>
        <div class='card card-info'>
            <h3>🧭 How the ISO-aligned calculation is executed</h3>
            <p>This section explains the workflow followed by the plugin so that receiver results remain traceable. The final receiver level is not obtained from a single direct subtraction. The plugin evaluates every source–receiver contribution within the calculation radius and then combines the contributions by energy summation.</p>
            <ol>
                <li><b>GIS inputs:</b> wind turbines or acoustic sources, receivers, receiver height, calculation radius, optional land-use layer and optional DTM/DSM are read from the QGIS project.</li>
                <li><b>Source acoustic state:</b> each source group receives an operational LwA value, either from a fixed value or from an LwA(ws) acoustic curve.</li>
                <li><b>Octave bands:</b> the ISO-aligned engine works with 8 octave bands. When no specific spectrum is available, the plugin builds one from a template/fallback and shifts it to reproduce the target LwA.</li>
                <li><b>Source selection:</b> for each receiver, only turbines inside the maximum calculation radius are considered. Receivers without sources inside the radius are marked as outside the radius.</li>
                <li><b>Path calculation:</b> for each source–receiver path, the plugin computes 3D distance, acoustic elevations, ground factor G/G_eff and, when available, topographic screening from the DTM/DSM.</li>
                <li><b>Band propagation:</b> each band is propagated with <code>Lp,b = Lw,b - Adiv - Aatm,b - Agr,b - Abar,b</code>.</li>
                <li><b>Energy summation:</b> octave bands are A-weighted and summed for each source, and all contributing sources are then summed to obtain the receiver total level in dB(A).</li>
                <li><b>Compliance check:</b> the total level is compared with the receiver limit or reference limit to compute the margin and compliance status.</li>
            </ol>
            <div class='formula'>{formula_receiver} = 10·log10(Σ_{sources_sub} 10^(LpA,{source_sub}/10))</div>
        </div>
        <div class='card'>
            <h3>🔎 Difference from Screening mode</h3>
            <p>The ISO-aligned mode is heavier but more traceable: it uses octave bands, final A-weighting, frequency-dependent atmospheric absorption, ground effect and optional DTM-based topographic screening. It is intended for preliminary technical reports and review of sensitive receivers.</p>
        </div>
    """
    html = re.sub(
        r"<h2>2\.\s*(?:HOW THE RESULT WAS GENERATED|CÓMO SE HA GENERADO EL RESULTADO|COMMENT LE RÉSULTAT A ÉTÉ GÉNÉRÉ)</h2>.*?<h2>3\.\s*(?:CRITICAL RECEIVER|RECEPTOR CRÍTICO|RÉCEPTEUR CRITIQUE)</h2>",
        clean_methodology + "\n        <h2>3. CRITICAL RECEIVER</h2>",
        html,
        flags=re.S,
    )

    clean_physics = """
        <h2>5. DETAILED PHYSICS AND CALCULATION TRACEABILITY</h2>
        <div class='card'>
            <h3>📖 Symbol glossary</h3>
            <table>
                <tr><th>Symbol</th><th>Meaning</th></tr>
                <tr><td><b>LwA</b></td><td>A-weighted sound power level of the source, in dB(A).</td></tr>
                <tr><td><b>Lw,b</b></td><td>Sound power level by octave band, in dB.</td></tr>
                <tr><td><b>LpA</b></td><td>A-weighted sound pressure level at the receiver, in dB(A).</td></tr>
                <tr><td><b>Adiv</b></td><td>Geometrical divergence attenuation due to distance.</td></tr>
                <tr><td><b>Aatm</b></td><td>Atmospheric absorption attenuation.</td></tr>
                <tr><td><b>Agr</b></td><td>Ground-effect attenuation.</td></tr>
                <tr><td><b>Abar</b></td><td>Topographic-screening attenuation when a DTM/DSM is used.</td></tr>
                <tr><td><b>G / G_eff</b></td><td>Ground factor, from hard ground near 0 to porous ground near 1.</td></tr>
            </table>
        </div>
        <div class='card'>
            <h3>📘 ISO-aligned propagation model</h3>
            <p>The ISO-aligned engine resolves propagation in octave bands from 63 Hz to 8000 Hz. The source spectrum may come from an imported/measured spectrum or from a template adjusted to the selected global LwA value.</p>
            <div class='formula'>Lp,b = Lw,b - Adiv - Aatm,b - Agr,b - Abar,b</div>
            <div class='formula'>LpA,total = 10·log10(Σ 10^((Lp,b + A_weight,b)/10))</div>
            <p><b>Geometrical divergence:</b> <code>Adiv = 20·log10(d) + 11</code>, using the 3D source–receiver distance.</p>
            <p><b>Atmospheric absorption:</b> calculated by octave band using a reference absorption table and simplified corrections for temperature, relative humidity and pressure.</p>
            <p><b>Ground effect:</b> calculated from a manual/global G value or from an effective G value derived from the land-use layer when available.</p>
            <p><b>Topographic screening:</b> when a DTM/DSM is available, the plugin checks the source–receiver path and estimates Abar for relevant terrain obstacles. Without DTM/DSM or without a relevant obstacle, Abar is set to 0 dB.</p>
            <p class='note'>The resulting receiver level includes multi-band and multi-source energy summation. It should not be interpreted as a direct subtraction from a single wind turbine.</p>
        </div>
    """
    html = re.sub(
        r"<h2>5\.\s*(?:DETAILED PHYSICS AND CALCULATION TRACEABILITY|FÍSICA DETALLADA Y TRAZABILIDAD DEL CÁLCULO|PHYSIQUE DÉTAILLÉE ET TRAÇABILITÉ DU CALCUL)</h2>.*?<h2>6\.\s*(?:ACOUSTIC SOURCE GROUPS|GRUPOS FUENTE ACÚSTICOS|GROUPES SOURCE ACOUSTIQUES)</h2>",
        clean_physics + "\n        <h2>6. ACOUSTIC SOURCE GROUPS</h2>",
        html,
        flags=re.S,
    )

    return html


def _noise_lang_code(lang: str) -> str:
    l = str(lang or '').lower()
    if l.startswith('de'):
        return 'de'
    if l.startswith('fr'):
        return 'fr'
    if l.startswith('en'):
        return 'en'
    return 'es'


def _noise_report_texts(code: str) -> Dict[str, str]:
    texts = {
        'es': {
            'title': 'INFORME TÉCNICO DE IMPACTO ACÚSTICO',
            'subtitle': 'Evaluación del ruido generado por aerogeneradores',
            'scope_title': 'Alcance de este informe — leer antes de usar los resultados',
            'what_is': 'Qué es: una evaluación acústica preliminar alineada con la metodología ISO 9613-2, orientada al diseño, la comparación de alternativas y el cribado de receptores sensibles.',
            'what_not': 'Qué no es: no es un informe acústico certificado y no sustituye a un estudio regulatorio definitivo realizado con software comercial validado.',
            'simplifications': 'Simplificaciones aplicadas en este modo:',
            's1': 'Absorción atmosférica Aatm mediante tabla de referencia con correcciones simplificadas de temperatura, humedad y presión; no es la formulación analítica completa de ISO 9613-1.',
            's2': 'Sin corrección meteorológica de largo plazo Cmet.',
            's3': 'Difracción topográfica de un único obstáculo dominante: sin difracción lateral ni pantallas múltiples.',
            's4': 'Resolución espectral en 8 bandas de octava de 63 a 8000 Hz, no en tercios de octava.',
            's5': 'Directividad de fuente Dc asumida igual a 0 dB.',
            'recommendation': 'Recomendación: para decisiones regulatorias críticas, valida los resultados con mediciones de campo o con software comercial certificado.',
            'exec': 'RESUMEN EJECUTIVO', 'turbines': 'Aerogeneradores', 'receivers': 'Receptores evaluados', 'max_level': 'Nivel máximo',
            'coverage': 'Cobertura del análisis', 'within': 'receptores dentro del radio', 'coverage_pct': 'cobertura', 'outside': 'receptores fuera del radio',
            'compliance': 'Cumplimiento regulatorio', 'exceed': 'receptores superan los límites', 'covered_compliance': 'cumplimiento sobre los receptores cubiertos', 'limit': 'Límite',
            'methodology': 'Metodología de cálculo', 'engine_used': 'Motor utilizado', 'source_groups': 'Grupos fuente acústicos', 'method': 'Método', 'raster_map': 'Mapa raster',
            'generated': 'CÓMO SE HA GENERADO EL RESULTADO', 'how_iso': 'Cómo se ha ejecutado el cálculo ISO-aligned',
            'flow_intro': 'Esta sección explica el flujo real seguido por el plugin para que el resultado por receptor sea trazable. El nivel final no procede de una resta única, sino de calcular todas las contribuciones fuente–receptor dentro del radio y sumarlas energéticamente.',
            'gis_inputs': 'Lectura de entradas SIG', 'gis_text': 'se leen aerogeneradores o fuentes acústicas, receptores, altura de receptor, radio máximo, capa de ocupación del suelo opcional y MDT/DSM opcional desde el proyecto QGIS.',
            'source_state': 'Estado acústico de la fuente', 'source_state_text': 'cada grupo fuente recibe un LwA operativo a partir de un valor fijo, de una curva acústica LwA(ws) o, si se ha importado un espectro OEM absoluto sin normalizar, de la suma ponderada A de ese espectro.',
            'bands': 'Bandas de octava', 'bands_text': 'el motor ISO-aligned trabaja con 8 bandas de octava. El espectro de cada grupo se resuelve por prioridad: valores introducidos a mano en la interfaz, CSV OEM importado, biblioteca de espectros y, en último término, una plantilla desplazada para reproducir el LwA objetivo.',
            'selection': 'Selección de fuentes', 'selection_text': 'para cada receptor se consideran solo los aerogeneradores situados dentro del radio máximo. Los receptores sin fuentes dentro del radio se marcan como fuera de radio.',
            'path': 'Cálculo de trayectoria', 'path_text': 'para cada trayectoria fuente–receptor se calcula la distancia 3D, las cotas acústicas, el factor de suelo G/G_eff y, si existe, el apantallamiento topográfico del MDT/DSM.',
            'propagation': 'Propagación por banda', 'propagation_text': 'cada banda se propaga con Lp,b = Lw,b - Adiv - Aatm,b - Agr,b - Abar,b.',
            'energy': 'Suma energética', 'energy_text': 'las bandas se ponderan A y se suman por fuente; después se suman todas las fuentes contribuyentes para obtener el nivel total del receptor en dB(A).',
            'check': 'Comprobación de límites', 'check_text': 'el nivel total se compara con el límite del receptor o con el límite de referencia para calcular margen y estado de cumplimiento.',
            'screening_diff': 'Diferencia respecto al modo Screening', 'screening_text': 'El modo ISO-aligned es más pesado pero más trazable: usa bandas de octava, ponderación A final, absorción atmosférica dependiente de la frecuencia, efecto de suelo y apantallamiento topográfico opcional basado en MDT/DSM.',
            'critical': 'RECEPTOR CRÍTICO', 'critical_title': 'Receptor crítico (mayor nivel sonoro)', 'receiver_id': 'ID receptor', 'total_level': 'Nivel total', 'applicable_limit': 'Límite aplicable', 'margin': 'Margen', 'dominant_model': 'Modelo dominante', 'source_group': 'Grupo fuente', 'contributors': 'Aerogeneradores contribuyentes dentro del radio', 'distance': 'Distancia', 'pass': 'CUMPLE', 'fail': 'SUPERA',
            'atten_break': 'Desglose de atenuaciones', 'atten_note': 'Los valores mostrados son magnitudes de atenuación. En la ecuación principal, estos términos se restan al nivel de fuente.', 'term': 'Término', 'value_db': 'Valor [dB]', 'description': 'Descripción', 'dom_lwa': 'LwA fuente dominante', 'sound_power': 'Potencia acústica del aerogenerador', 'geo_spread': 'Dispersión geométrica', 'air_abs': 'Absorción en el aire', 'ground_att': 'Atenuación debida al efecto del suelo', 'dtm_att': 'Atenuación debida al MDT en la trayectoria dominante', 'max_abar': 'Abar máximo entre contribuyentes', 'max_abar_desc': 'Abar máximo entre todos los aerogeneradores que contribuyen al receptor', 'ew_abar': 'Abar ponderado por energía', 'ew_desc': 'Media ponderada por la contribución acústica de cada aerogenerador', 'screened_paths': 'Trayectorias apantalladas', 'screened_desc': 'Número de aerogeneradores contribuyentes con Abar > 0 dB', 'result_level': 'NIVEL RESULTANTE', 'crit_note': 'El nivel resultante incluye suma energética multi-fuente y multi-banda; no es una resta directa desde un único aerogenerador.', 'dominant_band': 'Banda dominante', 'spectrum_origin': 'Origen del espectro',
            'stats': 'Estadísticos de atenuaciones (receptores cubiertos)', 'mean': 'Media [dB]', 'maximum': 'Máximo [dB]', 'ground_effect': 'efecto del suelo',
            'config': 'CONFIGURACIÓN Y PARÁMETROS', 'equation': 'Ecuación utilizada', 'interp_iso': 'Adiv representa la divergencia geométrica. Aatm se calcula por banda y depende de T, HR y presión mediante una formulación simplificada. Agr se aplica como término de suelo/terreno y Abar como apantallamiento topográfico básico cuando hay MDT disponible.', 'calc_params': 'Parámetros del cálculo', 'receiver_height': 'Altura del receptor', 'max_radius': 'Radio máximo', 'ground_mode': 'Modo suelo', 'global_ground': 'G utilizado', 'fallback_ground': 'G global de respaldo', 'mean_geff': 'G_eff medio', 'critical_geff': 'G_eff del receptor crítico', 'land_use': 'Ocupación del suelo', 'acoustic_scenario': 'Escenario acústico', 'temperature': 'Temperatura', 'humidity': 'Humedad relativa', 'pressure': 'Presión', 'review': 'Revisión recomendada', 'pressure_warning': 'la presión atmosférica introducida está fuera del rango típico usado como referencia en muchos estudios preliminares. Si no es una medición del emplazamiento, conviene comprobar si debería estar cerca de 101,325 kPa o ajustarse a la altitud.', 'paths_g': 'Trayectorias con G distinto del global', 'active_terms': 'Términos activos', 'active': 'activo', 'inactive': 'inactivo', 'yes': 'sí', 'no': 'no', 'simplified': 'simplificada', 'effective_g': 'G efectivo desde ocupación del suelo',
            'physics': 'FÍSICA DETALLADA Y TRAZABILIDAD DEL CÁLCULO', 'glossary': 'Glosario de símbolos', 'symbol': 'Símbolo', 'meaning': 'Significado', 'lwa_mean': 'Nivel de potencia acústica ponderado A de la fuente, en dB(A).', 'lwb_mean': 'Potencia acústica de la fuente por banda de octava, en dB.', 'lpa_mean': 'Nivel de presión acústica ponderado A en el receptor, en dB(A).', 'adiv_mean': 'Atenuación por divergencia geométrica debida a la distancia.', 'aatm_mean': 'Atenuación por absorción atmosférica.', 'agr_mean': 'Atenuación por efecto del suelo.', 'abar_mean': 'Atenuación por apantallamiento topográfico cuando se utiliza MDT/DSM.', 'g_mean': 'Factor de suelo, desde terreno duro próximo a 0 hasta terreno poroso próximo a 1.', 'model_title': 'Modelo de propagación ISO-aligned', 'model_text': 'El motor ISO-aligned resuelve la propagación en bandas de octava de 63 Hz a 8000 Hz. El espectro de fuente puede proceder de un espectro importado/medido o de una plantilla ajustada al LwA global seleccionado.', 'div_text': 'Divergencia geométrica: Adiv = 20·log10(d) + 11 usando la distancia 3D fuente–receptor.', 'atm_text': 'Absorción atmosférica: calculada por banda con tabla de referencia y correcciones simplificadas de temperatura, humedad relativa y presión.', 'ground_text': 'Efecto del suelo: calculado desde un G manual/global o desde un G_eff derivado de la capa de ocupación del suelo cuando está disponible.', 'screen_text': 'Apantallamiento topográfico: si hay MDT/DSM, el plugin revisa la trayectoria fuente–receptor y estima Abar para obstáculos relevantes. Sin MDT/DSM u obstáculo relevante, Abar se fija en 0 dB.',
            'source_section': 'GRUPOS FUENTE ACÚSTICOS', 'effective_lwa': 'LwA efectivo por grupo', 'model': 'modelo', 'park': 'parque', 'spectrum': 'espectro', 'not_available': 'No disponible',
            'receiver_dist': 'DISTRIBUCIÓN POR TIPO DE RECEPTOR', 'receivers_by_cat': 'Receptores por categoría', 'compliance_by_cat': 'Cumplimiento por categoría', 'exceed_limit': 'superan el límite', 'covered': 'cubiertos',
            'limits_recs': 'Límites y recomendaciones', 'fast_engine': 'Motor rápido: adecuado para screening preliminar y mapas ágiles.', 'iso_engine': 'Motor ISO-aligned: adecuado para estudios técnicos preliminares, comparaciones e iteración de diseño.', 'known_simp': 'Simplificaciones conocidas: Aatm simplificada; Agr y Abar con aproximaciones de base; directividad Dc asumida igual a 0 dB; Cmet no aplicada.', 'multi_models': 'Modelos múltiples: soportados mediante capas o grupos fuente independientes. Mezclar varios modelos en una sola capa mediante atributos no está activado en esta versión experimental.', 'iso_raster': 'Raster ISO + MDT: usa la misma lógica de apantallamiento topográfico que los receptores puntuales, pero puede ser costoso en mapas grandes.',
            'phys_title': 'Fundamentos físicos del modelo',
            'phys_db_title': 'La escala en decibelios',
            'phys_db_text': 'el decibelio es una escala logarítmica referida a la presión de 20 µPa (umbral de audición humano). Un aumento de +3 dB duplica la energía acústica y +10 dB se percibe aproximadamente como el doble de sonoridad. Por eso las atenuaciones y las sumas de fuentes no se comportan de forma lineal.',
            'phys_lwlp_title': 'Potencia acústica (Lw) frente a presión sonora (Lp)',
            'phys_lwlp_text': 'Lw describe la energía total que emite la fuente y es una propiedad del aerogenerador, independiente del entorno. Lp es lo que mediría un sonómetro en el receptor y depende de la distancia, la atmósfera, el suelo y los obstáculos. El modelo parte de Lw (dato de fabricante o curva LwA(ws)) y obtiene Lp restando las atenuaciones del camino de propagación.',
            'phys_div_title': 'Divergencia geométrica (Adiv)',
            'phys_div_text': 'una fuente puntual reparte su energía sobre una esfera de área 4·π·d²: al duplicar la distancia, la misma energía se distribuye sobre cuatro veces más superficie y el nivel cae 6 dB. El término +11 dB equivale a 10·log10(4·π). En las distancias típicas de un parque eólico es, con diferencia, la atenuación dominante.',
            'phys_atm_title': 'Absorción atmosférica (Aatm)',
            'phys_atm_text': 'el aire absorbe energía acústica por relajación molecular del oxígeno y del nitrógeno, un mecanismo que depende de la temperatura, la humedad relativa y la presión. La absorción crece fuertemente con la frecuencia: a gran distancia las bandas altas (2–8 kHz) prácticamente desaparecen y el ruido percibido se vuelve más grave. Esta dependencia espectral es el motivo de calcular por bandas de octava.',
            'phys_ground_title': 'Efecto del suelo (Agr)',
            'phys_ground_text': 'el sonido llega al receptor por el rayo directo y por la reflexión en el suelo, y ambos caminos interfieren entre sí. Un suelo duro (agua, asfalto, G≈0) refleja y puede incluso reforzar el nivel; un suelo poroso (hierba, cultivo, nieve, G≈1) absorbe parte de la energía. El resultado depende además de las alturas de fuente y receptor y de la distancia entre ambos.',
            'phys_bar_title': 'Apantallamiento topográfico (Abar)',
            'phys_bar_text': 'cuando el terreno interrumpe la línea de visión fuente–receptor, el sonido solo llega por difracción sobre el obstáculo. La atenuación crece con la diferencia entre el camino difractado y el directo (concepto del número de Fresnel) y es mayor a frecuencias altas, cuya longitud de onda es más corta y se difracta peor.',
            'phys_aw_title': 'Ponderación A',
            'phys_aw_text': 'el oído humano es menos sensible a las frecuencias bajas y muy altas. La curva de ponderación A corrige cada banda (por ejemplo −26,2 dB a 63 Hz y 0 dB a 1 kHz) para que el nivel total en dB(A) refleje la sonoridad percibida, que es la magnitud que regulan los límites normativos.',
            'phys_sum_title': 'Suma energética de fuentes',
            'phys_sum_text': 'los niveles en dB no se suman aritméticamente: se convierten a energía, se suman las energías y se vuelve a la escala logarítmica. Dos fuentes idénticas producen +3 dB, no el doble del nivel; diez fuentes idénticas, +10 dB. En la práctica el aerogenerador dominante (el más cercano o menos apantallado) controla el resultado y las fuentes lejanas apenas aportan.',
            'phys_example_title': 'Ejemplo orientativo',
            'phys_example_text': 'un aerogenerador con LwA = 105 dB(A) situado a 500 m sufre una divergencia Adiv = 20·log10(500) + 11 ≈ 65 dB, lo que deja unos 40 dB(A) antes de descontar la absorción atmosférica, el efecto del suelo y el apantallamiento. Este orden de magnitud explica por qué los receptores a varios cientos de metros suelen quedar entre 30 y 45 dB(A).',
            'phys_intro': 'Las ecuaciones de este apartado son exactamente las implementadas en el motor de cálculo del plugin, de modo que cualquier valor del informe puede reproducirse a mano a partir de ellas. Los símbolos están definidos en el glosario anterior.',
            'phys_impl_label': 'Implementación en el motor',
            'phys_spec_title': 'Construcción del espectro de fuente',
            'phys_spec_text': 'si el grupo no aporta un espectro medido, el motor parte de una plantilla relativa S_b con la forma típica de un aerogenerador moderno (máximo en bajas frecuencias) y le aplica un desplazamiento global Δ, calculado para que la suma ponderada A del espectro resultante reproduzca exactamente el LwA objetivo. El espectro final es así coherente con el dato de catálogo del fabricante. El espectro OEM del fabricante puede importarse (CSV) o teclearse en la interfaz: si es absoluto y no se normaliza, el LwA operativo del grupo se toma de su suma ponderada A, de modo que el ruido total se calcula a partir del espectro OEM; con la normalización activada o en modo curva, el espectro actúa solo como forma.',
            'phys_div_impl': 'se usa la distancia 3D fuente–receptor, incluyendo la diferencia de cotas acústicas entre el buje y el receptor, con una distancia mínima de cálculo configurable (25 m por defecto) que evita singularidades junto a la torre.',
            'phys_atm_impl': 'se parte de un coeficiente de referencia α_ref por banda, tabulado para 15 °C, 70 % de humedad y 101.325 kPa (tabla inferior), y se corrige con tres factores multiplicativos: +1 % por cada °C por encima de 15 °C, +0.3 % por cada punto porcentual de humedad que se aleja del óptimo del 50 %, y proporcionalidad inversa con la presión. La atenuación de la banda es α_b·d, lineal con la distancia, y por eso domina en las bandas altas a gran distancia.',
            'phys_ground_impl': 'siguiendo el esquema de regiones de ISO 9613-2, el trayecto se divide en región de fuente (los primeros min(30·h_s, d/3) metros), región media y región de receptor (los últimos min(30·h_r, d/3) metros). En cada región se evalúa un término base A_base dependiente de la banda y de la altura característica, se pondera por el factor G y se suman los tres. El total se limita a 10 dB; con G≈0 el término se anula y con G≈1 es máximo.',
            'phys_bar_impl': 'el motor muestrea el MDT a lo largo del trayecto con paso adaptativo (aproximadamente la resolución del ráster, mínimo 5 m, entre 50 y 1200 puntos), localiza el obstáculo dominante como el punto de mayor exceso h sobre la línea de visión y solo activa el término si ese exceso supera un umbral conservador de 1–3 m, para no reaccionar al ruido del propio MDT. Con la posición real del obstáculo (d₁, d₂) calcula la diferencia de caminos δ mediante la aproximación de pantalla delgada, el número de Fresnel C y la atenuación por tramos, limitada a 20 dB.',
            'phys_fast_title': 'Modo Screening (motor rápido)',
            'phys_fast_text': 'el motor rápido colapsa el cálculo espectral en una sola ecuación de banda ancha: usa directamente el LwA global, una absorción atmosférica lineal con coeficiente α fijo y un término empírico de suelo limitado a 6 dB, que crece con la distancia y se reduce con las alturas de buje y receptor. No aplica apantallamiento topográfico. Es coherente en tendencia con el modo ISO-aligned, pero menos fiable cerca de los límites normativos.',
            'phys_tbl_title': 'Constantes por banda utilizadas por el motor',
            'phys_tbl_template': 'S_b plantilla [dB]',
            'phys_example_extra': 'A esa misma distancia, la absorción atmosférica en la banda de 1 kHz añade aproximadamente 2 dB, mientras que en 8 kHz añadiría unos 45 dB: por eso el espectro recibido pierde las frecuencias altas.',
            'spec_used': 'Espectro utilizado por el grupo fuente',
            'spec_cols': 'S_b^ref es la forma espectral de referencia (si existe), A_b la ponderación A de cada banda y Lw,b el nivel final en dB que entra realmente en la ecuación por bandas.',
            'spec_delta_label': 'Δ aplicado',
            'spec_delta_text': 'desplazamiento constante que convierte la forma de referencia en el espectro final, Lw,b = S_b^ref + Δ; su valor absorbe el LwA objetivo y el ajuste de la ponderación A, por lo que es del orden del propio LwA.',
            'spec_lwa_check': 'Suma ponderada A del espectro utilizado',
            'band_hz': 'Banda [Hz]', 'spec_match': 'coincide con el LwA efectivo del grupo', 'spec_mismatch': 'difiere del LwA efectivo del grupo en',
        },
        'en': {
            'title': 'TECHNICAL NOISE IMPACT REPORT', 'subtitle': 'Assessment of wind turbine noise', 'scope_title': 'Scope of this report — read before using the results', 'what_is': 'What it is: a preliminary acoustic assessment aligned with ISO 9613-2 methodology, intended for design, alternative comparison and screening of sensitive receivers.', 'what_not': 'What it is not: it is not a certified acoustic report and does not replace a definitive regulatory study prepared with validated commercial software.', 'simplifications': 'Simplifications applied in this mode:', 's1': 'Atmospheric absorption Aatm using a reference table with simplified temperature, humidity and pressure corrections, not the complete analytical formulation of ISO 9613-1.', 's2': 'No long-term meteorological correction Cmet.', 's3': 'Topographic diffraction from a single dominant obstacle: no lateral diffraction or multiple barriers.', 's4': 'Spectral resolution in 8 octave bands from 63 to 8000 Hz, not in one-third octave bands.', 's5': 'Source directivity Dc assumed equal to 0 dB.', 'recommendation': 'Recommendation: for critical regulatory decisions, validate the results with field measurements or certified commercial software.', 'exec': 'EXECUTIVE SUMMARY', 'turbines': 'Wind turbines', 'receivers': 'Receivers evaluated', 'max_level': 'Maximum level', 'coverage': 'Analysis coverage', 'within': 'receivers within the radius', 'coverage_pct': 'coverage', 'outside': 'receivers outside the radius', 'compliance': 'Regulatory compliance', 'exceed': 'receivers exceed the limits', 'covered_compliance': 'compliance over covered receivers', 'limit': 'Limit', 'methodology': 'Calculation methodology', 'engine_used': 'Engine used', 'source_groups': 'Acoustic source groups', 'method': 'Method', 'raster_map': 'Raster map', 'generated': 'HOW THE RESULT WAS GENERATED', 'how_iso': 'How the ISO-aligned calculation is executed', 'flow_intro': 'This section explains the workflow followed by the plugin so that receiver results remain traceable. The final level is not obtained from a single direct subtraction; the plugin evaluates all source–receiver contributions within the calculation radius and combines them by energy summation.', 'gis_inputs': 'GIS inputs', 'gis_text': 'wind turbines or acoustic sources, receivers, receiver height, calculation radius, optional land-use layer and optional DTM/DSM are read from the QGIS project.', 'source_state': 'Source acoustic state', 'source_state_text': 'each source group receives an operational LwA value from a fixed value, from an LwA(ws) acoustic curve or, when an absolute OEM spectrum has been imported without normalization, from the A-weighted sum of that spectrum.', 'bands': 'Octave bands', 'bands_text': 'the ISO-aligned engine works with 8 octave bands. The spectrum of each group is resolved by priority: values typed manually in the interface, an imported OEM CSV, the spectrum library and, as a last resort, a template shifted to reproduce the target LwA.', 'selection': 'Source selection', 'selection_text': 'for each receiver, only turbines inside the maximum calculation radius are considered. Receivers without sources inside the radius are marked as outside the radius.', 'path': 'Path calculation', 'path_text': 'for each source–receiver path, the plugin computes 3D distance, acoustic elevations, ground factor G/G_eff and, when available, DTM/DSM topographic screening.', 'propagation': 'Band propagation', 'propagation_text': 'each band is propagated with Lp,b = Lw,b - Adiv - Aatm,b - Agr,b - Abar,b.', 'energy': 'Energy summation', 'energy_text': 'octave bands are A-weighted and summed for each source; all contributing sources are then summed to obtain the receiver total level in dB(A).', 'check': 'Compliance check', 'check_text': 'the total level is compared with the receiver limit or reference limit to compute the margin and compliance status.', 'screening_diff': 'Difference from Screening mode', 'screening_text': 'The ISO-aligned mode is heavier but more traceable: it uses octave bands, final A-weighting, frequency-dependent atmospheric absorption, ground effect and optional DTM-based topographic screening.', 'critical': 'CRITICAL RECEIVER', 'critical_title': 'Critical receiver (highest sound level)', 'receiver_id': 'Receiver ID', 'total_level': 'Total level', 'applicable_limit': 'Applicable limit', 'margin': 'Margin', 'dominant_model': 'Dominant model', 'source_group': 'Source group', 'contributors': 'Contributing wind turbines within the radius', 'distance': 'Distance', 'pass': 'PASS', 'fail': 'EXCEEDS', 'atten_break': 'Attenuation breakdown', 'atten_note': 'The values shown are attenuation magnitudes. In the main equation, these terms are subtracted from the source level.', 'term': 'Term', 'value_db': 'Value [dB]', 'description': 'Description', 'dom_lwa': 'Dominant source LwA', 'sound_power': 'Wind turbine sound power', 'geo_spread': 'Geometrical spreading', 'air_abs': 'Air absorption', 'ground_att': 'Attenuation due to ground effect', 'dtm_att': 'Attenuation due to the DTM along the dominant path', 'max_abar': 'Maximum Abar among contributors', 'max_abar_desc': 'Maximum Abar among all wind turbines contributing to the receiver', 'ew_abar': 'Energy-weighted Abar', 'ew_desc': 'Average weighted by the acoustic contribution of each wind turbine', 'screened_paths': 'Screened paths', 'screened_desc': 'Number of contributing wind turbines with Abar > 0 dB', 'result_level': 'RESULTING LEVEL', 'crit_note': 'The resulting level includes multi-source and multi-band energy summation; it is not a direct subtraction from a single wind turbine.', 'dominant_band': 'Dominant band', 'spectrum_origin': 'Spectrum source', 'stats': 'Attenuation statistics (covered receivers)', 'mean': 'Mean [dB]', 'maximum': 'Maximum [dB]', 'ground_effect': 'ground effect', 'config': 'CONFIGURATION AND PARAMETERS', 'equation': 'Equation used', 'interp_iso': 'Adiv represents geometrical divergence. Aatm is calculated by band and depends on T, RH and pressure through a simplified formulation. Agr is applied as the ground/terrain term and Abar as basic topographic screening when a DTM is available.', 'calc_params': 'Calculation parameters', 'receiver_height': 'Receiver height', 'max_radius': 'Maximum radius', 'ground_mode': 'Ground mode', 'global_ground': 'G used', 'fallback_ground': 'fallback global G', 'mean_geff': 'Mean G_eff used', 'critical_geff': 'Critical receiver G_eff used', 'land_use': 'Land use', 'acoustic_scenario': 'Acoustic scenario', 'temperature': 'Temperature', 'humidity': 'Relative humidity', 'pressure': 'Pressure', 'review': 'Recommended review', 'pressure_warning': 'the atmospheric pressure entered is outside the typical range used as reference in many preliminary studies. If it is not a site measurement, check whether it should be close to 101.325 kPa or adjusted to altitude.', 'paths_g': 'Paths with G different from the global value', 'active_terms': 'Active terms', 'active': 'active', 'inactive': 'inactive', 'yes': 'yes', 'no': 'no', 'simplified': 'simplified', 'effective_g': 'Effective G from land use', 'physics': 'DETAILED PHYSICS AND CALCULATION TRACEABILITY', 'glossary': 'Symbol glossary', 'symbol': 'Symbol', 'meaning': 'Meaning', 'lwa_mean': 'A-weighted sound power level of the source, in dB(A).', 'lwb_mean': 'Sound power level by octave band, in dB.', 'lpa_mean': 'A-weighted sound pressure level at the receiver, in dB(A).', 'adiv_mean': 'Geometrical divergence attenuation due to distance.', 'aatm_mean': 'Atmospheric absorption attenuation.', 'agr_mean': 'Ground-effect attenuation.', 'abar_mean': 'Topographic-screening attenuation when a DTM/DSM is used.', 'g_mean': 'Ground factor, from hard ground near 0 to porous ground near 1.', 'model_title': 'ISO-aligned propagation model', 'model_text': 'The ISO-aligned engine resolves propagation in octave bands from 63 Hz to 8000 Hz. The source spectrum may come from an imported/measured spectrum or from a template adjusted to the selected global LwA value.', 'div_text': 'Geometrical divergence: Adiv = 20·log10(d) + 11 using the 3D source–receiver distance.', 'atm_text': 'Atmospheric absorption: calculated by octave band using a reference absorption table and simplified corrections for temperature, relative humidity and pressure.', 'ground_text': 'Ground effect: calculated from a manual/global G value or from an effective G value derived from the land-use layer when available.', 'screen_text': 'Topographic screening: when a DTM/DSM is available, the plugin checks the source–receiver path and estimates Abar for relevant terrain obstacles. Without DTM/DSM or without a relevant obstacle, Abar is set to 0 dB.', 'source_section': 'ACOUSTIC SOURCE GROUPS', 'effective_lwa': 'Effective LwA by group', 'model': 'model', 'park': 'park', 'spectrum': 'spectrum', 'not_available': 'Not available', 'receiver_dist': 'DISTRIBUTION BY RECEIVER TYPE', 'receivers_by_cat': 'Receivers by category', 'compliance_by_cat': 'Compliance by category', 'exceed_limit': 'exceed the limit', 'covered': 'covered', 'limits_recs': 'Limits and recommendations', 'fast_engine': 'Fast engine: suitable for preliminary screening and quick maps.', 'iso_engine': 'ISO-aligned engine: suitable for preliminary technical studies, comparisons and design iteration.', 'known_simp': 'Known simplifications: simplified Aatm; Agr and Abar use basic approximations; source directivity Dc is assumed equal to 0 dB; Cmet is not applied.', 'multi_models': 'Multiple models: supported through independent layers or source groups. Mixing several models in a single layer through attributes is not enabled in this experimental version.', 'iso_raster': 'ISO raster + DEM: uses the same topographic-screening logic as point receivers, but can be costly on large maps.', 'phys_title': 'Physical foundations of the model', 'phys_db_title': 'The decibel scale', 'phys_db_text': 'the decibel is a logarithmic scale referenced to a pressure of 20 µPa (the human hearing threshold). An increase of +3 dB doubles the acoustic energy and +10 dB is perceived as roughly twice as loud. This is why attenuations and source additions do not behave linearly.', 'phys_lwlp_title': 'Sound power (Lw) versus sound pressure (Lp)', 'phys_lwlp_text': 'Lw describes the total energy emitted by the source and is a property of the wind turbine, independent of its surroundings. Lp is what a sound level meter would measure at the receiver and depends on distance, the atmosphere, the ground and obstacles. The model starts from Lw (manufacturer data or an LwA(ws) curve) and obtains Lp by subtracting the attenuations along the propagation path.', 'phys_div_title': 'Geometrical divergence (Adiv)', 'phys_div_text': 'a point source spreads its energy over a sphere of area 4·π·d²: doubling the distance spreads the same energy over four times the surface, so the level drops by 6 dB. The +11 dB term equals 10·log10(4·π). At typical wind-farm distances this is by far the dominant attenuation.', 'phys_atm_title': 'Atmospheric absorption (Aatm)', 'phys_atm_text': 'air absorbs acoustic energy through molecular relaxation of oxygen and nitrogen, a mechanism that depends on temperature, relative humidity and pressure. Absorption grows strongly with frequency: at long range the high bands (2–8 kHz) practically vanish and the perceived noise becomes duller. This spectral dependence is the reason for computing in octave bands.', 'phys_ground_title': 'Ground effect (Agr)', 'phys_ground_text': 'sound reaches the receiver via the direct ray and via a ground reflection, and both paths interfere with each other. Hard ground (water, asphalt, G≈0) reflects and may even reinforce the level; porous ground (grass, farmland, snow, G≈1) absorbs part of the energy. The outcome also depends on source and receiver heights and on the distance between them.', 'phys_bar_title': 'Topographic screening (Abar)', 'phys_bar_text': 'when the terrain interrupts the source–receiver line of sight, sound only arrives by diffraction over the obstacle. The attenuation grows with the difference between the diffracted and direct paths (the Fresnel-number concept) and is larger at high frequencies, whose shorter wavelengths diffract less easily.', 'phys_aw_title': 'A-weighting', 'phys_aw_text': 'the human ear is less sensitive to low and very high frequencies. The A-weighting curve corrects each band (for example −26.2 dB at 63 Hz and 0 dB at 1 kHz) so that the total dB(A) level reflects perceived loudness, which is the quantity regulated by statutory limits.', 'phys_sum_title': 'Energy summation of sources', 'phys_sum_text': 'levels in dB are not added arithmetically: they are converted to energy, the energies are summed and the result is converted back to the logarithmic scale. Two identical sources produce +3 dB, not double the level; ten identical sources, +10 dB. In practice the dominant turbine (the closest or least screened one) controls the result and distant sources contribute very little.', 'phys_example_title': 'Illustrative example', 'phys_example_text': 'a wind turbine with LwA = 105 dB(A) at 500 m undergoes a divergence Adiv = 20·log10(500) + 11 ≈ 65 dB, leaving about 40 dB(A) before subtracting atmospheric absorption, ground effect and screening. This order of magnitude explains why receivers several hundred metres away typically end up between 30 and 45 dB(A).', 'phys_intro': 'The equations in this section are exactly the ones implemented in the plugin calculation engine, so any value in the report can be reproduced by hand from them. The symbols are defined in the glossary above.', 'phys_impl_label': 'Engine implementation', 'phys_spec_title': 'Construction of the source spectrum', 'phys_spec_text': 'when a group provides no measured spectrum, the engine starts from a relative template S_b with the typical shape of a modern wind turbine (peaking at low frequencies) and applies a global shift Δ, computed so that the A-weighted sum of the resulting spectrum reproduces exactly the target LwA. The final spectrum is therefore consistent with the manufacturer catalogue value. The manufacturer OEM spectrum can be imported (CSV) or typed in the interface: if it is absolute and not normalized, the group operational LwA is taken from its A-weighted sum, so the total noise is calculated from the OEM spectrum; with normalization enabled or in curve mode, the spectrum acts as a shape only.', 'phys_div_impl': 'the 3D source–receiver distance is used, including the acoustic height difference between hub and receiver, with a configurable minimum calculation distance (25 m by default) that avoids singularities next to the tower.', 'phys_atm_impl': 'the engine starts from a per-band reference coefficient α_ref, tabulated for 15 °C, 70 % humidity and 101.325 kPa (table below), and corrects it with three multiplicative factors: +1 % per °C above 15 °C, +0.3 % per percentage point of humidity away from the 50 % optimum, and inverse proportionality with pressure. The band attenuation is α_b·d, linear with distance, which is why it dominates in the high bands at long range.', 'phys_ground_impl': 'following the ISO 9613-2 region scheme, the path is split into a source region (the first min(30·h_s, d/3) metres), a middle region and a receiver region (the last min(30·h_r, d/3) metres). In each region a base term A_base depending on the band and on the characteristic height is evaluated, weighted by the ground factor G, and the three are added. The total is capped at 10 dB; with G≈0 the term vanishes and with G≈1 it is maximal.', 'phys_bar_impl': 'the engine samples the DTM along the path with an adaptive step (roughly the raster resolution, minimum 5 m, between 50 and 1200 points), locates the dominant obstacle as the point of largest excess h above the line of sight and only activates the term when that excess exceeds a conservative 1–3 m threshold, to avoid reacting to DTM noise. Using the real obstacle position (d₁, d₂) it computes the path difference δ with the thin-screen approximation, the Fresnel number C and the piecewise attenuation, capped at 20 dB.', 'phys_fast_title': 'Screening mode (fast engine)', 'phys_fast_text': 'the fast engine collapses the spectral calculation into a single broadband equation: it uses the global LwA directly, a linear atmospheric absorption with a fixed coefficient α and an empirical ground term capped at 6 dB, which grows with distance and shrinks with hub and receiver heights. It applies no topographic screening. It is consistent in trend with the ISO-aligned mode but less reliable near statutory limits.', 'phys_tbl_title': 'Per-band constants used by the engine', 'phys_tbl_template': 'S_b template [dB]', 'phys_example_extra': 'At that same distance, atmospheric absorption in the 1 kHz band adds about 2 dB, while at 8 kHz it would add some 45 dB — which is why the received spectrum loses its high frequencies.', 'spec_used': 'Spectrum used by the source group', 'spec_cols': 'S_b^ref is the reference spectral shape (when it exists), A_b the A-weighting of each band and Lw,b the final level in dB that actually enters the band equation.', 'spec_delta_label': 'Δ applied', 'spec_delta_text': 'constant shift that turns the reference shape into the final spectrum, Lw,b = S_b^ref + Δ; its value absorbs the target LwA and the A-weighting adjustment, so it is of the order of the LwA itself.', 'spec_lwa_check': 'A-weighted sum of the spectrum used', 'band_hz': 'Band [Hz]', 'spec_match': 'matches the effective group LwA', 'spec_mismatch': 'differs from the effective group LwA by',
        },
        'fr': {
            'title': 'RAPPORT TECHNIQUE D’IMPACT ACOUSTIQUE', 'subtitle': 'Évaluation du bruit généré par les éoliennes', 'scope_title': 'Portée de ce rapport — à lire avant d’utiliser les résultats', 'what_is': 'Ce que c’est : une évaluation acoustique préliminaire alignée avec la méthodologie ISO 9613-2, destinée à la conception, à la comparaison d’alternatives et au criblage des récepteurs sensibles.', 'what_not': 'Ce que ce n’est pas : ce n’est pas un rapport acoustique certifié et il ne remplace pas une étude réglementaire définitive réalisée avec un logiciel commercial validé.', 'simplifications': 'Simplifications appliquées dans ce mode :', 's1': 'Absorption atmosphérique Aatm au moyen d’une table de référence avec corrections simplifiées de température, humidité et pression, et non de la formulation analytique complète de l’ISO 9613-1.', 's2': 'Sans correction météorologique de long terme Cmet.', 's3': 'Diffraction topographique d’un obstacle dominant unique : pas de diffraction latérale ni d’écrans multiples.', 's4': 'Résolution spectrale en 8 bandes d’octave de 63 à 8000 Hz, pas en tiers d’octave.', 's5': 'Directivité de source Dc supposée égale à 0 dB.', 'recommendation': 'Recommandation : pour les décisions réglementaires critiques, valider les résultats avec des mesures de terrain ou un logiciel commercial certifié.', 'exec': 'RÉSUMÉ EXÉCUTIF', 'turbines': 'Éoliennes', 'receivers': 'Récepteurs évalués', 'max_level': 'Niveau maximal', 'coverage': 'Couverture de l’analyse', 'within': 'récepteurs dans le rayon', 'coverage_pct': 'couverture', 'outside': 'récepteurs hors rayon', 'compliance': 'Conformité réglementaire', 'exceed': 'récepteurs dépassent les limites', 'covered_compliance': 'conformité sur les récepteurs couverts', 'limit': 'Limite', 'methodology': 'Méthodologie de calcul', 'engine_used': 'Moteur utilisé', 'source_groups': 'Groupes source acoustiques', 'method': 'Méthode', 'raster_map': 'Carte raster', 'generated': 'COMMENT LE RÉSULTAT A ÉTÉ GÉNÉRÉ', 'how_iso': 'Comment le calcul ISO-aligned a été exécuté', 'flow_intro': 'Cette section explique le flux suivi par le plugin afin que le résultat par récepteur soit traçable. Le niveau final ne provient pas d’une soustraction unique : le plugin évalue toutes les contributions source–récepteur dans le rayon de calcul et les combine par sommation énergétique.', 'gis_inputs': 'Entrées SIG', 'gis_text': 'les éoliennes ou sources acoustiques, les récepteurs, la hauteur de récepteur, le rayon de calcul, la couche d’occupation du sol optionnelle et le MDT/DSM optionnel sont lus depuis le projet QGIS.', 'source_state': 'État acoustique de la source', 'source_state_text': 'chaque groupe source reçoit une valeur LwA opérationnelle issue d’une valeur fixe, d’une courbe acoustique LwA(ws) ou, si un spectre OEM absolu a été importé sans normalisation, de la somme pondérée A de ce spectre.', 'bands': 'Bandes d’octave', 'bands_text': 'le moteur ISO-aligned travaille avec 8 bandes d’octave. Le spectre de chaque groupe est résolu par priorité : valeurs saisies à la main dans l’interface, CSV OEM importé, bibliothèque de spectres et, en dernier recours, un gabarit décalé pour reproduire le LwA cible.', 'selection': 'Sélection des sources', 'selection_text': 'pour chaque récepteur, seules les éoliennes dans le rayon maximal sont prises en compte. Les récepteurs sans source dans le rayon sont marqués hors rayon.', 'path': 'Calcul de trajet', 'path_text': 'pour chaque trajet source–récepteur, le plugin calcule la distance 3D, les altitudes acoustiques, le facteur de sol G/G_eff et, si disponible, l’écran topographique issu du MDT/DSM.', 'propagation': 'Propagation par bande', 'propagation_text': 'chaque bande est propagée avec Lp,b = Lw,b - Adiv - Aatm,b - Agr,b - Abar,b.', 'energy': 'Sommation énergétique', 'energy_text': 'les bandes d’octave sont pondérées A et sommées pour chaque source ; toutes les sources contributrices sont ensuite sommées pour obtenir le niveau total du récepteur en dB(A).', 'check': 'Vérification de conformité', 'check_text': 'le niveau total est comparé à la limite du récepteur ou à la limite de référence afin de calculer la marge et l’état de conformité.', 'screening_diff': 'Différence avec le mode Screening', 'screening_text': 'Le mode ISO-aligned est plus lourd mais plus traçable : il utilise les bandes d’octave, la pondération A finale, l’absorption atmosphérique dépendante de la fréquence, l’effet de sol et l’écran topographique optionnel basé sur MDT/DSM.', 'critical': 'RÉCEPTEUR CRITIQUE', 'critical_title': 'Récepteur critique (niveau sonore le plus élevé)', 'receiver_id': 'ID récepteur', 'total_level': 'Niveau total', 'applicable_limit': 'Limite applicable', 'margin': 'Marge', 'dominant_model': 'Modèle dominant', 'source_group': 'Groupe source', 'contributors': 'Éoliennes contributrices dans le rayon', 'distance': 'Distance', 'pass': 'CONFORME', 'fail': 'DÉPASSE', 'atten_break': 'Décomposition des atténuations', 'atten_note': 'Les valeurs affichées sont les amplitudes d’atténuation. Dans l’équation principale, ces termes sont soustraits au niveau de source.', 'term': 'Terme', 'value_db': 'Valeur [dB]', 'description': 'Description', 'dom_lwa': 'LwA source dominante', 'sound_power': 'Puissance acoustique de l’éolienne', 'geo_spread': 'Dispersion géométrique', 'air_abs': 'Absorption dans l’air', 'ground_att': 'Atténuation due à l’effet de sol', 'dtm_att': 'Atténuation due au MDT sur le trajet dominant', 'max_abar': 'Abar maximal des contributeurs', 'max_abar_desc': 'Abar maximal parmi toutes les éoliennes qui contribuent au récepteur', 'ew_abar': 'Abar pondéré par énergie', 'ew_desc': 'Moyenne pondérée par la contribution acoustique de chaque éolienne', 'screened_paths': 'Trajets écrantés', 'screened_desc': 'Nombre d’éoliennes contributrices avec Abar > 0 dB', 'result_level': 'NIVEAU RÉSULTANT', 'crit_note': 'Le niveau résultant inclut une sommation énergétique multi-source et multi-bande ; ce n’est pas une soustraction directe depuis une seule éolienne.', 'dominant_band': 'Bande dominante', 'spectrum_origin': 'Origine du spectre', 'stats': 'Statistiques des atténuations (récepteurs couverts)', 'mean': 'Moyenne [dB]', 'maximum': 'Maximum [dB]', 'ground_effect': 'effet de sol', 'config': 'CONFIGURATION ET PARAMÈTRES', 'equation': 'Équation utilisée', 'interp_iso': 'Adiv représente la divergence géométrique. Aatm est calculé par bande et dépend de T, HR et de la pression au moyen d’une formulation simplifiée. Agr est appliqué comme terme de sol/terrain et Abar comme écran topographique de base lorsqu’un MDT est disponible.', 'calc_params': 'Paramètres du calcul', 'receiver_height': 'Hauteur du récepteur', 'max_radius': 'Rayon maximal', 'ground_mode': 'Mode sol', 'global_ground': 'G utilisé', 'fallback_ground': 'G global de secours', 'mean_geff': 'G_eff moyen', 'critical_geff': 'G_eff du récepteur critique', 'land_use': 'Occupation du sol', 'acoustic_scenario': 'Scénario acoustique', 'temperature': 'Température', 'humidity': 'Humidité relative', 'pressure': 'Pression', 'review': 'Révision recommandée', 'pressure_warning': 'la pression atmosphérique saisie est hors de la plage typique utilisée comme référence dans de nombreuses études préliminaires. Si ce n’est pas une mesure du site, vérifier si elle devrait être proche de 101,325 kPa ou ajustée à l’altitude.', 'paths_g': 'Trajets avec G différent du global', 'active_terms': 'Termes actifs', 'active': 'actif', 'inactive': 'inactif', 'yes': 'oui', 'no': 'non', 'simplified': 'simplifiée', 'effective_g': 'G effectif depuis l’occupation du sol', 'physics': 'PHYSIQUE DÉTAILLÉE ET TRAÇABILITÉ DU CALCUL', 'glossary': 'Glossaire des symboles', 'symbol': 'Symbole', 'meaning': 'Signification', 'lwa_mean': 'Niveau de puissance acoustique pondéré A de la source, en dB(A).', 'lwb_mean': 'Puissance acoustique de la source par bande d’octave, en dB.', 'lpa_mean': 'Niveau de pression acoustique pondéré A au récepteur, en dB(A).', 'adiv_mean': 'Atténuation par divergence géométrique due à la distance.', 'aatm_mean': 'Atténuation par absorption atmosphérique.', 'agr_mean': 'Atténuation par effet de sol.', 'abar_mean': 'Atténuation par écran topographique lorsqu’un MDT/DSM est utilisé.', 'g_mean': 'Facteur de sol, du sol dur proche de 0 au sol poreux proche de 1.', 'model_title': 'Modèle de propagation ISO-aligned', 'model_text': 'Le moteur ISO-aligned résout la propagation en bandes d’octave de 63 Hz à 8000 Hz. Le spectre de source peut provenir d’un spectre importé/mesuré ou d’un gabarit ajusté au LwA global sélectionné.', 'div_text': 'Divergence géométrique : Adiv = 20·log10(d) + 11 avec la distance 3D source–récepteur.', 'atm_text': 'Absorption atmosphérique : calculée par bande avec une table de référence et des corrections simplifiées de température, humidité relative et pression.', 'ground_text': 'Effet de sol : calculé à partir d’un G manuel/global ou d’un G_eff dérivé de la couche d’occupation du sol lorsqu’elle est disponible.', 'screen_text': 'Écran topographique : si un MDT/DSM est disponible, le plugin vérifie le trajet source–récepteur et estime Abar pour les obstacles pertinents. Sans MDT/DSM ou sans obstacle pertinent, Abar vaut 0 dB.', 'source_section': 'GROUPES SOURCE ACOUSTIQUES', 'effective_lwa': 'LwA effectif par groupe', 'model': 'modèle', 'park': 'parc', 'spectrum': 'spectre', 'not_available': 'Non disponible', 'receiver_dist': 'DISTRIBUTION PAR TYPE DE RÉCEPTEUR', 'receivers_by_cat': 'Récepteurs par catégorie', 'compliance_by_cat': 'Conformité par catégorie', 'exceed_limit': 'dépassent la limite', 'covered': 'couverts', 'limits_recs': 'Limites et recommandations', 'fast_engine': 'Moteur rapide : adapté au criblage préliminaire et aux cartes agiles.', 'iso_engine': 'Moteur ISO-aligned : adapté aux études techniques préliminaires, comparaisons et itérations de conception.', 'known_simp': 'Simplifications connues : Aatm simplifié ; Agr et Abar avec approximations de base ; directivité Dc supposée égale à 0 dB ; Cmet non appliquée.', 'multi_models': 'Modèles multiples : pris en charge au moyen de couches ou groupes source indépendants. Mélanger plusieurs modèles dans une seule couche via attributs n’est pas activé dans cette version expérimentale.', 'iso_raster': 'Raster ISO + MNT : utilise la même logique d’écran topographique que les récepteurs ponctuels, mais peut être coûteux sur de grandes cartes.', 'phys_title': 'Fondements physiques du modèle', 'phys_db_title': 'L’échelle en décibels', 'phys_db_text': 'le décibel est une échelle logarithmique référencée à une pression de 20 µPa (seuil d’audition humain). Une augmentation de +3 dB double l’énergie acoustique et +10 dB est perçu environ comme deux fois plus fort. C’est pourquoi les atténuations et les additions de sources ne se comportent pas de façon linéaire.', 'phys_lwlp_title': 'Puissance acoustique (Lw) et pression acoustique (Lp)', 'phys_lwlp_text': 'Lw décrit l’énergie totale émise par la source ; c’est une propriété de l’éolienne, indépendante de l’environnement. Lp est ce que mesurerait un sonomètre au récepteur et dépend de la distance, de l’atmosphère, du sol et des obstacles. Le modèle part de Lw (donnée constructeur ou courbe LwA(ws)) et obtient Lp en soustrayant les atténuations le long du trajet de propagation.', 'phys_div_title': 'Divergence géométrique (Adiv)', 'phys_div_text': 'une source ponctuelle répartit son énergie sur une sphère d’aire 4·π·d² : en doublant la distance, la même énergie se répartit sur quatre fois plus de surface et le niveau chute de 6 dB. Le terme +11 dB équivaut à 10·log10(4·π). Aux distances typiques d’un parc éolien, c’est de loin l’atténuation dominante.', 'phys_atm_title': 'Absorption atmosphérique (Aatm)', 'phys_atm_text': 'l’air absorbe l’énergie acoustique par relaxation moléculaire de l’oxygène et de l’azote, un mécanisme qui dépend de la température, de l’humidité relative et de la pression. L’absorption augmente fortement avec la fréquence : à grande distance, les bandes hautes (2–8 kHz) disparaissent pratiquement et le bruit perçu devient plus grave. Cette dépendance spectrale justifie le calcul par bandes d’octave.', 'phys_ground_title': 'Effet de sol (Agr)', 'phys_ground_text': 'le son atteint le récepteur par le rayon direct et par la réflexion sur le sol ; les deux trajets interfèrent. Un sol dur (eau, asphalte, G≈0) réfléchit et peut même renforcer le niveau ; un sol poreux (herbe, terres agricoles, neige, G≈1) absorbe une partie de l’énergie. Le résultat dépend aussi des hauteurs de la source et du récepteur et de la distance entre eux.', 'phys_bar_title': 'Écran topographique (Abar)', 'phys_bar_text': 'lorsque le terrain interrompt la ligne de visée source–récepteur, le son n’arrive que par diffraction au-dessus de l’obstacle. L’atténuation croît avec la différence entre le trajet diffracté et le trajet direct (concept du nombre de Fresnel) et elle est plus forte aux hautes fréquences, dont la longueur d’onde plus courte se diffracte moins bien.', 'phys_aw_title': 'Pondération A', 'phys_aw_text': 'l’oreille humaine est moins sensible aux basses et aux très hautes fréquences. La courbe de pondération A corrige chaque bande (par exemple −26,2 dB à 63 Hz et 0 dB à 1 kHz) pour que le niveau total en dB(A) reflète la sonie perçue, qui est la grandeur encadrée par les limites réglementaires.', 'phys_sum_title': 'Sommation énergétique des sources', 'phys_sum_text': 'les niveaux en dB ne s’additionnent pas arithmétiquement : ils sont convertis en énergie, les énergies sont sommées puis reconverties en échelle logarithmique. Deux sources identiques produisent +3 dB, et non le double du niveau ; dix sources identiques, +10 dB. En pratique, l’éolienne dominante (la plus proche ou la moins écrantée) contrôle le résultat et les sources lointaines contribuent très peu.', 'phys_example_title': 'Exemple indicatif', 'phys_example_text': 'une éolienne avec LwA = 105 dB(A) à 500 m subit une divergence Adiv = 20·log10(500) + 11 ≈ 65 dB, ce qui laisse environ 40 dB(A) avant de déduire l’absorption atmosphérique, l’effet de sol et l’écran topographique. Cet ordre de grandeur explique pourquoi les récepteurs situés à plusieurs centaines de mètres se retrouvent généralement entre 30 et 45 dB(A).', 'phys_intro': 'Les équations de cette section sont exactement celles implémentées dans le moteur de calcul du plugin ; toute valeur du rapport peut donc être reproduite à la main à partir d’elles. Les symboles sont définis dans le glossaire ci-dessus.', 'phys_impl_label': 'Implémentation dans le moteur', 'phys_spec_title': 'Construction du spectre de source', 'phys_spec_text': 'lorsqu’un groupe ne fournit pas de spectre mesuré, le moteur part d’un gabarit relatif S_b ayant la forme typique d’une éolienne moderne (maximum aux basses fréquences) et applique un décalage global Δ, calculé pour que la somme pondérée A du spectre résultant reproduise exactement le LwA cible. Le spectre final est ainsi cohérent avec la valeur catalogue du constructeur. Le spectre OEM du constructeur peut être importé (CSV) ou saisi dans l’interface : s’il est absolu et non normalisé, le LwA opérationnel du groupe est pris de sa somme pondérée A, de sorte que le bruit total est calculé à partir du spectre OEM ; avec la normalisation activée ou en mode courbe, le spectre n’agit que comme forme.', 'phys_div_impl': 'la distance 3D source–récepteur est utilisée, en incluant la différence de hauteurs acoustiques entre le moyeu et le récepteur, avec une distance minimale de calcul configurable (25 m par défaut) qui évite les singularités au pied de la tour.', 'phys_atm_impl': 'le moteur part d’un coefficient de référence α_ref par bande, tabulé pour 15 °C, 70 % d’humidité et 101,325 kPa (tableau ci-dessous), et le corrige par trois facteurs multiplicatifs : +1 % par °C au-dessus de 15 °C, +0,3 % par point d’humidité s’écartant de l’optimum de 50 %, et proportionnalité inverse avec la pression. L’atténuation de la bande vaut α_b·d, linéaire avec la distance, ce qui explique sa domination dans les bandes hautes à grande distance.', 'phys_ground_impl': 'selon le schéma de régions de l’ISO 9613-2, le trajet est divisé en région source (les premiers min(30·h_s, d/3) mètres), région médiane et région récepteur (les derniers min(30·h_r, d/3) mètres). Dans chaque région, un terme de base A_base dépendant de la bande et de la hauteur caractéristique est évalué, pondéré par le facteur de sol G, puis les trois sont additionnés. Le total est plafonné à 10 dB ; avec G≈0 le terme s’annule et avec G≈1 il est maximal.', 'phys_bar_impl': 'le moteur échantillonne le MNT le long du trajet avec un pas adaptatif (environ la résolution du raster, minimum 5 m, entre 50 et 1200 points), localise l’obstacle dominant comme le point de plus grand excès h au-dessus de la ligne de visée et n’active le terme que si cet excès dépasse un seuil conservateur de 1–3 m, afin de ne pas réagir au bruit du MNT lui-même. Avec la position réelle de l’obstacle (d₁, d₂), il calcule la différence de trajets δ par l’approximation de l’écran mince, le nombre de Fresnel C et l’atténuation par morceaux, plafonnée à 20 dB.', 'phys_fast_title': 'Mode Screening (moteur rapide)', 'phys_fast_text': 'le moteur rapide réduit le calcul spectral à une seule équation large bande : il utilise directement le LwA global, une absorption atmosphérique linéaire à coefficient α fixe et un terme de sol empirique plafonné à 6 dB, croissant avec la distance et décroissant avec les hauteurs du moyeu et du récepteur. Il n’applique aucun écran topographique. Il est cohérent en tendance avec le mode ISO-aligned, mais moins fiable près des limites réglementaires.', 'phys_tbl_title': 'Constantes par bande utilisées par le moteur', 'phys_tbl_template': 'S_b gabarit [dB]', 'phys_example_extra': 'À cette même distance, l’absorption atmosphérique dans la bande de 1 kHz ajoute environ 2 dB, alors qu’à 8 kHz elle ajouterait quelque 45 dB — c’est pourquoi le spectre reçu perd ses hautes fréquences.', 'spec_used': 'Spectre utilisé par le groupe source', 'spec_cols': 'S_b^ref est la forme spectrale de référence (si elle existe), A_b la pondération A de chaque bande et Lw,b le niveau final en dB qui entre réellement dans l’équation par bandes.', 'spec_delta_label': 'Δ appliqué', 'spec_delta_text': 'décalage constant qui transforme la forme de référence en spectre final, Lw,b = S_b^ref + Δ ; sa valeur absorbe le LwA cible et l’ajustement de la pondération A, elle est donc de l’ordre du LwA lui-même.', 'spec_lwa_check': 'Somme pondérée A du spectre utilisé', 'band_hz': 'Bande [Hz]', 'spec_match': 'coïncide avec le LwA effectif du groupe', 'spec_mismatch': 'diffère du LwA effectif du groupe de',
        },
        'de': {
            'title': 'TECHNISCHER BERICHT ZUR SCHALLIMMISSION', 'subtitle': 'Bewertung der durch Windenergieanlagen verursachten Geräusche', 'scope_title': 'Geltungsbereich dieses Berichts — vor Nutzung der Ergebnisse lesen', 'what_is': 'Was es ist: eine vorläufige akustische Bewertung nach der Methodik ISO 9613-2, vorgesehen für Entwurf, Variantenvergleich und Screening empfindlicher Rezeptoren.', 'what_not': 'Was es nicht ist: kein zertifizierter Schallbericht und kein Ersatz für eine endgültige behördliche Studie mit validierter kommerzieller Software.', 'simplifications': 'In diesem Modus angewendete Vereinfachungen:', 's1': 'Atmosphärische Absorption Aatm über eine Referenztabelle mit vereinfachten Korrekturen für Temperatur, Feuchte und Druck; nicht die vollständige analytische Formulierung der ISO 9613-1.', 's2': 'Keine langfristige meteorologische Korrektur Cmet.', 's3': 'Topografische Beugung an einem einzelnen dominanten Hindernis: keine seitliche Beugung und keine Mehrfachschirme.', 's4': 'Spektrale Auflösung in 8 Oktavbändern von 63 bis 8000 Hz, nicht in Terzbändern.', 's5': 'Quellrichtwirkung Dc wird mit 0 dB angesetzt.', 'recommendation': 'Empfehlung: Für kritische regulatorische Entscheidungen sollten die Ergebnisse mit Feldmessungen oder zertifizierter kommerzieller Software validiert werden.', 'exec': 'KURZFASSUNG', 'turbines': 'Windenergieanlagen', 'receivers': 'Bewertete Rezeptoren', 'max_level': 'Maximaler Pegel', 'coverage': 'Abdeckung der Analyse', 'within': 'Rezeptoren innerhalb des Radius', 'coverage_pct': 'Abdeckung', 'outside': 'Rezeptoren außerhalb des Radius', 'compliance': 'Regulatorische Konformität', 'exceed': 'Rezeptoren überschreiten die Grenzwerte', 'covered_compliance': 'Konformität der abgedeckten Rezeptoren', 'limit': 'Grenzwert', 'methodology': 'Berechnungsmethodik', 'engine_used': 'Verwendeter Rechenkern', 'source_groups': 'Akustische Quellgruppen', 'method': 'Methode', 'raster_map': 'Rasterkarte', 'generated': 'WIE DAS ERGEBNIS ERZEUGT WURDE', 'how_iso': 'Ablauf der ISO-aligned-Berechnung', 'flow_intro': 'Dieser Abschnitt beschreibt den vom Plugin verwendeten Ablauf, damit die Ergebnisse je Rezeptor nachvollziehbar bleiben. Der Endpegel entsteht nicht durch eine einzelne Subtraktion, sondern durch die Berechnung aller Quelle–Rezeptor-Beiträge innerhalb des Berechnungsradius und deren energetische Summierung.', 'gis_inputs': 'GIS-Eingaben', 'gis_text': 'Windenergieanlagen oder akustische Quellen, Rezeptoren, Rezeptorhöhe, Berechnungsradius sowie optionale Landnutzungs- und DGM/DOM-Layer werden aus dem QGIS-Projekt gelesen.', 'source_state': 'Akustischer Zustand der Quelle', 'source_state_text': 'jede Quellgruppe erhält einen betrieblichen LwA-Wert aus einem festen Wert, aus einer akustischen Kurve LwA(ws) oder, wenn ein absolutes OEM-Spektrum ohne Normierung importiert wurde, aus der A-bewerteten Summe dieses Spektrums.', 'bands': 'Oktavbänder', 'bands_text': 'der ISO-aligned-Rechenkern arbeitet mit 8 Oktavbändern. Das Spektrum jeder Gruppe wird nach Priorität aufgelöst: manuell in der Oberfläche eingegebene Werte, importierte OEM-CSV, Spektrenbibliothek und zuletzt eine Vorlage, die auf den Ziel-LwA verschoben wird.', 'selection': 'Quellenauswahl', 'selection_text': 'für jeden Rezeptor werden nur Windenergieanlagen innerhalb des maximalen Berechnungsradius berücksichtigt. Rezeptoren ohne Quellen innerhalb des Radius werden als außerhalb des Radius markiert.', 'path': 'Pfadberechnung', 'path_text': 'für jeden Quelle–Rezeptor-Pfad berechnet das Plugin 3D-Abstand, akustische Höhen, den Bodenfaktor G/G_eff und gegebenenfalls topografische Abschirmung aus DGM/DOM.', 'propagation': 'Ausbreitung je Band', 'propagation_text': 'jedes Band wird mit Lp,b = Lw,b - Adiv - Aatm,b - Agr,b - Abar,b berechnet.', 'energy': 'Energetische Summierung', 'energy_text': 'die Oktavbänder werden A-bewertet und je Quelle summiert; danach werden alle beitragenden Quellen zum Gesamtpegel des Rezeptors in dB(A) summiert.', 'check': 'Grenzwertprüfung', 'check_text': 'der Gesamtpegel wird mit dem Grenzwert des Rezeptors oder dem Referenzgrenzwert verglichen, um Abstand zum Grenzwert und Konformitätsstatus zu berechnen.', 'screening_diff': 'Unterschied zum Screening-Modus', 'screening_text': 'Der ISO-aligned-Modus ist rechenintensiver, aber nachvollziehbarer: Er nutzt Oktavbänder, abschließende A-Bewertung, frequenzabhängige atmosphärische Absorption, Bodeneffekt und optionale DGM-basierte topografische Abschirmung.', 'critical': 'KRITISCHER REZEPTOR', 'critical_title': 'Kritischer Rezeptor (höchster Schallpegel)', 'receiver_id': 'ID Rezeptor', 'total_level': 'Gesamtpegel', 'applicable_limit': 'Anwendbarer Grenzwert', 'margin': 'Abstand zum Grenzwert', 'dominant_model': 'Dominantes Modell', 'source_group': 'Quellgruppe', 'contributors': 'Beitragende Windenergieanlagen innerhalb des Radius', 'distance': 'Entfernung', 'pass': 'ERFÜLLT', 'fail': 'ÜBERSCHREITET', 'atten_break': 'Dämpfungsaufschlüsselung', 'atten_note': 'Die angezeigten Werte sind Dämpfungsbeträge. In der Hauptgleichung werden diese Terme vom Quellpegel abgezogen.', 'term': 'Term', 'value_db': 'Wert [dB]', 'description': 'Beschreibung', 'dom_lwa': 'LwA der dominanten Quelle', 'sound_power': 'Schallleistungspegel der Windenergieanlage', 'geo_spread': 'Geometrische Ausbreitung', 'air_abs': 'Luftabsorption', 'ground_att': 'Dämpfung durch Bodeneffekt', 'dtm_att': 'Dämpfung durch DGM entlang des dominanten Pfads', 'max_abar': 'Maximaler Abar der Beitragenden', 'max_abar_desc': 'Maximaler Abar-Wert unter allen Windenergieanlagen, die zum Rezeptor beitragen', 'ew_abar': 'Energetisch gewichteter Abar', 'ew_desc': 'Mittelwert, gewichtet nach akustischem Beitrag jeder Windenergieanlage', 'screened_paths': 'Abgeschirmte Pfade', 'screened_desc': 'Anzahl beitragender Windenergieanlagen mit Abar > 0 dB', 'result_level': 'ERGEBNISPEGEL', 'crit_note': 'Der Ergebnispegel enthält die energetische Summierung über mehrere Quellen und Frequenzbänder; dies ist keine direkte Subtraktion von einer einzelnen Windenergieanlage.', 'dominant_band': 'Dominantes Frequenzband', 'spectrum_origin': 'Spektrumquelle', 'stats': 'Dämpfungsstatistik (abgedeckte Rezeptoren)', 'mean': 'Mittel [dB]', 'maximum': 'Maximum [dB]', 'ground_effect': 'Bodeneffekt', 'config': 'KONFIGURATION UND PARAMETER', 'equation': 'Verwendete Gleichung', 'interp_iso': 'Adiv steht für geometrische Divergenz. Aatm wird je Band berechnet und hängt über eine vereinfachte Formulierung von T, relativer Feuchte und Druck ab. Agr wird als Boden-/Geländeterm angewendet und Abar als einfache topografische Abschirmung, wenn ein DGM verfügbar ist.', 'calc_params': 'Berechnungsparameter', 'receiver_height': 'Höhe des Rezeptors', 'max_radius': 'Maximaler Radius', 'ground_mode': 'Bodenmodus', 'global_ground': 'Verwendeter G-Wert', 'fallback_ground': 'Globaler Ersatz-G-Wert', 'mean_geff': 'Mittlerer G_eff-Wert', 'critical_geff': 'G_eff-Wert des kritischen Rezeptors', 'land_use': 'Landnutzung', 'acoustic_scenario': 'Akustisches Szenario', 'temperature': 'Temperatur', 'humidity': 'Relative Luftfeuchte', 'pressure': 'Druck', 'review': 'Empfohlene Prüfung', 'pressure_warning': 'der eingegebene atmosphärische Druck liegt außerhalb des typischen Referenzbereichs vieler Vorstudien. Falls es sich nicht um eine Standortmessung handelt, sollte geprüft werden, ob der Wert nahe 101,325 kPa liegen oder höhenkorrigiert werden sollte.', 'paths_g': 'Pfade mit G abweichend vom globalen Wert', 'active_terms': 'Aktive Terme', 'active': 'aktiv', 'inactive': 'inaktiv', 'yes': 'ja', 'no': 'nein', 'simplified': 'vereinfacht', 'effective_g': 'Effektiver G-Wert aus Landnutzung', 'physics': 'DETAILLIERTE PHYSIK UND BERECHNUNGSNACHVOLLZIEHBARKEIT', 'glossary': 'Symbolglossar', 'symbol': 'Symbol', 'meaning': 'Bedeutung', 'lwa_mean': 'A-bewerteter Schallleistungspegel der Quelle in dB(A).', 'lwb_mean': 'Schallleistungspegel je Oktavband in dB.', 'lpa_mean': 'A-bewerteter Schalldruckpegel am Rezeptor in dB(A).', 'adiv_mean': 'Dämpfung durch geometrische Divergenz infolge der Entfernung.', 'aatm_mean': 'Dämpfung durch atmosphärische Absorption.', 'agr_mean': 'Dämpfung durch Bodeneffekt.', 'abar_mean': 'Dämpfung durch topografische Abschirmung bei Verwendung eines DGM/DOM.', 'g_mean': 'Bodenfaktor: von hartem Boden nahe 0 bis zu porösem Boden nahe 1.', 'model_title': 'ISO-aligned-Ausbreitungsmodell', 'model_text': 'Der ISO-aligned-Rechenkern löst die Ausbreitung in Oktavbändern von 63 Hz bis 8000 Hz. Das Quellspektrum kann aus einem importierten/gemessenen Spektrum stammen oder aus einer auf den gewählten globalen LwA-Wert angepassten Vorlage.', 'div_text': 'Geometrische Divergenz: Adiv = 20·log10(d) + 11 mit dem 3D-Abstand Quelle–Rezeptor.', 'atm_text': 'Atmosphärische Absorption: je Oktavband mit einer Referenztabelle und vereinfachten Korrekturen für Temperatur, relative Feuchte und Druck berechnet.', 'ground_text': 'Bodeneffekt: berechnet aus einem manuellen/globalen G-Wert oder aus einem aus der Landnutzung abgeleiteten G_eff-Wert, sofern verfügbar.', 'screen_text': 'Topografische Abschirmung: Wenn ein DGM/DOM verfügbar ist, prüft das Plugin den Quelle–Rezeptor-Pfad und schätzt Abar für relevante Geländeobjekte. Ohne DGM/DOM oder relevantes Hindernis wird Abar auf 0 dB gesetzt.', 'source_section': 'AKUSTISCHE QUELLGRUPPEN', 'effective_lwa': 'Effektiver LwA je Gruppe', 'model': 'Modell', 'park': 'Park', 'spectrum': 'Spektrum', 'not_available': 'Nicht verfügbar', 'receiver_dist': 'VERTEILUNG NACH REZEPTORTYP', 'receivers_by_cat': 'Rezeptoren nach Kategorie', 'compliance_by_cat': 'Konformität nach Kategorie', 'exceed_limit': 'überschreiten den Grenzwert', 'covered': 'abgedeckt', 'limits_recs': 'Grenzwerte und Empfehlungen', 'fast_engine': 'Schneller Rechenkern: geeignet für vorläufiges Screening und schnelle Karten.', 'iso_engine': 'ISO-aligned-Rechenkern: geeignet für vorläufige technische Studien, Vergleiche und Entwurfsiterationen.', 'known_simp': 'Bekannte Vereinfachungen: vereinfachtes Aatm; Agr und Abar verwenden Grundnäherungen; Quellrichtwirkung Dc wird mit 0 dB angenommen; Cmet wird nicht angewendet.', 'multi_models': 'Mehrere Modelle: unterstützt über unabhängige Layer oder Quellgruppen. Das Mischen mehrerer Modelle in einem einzelnen Layer über Attribute ist in dieser experimentellen Version nicht aktiviert.', 'iso_raster': 'ISO-Raster + DGM: verwendet die gleiche Logik der topografischen Abschirmung wie Punktrezeptoren, kann bei großen Karten jedoch rechenintensiv sein.', 'phys_title': 'Physikalische Grundlagen des Modells', 'phys_db_title': 'Die Dezibel-Skala', 'phys_db_text': 'das Dezibel ist eine logarithmische Skala mit der Bezugsgröße 20 µPa (Hörschwelle des Menschen). Eine Erhöhung um +3 dB verdoppelt die Schallenergie, und +10 dB werden etwa als doppelt so laut wahrgenommen. Deshalb verhalten sich Dämpfungen und die Addition von Quellen nicht linear.', 'phys_lwlp_title': 'Schallleistung (Lw) und Schalldruck (Lp)', 'phys_lwlp_text': 'Lw beschreibt die insgesamt abgestrahlte Energie der Quelle und ist eine Eigenschaft der Windenergieanlage, unabhängig von der Umgebung. Lp ist der Wert, den ein Schallpegelmesser am Rezeptor messen würde; er hängt von Entfernung, Atmosphäre, Boden und Hindernissen ab. Das Modell geht von Lw aus (Herstellerangabe oder LwA(ws)-Kurve) und berechnet Lp durch Abzug der Dämpfungen entlang des Ausbreitungspfads.', 'phys_div_title': 'Geometrische Divergenz (Adiv)', 'phys_div_text': 'eine Punktquelle verteilt ihre Energie auf eine Kugel mit der Fläche 4·π·d²: Bei Verdopplung der Entfernung verteilt sich dieselbe Energie auf die vierfache Fläche, und der Pegel sinkt um 6 dB. Der Term +11 dB entspricht 10·log10(4·π). Bei typischen Windpark-Entfernungen ist dies mit Abstand die dominierende Dämpfung.', 'phys_atm_title': 'Atmosphärische Absorption (Aatm)', 'phys_atm_text': 'die Luft absorbiert Schallenergie durch molekulare Relaxation von Sauerstoff und Stickstoff; dieser Mechanismus hängt von Temperatur, relativer Feuchte und Druck ab. Die Absorption steigt stark mit der Frequenz: Über große Entfernungen verschwinden die hohen Bänder (2–8 kHz) praktisch, und das wahrgenommene Geräusch wird dumpfer. Diese spektrale Abhängigkeit ist der Grund für die Berechnung in Oktavbändern.', 'phys_ground_title': 'Bodeneffekt (Agr)', 'phys_ground_text': 'der Schall erreicht den Rezeptor über den Direktstrahl und über die Bodenreflexion; beide Pfade interferieren miteinander. Harter Boden (Wasser, Asphalt, G≈0) reflektiert und kann den Pegel sogar verstärken; poröser Boden (Gras, Ackerland, Schnee, G≈1) absorbiert einen Teil der Energie. Das Ergebnis hängt außerdem von den Höhen von Quelle und Rezeptor sowie von deren Abstand ab.', 'phys_bar_title': 'Topografische Abschirmung (Abar)', 'phys_bar_text': 'unterbricht das Gelände die Sichtlinie Quelle–Rezeptor, erreicht der Schall den Rezeptor nur durch Beugung über das Hindernis. Die Dämpfung wächst mit der Differenz zwischen gebeugtem und direktem Weg (Konzept der Fresnel-Zahl) und ist bei hohen Frequenzen größer, deren kürzere Wellenlängen schlechter gebeugt werden.', 'phys_aw_title': 'A-Bewertung', 'phys_aw_text': 'das menschliche Gehör ist bei tiefen und sehr hohen Frequenzen weniger empfindlich. Die A-Bewertungskurve korrigiert jedes Band (z. B. −26,2 dB bei 63 Hz und 0 dB bei 1 kHz), damit der Gesamtpegel in dB(A) die wahrgenommene Lautheit widerspiegelt – die Größe, die von den gesetzlichen Grenzwerten geregelt wird.', 'phys_sum_title': 'Energetische Summierung der Quellen', 'phys_sum_text': 'Pegel in dB werden nicht arithmetisch addiert: Sie werden in Energie umgerechnet, die Energien summiert und das Ergebnis in die logarithmische Skala zurückgeführt. Zwei identische Quellen ergeben +3 dB, nicht den doppelten Pegel; zehn identische Quellen +10 dB. In der Praxis bestimmt die dominante Windenergieanlage (die nächstgelegene oder am wenigsten abgeschirmte) das Ergebnis, während entfernte Quellen kaum beitragen.', 'phys_example_title': 'Anschauliches Beispiel', 'phys_example_text': 'eine Windenergieanlage mit LwA = 105 dB(A) in 500 m Entfernung erfährt eine Divergenz von Adiv = 20·log10(500) + 11 ≈ 65 dB; es verbleiben also etwa 40 dB(A), bevor atmosphärische Absorption, Bodeneffekt und Abschirmung abgezogen werden. Diese Größenordnung erklärt, warum Rezeptoren in mehreren hundert Metern Entfernung typischerweise zwischen 30 und 45 dB(A) liegen.', 'phys_intro': 'Die Gleichungen in diesem Abschnitt sind exakt diejenigen, die im Rechenkern des Plugins implementiert sind; jeder Wert des Berichts lässt sich daher von Hand aus ihnen reproduzieren. Die Symbole sind im Glossar oben definiert.', 'phys_impl_label': 'Umsetzung im Rechenkern', 'phys_spec_title': 'Aufbau des Quellspektrums', 'phys_spec_text': 'liefert eine Gruppe kein gemessenes Spektrum, geht der Rechenkern von einer relativen Vorlage S_b mit der typischen Form einer modernen Windenergieanlage aus (Maximum bei tiefen Frequenzen) und wendet eine globale Verschiebung Δ an, die so berechnet wird, dass die A-bewertete Summe des resultierenden Spektrums exakt den Ziel-LwA reproduziert. Das endgültige Spektrum ist damit konsistent mit dem Katalogwert des Herstellers. Das OEM-Spektrum des Herstellers kann importiert (CSV) oder in der Oberfläche eingegeben werden: Ist es absolut und wird nicht normiert, wird der betriebliche LwA der Gruppe aus seiner A-bewerteten Summe übernommen, sodass der Gesamtschall aus dem OEM-Spektrum berechnet wird; bei aktivierter Normierung oder im Kurvenmodus wirkt das Spektrum nur als Form.', 'phys_div_impl': 'verwendet wird der 3D-Abstand Quelle–Rezeptor einschließlich der akustischen Höhendifferenz zwischen Nabe und Rezeptor, mit einer konfigurierbaren Mindestrechendistanz (standardmäßig 25 m), die Singularitäten direkt am Turm vermeidet.', 'phys_atm_impl': 'der Rechenkern geht von einem Referenzkoeffizienten α_ref je Band aus, tabelliert für 15 °C, 70 % Feuchte und 101,325 kPa (Tabelle unten), und korrigiert ihn mit drei multiplikativen Faktoren: +1 % je °C über 15 °C, +0,3 % je Prozentpunkt Feuchteabweichung vom Optimum von 50 % sowie umgekehrte Proportionalität zum Druck. Die Banddämpfung beträgt α_b·d, linear mit der Entfernung — deshalb dominiert sie bei großen Entfernungen in den hohen Bändern.', 'phys_ground_impl': 'nach dem Regionenschema der ISO 9613-2 wird der Pfad in eine Quellregion (die ersten min(30·h_s, d/3) Meter), eine Mittelregion und eine Rezeptorregion (die letzten min(30·h_r, d/3) Meter) unterteilt. In jeder Region wird ein von Band und charakteristischer Höhe abhängiger Basisterm A_base ausgewertet, mit dem Bodenfaktor G gewichtet, und die drei Terme werden addiert. Die Summe ist auf 10 dB begrenzt; mit G≈0 verschwindet der Term, mit G≈1 ist er maximal.', 'phys_bar_impl': 'der Rechenkern tastet das DGM entlang des Pfads mit adaptivem Schritt ab (etwa Rasterauflösung, mindestens 5 m, zwischen 50 und 1200 Punkten), bestimmt das dominante Hindernis als Punkt des größten Überstands h über der Sichtlinie und aktiviert den Term nur, wenn dieser Überstand eine konservative Schwelle von 1–3 m überschreitet, um nicht auf DGM-Rauschen zu reagieren. Mit der realen Hindernisposition (d₁, d₂) berechnet er die Wegdifferenz δ in Dünnschirm-Näherung, die Fresnel-Zahl C und die abschnittsweise Dämpfung, begrenzt auf 20 dB.', 'phys_fast_title': 'Screening-Modus (schneller Rechenkern)', 'phys_fast_text': 'der schnelle Rechenkern fasst die spektrale Berechnung in einer einzigen Breitbandgleichung zusammen: Er verwendet direkt den globalen LwA, eine lineare atmosphärische Absorption mit festem Koeffizienten α und einen empirischen Bodenterm mit Obergrenze 6 dB, der mit der Entfernung wächst und mit Naben- und Rezeptorhöhe abnimmt. Topografische Abschirmung wird nicht angewendet. Er ist in der Tendenz konsistent mit dem ISO-aligned-Modus, aber nahe gesetzlicher Grenzwerte weniger zuverlässig.', 'phys_tbl_title': 'Vom Rechenkern verwendete Konstanten je Band', 'phys_tbl_template': 'S_b Vorlage [dB]', 'phys_example_extra': 'Bei derselben Entfernung fügt die atmosphärische Absorption im 1-kHz-Band etwa 2 dB hinzu, bei 8 kHz wären es rund 45 dB — deshalb verliert das empfangene Spektrum seine hohen Frequenzen.', 'spec_used': 'Von der Quellgruppe verwendetes Spektrum', 'spec_cols': 'S_b^ref ist die spektrale Referenzform (falls vorhanden), A_b die A-Bewertung jedes Bands und Lw,b der endgültige Pegel in dB, der tatsächlich in die Bandgleichung eingeht.', 'spec_delta_label': 'Angewendetes Δ', 'spec_delta_text': 'konstante Verschiebung, die die Referenzform in das endgültige Spektrum überführt, Lw,b = S_b^ref + Δ; ihr Wert enthält den Ziel-LwA und die Anpassung der A-Bewertung und liegt daher in der Größenordnung des LwA selbst.', 'spec_lwa_check': 'A-bewertete Summe des verwendeten Spektrums', 'band_hz': 'Band [Hz]', 'spec_match': 'stimmt mit dem effektiven LwA der Gruppe überein', 'spec_mismatch': 'weicht vom effektiven LwA der Gruppe ab um',
        },
    }
    # Final report wording layer: keeps the consultancy report concise,
    # prevents misleading compliance statements when receiver coverage is low,
    # and avoids exposing internal placeholders such as '-' for absolute OEM spectra.
    texts['es'].update({
        'physics': 'ANEXO TÉCNICO: FÍSICA Y TRAZABILIDAD DEL CÁLCULO',
        'run_acoustic_scenario': 'Escenario de esta corrida',
        'coverage_low_title': 'Revisar cobertura antes de usar el cumplimiento',
        'coverage_low_text': 'Solo {covered} de {total} receptores ({pct:.1f}%) están dentro del radio de cálculo. El cumplimiento se informa únicamente para los receptores cubiertos; los receptores fuera de radio no se han evaluado frente al límite.',
        'compliance_detail': '{exceed} de {covered} receptores cubiertos superan el límite seleccionado. {outside} receptores quedaron fuera del radio y no se incluyen en este porcentaje.',
        'covered_receivers_exceed': 'receptores cubiertos superan el límite',
        'total_receivers': 'receptores totales',
        'raster_note': 'Los resultados en receptores puntuales se calculan de forma independiente a la resolución del raster. La resolución del mapa solo afecta a la visualización raster generada.',
        'meteo_warning': 'Revisa los parámetros meteorológicos: temperatura, humedad o presión están fuera de un rango típico de referencia. Estos valores afectan sobre todo a la absorción atmosférica y deberían venir de una hipótesis o medición defendible.',
        'enabled': 'habilitado',
        'layer': 'capa',
        'terrain_model': 'modelo del terreno',
        'spec_cols_template': 'La tabla muestra la forma espectral de referencia S_b^ref, la ponderación A de cada banda y el nivel final Lw,b que entra realmente en la ecuación por bandas.',
        'spec_cols_absolute': 'Este grupo utiliza un espectro final absoluto/importado. La tabla muestra la ponderación A y el nivel Lw,b que entra realmente en la ecuación por bandas; no se utiliza una plantilla S_b^ref interna visible.',
        'imported_lwb': 'Lw,b importado/final [dB]',
        'phys_intro': 'Las ecuaciones de este apartado describen la lógica de cálculo implementada por el plugin en este modo preliminar ISO-aligned. Los símbolos están definidos en el glosario anterior.',
        'phys_spec_text': 'si el grupo no aporta un espectro absoluto medido/importado, el motor parte de una plantilla relativa S_b^ref y calcula un desplazamiento Δ para que la suma ponderada A reproduzca el LwA objetivo. Si se importa o se introduce un espectro OEM absoluto sin normalización, ese espectro se usa directamente como Lw,b y el LwA operativo se obtiene de su suma ponderada A.',
        'pressure_warning': 'la presión atmosférica introducida está fuera del rango típico usado como referencia en muchos estudios preliminares. Si no es una medición del emplazamiento, conviene comprobar si debería estar cerca de 101,325 kPa o ajustarse a la altitud.',
    })
    texts['en'].update({
        'physics': 'TECHNICAL APPENDIX: CALCULATION PHYSICS AND TRACEABILITY',
        'run_acoustic_scenario': 'Scenario in this run',
        'coverage_low_title': 'Review coverage before using the compliance result',
        'coverage_low_text': 'Only {covered} of {total} receivers ({pct:.1f}%) are inside the calculation radius. Compliance is reported only for covered receivers; receivers outside the radius have not been assessed against the limit.',
        'compliance_detail': '{exceed} of {covered} covered receivers exceed the selected limit. {outside} receivers were outside the calculation radius and are not included in this percentage.',
        'covered_receivers_exceed': 'covered receivers exceed the limit',
        'total_receivers': 'total receivers',
        'raster_note': 'Point receiver results are calculated independently from the raster resolution. The raster resolution only affects the generated map visualization.',
        'meteo_warning': 'Review the meteorological inputs: temperature, humidity or pressure are outside a typical reference range. These values mainly affect atmospheric absorption and should come from a defensible assumption or site measurement.',
        'enabled': 'enabled',
        'layer': 'layer',
        'terrain_model': 'terrain model',
        'spec_cols_template': 'The table shows the reference spectral shape S_b^ref, the A-weighting of each band and the final Lw,b level that actually enters the band equation.',
        'spec_cols_absolute': 'This group uses a final absolute/imported spectrum. The table shows the A-weighting and the Lw,b level that actually enters the band equation; no visible internal S_b^ref template is used.',
        'imported_lwb': 'imported/final Lw,b [dB]',
        'phys_intro': 'The equations in this section describe the calculation logic implemented by the plugin in this preliminary ISO-aligned mode. The symbols are defined in the glossary above.',
        'phys_spec_text': 'if the group does not provide an absolute measured/imported spectrum, the engine starts from a relative S_b^ref template and computes a shift Δ so that the A-weighted sum reproduces the target LwA. If an absolute OEM spectrum is imported or typed without normalization, that spectrum is used directly as Lw,b and the operating LwA is taken from its A-weighted sum.',
        'pressure_warning': 'the atmospheric pressure entered is outside the typical range used as reference in many preliminary studies. If it is not a site measurement, check whether it should be close to 101.325 kPa or adjusted to altitude.',
    })
    texts['fr'].update({
        'physics': 'ANNEXE TECHNIQUE : PHYSIQUE ET TRAÇABILITÉ DU CALCUL',
        'run_acoustic_scenario': 'Scénario de ce calcul',
        'coverage_low_title': 'Vérifier la couverture avant d’utiliser le résultat de conformité',
        'coverage_low_text': 'Seuls {covered} récepteurs sur {total} ({pct:.1f} %) sont dans le rayon de calcul. La conformité est indiquée uniquement pour les récepteurs couverts ; les récepteurs hors rayon n’ont pas été évalués par rapport à la limite.',
        'compliance_detail': '{exceed} récepteurs couverts sur {covered} dépassent la limite sélectionnée. {outside} récepteurs étaient hors rayon de calcul et ne sont pas inclus dans ce pourcentage.',
        'covered_receivers_exceed': 'récepteurs couverts dépassent la limite',
        'total_receivers': 'récepteurs au total',
        'raster_note': 'Les résultats aux récepteurs ponctuels sont calculés indépendamment de la résolution du raster. La résolution du raster n’affecte que la visualisation cartographique générée.',
        'meteo_warning': 'Vérifiez les paramètres météorologiques : température, humidité ou pression sont hors d’une plage de référence typique. Ces valeurs affectent surtout l’absorption atmosphérique et doivent provenir d’une hypothèse ou mesure défendable.',
        'enabled': 'activé',
        'layer': 'couche',
        'terrain_model': 'modèle de terrain',
        'spec_cols_template': 'Le tableau montre la forme spectrale de référence S_b^ref, la pondération A de chaque bande et le niveau final Lw,b qui entre réellement dans l’équation par bandes.',
        'spec_cols_absolute': 'Ce groupe utilise un spectre final absolu/importé. Le tableau montre la pondération A et le niveau Lw,b qui entre réellement dans l’équation par bandes ; aucun gabarit interne visible S_b^ref n’est utilisé.',
        'imported_lwb': 'Lw,b importé/final [dB]',
        'phys_intro': 'Les équations de cette section décrivent la logique de calcul implémentée par le plugin dans ce mode préliminaire ISO-aligned. Les symboles sont définis dans le glossaire ci-dessus.',
        'phys_spec_text': 'si le groupe ne fournit pas de spectre absolu mesuré/importé, le moteur part d’un gabarit relatif S_b^ref et calcule un décalage Δ afin que la somme pondérée A reproduise le LwA cible. Si un spectre OEM absolu est importé ou saisi sans normalisation, ce spectre est utilisé directement comme Lw,b et le LwA opérationnel est déduit de sa somme pondérée A.',
        'pressure_warning': 'la pression atmosphérique saisie est hors de la plage typique utilisée comme référence dans de nombreuses études préliminaires. S’il ne s’agit pas d’une mesure du site, vérifiez si elle devrait être proche de 101,325 kPa ou ajustée à l’altitude.',
    })
    texts['de'].update({
        'physics': 'TECHNISCHER ANHANG: PHYSIK UND NACHVOLLZIEHBARKEIT DER BERECHNUNG',
        'run_acoustic_scenario': 'Szenario in diesem Lauf',
        'coverage_low_title': 'Abdeckung prüfen, bevor das Konformitätsergebnis verwendet wird',
        'coverage_low_text': 'Nur {covered} von {total} Rezeptoren ({pct:.1f} %) liegen innerhalb des Berechnungsradius. Die Konformität wird nur für abgedeckte Rezeptoren angegeben; Rezeptoren außerhalb des Radius wurden nicht gegen den Grenzwert bewertet.',
        'compliance_detail': '{exceed} von {covered} abgedeckten Rezeptoren überschreiten den ausgewählten Grenzwert. {outside} Rezeptoren lagen außerhalb des Berechnungsradius und sind in diesem Prozentsatz nicht enthalten.',
        'covered_receivers_exceed': 'abgedeckte Rezeptoren überschreiten den Grenzwert',
        'total_receivers': 'Rezeptoren insgesamt',
        'raster_note': 'Punktrezeptor-Ergebnisse werden unabhängig von der Rasterauflösung berechnet. Die Rasterauflösung beeinflusst nur die erzeugte Kartenvisualisierung.',
        'meteo_warning': 'Prüfen Sie die meteorologischen Eingaben: Temperatur, Feuchte oder Druck liegen außerhalb eines typischen Referenzbereichs. Diese Werte beeinflussen vor allem die atmosphärische Absorption und sollten aus einer belastbaren Annahme oder Standortmessung stammen.',
        'enabled': 'aktiviert',
        'layer': 'Layer',
        'terrain_model': 'Geländemodell',
        'spec_cols_template': 'Die Tabelle zeigt die spektrale Referenzform S_b^ref, die A-Bewertung jedes Bands und den endgültigen Pegel Lw,b, der tatsächlich in die Bandgleichung eingeht.',
        'spec_cols_absolute': 'Diese Gruppe verwendet ein endgültiges absolutes/importiertes Spektrum. Die Tabelle zeigt die A-Bewertung und den Pegel Lw,b, der tatsächlich in die Bandgleichung eingeht; es wird keine sichtbare interne S_b^ref-Vorlage verwendet.',
        'imported_lwb': 'importiertes/endgültiges Lw,b [dB]',
        'phys_intro': 'Die Gleichungen in diesem Abschnitt beschreiben die im Plugin implementierte Berechnungslogik in diesem vorläufigen ISO-aligned-Modus. Die Symbole sind im Glossar oben definiert.',
        'phys_spec_text': 'liefert die Gruppe kein absolutes gemessenes/importiertes Spektrum, startet der Rechenkern mit einer relativen Vorlage S_b^ref und berechnet eine Verschiebung Δ, sodass die A-bewertete Summe den Ziel-LwA reproduziert. Wird ein absolutes OEM-Spektrum ohne Normierung importiert oder eingegeben, wird dieses Spektrum direkt als Lw,b verwendet und der betriebliche LwA aus seiner A-bewerteten Summe abgeleitet.',
        'pressure_warning': 'der eingegebene atmosphärische Druck liegt außerhalb des typischen Referenzbereichs vieler Vorstudien. Falls es sich nicht um eine Standortmessung handelt, prüfen Sie, ob er näher bei 101,325 kPa liegen oder an die Höhe angepasst werden sollte.',
    })
    return texts.get(code, texts['es'])


def _noise_clean_value(value, code: str) -> str:
    """Normalize dynamic labels used in the noise report.

    Static report sentences are rendered natively by ``_render_native_noise_report``.
    Some values, however, come from calculation metadata, CSV model names or older
    report payloads and may already contain Spanish/French/English/German fragments.
    This helper keeps those dynamic values from reintroducing language mixing in the
    four supported UI languages.
    """
    code = _noise_lang_code(code)
    txt = str(value if value is not None else '').strip()
    if not txt:
        return txt

    exact = {
        '': {'es': '', 'en': '', 'fr': '', 'de': ''},
        '-': {'es': '-', 'en': '-', 'fr': '-', 'de': '-'},
        'no_type': {'es': 'sin categoría', 'en': 'uncategorized', 'fr': 'sans catégorie', 'de': 'ohne Kategorie'},
        'sin tipo': {'es': 'sin categoría', 'en': 'uncategorized', 'fr': 'sans catégorie', 'de': 'ohne Kategorie'},
        'sans type': {'es': 'sin categoría', 'en': 'uncategorized', 'fr': 'sans catégorie', 'de': 'ohne Kategorie'},
        'ohne Typ': {'es': 'sin categoría', 'en': 'uncategorized', 'fr': 'sans catégorie', 'de': 'ohne Kategorie'},
        'global': {'es': 'global', 'en': 'global', 'fr': 'global', 'de': 'global'},
        'no generado': {'es': 'no generado', 'en': 'not generated', 'fr': 'non générée', 'de': 'nicht erzeugt'},
        'not generated': {'es': 'no generado', 'en': 'not generated', 'fr': 'non générée', 'de': 'nicht erzeugt'},
        'non généré': {'es': 'no generado', 'en': 'not generated', 'fr': 'non générée', 'de': 'nicht erzeugt'},
        'non générée': {'es': 'no generado', 'en': 'not generated', 'fr': 'non générée', 'de': 'nicht erzeugt'},
        'nicht erzeugt': {'es': 'no generado', 'en': 'not generated', 'fr': 'non générée', 'de': 'nicht erzeugt'},
        'sí': {'es': 'sí', 'en': 'yes', 'fr': 'oui', 'de': 'ja'},
        'si': {'es': 'sí', 'en': 'yes', 'fr': 'oui', 'de': 'ja'},
        'yes': {'es': 'sí', 'en': 'yes', 'fr': 'oui', 'de': 'ja'},
        'oui': {'es': 'sí', 'en': 'yes', 'fr': 'oui', 'de': 'ja'},
        'ja': {'es': 'sí', 'en': 'yes', 'fr': 'oui', 'de': 'ja'},
        'no': {'es': 'no', 'en': 'no', 'fr': 'non', 'de': 'nein'},
        'non': {'es': 'no', 'en': 'no', 'fr': 'non', 'de': 'nein'},
        'nein': {'es': 'no', 'en': 'no', 'fr': 'non', 'de': 'nein'},
    }
    low_exact = txt.strip()
    for key, mapping in exact.items():
        if low_exact.lower() == key.lower():
            return mapping.get(code, mapping['es'])

    replacements = [
        ('Plantilla:', {'es': 'Plantilla:', 'en': 'Template:', 'fr': 'Gabarit :', 'de': 'Vorlage:'}),
        ('plantilla:', {'es': 'Plantilla:', 'en': 'Template:', 'fr': 'Gabarit :', 'de': 'Vorlage:'}),
        ('Template:', {'es': 'Plantilla:', 'en': 'Template:', 'fr': 'Gabarit :', 'de': 'Vorlage:'}),
        ('Gabarit :', {'es': 'Plantilla:', 'en': 'Template:', 'fr': 'Gabarit :', 'de': 'Vorlage:'}),
        ('Vorlage:', {'es': 'Plantilla:', 'en': 'Template:', 'fr': 'Gabarit :', 'de': 'Vorlage:'}),
        ('generic curve', {'es': 'curva genérica', 'en': 'generic curve', 'fr': 'courbe générique', 'de': 'generische Kennlinie'}),
        ('genérica curve', {'es': 'curva genérica', 'en': 'generic curve', 'fr': 'courbe générique', 'de': 'generische Kennlinie'}),
        ('curva genérica', {'es': 'curva genérica', 'en': 'generic curve', 'fr': 'courbe générique', 'de': 'generische Kennlinie'}),
        ('courbe générique', {'es': 'curva genérica', 'en': 'generic curve', 'fr': 'courbe générique', 'de': 'generische Kennlinie'}),
        ('generische Kennlinie', {'es': 'curva genérica', 'en': 'generic curve', 'fr': 'courbe générique', 'de': 'generische Kennlinie'}),
        ('sin nombre', {'es': 'sin nombre', 'en': 'unnamed', 'fr': 'sans nom', 'de': 'ohne Namen'}),
        ('unnamed', {'es': 'sin nombre', 'en': 'unnamed', 'fr': 'sans nom', 'de': 'ohne Namen'}),
        ('sans nom', {'es': 'sin nombre', 'en': 'unnamed', 'fr': 'sans nom', 'de': 'ohne Namen'}),
        ('ohne Namen', {'es': 'sin nombre', 'en': 'unnamed', 'fr': 'sans nom', 'de': 'ohne Namen'}),
        ('desde capa', {'es': 'desde capa', 'en': 'from layer', 'fr': 'depuis couche', 'de': 'aus Layer'}),
        ('from layer', {'es': 'desde capa', 'en': 'from layer', 'fr': 'depuis couche', 'de': 'aus Layer'}),
        ('depuis couche', {'es': 'desde capa', 'en': 'from layer', 'fr': 'depuis couche', 'de': 'aus Layer'}),
        ('aus Layer', {'es': 'desde capa', 'en': 'from layer', 'fr': 'depuis couche', 'de': 'aus Layer'}),
        ('Manual (interfaz)', {'es': 'Manual (interfaz)', 'en': 'Manual (UI)', 'fr': 'Manuel (interface)', 'de': 'Manuell (Oberfläche)'}),
        ('forma relativa normalizada a LwA', {'es': 'forma relativa normalizada a LwA', 'en': 'relative shape normalized to LwA', 'fr': 'forme relative normalisée au LwA', 'de': 'relative Form auf LwA normiert'}),
        ('normalizado a LwA del grupo', {'es': 'normalizado a LwA del grupo', 'en': 'normalized to the group LwA', 'fr': 'normalisé au LwA du groupe', 'de': 'auf den Gruppen-LwA normiert'}),
        ('espectro OEM absoluto', {'es': 'espectro OEM absoluto', 'en': 'absolute OEM spectrum', 'fr': 'spectre OEM absolu', 'de': 'absolutes OEM-Spektrum'}),
        ('convertido desde LwA por banda', {'es': 'convertido desde LwA por banda', 'en': 'converted from per-band LwA', 'fr': 'converti depuis le LwA par bande', 'de': 'aus LwA je Band umgerechnet'}),
        ('LwA(A) del espectro =', {'es': 'LwA(A) del espectro =', 'en': 'spectrum LwA(A) =', 'fr': 'LwA(A) du spectre =', 'de': 'LwA(A) des Spektrums ='}),
        ('Sin espectro cargado (fallback LwA si motor ISO)', {'es': 'Sin espectro cargado (fallback LwA si motor ISO)', 'en': 'No spectrum loaded (LwA fallback in ISO engine)', 'fr': 'Aucun spectre chargé (repli LwA en moteur ISO)', 'de': 'Kein Spektrum geladen (LwA-Fallback im ISO-Rechenkern)'}),
        ('Biblioteca:', {'es': 'Biblioteca:', 'en': 'Library:', 'fr': 'Bibliothèque :', 'de': 'Bibliothek:'}),
    ]
    for old, mapping in replacements:
        txt = re.sub(re.escape(old), mapping.get(code, mapping['es']), txt, flags=re.IGNORECASE)

    txt = txt.replace('  ', ' ').strip()
    return txt


def _noise_yes_no(value: bool, code: str) -> str:
    t = _noise_report_texts(code)
    return t['yes'] if bool(value) else t['no']


def _noise_engine_label(engine: str, code: str) -> str:
    engine = str(engine or '').lower()
    if engine == 'iso_aligned':
        return {'es': 'ISO-aligned por bandas', 'en': 'ISO-aligned by bands', 'fr': 'ISO-aligned par bandes', 'de': 'ISO-aligned nach Bändern'}.get(code, 'ISO-aligned por bandas')
    return {'es': 'Rápido LwA global', 'en': 'Fast global LwA', 'fr': 'Rapide LwA global', 'de': 'Schnellmodus mit globalem LwA'}.get(code, 'Rápido LwA global')


def _noise_acoustic_text(acoustic: dict, code: str) -> str:
    acoustic = acoustic or {}
    mode = str(acoustic.get('mode') or 'fixed').lower()
    if mode == 'curve':
        if bool(acoustic.get('use_curve_worst_case', False)):
            return {'es': 'Curvas acústicas LwA(ws) en caso más desfavorable', 'en': 'LwA(ws) acoustic curves, worst-case value', 'fr': 'Courbes acoustiques LwA(ws) en cas le plus défavorable', 'de': 'Akustische LwA(ws)-Kurven im ungünstigsten Fall'}.get(code)
        try:
            ws = float(acoustic.get('eval_ws_m_s'))
            return {'es': f'Curvas acústicas LwA(ws) a {ws:.1f} m/s', 'en': f'LwA(ws) acoustic curves at {ws:.1f} m/s', 'fr': f'Courbes acoustiques LwA(ws) à {ws:.1f} m/s', 'de': f'Akustische LwA(ws)-Kurven bei {ws:.1f} m/s'}.get(code)
        except Exception:
            return {'es': 'Curvas acústicas LwA(ws)', 'en': 'LwA(ws) acoustic curves', 'fr': 'Courbes acoustiques LwA(ws)', 'de': 'Akustische LwA(ws)-Kurven'}.get(code)
    return {'es': 'LwA fijo por grupo de fuente acústica', 'en': 'fixed LwA per acoustic source group', 'fr': 'LwA fixe par groupe de source acoustique', 'de': 'fester LwA je akustischer Quellgruppe'}.get(code)


def _render_native_noise_report(lang: str, ctx: dict) -> str:
    code = _noise_lang_code(lang)
    t = _noise_report_texts(code)
    esc = lambda v: _html.escape(str(v if v is not None else ''))
    f2 = lambda v: f"{float(v):.2f}" if _is_number(v) else '-'
    f1 = lambda v: f"{float(v):.1f}" if _is_number(v) else '-'
    n_sources = int(ctx.get('n_sources', 0) or 0)
    n_receivers = int(ctx.get('n_receivers', 0) or 0)
    n_with = int(ctx.get('n_with', 0) or 0)
    n_without = int(ctx.get('n_without', max(0, n_receivers - n_with)) or 0)
    n_exceed = int(ctx.get('n_exceed', 0) or 0)
    max_noise = float(ctx.get('max_noise', 0.0) or 0.0)
    n_models = int(ctx.get('n_models', 0) or 0)
    engine = str(ctx.get('engine') or 'fast')
    engine_label = _noise_engine_label(engine, code)
    radius = float(ctx.get('radius', 0.0) or 0.0)
    rec_h = float(ctx.get('rec_h', 0.0) or 0.0)
    g = float(ctx.get('g', 0.0) or 0.0)
    g_eff_stats = ctx.get('g_eff_stats') or {}
    ground_diag = ctx.get('ground_diag') or {}
    report = ctx.get('report') or {}
    grid_diag = ctx.get('grid_diag') or {}
    limit_stats = ctx.get('limit_stats') or {}
    acoustic = ctx.get('acoustic') or {}
    spectrum_rows = ctx.get('spectrum_rows') or []
    terms = ctx.get('terms') or {}
    receiver_type_counts = ctx.get('receiver_type_counts') or {}
    compliance = ctx.get('compliance') or {}
    ground_mode = str(ctx.get('ground_mode') or 'global')
    landuse_layer_name = str(ctx.get('landuse_layer_name') or '')
    dem_layer_name = str(ctx.get('dem_layer_name') or '')
    dem_used = bool(ctx.get('dem_used', False))
    temp_c = float(ctx.get('temp_c', 15.0) or 15.0)
    hum_pct = float(ctx.get('hum_pct', 70.0) or 70.0)
    pressure_kpa = float(ctx.get('pressure_kpa', 101.325) or 101.325)
    alpha = float(ctx.get('alpha', 0.0) or 0.0)
    equation = str(ctx.get('equation') or ('Lp,b = Lw,b - Adiv - Aatm,b - Agr,b - Abar,b' if engine == 'iso_aligned' else 'Lp = LwA - Adiv - Aatm - Aground'))
    try:
        now_txt = esc(ctx.get('now_txt') or QtCore.QDateTime.currentDateTime().toString('dd/MM/yyyy - HH:mm:ss'))
    except Exception:
        now_txt = esc(ctx.get('now_txt') or '')
    coverage_pct = (100.0 * n_with / max(1, n_receivers))
    covered_compliance = (100.0 * (n_with - n_exceed) / max(1, n_with))
    limit_min = float(limit_stats.get('min', limit_stats.get('max', 0.0)) or 0.0)
    limit_max = float(limit_stats.get('max', limit_min) or limit_min)
    if abs(limit_min - limit_max) < 1e-9:
        limit_txt = f"{limit_max:.1f} dB(A)"
    else:
        limit_txt = f"{limit_min:.1f}–{limit_max:.1f} dB(A)"
    if ctx.get('grid_txt'):
        # Rebuild native raster text when possible instead of reusing translated fragments.
        pass
    if report.get('grid_layer') is not None or bool(ctx.get('self')._res.get('grid_layer') is not None if ctx.get('self') is not None else False):
        grid_txt = {
            'es': f"sí · resolución solicitada {float(grid_diag.get('requested_resolution_m',0.0)):.1f} m · efectiva {float(grid_diag.get('effective_resolution_m',0.0)):.1f} m",
            'en': f"yes · requested resolution {float(grid_diag.get('requested_resolution_m',0.0)):.1f} m · effective {float(grid_diag.get('effective_resolution_m',0.0)):.1f} m",
            'fr': f"oui · résolution demandée {float(grid_diag.get('requested_resolution_m',0.0)):.1f} m · effective {float(grid_diag.get('effective_resolution_m',0.0)):.1f} m",
            'de': f"ja · angeforderte Auflösung {float(grid_diag.get('requested_resolution_m',0.0)):.1f} m · effektive Auflösung {float(grid_diag.get('effective_resolution_m',0.0)):.1f} m",
        }[code]
    else:
        grid_txt = {'es': 'no generado', 'en': 'not generated', 'fr': 'non générée', 'de': 'nicht erzeugt'}[code]
    method_txt = {'es': 'Propagación por bandas de octava según la metodología ISO-aligned' if engine == 'iso_aligned' else 'Cálculo acústico simplificado para screening', 'en': 'Octave-band propagation according to the ISO-aligned methodology' if engine == 'iso_aligned' else 'Simplified acoustic calculation for screening', 'fr': 'Propagation par bandes d’octave selon la méthodologie ISO-aligned' if engine == 'iso_aligned' else 'Calcul acoustique simplifié pour le criblage', 'de': 'Oktavband-Ausbreitung nach der ISO-aligned-Methodik' if engine == 'iso_aligned' else 'Vereinfachte akustische Screening-Berechnung'}[code]
    if ground_mode == 'landuse':
        ground_txt = {'es': f"desde capa ({landuse_layer_name or 'sin nombre'})", 'en': f"from layer ({landuse_layer_name or 'unnamed'})", 'fr': f"depuis couche ({landuse_layer_name or 'sans nom'})", 'de': f"aus Layer ({landuse_layer_name or 'ohne Namen'})"}[code]
    else:
        ground_txt = {'es': 'global', 'en': 'global', 'fr': 'global', 'de': 'global'}[code]
    dem_default_name = {'es': 'MDT/DSM', 'en': 'DTM/DSM', 'fr': 'MDT/DSM', 'de': 'DGM/DOM'}[code]
    dem_param_label = dem_default_name
    _dem_name_clean = _noise_clean_value(dem_layer_name or '', code)
    if dem_used:
        if str(dem_layer_name or '').strip().lower() in ('', 'output', 'salida', 'result', 'resultado'):
            dem_txt = t['enabled']
        else:
            dem_txt = f"{t['enabled']} · {t['layer']}: {esc(_dem_name_clean)}"
    else:
        dem_txt = t['no']
    landuse_txt = (f"{t['yes']} · {esc(_noise_clean_value(landuse_layer_name or '', code))}" if bool(report.get('landuse_used', False)) else t['no'])
    acoustic_txt = _noise_acoustic_text(acoustic, code)
    pressure_warning = ''
    meteo_warning = ''
    if engine == 'iso_aligned' and (pressure_kpa < 85.0 or pressure_kpa > 105.0):
        pressure_warning = f"<p class='note'><b>{t['review']}:</b> {t['pressure_warning']}</p>"
    if engine == 'iso_aligned' and (temp_c < 0.0 or temp_c > 35.0 or hum_pct < 20.0 or hum_pct > 95.0 or pressure_kpa < 85.0 or pressure_kpa > 105.0):
        meteo_warning = f"<p class='note'><b>{t['review']}:</b> {t['meteo_warning']}</p>"

    # Critical receiver data.
    has_crit = bool(ctx.get('crit'))
    crit_id = ctx.get('crit_id', '-')
    crit_level = ctx.get('crit_level', max_noise)
    crit_limit = ctx.get('crit_limit', 0.0)
    crit_margin = ctx.get('crit_margin', 0.0)
    crit_model = _noise_clean_value(ctx.get('crit_model', '-'), code)
    crit_group = _noise_clean_value(ctx.get('crit_group', '-'), code)
    crit_n_turb = ctx.get('crit_n_turb', '-')
    crit_dist = ctx.get('crit_dist', 0.0)
    crit_lwa = ctx.get('crit_lwa', 0.0)
    crit_adiv_txt = ctx.get('crit_adiv_txt', '-')
    crit_aatm_txt = ctx.get('crit_aatm_txt', '-')
    crit_agr_txt = ctx.get('crit_agr_txt', '-')
    crit_abar_txt = ctx.get('crit_abar_txt', '-')
    crit_abar_max_txt = ctx.get('crit_abar_max_txt', '-')
    crit_abar_ew_txt = ctx.get('crit_abar_ew_txt', '-')
    crit_abar_screen_n = ctx.get('crit_abar_screen_n', 0)
    crit_freq = ctx.get('crit_freq', '-')
    crit_spec_src = _noise_clean_value(ctx.get('crit_spec_src', '-'), code)
    if str(crit_spec_src or '').strip() in ('', '-'):
        _crit_group_raw = str(ctx.get('crit_group') or '').strip()
        _crit_group_norm = _noise_clean_value(_crit_group_raw, code).strip().lower()
        _spec_candidates = []
        for _sp in spectrum_rows:
            _sp_group_raw = str(_sp.get('group_name') or '').strip()
            _sp_group_norm = _noise_clean_value(_sp_group_raw, code).strip().lower()
            if (not _crit_group_raw) or (_sp_group_raw == _crit_group_raw) or (_sp_group_norm == _crit_group_norm):
                _src_candidate = str(_sp.get('spectrum_source') or '').strip()
                if _src_candidate:
                    _spec_candidates.append(_src_candidate)
        if not _spec_candidates and len(spectrum_rows) == 1:
            _src_candidate = str((spectrum_rows[0] or {}).get('spectrum_source') or '').strip()
            if _src_candidate:
                _spec_candidates.append(_src_candidate)
        if _spec_candidates:
            crit_spec_src = _noise_clean_value(_spec_candidates[0], code)
    adiv_term_label = {'es': 'Adiv (divergencia geométrica)', 'en': 'Adiv (geometrical divergence)', 'fr': 'Adiv (divergence géométrique)', 'de': 'Adiv (geometrische Divergenz)'}[code]
    aatm_term_label = {'es': 'Aatm (atmosférica)', 'en': 'Aatm (atmospheric)', 'fr': 'Aatm (atmosphérique)', 'de': 'Aatm (atmosphärisch)'}[code]
    agr_term_label = {'es': 'Agr (suelo)', 'en': 'Agr (ground)', 'fr': 'Agr (sol)', 'de': 'Agr (Boden)'}[code]
    abar_dom_label = {'es': 'Abar trayectoria dominante', 'en': 'Abar dominant path', 'fr': 'Abar trajet dominant', 'de': 'Abar dominanter Pfad'}[code]
    formula_receiver = {'es': 'LpA,receptor', 'en': 'LpA,receiver', 'fr': 'LpA,récepteur', 'de': 'LpA,Rezeptor'}[code]
    sources_sub = {'es': 'fuentes', 'en': 'sources', 'fr': 'sources', 'de': 'Quellen'}[code]
    source_sub = {'es': 'fuente', 'en': 'source', 'fr': 'source', 'de': 'Quelle'}[code]
    model_count_label = {'es': 'modelo(s)', 'en': 'model(s)', 'fr': 'modèle(s)', 'de': 'Modell(e)'}[code]
    crit_g_eff = float(ctx.get('crit_g_eff', g) or g)
    status_text = t['pass'] if float(crit_margin or 0.0) <= 0 else t['fail']
    status_badge = 'badge-success' if float(crit_margin or 0.0) <= 0 else 'badge-danger'
    card_class = 'card-success' if float(crit_margin or 0.0) <= 0 else 'card-danger'
    if str(crit_spec_src or '').strip() in ('', '-'):
        critical_spectrum_line = f"<p style='margin-top:16px;'><b>{t['dominant_band']}:</b> {esc(crit_freq)} Hz</p>"
    else:
        critical_spectrum_line = f"<p style='margin-top:16px;'><b>{t['dominant_band']}:</b> {esc(crit_freq)} Hz &nbsp;&nbsp;&nbsp; <b>{t['spectrum_origin']}:</b> {esc(crit_spec_src)}</p>"

    if has_crit:
        critical_html = f"""
        <div class='{card_class}'>
            <h3>🎯 {t['critical_title']}</h3>
            <table style='margin-bottom:20px;'>
                <tr><td style='width:50%; padding-right:20px;'>
                    <p><b>{t['receiver_id']}:</b> {esc(crit_id)}</p>
                    <p><b>{t['total_level']}:</b> <span style='font-size:28px; font-weight:bold; color:{'#dc3545' if float(crit_margin or 0.0) > 0 else '#28a745'};'>{float(crit_level):.2f} dB(A)</span></p>
                    <p><b>{t['applicable_limit']}:</b> {float(crit_limit):.2f} dB(A)</p>
                    <p><b>{t['margin']}:</b> {float(crit_margin):+.2f} dB <span class='{status_badge}'>{status_text}</span></p>
                </td><td style='width:50%;'>
                    <p><b>{t['dominant_model']}:</b> {esc(crit_model)}</p>
                    <p><b>{t['source_group']}:</b> {esc(crit_group)}</p>
                    <p><b>{t['contributors']}:</b> {esc(crit_n_turb)}</p>
                    <p><b>{t['distance']}:</b> {float(crit_dist):.1f} m</p>
                </td></tr>
            </table>
            <h4>📊 {t['atten_break']}</h4>
            <p style='margin:6px 0 10px 0; color:#495057;'><i>{t['atten_note']}</i></p>
            <table style='margin:16px 0;'>
                <tr><th>{t['term']}</th><th style='text-align:right;'>{t['value_db']}</th><th>{t['description']}</th></tr>
                <tr style='background:#e3f2fd;'><td><b>{t['dom_lwa']}</b></td><td style='text-align:right;'><b>{float(crit_lwa):.2f}</b></td><td>{t['sound_power']}</td></tr>
                <tr><td>{adiv_term_label}</td><td style='text-align:right;'>{esc(crit_adiv_txt)}</td><td>{t['geo_spread']}</td></tr>
                <tr><td>{aatm_term_label}</td><td style='text-align:right;'>{esc(crit_aatm_txt)}</td><td>{t['air_abs']}</td></tr>
                <tr><td>{agr_term_label}</td><td style='text-align:right;'>{esc(crit_agr_txt)}</td><td>{t['ground_att']} (G_eff={crit_g_eff:.2f})</td></tr>
                <tr><td>{abar_dom_label}</td><td style='text-align:right;'>{esc(crit_abar_txt)}</td><td>{t['dtm_att']}</td></tr>
                <tr><td>{t['max_abar']}</td><td style='text-align:right;'>{esc(crit_abar_max_txt)}</td><td>{t['max_abar_desc']}</td></tr>
                <tr><td>{t['ew_abar']}</td><td style='text-align:right;'>{esc(crit_abar_ew_txt)}</td><td>{t['ew_desc']}</td></tr>
                <tr><td>{t['screened_paths']}</td><td style='text-align:right;'>{esc(crit_abar_screen_n)}/{esc(crit_n_turb)}</td><td>{t['screened_desc']}</td></tr>
                <tr style='background:#1e3a5f; color:white; font-weight:bold;'><td>{t['result_level']}</td><td style='text-align:right;'>{float(crit_level):.2f}</td><td>dB(A)</td></tr>
            </table>
            <p style='margin:6px 0 10px 0; color:#495057;'><i>{t['crit_note']}</i></p>
            {critical_spectrum_line}
        </div>
        """
    else:
        critical_html = f"<div class='card'><p>{t['not_available']}.</p></div>"

    def stats_table(adiv_stats, aatm_stats, aground_stats, abar_stats):
        return f"""
        <div class='card'>
            <h3>📊 {t['stats']}</h3>
            <table>
                <tr><th>{t['term']}</th><th style='text-align:right;'>{t['mean']}</th><th style='text-align:right;'>{t['maximum']}</th></tr>
                <tr><td>{adiv_term_label}</td><td style='text-align:right;'>{f2((adiv_stats or {}).get('mean', 0.0))}</td><td style='text-align:right;'>{f2((adiv_stats or {}).get('max', 0.0))}</td></tr>
                <tr><td>{aatm_term_label}</td><td style='text-align:right;'>{f2((aatm_stats or {}).get('mean', 0.0))}</td><td style='text-align:right;'>{f2((aatm_stats or {}).get('max', 0.0))}</td></tr>
                <tr><td>Agr/Aground ({t['ground_effect']})</td><td style='text-align:right;'>{f2((aground_stats or {}).get('mean', 0.0))}</td><td style='text-align:right;'>{f2((aground_stats or {}).get('max', 0.0))}</td></tr>
                <tr><td>Abar</td><td style='text-align:right;'>{f2((abar_stats or {}).get('mean', 0.0))}</td><td style='text-align:right;'>{f2((abar_stats or {}).get('max', 0.0))}</td></tr>
            </table>
        </div>"""

    params = [
        (t['engine_used'], engine_label),
        (t['receiver_height'], f"{rec_h:.1f} m"),
        (t['max_radius'], f"{radius:.0f} m"),
        (t['ground_mode'], ground_txt),
        (t['fallback_ground'] if ground_mode == 'landuse' else t['global_ground'], f"{g:.2f}"),
        (t['mean_geff'], f"{float(g_eff_stats.get('mean', g)):.2f}"),
        (t['critical_geff'], f"{float(g_eff_stats.get('critical', g)):.2f}"),
        (dem_param_label, dem_txt),
        (t['land_use'], landuse_txt),
        (t['acoustic_scenario'], acoustic_txt),
    ]
    if engine == 'iso_aligned':
        params += [(t['temperature'], f"{temp_c:.1f} °C"), (t['humidity'], f"{hum_pct:.1f} %"), (t['pressure'], f"{pressure_kpa:.3f} kPa")]
    else:
        params += [('α', f"{alpha:.4f} dB/m")]
    param_lines = ''.join(f"<li><b>{esc(k)}:</b> {esc(v)}</li>" for k, v in params)
    term_lines = ''.join([
        f"<li><b>Adiv:</b> {t['active'] if terms.get('Adiv', True) else t['inactive']}</li>",
        f"<li><b>Aatm:</b> {t['active'] if terms.get('Aatm', True) else t['inactive']} ({({'es': 'T, HR, P ', 'en': 'T, RH, P ', 'fr': 'T, HR, P ', 'de': 'T, rF, P '}[code]) + t['simplified'] if engine == 'iso_aligned' else 'α·distance'})</li>",
        f"<li><b>Agr/Aground:</b> {t['active'] if terms.get('Agr', True) else t['inactive']}</li>",
        f"<li><b>Abar:</b> {t['active'] if terms.get('Abar', False) else t['inactive']}</li>",
        f"<li><b>{t['effective_g']}:</b> {t['yes'] if terms.get('landuse_g', False) else t['no']}</li>",
    ])
    eff_items = []
    for d in list(acoustic.get('effective_models') or []):
        group_name = _noise_clean_value(d.get('name') or 'Group', code)
        park_name = _noise_clean_value(str(d.get('park_name') or '').strip(), code)
        model_name = _noise_clean_value(str(d.get('model_name') or '').strip(), code)
        spec_src = ''
        for sp in spectrum_rows:
            if str(sp.get('group_name') or '') == group_name:
                spec_src = str(sp.get('spectrum_source') or '')
                break
        try:
            line = f"<li><b>{esc(group_name)}</b>: {float(d.get('lwa_effective')):.2f} dB(A)"
        except Exception:
            line = f"<li><b>{esc(group_name)}</b>: {t['not_available']}"
        extra = []
        if model_name:
            extra.append(f"{t['model']} {esc(model_name)}")
        if park_name:
            extra.append(f"{t['park']} {esc(park_name)}")
        if spec_src:
            extra.append(f"{t['spectrum']} {esc(_noise_clean_value(spec_src, code))}")
        if extra:
            line += ' · ' + ' · '.join(extra)
        line += '</li>'
        eff_items.append(line)
    eff_html = ''.join(eff_items) if eff_items else f"<li>{t['not_available']}</li>"
    rec_types_html = ''.join([f"<li><b>{esc(_noise_clean_value(k, code))}:</b> {int(v)}</li>" for k, v in sorted((receiver_type_counts or {}).items())]) or f"<li>{t['not_available']}</li>"
    compliance_html = ''
    for k, v in sorted((compliance or {}).items()):
        vv = v or {}
        _ex = int(vv.get('exceed', 0) or 0)
        _tot = int(vv.get('total', 0) or 0)
        if vv.get('covered') is not None:
            _cov = int(vv.get('covered', 0) or 0)
            compliance_html += f"<li><b>{esc(_noise_clean_value(k, code))}:</b> {_ex}/{_cov} {t['covered_receivers_exceed']} · {t['total_receivers']}: {_tot}</li>"
        else:
            compliance_html += f"<li><b>{esc(_noise_clean_value(k, code))}:</b> {_ex}/{_tot} {t['exceed_limit']}</li>"
    compliance_html = compliance_html or f"<li>{t['not_available']}</li>"

    # ------------------------------------------------------------------
    # Physics deep-dive card. The equations are language-neutral HTML
    # shared by the four languages so they always stay identical to the
    # formulas actually implemented in the calculation engine
    # (noise_core/noise_engine_iso.py and noise_engine_fast.py).
    # ------------------------------------------------------------------
    _phys_bands = [63, 125, 250, 500, 1000, 2000, 4000, 8000]
    _alpha_ref = {63: 0.0001, 125: 0.0003, 250: 0.0008, 500: 0.0020, 1000: 0.0040, 2000: 0.0095, 4000: 0.0280, 8000: 0.0900}
    _a_weight = {63: -26.2, 125: -16.1, 250: -8.6, 500: -3.2, 1000: 0.0, 2000: 1.2, 4000: 1.0, 8000: -1.1}
    _s_template = {63: -3.0, 125: -1.5, 250: -2.0, 500: -4.0, 1000: -6.0, 2000: -9.0, 4000: -13.0, 8000: -17.0}

    def _eq(expr: str) -> str:
        return f"<div class='eq'>{expr}</div>"

    # ------------------------------------------------------------------
    # Section 2.1 — per-group spectrum traceability tables. They show the
    # exact Lw,b that enters the band equation for each acoustic source
    # group, its origin (manual/OEM CSV/library/template), the reference
    # shape and Δ when a template was fitted, and the A-weighted sum of the
    # spectrum as a cross-check against the effective group LwA.
    # ------------------------------------------------------------------
    _eff_lwa_by_group = {}
    for _em in list((acoustic or {}).get('effective_models') or []):
        try:
            _eff_lwa_by_group[str(_em.get('name') or '')] = float(_em.get('lwa_effective'))
        except Exception:
            continue
    spectrum_tables_html = ''
    for _sp in spectrum_rows:
        try:
            _grp = esc(_noise_clean_value(str(_sp.get('group_name') or ''), code))
            _mdl = esc(_noise_clean_value(str(_sp.get('model_name') or _sp.get('group_name') or ''), code))
            _src_txt = esc(_noise_clean_value(str(_sp.get('spectrum_source') or ''), code))
            _lw = {int(k): float(v) for k, v in (_sp.get('lw_octave') or {}).items()}
            if not _lw:
                continue
            _sref = {int(k): float(v) for k, v in (_sp.get('spectrum_template_ref') or {}).items()}
            try:
                _delta = float(_sp.get('spectrum_delta_db'))
                if _delta != _delta:
                    _delta = None
            except Exception:
                _delta = None
            _rows_html = ''
            _has_sref = bool(_sref)
            for _f in _phys_bands:
                _lw_txt = f"{_lw[_f]:.2f}" if _f in _lw else '-'
                if _has_sref:
                    _sref_txt = f"{_sref[_f]:+.2f}" if _f in _sref else '-'
                    _rows_html += f"<tr><td>{_f}</td><td style='text-align:right;'>{_sref_txt}</td><td style='text-align:right;'>{_a_weight[_f]:+.1f}</td><td style='text-align:right;'><b>{_lw_txt}</b></td></tr>"
                else:
                    _rows_html += f"<tr><td>{_f}</td><td style='text-align:right;'>{_a_weight[_f]:+.1f}</td><td style='text-align:right;'><b>{_lw_txt}</b></td></tr>"
            _delta_html = f"<p><b>{t['spec_delta_label']}:</b> {_delta:+.2f} dB · {t['spec_delta_text']}</p>" if (_delta is not None and _has_sref) else ''
            if _has_sref:
                _spec_cols_text = t['spec_cols_template']
                _spec_header = f"<tr><th>{t['band_hz']}</th><th style='text-align:right;'>S<sub>b</sub><sup>ref</sup> [dB]</th><th style='text-align:right;'>A<sub>b</sub> [dB]</th><th style='text-align:right;'>L<sub>w,b</sub> [dB]</th></tr>"
            else:
                _spec_cols_text = t['spec_cols_absolute']
                _spec_header = f"<tr><th>{t['band_hz']}</th><th style='text-align:right;'>A<sub>b</sub> [dB]</th><th style='text-align:right;'>{t['imported_lwb']}</th></tr>"
            try:
                _lwa_check = 10.0 * math.log10(sum(10.0 ** ((_lw[_f] + _a_weight[_f]) / 10.0) for _f in _phys_bands if _f in _lw))
                _eff_lwa = _eff_lwa_by_group.get(str(_sp.get('group_name') or ''))
                if _eff_lwa is not None and _eff_lwa == _eff_lwa:
                    if abs(_lwa_check - _eff_lwa) <= 0.05:
                        _match_html = f" · ✓ {t['spec_match']}"
                    else:
                        _match_html = f" · ⚠ {t['spec_mismatch']} {abs(_lwa_check - _eff_lwa):.2f} dB"
                else:
                    _match_html = ''
                _check_html = f"<p><b>{t['spec_lwa_check']}:</b> {_lwa_check:.2f} dB(A){_match_html}</p>"
            except Exception:
                _check_html = ''
            spectrum_tables_html += (
                f"<div class='card'><h3>🎼 2.1 {t['spec_used']}: {_grp}</h3>"
                f"<p><b>{t['source_group']}:</b> {_grp} · {t['model']} {_mdl} · <b>{t['spectrum_origin']}:</b> {_src_txt}</p>"
                f"<p>{_spec_cols_text}</p>{_delta_html}"
                f"<table>{_spec_header}{_rows_html}</table>"
                f"{_check_html}</div>"
            )
        except Exception:
            continue

    def _band_row(label: str, values: Dict[int, float], fmt) -> str:
        cells = ''.join(f"<td style='text-align:right;'>{fmt(values[f])}</td>" for f in _phys_bands)
        return f"<tr><td><b>{label}</b></td>{cells}</tr>"

    _band_header = ''.join(f"<th style='text-align:right;'>{f}</th>" for f in _phys_bands)
    band_constants_table = (
        "<table><tr><th>f [Hz]</th>" + _band_header + "</tr>"
        + _band_row('α<sub>ref</sub> [dB/m]', _alpha_ref, lambda v: f"{v:.4f}")
        + _band_row('A<sub>b</sub> [dB]', _a_weight, lambda v: f"{v:+.1f}")
        + _band_row(t['phys_tbl_template'], _s_template, lambda v: f"{v:+.1f}")
        + "</table>"
    )

    EQ_DB = "L = 10·log<sub>10</sub>(p² / p₀²) = 20·log<sub>10</sub>(p / p₀) dB, &nbsp; p₀ = 20 µPa"
    EQ_MODEL = "L<sub>p,b</sub> = L<sub>w,b</sub> + D<sub>c</sub> − A<sub>div</sub> − A<sub>atm,b</sub> − A<sub>gr,b</sub> − A<sub>bar,b</sub> &nbsp;&nbsp; (D<sub>c</sub> = 0 dB)"
    EQ_SPEC = "Template case: L<sub>w,b</sub> = S<sub>b</sub><sup>ref</sup> + Δ, &nbsp;&nbsp; Δ = L<sub>wA,target</sub> − 10·log<sub>10</sub> Σ<sub>b</sub> 10<sup>((S<sub>b</sub><sup>ref</sup> + A<sub>b</sub>) / 10)</sup>. &nbsp; Absolute OEM case: L<sub>w,b</sub> = L<sub>w,b,OEM</sub>"
    EQ_DIV = "A<sub>div</sub> = 20·log<sub>10</sub>(d / d₀) + 11 = 10·log<sub>10</sub>(4·π·d²) dB, &nbsp; d₀ = 1 m, &nbsp; d ≥ d<sub>min</sub>"
    EQ_ATM = "A<sub>atm,b</sub> = α<sub>b</sub>·d, &nbsp;&nbsp; α<sub>b</sub> = α<sub>ref,b</sub> · [1 + 0.01·(T − 15)] · [1 + 0.003·|HR − 50|] · (101.325 / P)"
    EQ_GR = "A<sub>gr,b</sub> = A<sub>s</sub> + A<sub>m</sub> + A<sub>r</sub> = G·[A<sub>base,b</sub>(h<sub>s</sub>) + A<sub>base,b</sub>(h̄) + A<sub>base,b</sub>(h<sub>r</sub>)], &nbsp; 0 ≤ A<sub>gr,b</sub> ≤ 10 dB"
    EQ_GR_REG = "d<sub>s</sub> = min(30·h<sub>s</sub>, d/3), &nbsp; d<sub>r</sub> = min(30·h<sub>r</sub>, d/3), &nbsp; h̄ = (h<sub>s</sub> + h<sub>r</sub>) / 2"
    EQ_GR_BASE = "A<sub>base,b</sub>(h) = 1.5 dB (f ≤ 500 Hz); &nbsp; k<sub>b</sub>·(1 − e<sup>−h/10</sup>) dB, &nbsp; k<sub>b</sub> = 1.5 / 3 / 6 / 12 (1 / 2 / 4 / 8 kHz)"
    EQ_BAR_GEOM = "δ = (h² / 2)·(1/d₁ + 1/d₂), &nbsp;&nbsp; C = 2·δ / λ = 2·f·δ / c, &nbsp; c = 343 m/s"
    EQ_BAR = "A<sub>bar</sub> = 10·log<sub>10</sub>(3 + 20·C), −2 &lt; C ≤ 0; &nbsp; 10·log<sub>10</sub>(3 + 80·C), 0 &lt; C ≤ 3.5; &nbsp; 10·log<sub>10</sub>(3 + 280·C), C &gt; 3.5; &nbsp; 0 ≤ A<sub>bar</sub> ≤ 20 dB"
    EQ_AW = "L<sub>pA,b</sub> = L<sub>p,b</sub> + A<sub>b</sub>"
    EQ_SUM = "L<sub>pA,i</sub> = 10·log<sub>10</sub> Σ<sub>b</sub> 10<sup>L<sub>pA,b</sub> / 10</sup>, &nbsp;&nbsp; L<sub>pA,total</sub> = 10·log<sub>10</sub> Σ<sub>i</sub> 10<sup>L<sub>pA,i</sub> / 10</sup>"
    EQ_FAST = "L<sub>p</sub> = L<sub>wA</sub> − A<sub>div</sub> − α·d − A<sub>gnd</sub>, &nbsp;&nbsp; A<sub>gnd</sub> = min(6, G · 3·log<sub>10</sub>(1 + d<sub>xy</sub>/100) · [1 + (h<sub>s</sub> + h<sub>r</sub>)/80]<sup>−1</sup>) dB"
    EQ_EXAMPLE = "A<sub>div</sub> = 20·log<sub>10</sub>(500) + 11 ≈ 65 dB, &nbsp; A<sub>atm,1 kHz</sub> ≈ 0.004·500 = 2 dB, &nbsp; A<sub>atm,8 kHz</sub> ≈ 0.090·500 = 45 dB"

    coverage_note_html = ''
    if n_receivers > 0 and n_without > 0:
        coverage_note_html = (
            f"<p class='note'><b>{t['coverage_low_title']}:</b> "
            f"{t['coverage_low_text'].format(covered=n_with, total=n_receivers, pct=coverage_pct)}</p>"
        )
    coverage_card_html = (
        f"<div class='card'><h3>📍 {t['coverage']}</h3>"
        f"<p>{n_with} {t['within']}</p><p>{coverage_pct:.1f}% {t['coverage_pct']}</p>"
        f"<p>{n_without} {t['outside']}</p>{coverage_note_html}</div>"
    )
    compliance_card_html = (
        f"<div class='card'><h3>✓ {t['compliance']}</h3>"
        f"<p>{t['compliance_detail'].format(exceed=n_exceed, covered=n_with, outside=n_without)}</p>"
        f"<p>{covered_compliance:.1f}% {t['covered_compliance']}</p><p>{t['limit']}: {limit_txt}</p></div>"
    )
    raster_note_html = ''
    if report.get('grid_layer') is not None or bool(ctx.get('self')._res.get('grid_layer') is not None if ctx.get('self') is not None else False):
        raster_note_html = f"<p class='note'>{t['raster_note']}</p>"
    flow_source_state_line = (
        f"<li><b>{t['source_state']}:</b> {t['source_state_text']} "
        f"<b>{t['run_acoustic_scenario']}:</b> {esc(acoustic_txt)}.</li>"
    )

    physics_card = (
        f"<div class='card'><h3>🧪 {t['phys_title']}</h3>"
        f"<p class='note'>{t['phys_intro']}</p>"
        f"<p><b>{t['phys_db_title']}:</b> {t['phys_db_text']}</p>{_eq(EQ_DB)}"
        f"<p><b>{t['phys_lwlp_title']}:</b> {t['phys_lwlp_text']}</p>{_eq(EQ_MODEL)}"
        f"<p><b>{t['phys_spec_title']}:</b> {t['phys_spec_text']}</p>{_eq(EQ_SPEC)}"
        f"<p><b>{t['phys_div_title']}:</b> {t['phys_div_text']} <i>{t['phys_impl_label']}:</i> {t['phys_div_impl']}</p>{_eq(EQ_DIV)}"
        f"<p><b>{t['phys_atm_title']}:</b> {t['phys_atm_text']} <i>{t['phys_impl_label']}:</i> {t['phys_atm_impl']}</p>{_eq(EQ_ATM)}"
        f"<p><b>{t['phys_ground_title']}:</b> {t['phys_ground_text']} <i>{t['phys_impl_label']}:</i> {t['phys_ground_impl']}</p>{_eq(EQ_GR)}{_eq(EQ_GR_REG)}{_eq(EQ_GR_BASE)}"
        f"<p><b>{t['phys_bar_title']}:</b> {t['phys_bar_text']} <i>{t['phys_impl_label']}:</i> {t['phys_bar_impl']}</p>{_eq(EQ_BAR_GEOM)}{_eq(EQ_BAR)}"
        f"<p><b>{t['phys_aw_title']}:</b> {t['phys_aw_text']}</p>{_eq(EQ_AW)}"
        f"<p><b>{t['phys_sum_title']}:</b> {t['phys_sum_text']}</p>{_eq(EQ_SUM)}"
        f"<p><b>{t['phys_fast_title']}:</b> {t['phys_fast_text']}</p>{_eq(EQ_FAST)}"
        f"<h4>📐 {t['phys_tbl_title']}</h4>{band_constants_table}"
        f"<p class='note'><b>{t['phys_example_title']}:</b> {t['phys_example_text']} {t['phys_example_extra']}</p>{_eq(EQ_EXAMPLE)}"
        f"</div>"
    )

    return f"""
    <html><head><meta charset='utf-8'><style>
        body {{ font-family: 'Segoe UI', Arial, sans-serif; color:#212529; line-height:1.5; margin:0; padding:20px; background:#f8f9fa; }}
        h1 {{ margin:0 0 8px 0; font-size:32px; }} h2 {{ color:#1e3a5f; border-bottom:3px solid #2c5aa0; padding-bottom:8px; margin-top:28px; }} h3 {{ color:#2c5aa0; margin-top:0; }}
        .header {{ background-color:#1e3a5f; color:white; padding:28px; border-radius:10px; margin-bottom:20px; }}
        .card, .card-info, .card-success, .card-danger {{ background:white; padding:18px; margin:16px 0; border-radius:8px; box-shadow:0 2px 6px rgba(0,0,0,.08); }}
        .card-info {{ border-left:5px solid #17a2b8; }} .card-success {{ border-left:5px solid #28a745; }} .card-danger {{ border-left:5px solid #dc3545; }}
        .scope {{ background:#fff3cd; border-left:5px solid #ffc107; padding:16px 20px; border-radius:8px; margin:16px 0 22px 0; }} .note {{ background:#eef6ff; border-left:4px solid #2c5aa0; padding:10px 12px; border-radius:5px; }}
        .kpi-label {{ font-size:12px; color:#5a6b7f; letter-spacing:1px; }} .kpi-num {{ font-size:30px; font-weight:700; color:#1e3a5f; }}
        table {{ width:100%; border-collapse:collapse; margin:12px 0; }} th {{ background:#1e3a5f; color:white; padding:10px; text-align:left; }} td {{ border-bottom:1px solid #dee2e6; padding:9px; }}
        .formula {{ font-family:Consolas,monospace; background:#eef3fb; border:1px solid #cfd8e3; border-radius:6px; padding:10px; margin:10px 0; }} .eq {{ font-family:'Cambria Math','Georgia','Times New Roman',serif; font-size:15px; text-align:center; background:#f4f7fc; padding:10px; margin:10px 0; }} .badge-success {{ background:#28a745; color:white; padding:3px 8px; border-radius:4px; }} .badge-danger {{ background:#dc3545; color:white; padding:3px 8px; border-radius:4px; }}
    </style></head><body>
        <table width='100%' bgcolor='#1e3a5f' cellpadding='18' cellspacing='0' style='margin-bottom:20px; border-collapse:separate;'><tr><td style='border-bottom:none;'><span style='color:#ffffff; font-size:28px; font-weight:bold;'>📊 {t['title']}</span><br/><span style='color:#d8e4f5; font-size:14px;'>{t['subtitle']}</span><br/><span style='color:#d8e4f5; font-size:13px;'>📅 {now_txt}</span></td></tr></table>
        <div class='scope'><h3>⚠️ {t['scope_title']}</h3><p>{t['what_is']}</p><p>{t['what_not']}</p><p><b>{t['simplifications']}</b></p><ul><li>{t['s1']}</li><li>{t['s2']}</li><li>{t['s3']}</li><li>{t['s4']}</li><li>{t['s5']}</li></ul><p><b>{t['recommendation']}</b></p></div>
        <h2>1. {t['exec']}</h2><table width='100%' cellpadding='14' cellspacing='8' style='margin:16px 0; border-collapse:separate;'><tr><td width='33%' align='center' bgcolor='#ffffff' style='border-bottom:none;'><span class='kpi-label'>{t['turbines'].upper()}</span><br/><span class='kpi-num'>{n_sources}</span></td><td width='33%' align='center' bgcolor='#ffffff' style='border-bottom:none;'><span class='kpi-label'>{t['receivers'].upper()}</span><br/><span class='kpi-num'>{n_receivers}</span></td><td width='34%' align='center' bgcolor='#ffffff' style='border-bottom:none;'><span class='kpi-label'>{t['max_level'].upper()} (dB(A))</span><br/><span class='kpi-num'>{max_noise:.1f}</span></td></tr></table>
        {coverage_card_html}{compliance_card_html}
        <div class='card'><h3>🎯 {t['methodology']}</h3><p><b>{t['engine_used']}:</b> {engine_label}</p><p><b>{t['source_groups']}:</b> {n_models} {model_count_label}</p><p><b>{t['method']}:</b> {method_txt}</p><p><b>{t['raster_map']}:</b> {grid_txt}</p>{raster_note_html}</div>
        <h2>2. {t['generated']}</h2><div class='card-info'><h3>🧭 {t['how_iso']}</h3><p>{t['flow_intro']}</p><ol><li><b>{t['gis_inputs']}:</b> {t['gis_text']}</li>{flow_source_state_line}<li><b>{t['bands']}:</b> {t['bands_text']}</li><li><b>{t['selection']}:</b> {t['selection_text']}</li><li><b>{t['path']}:</b> {t['path_text']}</li><li><b>{t['propagation']}:</b> {t['propagation_text']}</li><li><b>{t['energy']}:</b> {t['energy_text']}</li><li><b>{t['check']}:</b> {t['check_text']}</li></ol><div class='formula'>{formula_receiver} = 10·log10(Σ_{sources_sub} 10^(LpA,{source_sub}/10))</div></div>{spectrum_tables_html}<div class='card'><h3>🔎 {t['screening_diff']}</h3><p>{t['screening_text']}</p></div>
        <h2>3. {t['critical']}</h2>{critical_html}{stats_table(ctx.get('adiv_stats'), ctx.get('aatm_stats'), ctx.get('aground_stats'), ctx.get('abar_stats'))}
        <h2>4. {t['config']}</h2><div class='card'><h3>⚙️ {t['equation']}</h3><div class='formula'>{esc(equation)}</div><p>{t['interp_iso']}</p></div><div class='card'><h3>📋 {t['calc_params']}</h3><ul>{param_lines}</ul>{pressure_warning}{meteo_warning}<p><b>{t['paths_g']}:</b> {int((ground_diag or {}).get('from_landuse_count',0))} ({float((ground_diag or {}).get('from_landuse_pct',0.0)):.1f}%)</p></div><div class='card'><h3>✓ {t['active_terms']}</h3><ul>{term_lines}</ul></div>
        <h2>5. {t['physics']}</h2><div class='card'><h3>📖 {t['glossary']}</h3><table><tr><th>{t['symbol']}</th><th>{t['meaning']}</th></tr><tr><td><b>LwA</b></td><td>{t['lwa_mean']}</td></tr><tr><td><b>Lw,b</b></td><td>{t['lwb_mean']}</td></tr><tr><td><b>LpA</b></td><td>{t['lpa_mean']}</td></tr><tr><td><b>Adiv</b></td><td>{t['adiv_mean']}</td></tr><tr><td><b>Aatm</b></td><td>{t['aatm_mean']}</td></tr><tr><td><b>Agr</b></td><td>{t['agr_mean']}</td></tr><tr><td><b>Abar</b></td><td>{t['abar_mean']}</td></tr><tr><td><b>G / G_eff</b></td><td>{t['g_mean']}</td></tr></table></div><div class='card'><h3>📘 {t['model_title']}</h3><p>{t['model_text']}</p><div class='formula'>Lp,b = Lw,b - Adiv - Aatm,b - Agr,b - Abar,b</div><div class='formula'>LpA,total = 10·log10(Σ 10^((Lp,b + A_weight,b)/10))</div><p>{t['div_text']}</p><p>{t['atm_text']}</p><p>{t['ground_text']}</p><p>{t['screen_text']}</p><p class='note'>{t['crit_note']}</p></div>{physics_card}
        <h2>6. {t['source_section']}</h2><div class='card'><h3>⚡ {t['effective_lwa']}</h3><ul>{eff_html}</ul></div>
        <h2>7. {t['receiver_dist']}</h2><table width='100%' cellpadding='0' cellspacing='8' style='border-collapse:separate;'><tr><td width='50%' style='vertical-align:top; border-bottom:none;'><div class='card'><h3>📍 {t['receivers_by_cat']}</h3><ul>{rec_types_html}</ul></div></td><td width='50%' style='vertical-align:top; border-bottom:none;'><div class='card'><h3>✓ {t['compliance_by_cat']}</h3><ul>{compliance_html}</ul></div></td></tr></table>
        <div class='scope'><h3>⚠️ {t['limits_recs']}</h3><p>{t['fast_engine']}</p><p>{t['iso_engine']}</p><p>{t['known_simp']}</p><p>{t['multi_models']}</p><p>{t['iso_raster']}</p><p><b>{t['recommendation']}</b></p></div>
    </body></html>
    """


def _is_number(v) -> bool:
    try:
        f = float(v)
        return f == f
    except Exception:
        return False


class NoiseResultsDialog(QtWidgets.QDialog):
    def __init__(self, parent=None, result: Dict[str, object] | None = None):
        install_runtime_i18n_patches()
        super().__init__(parent)
        self._res = result or {}
        self.setWindowTitle("Schall · Technische Übersicht" if str(current_language()).lower().startswith("de") else _tr("Ruido · Resumen técnico"))
        self.setModal(True)
        self._resize_to_screen()
        self._build_ui()
        apply_i18n(self)
        self._fill()

    def _resize_to_screen(self):
        fit_to_screen(self, preferred=(1100, 820), minimum=(680, 460), max_ratio=(0.92, 0.90))

    def _build_ui(self):
        root = QtWidgets.QVBoxLayout(self)
        root.setContentsMargins(12, 12, 12, 12)
        root.setSpacing(8)

        header = QtWidgets.QHBoxLayout()
        title = QtWidgets.QLabel("Schall · Berechnungsübersicht" if str(current_language()).lower().startswith("de") else _tr("Ruido · Resumen del cálculo"))
        title.setStyleSheet("font-size:20px; font-weight:700; color:#103b67;")
        header.addWidget(title, 1)
        header.addStretch(1)
        logo = QtWidgets.QLabel(self)
        logo_path = os.path.join(os.path.dirname(__file__), "assets", "velantiswind_logo.png")
        if os.path.exists(logo_path):
            pix = QtGui.QPixmap(logo_path)
            if not pix.isNull():
                logo.setPixmap(pix.scaled(180, 180, QtCore.Qt.KeepAspectRatio, QtCore.Qt.SmoothTransformation))
                logo.setToolTip("Velantis Wind")
        header.addWidget(logo, 0, QtCore.Qt.AlignRight | QtCore.Qt.AlignTop)
        root.addLayout(header)

        self.tabs = QtWidgets.QTabWidget(self)
        root.addWidget(self.tabs, 1)

        self.page_summary = QtWidgets.QTextBrowser(self)
        self.tabs.addTab(self.page_summary, _tr("Resumen"))

        self.tbl_models = QtWidgets.QTableWidget(0, 6, self)
        self.tbl_models.setHorizontalHeaderLabels([_tr("Modelo WT"), _tr("Aerogeneradores"), "LwA eff.", "HH", "D", _tr("Notas")])
        configure_table(self.tbl_models, stretch_columns=(0, 5))
        self.tbl_models.horizontalHeader().setStretchLastSection(True)
        self.tabs.addTab(self.tbl_models, _tr("Modelos"))

        self.tbl_top = QtWidgets.QTableWidget(0, len(CONSULTANCY_RECEIVER_HEADERS), self)
        self.tbl_top.setHorizontalHeaderLabels(CONSULTANCY_RECEIVER_HEADERS)
        self.tbl_top.setToolTip(
            _tr("Tabla sintética para consultoría: resultados acústicos por receptor, ")
            + _tr("cumplimiento, fuente dominante y atenuaciones principales. ")
            + _tr("Los diagnósticos internos MDT por pares se conservan en memoria, pero no se muestran por defecto.")
        )
        configure_table(self.tbl_top, stretch_columns=(0, 1, 9, 10, 11, 22))
        self.tbl_top.horizontalHeader().setStretchLastSection(True)
        self.tabs.addTab(self.tbl_top, _tr("Top receptores"))

        # Internal MDT screening table kept for compatibility with helper methods,
        # but no longer exposed as a default consultancy tab/export.
        self.tbl_mdt = QtWidgets.QTableWidget(0, 27, self)
        self.tbl_mdt.setHorizontalHeaderLabels([
            _tr('ID receptor'), _tr('nivel total dB(A)'), _tr('nº aerogeneradores'), 'Abar max contrib. dB',
            _tr('Abar ponderado dB'), _tr('aerogeneradores apantallados'), _tr('estado MDT dom.'), 'Abar dom. dB',
            _tr('ID fuente Abar max'), _tr('estado Abar max'), _tr('obs. Abar max m'),
            _tr('umbral Abar max m'), _tr('d1 Abar max m'), _tr('d2 Abar max m'),
            _tr('ID fuente obstáculo max'), _tr('estado obstáculo max'), _tr('obs. obstáculo max m'),
            _tr('umbral obstáculo max m'), _tr('d1 obstáculo max m'), _tr('d2 obstáculo max m'),
            _tr('z terreno receptor m'), _tr('h receptor m'), _tr('z acústica receptor m'),
            _tr('z terreno aerogenerador dom. m'), _tr('z acústica aerogenerador dom. m'),
            _tr('z terreno aerogenerador Abar max m'), _tr('z acústica aerogenerador Abar max m')
        ])

        self.tbl_layers = QtWidgets.QTableWidget(0, 2, self)
        self.tbl_layers.setHorizontalHeaderLabels([_tr("Capa"), _tr("Estado")])
        self.tbl_layers.horizontalHeader().setStretchLastSection(True)
        self.tabs.addTab(self.tbl_layers, _tr("Capas creadas"))

        btns = QtWidgets.QHBoxLayout()
        self.btn_export_summary = QtWidgets.QPushButton(_tr("Exportar informe…"))
        self.btn_export_summary.setToolTip(_tr("Guarda el resumen técnico en HTML o TXT."))
        self.btn_export_summary.clicked.connect(self._export_summary)
        self.btn_export_receivers = QtWidgets.QPushButton(_tr("Exportar receptores CSV…"))
        self.btn_export_receivers.setToolTip(_tr("Guarda una tabla limpia con una fila por receptor y las columnas necesarias para consultoría."))
        self.btn_export_receivers.clicked.connect(self._export_receivers_csv)
        self.btn_export_exceed = QtWidgets.QPushButton(_tr("Exportar excedencias CSV…"))
        self.btn_export_exceed.setToolTip(_tr("Guarda únicamente los receptores que superan su límite acústico."))
        self.btn_export_exceed.clicked.connect(self._export_exceedances_csv)
        self.btn_export_xlsx = QtWidgets.QPushButton(_tr("Exportar paquete XLSX…"))
        self.btn_export_xlsx.setToolTip(_tr("Guarda resumen, modelos, receptores y excedencias en un único libro Excel."))
        self.btn_export_xlsx.clicked.connect(self._export_package_xlsx)
        btns.addWidget(self.btn_export_summary)
        btns.addWidget(self.btn_export_receivers)
        btns.addWidget(self.btn_export_exceed)
        btns.addWidget(self.btn_export_xlsx)
        btns.addStretch(1)
        close_btn = QtWidgets.QPushButton(_tr("Cerrar"))
        close_btn.setMinimumHeight(34)
        close_btn.clicked.connect(self.accept)
        btns.addWidget(close_btn)

        if str(current_language()).lower().startswith("de"):
            self.tabs.setTabText(0, "Übersicht")
            self.tabs.setTabText(1, "Modelle")
            self.tabs.setTabText(2, "Top-Rezeptoren")
            self.tabs.setTabText(3, "Erzeugte Layer")
            self.tbl_models.setHorizontalHeaderLabels(["WT-Modell", "Windturbinen", "LwA eff.", "NH", "D", "Notizen"])
            self.tbl_top.setHorizontalHeaderLabels([
                "Rezeptor-ID", "Typ", "Gesamtpegel dB(A)", "Grenzwert dB(A)",
                "Abstand zum Grenzwert dB", "Status", "überschreitet Grenzwert",
                "Anz. Windturbinen", "nächste Windturbine (m)", "dominantes Modell",
                "dom. Quellgruppe", "dom. Park", "LwA dom. Quelle dB(A)",
                "Adiv dB", "Aatm dB", "Agr/Aground dB", "Abar max. dB",
                "G Boden", "Bodenmodus", "Rezeptorhöhe m", "z Gelände Rezeptor m",
                "z akustisch Rezeptor m", "dominanter Quell-Layer"
            ])
            self.tbl_layers.setHorizontalHeaderLabels(["Layer", "Status"])
            self.btn_export_summary.setText("Bericht exportieren…")
            self.btn_export_summary.setToolTip("Speichert die technische Übersicht als HTML oder TXT.")
            self.btn_export_receivers.setText("Rezeptoren als CSV exportieren…")
            self.btn_export_receivers.setToolTip("Speichert eine saubere Tabelle mit einer Zeile pro Rezeptor.")
            self.btn_export_exceed.setText("Überschreitungen als CSV exportieren…")
            self.btn_export_exceed.setToolTip("Speichert nur Rezeptoren, die ihren akustischen Grenzwert überschreiten.")
            self.btn_export_xlsx.setText("XLSX-Paket exportieren…")
            self.btn_export_xlsx.setToolTip("Speichert Übersicht, Modelle, Rezeptoren und Überschreitungen in einer Excel-Datei.")
            close_btn.setText("Schließen")
        root.addLayout(btns)

    def _fill(self):
        self._fill_summary()
        self._fill_models()
        self._fill_top_receivers()
        self._fill_mdt_screening()
        self._fill_layers()

    def _payload_top_receivers(self) -> List[Dict[str, object]]:
        rows = self._res.get("top_receivers") or []
        out: List[Dict[str, object]] = []
        for row in rows:
            if isinstance(row, dict):
                out.append(row)
        def _noise(d):
            try:
                return float(d.get("noise_dba") or d.get("total_level_dba") or 0.0)
            except Exception:
                return -1.0e99
        out.sort(key=_noise, reverse=True)
        return out


    def _payload_receiver_rows(self) -> List[Dict[str, object]]:
        rows = self._res.get("receiver_rows") or []
        out: List[Dict[str, object]] = []
        for row in rows:
            if isinstance(row, dict):
                out.append(row)
        if out:
            return out
        # Fallback to visible top rows if the full receiver payload is absent.
        return self._payload_top_receivers()


    def _attenuation_stats_from_payload_rows(self) -> Dict[str, Dict[str, float]]:
        """Compute attenuation statistics from stable named receiver rows.

        The HTML report historically used the precomputed ``*_stats`` entries
        in ``self._res``.  When the calculation is returned by a background
        QgsTask, those entries can remain zero if the QGIS memory layer cannot
        be read at the exact moment the dialog is built, even though
        ``receiver_rows`` and the Top receivers table contain the correct
        values.  This fallback derives the statistics directly from the named
        payload used by the CSV/XLSX exports.
        """
        rows = self._payload_receiver_rows()

        def _f(d: Dict[str, object], *keys: str):
            for key in keys:
                try:
                    v = d.get(key)
                except Exception:
                    v = None
                if v is None:
                    continue
                txt = str(v).strip()
                if txt == '' or txt.lower() in ('none', 'nan', 'n/a'):
                    continue
                try:
                    x = float(txt.replace(',', '.'))
                except Exception:
                    continue
                if x == x:
                    return x
            return None

        def _covered(d: Dict[str, object]) -> bool:
            nsrc = _f(d, 'n_src', 'turbines_in_radius', 'no. turbines')
            if nsrc is not None:
                return nsrc > 0
            covered = _f(d, 'covered')
            if covered is not None:
                return covered > 0
            noise = _f(d, 'noise_dba', 'total_level_dba', 'total level dB(A)')
            return bool(noise is not None and noise > 0)

        vals = {
            'adiv': [],
            'aatm': [],
            'aground': [],
            'abar': [],
        }
        for d in rows:
            if not isinstance(d, dict) or not _covered(d):
                continue
            for name, keys in {
                'adiv': ('adiv_db', 'divergence_loss_db', 'Adiv loss dB', 'pérdida Adiv dB'),
                'aatm': ('aatm_db', 'atmospheric_loss_db', 'Aatm loss dB', 'pérdida Aatm dB'),
                'aground': ('aground_db', 'ground_loss_db', 'Agr/Aground loss dB', 'pérdida Agr/Aground dB'),
                'abar': ('abar_max_db', 'barrier_loss_max_contributors_db', 'Abar max contrib. dB', 'abar_db', 'Abar dom. dB'),
            }.items():
                x = _f(d, *keys)
                if x is not None:
                    vals[name].append(float(x))

        def _stat(seq: List[float]) -> Dict[str, float]:
            if not seq:
                return {'mean': 0.0, 'max': 0.0}
            return {'mean': sum(seq) / float(len(seq)), 'max': max(seq)}

        return {name: _stat(seq) for name, seq in vals.items()}

    def _prefer_payload_stats_if_needed(self, current: Dict[str, object], fallback: Dict[str, float]) -> Dict[str, float]:
        """Use payload-derived stats when the current report stats are empty/zero."""
        try:
            cur_max = float((current or {}).get('max', 0.0) or 0.0)
        except Exception:
            cur_max = 0.0
        try:
            fb_max = float((fallback or {}).get('max', 0.0) or 0.0)
        except Exception:
            fb_max = 0.0
        if fb_max > 0.0 and cur_max <= 0.0:
            return dict(fallback or {})
        return dict(current or {})

    def _infer_critical_receiver_from_layer(self) -> Dict[str, object]:
        """Return the highest-noise receiver as a dict using current layer fields."""
        layer = self._res.get("result_layer")
        payload_rows = self._payload_top_receivers()
        if not isinstance(layer, QgsVectorLayer):
            return dict(payload_rows[0]) if payload_rows else {}
        best_feat = None
        best_level = -1.0e99
        level_keys = ("noise_dba", "total_level_dba", "nivel_total_dba")
        try:
            iterator = layer.getFeatures()
        except Exception:
            return {}
        for feat in iterator:
            level = None
            for key in level_keys:
                try:
                    level = float(feat[key])
                    break
                except Exception:
                    continue
            if level is None:
                continue
            try:
                if level != level:
                    continue
            except Exception:
                continue
            if best_feat is None or level > best_level:
                best_feat = feat
                best_level = level
        if best_feat is None:
            return dict(payload_rows[0]) if payload_rows else {}
        row: Dict[str, object] = {"fid": best_feat.id(), "rec_id": best_feat.id()}
        try:
            for fld in layer.fields():
                name = fld.name()
                try:
                    row[name] = best_feat[name]
                except Exception:
                    pass
        except Exception:
            pass
        if not row.get("rec_id"):
            row["rec_id"] = best_feat.id()
        return row

    def _fill_summary(self):
        n_sources = int(self._res.get("n_sources", 0))
        n_receivers = int(self._res.get("n_receivers", 0))
        n_with = int(self._res.get("n_receivers_with_sources", 0))
        n_without = int(self._res.get("n_uncovered_receivers", max(0, n_receivers - n_with)))
        n_exceed = int(self._res.get("n_receivers_exceeding_limit", 0))
        max_noise = float(self._res.get("max_noise_dba", 0.0))
        model_diag = self._res.get("model_diag", {}) or {}
        n_models = len(model_diag)
        limit_stats = self._res.get('limit_stats') or self._infer_limit_stats_from_layer()
        acoustic = self._res.get('acoustic_scenario', {}) or {}
        crit_raw = self._res.get('critical_receiver') or {}
        crit_layer = self._infer_critical_receiver_from_layer()

        def _has_value(v):
            if v is None:
                return False
            try:
                if isinstance(v, float) and v != v:
                    return False
            except Exception:
                pass
            return str(v).strip() != ''

        # Merge stored critical-receiver metadata with a robust fallback read directly
        # from the result layer. This avoids visual summaries falling back to 0.00
        # when the engine changes field names.
        crit = dict(crit_layer or {})
        for _k, _v in dict(crit_raw or {}).items():
            if _has_value(_v):
                crit[_k] = _v
        payload_att_stats = self._attenuation_stats_from_payload_rows()
        adiv_stats = self._prefer_payload_stats_if_needed(self._res.get('adiv_stats') or {}, payload_att_stats.get('adiv') or {})
        aatm_stats = self._prefer_payload_stats_if_needed(self._res.get('aatm_stats') or {}, payload_att_stats.get('aatm') or {})
        aground_stats = self._prefer_payload_stats_if_needed(self._res.get('aground_stats') or {}, payload_att_stats.get('aground') or {})
        abar_stats = self._prefer_payload_stats_if_needed(self._res.get('abar_stats') or {}, payload_att_stats.get('abar') or {})
        g_eff_stats = self._res.get('g_eff_stats') or {}
        ground_diag = self._res.get('ground_diag') or {}
        receiver_type_counts = self._res.get('receiver_type_counts') or {}
        grid_diag = self._res.get('grid_diag') or {}
        report = self._res.get('report_meta') or {}
        ground_mode = str(report.get('ground_mode') or self._res.get('ground_mode') or 'global')
        landuse_layer_name = str(report.get('landuse_layer_name') or self._res.get('landuse_layer_name') or '')
        dem_layer_name = str(report.get('dem_layer_name') or self._res.get('dem_layer_name') or '')
        dem_used = bool(report.get('dem_used', self._res.get('dem_used', False)))
        engine = str(report.get('engine') or ('iso_aligned' if str(self._res.get('method') or '').startswith('iso_') else 'fast'))
        engine_label = str(report.get('engine_label') or ('ISO-aligned par bandes' if engine == 'iso_aligned' else 'Rapide LwA global'))
        equation = str(report.get('equation') or ('Lp,b = Lw,b - Adiv - Aatm,b - Agr,b - Abar,b' if engine == 'iso_aligned' else 'Lp = LwA - Adiv - Aatm - Aground'))
        alpha = float(report.get('alpha_db_per_m', self._res.get('alpha_db_per_m', 0.0)))
        g = float(report.get('ground_factor_g', self._res.get('ground_factor_g', 0.0)))
        rec_h = float(report.get('receiver_height_m', self._res.get('receiver_height_m', 0.0)))
        radius = float(report.get('max_radius_m', self._res.get('max_radius_m', 0.0)))
        temp_c = float(report.get('temperature_c', 15.0))
        hum_pct = float(report.get('humidity_percent', 70.0))
        pressure_kpa = float(report.get('pressure_kpa', 101.325))
        terms = report.get('active_terms') or {}
        spectrum_rows = report.get('spectrum_sources') or []

        if str(acoustic.get('mode') or 'fixed') == 'curve':
            if bool(acoustic.get('use_curve_worst_case', False)):
                acoustic_txt = 'Courbes acoustiques LwA(ws) en cas le plus défavorable'
            else:
                try:
                    acoustic_txt = f"Courbes acoustiques LwA(ws) à {float(acoustic.get('eval_ws_m_s')):.1f} m/s"
                except Exception:
                    acoustic_txt = 'Courbes acoustiques LwA(ws)'
        else:
            acoustic_txt = 'LwA fixe par groupe de source acoustique'

        eff_lines = []
        for d in list(acoustic.get('effective_models') or []):
            group_name = str(d.get('name') or 'Groupe')
            park_name = str(d.get('park_name') or '').strip()
            model_name = str(d.get('model_name') or '').strip()
            spec_src = ''
            for sp in spectrum_rows:
                if str(sp.get('group_name') or '') == group_name:
                    spec_src = str(sp.get('spectrum_source') or '')
                    break
            try:
                line = f"<li><b>{group_name}</b>: {float(d.get('lwa_effective')):.2f} dB(A)"
            except Exception:
                line = f"<li><b>{group_name}</b>: sans valeur"
            extra = []
            if model_name:
                extra.append(f"modèle {model_name}")
            if park_name:
                extra.append(f"parc {park_name}")
            if str(d.get('curve_note') or '').strip():
                extra.append(str(d.get('curve_note')))
            if spec_src:
                extra.append(f"spectre {spec_src}")
            if extra:
                line += " · " + " · ".join(extra)
            line += "</li>"
            eff_lines.append(line)

        spectrum_detail_blocks = []
        for sp in spectrum_rows:
            group_name = str(sp.get('group_name') or 'Groupe')
            model_name = str(sp.get('model_name') or group_name)
            spec_src = str(sp.get('spectrum_source') or '')
            lw_oct = {int(k): float(v) for k, v in (sp.get('lw_octave') or {}).items()}
            sref = {int(k): float(v) for k, v in (sp.get('spectrum_template_ref') or {}).items()}
            try:
                delta_db = float(sp.get('spectrum_delta_db'))
                if not (delta_db == delta_db):
                    delta_db = None
            except Exception:
                delta_db = None
            rows = []
            for f in OCTAVE_BANDS:
                sref_txt = '-'
                if f in sref:
                    sref_txt = f"{sref[f]:.2f}"
                lw_txt = '-'
                if f in lw_oct:
                    lw_txt = f"{lw_oct[f]:.2f}"
                a_txt = f"{float(A_WEIGHTING.get(f, 0.0)):.1f}"
                rows.append(f"<tr><td>{f}</td><td style='text-align:right;'>{sref_txt}</td><td style='text-align:right;'>{a_txt}</td><td style='text-align:right;'>{lw_txt}</td></tr>")
            delta_line = ''
            if delta_db is not None:
                delta_line = f"<p><b>Δ appliqué :</b> {delta_db:.2f} dB. Décalage constant qui transforme la forme de référence en spectre final, <code>Lw,b = S_b^ref + Δ</code> ; sa valeur absorbe le LwA cible et l’ajustement de la pondération A.</p>"
            origin_line = '<p><b>Interprétation :</b> le spectre final <code>Lw,b</code> est celui qui entre réellement dans l’équation par bandes. Si <code>S_b^ref</code> existe, il correspond à la forme de référence avant l’ajustement global <code>Δ</code>.</p>' if sref else '<p><b>Interprétation :</b> pour ce groupe, aucune forme interne visible n’a été utilisée ; le spectre final <code>Lw,b</code> provient directement du spectre chargé/importé ou d’une bibliothèque externe.</p>'
            spectrum_detail_blocks.append(f"""
                <div class='card'>
                    <h4>2.1 Spectre utilisé par le groupe source : {group_name}</h4>
                    <p><b>Modèle :</b> {model_name} · <b>Origine du spectre :</b> {spec_src or '-'}.</p>
                    <p><b>Ce que représente chaque colonne :</b> <code>S_b^ref</code> est la forme spectrale de référence (si elle existe), <code>A_weight,b</code> la pondération A de chaque bande et <code>Lw,b</code> le niveau final en dB réellement utilisé par le calcul.</p>
                    {delta_line}
                    <table>
                        <tr><th>Bande [Hz]</th><th style='text-align:right;'>S_b^ref [dB]</th><th style='text-align:right;'>A_weight,b [dB]</th><th style='text-align:right;'>Lw,b final [dB]</th></tr>
                        {''.join(rows)}
                    </table>
                    {origin_line}
                </div>
            """)
        spectrum_detail_html = ''.join(spectrum_detail_blocks)

        def _fmt_equation_term(value: float) -> str:
            try:
                v = float(value)
            except Exception:
                return '-'
            if v != v:
                return '-'
            if abs(v) < 0.005:
                return '0.00'
            return f"{v:.2f}"

        if crit:
            def _crit_value(*keys, default=None):
                for key in keys:
                    try:
                        val = crit.get(key)
                    except Exception:
                        val = None
                    if val is None:
                        continue
                    try:
                        if isinstance(val, float) and val != val:
                            continue
                    except Exception:
                        pass
                    if str(val).strip() == '':
                        continue
                    return val
                return default

            def _crit_float(*keys, default=0.0):
                val = _crit_value(*keys, default=None)
                if val is None:
                    return float(default)
                try:
                    f = float(val)
                    if f != f:
                        return float(default)
                    return f
                except Exception:
                    return float(default)

            crit_id = _crit_value('rec_id', 'fid', default='-')
            crit_level = _crit_float('nivel_total_dba', 'total_level_dba', 'noise_dba', default=max_noise)
            crit_limit = _crit_float('limite_aplicado_dba', 'limit_dba', default=45.0)
            crit_margin = _crit_float('margen_limite_db', 'limit_margin_db', 'margin_db', default=crit_level - crit_limit)
            crit_model = _crit_value('modelo_dominante', 'dominant_model', 'dom_model', default='-')
            crit_group = _crit_value('grupo_fuente_dominante', 'dominant_source_group', 'dom_group', default='-')
            crit_n_turb = _crit_value('n_turbinas_en_radio', 'turbines_in_radius', 'n_src', default='-')
            crit_lwa = _crit_float('lwa_fuente_dom_dba', 'source_lwa_dba', 'src_lwa', default=0.0)
            crit_dist = _crit_float('dist_fuente_dom_3d_m', 'source_receiver_3d_m', 'dist3d_m', 'near_m', default=0.0)
            crit_adiv = _crit_float('perdida_divergencia_db', 'divergence_loss_db', 'adiv_db', default=0.0)
            crit_aatm = _crit_float('perdida_atmosferica_db', 'atmospheric_loss_db', 'aatm_db', default=0.0)
            crit_agr = _crit_float('perdida_suelo_db', 'ground_loss_db', 'aground_db', default=0.0)
            crit_abar = _crit_float('perdida_barrera_db', 'barrier_loss_db', 'abar_db', default=0.0)
            crit_abar_max = _crit_float('perdida_barrera_max_db', 'barrier_loss_max_contributors_db', 'abar_max_db', default=crit_abar)
            crit_abar_mean = _crit_float('perdida_barrera_media_db', 'barrier_loss_mean_contributors_db', 'abar_mean_db', default=crit_abar)
            crit_abar_ew = _crit_float('perdida_barrera_ponderada_db', 'barrier_loss_energy_weighted_db', 'abar_ew_db', default=crit_abar)
            crit_abar_screen_n = _crit_value('n_fuentes_apantalladas', 'barrier_screened_sources_n', 'abar_screen_n', default=0)
            try:
                crit_abar_screen_n = int(crit_abar_screen_n or 0)
            except Exception:
                crit_abar_screen_n = 0
            crit_g_eff = _crit_float('factor_suelo_g', 'ground_factor_g', 'ground_g', default=float(g_eff_stats.get('critical', g)))
            crit_freq = _crit_value('banda_dominante_hz', 'dominant_band_hz', 'dom_freq', default='-')
            crit_spec_src = _crit_value('origen_espectro', 'spectrum_source', 'spec_src', default='-')
            crit_abar_state = str(_crit_value('mdt_abar_state', 'abar_state', default='') or '').strip()
            crit_obs_h = _crit_float('mdt_obstacle_height_m', 'obs_h_m', default=0.0)
            crit_obs_d1 = _crit_float('mdt_d1_m', 'obs_d1_m', default=0.0)
            crit_obs_d2 = _crit_float('mdt_d2_m', 'obs_d2_m', default=0.0)
            crit_obs_thr = _crit_float('mdt_obstacle_threshold_m', 'obs_thr_m', default=0.0)
            crit_src_z = _crit_float('dominant_source_ground_z_m', 'src_z_m', default=float('nan'))
            crit_hub_h = _crit_float('dominant_source_hub_height_m', 'hub_h_m', default=float('nan'))
            crit_src_ac_z = _crit_float('dominant_source_acoustic_z_m', 'src_ac_z_m', default=float('nan'))
            crit_rec_z = _crit_float('receiver_ground_z_m', 'rec_z_m', default=float('nan'))
            crit_rec_h = _crit_float('receiver_height_agl_m', 'rec_h_m', default=float('nan'))
            crit_rec_ac_z = _crit_float('receiver_acoustic_z_m', 'rec_ac_z_m', default=float('nan'))
            crit_maxab_src = _crit_value('max_abar_source_index', 'maxab_src', default='-')
            crit_maxab_state = str(_crit_value('max_abar_mdt_state', 'maxab_state', default='') or '').strip()
            crit_maxab_obs_h = _crit_float('max_abar_obstacle_height_m', 'maxab_obs_h', default=0.0)
            crit_maxab_d1 = _crit_float('max_abar_source_obstacle_m', 'maxab_d1', default=0.0)
            crit_maxab_d2 = _crit_float('max_abar_obstacle_receiver_m', 'maxab_d2', default=0.0)
            
            status_badge = 'badge-success' if crit_margin <= 0 else 'badge-danger'
            status_text = 'CONFORME' if crit_margin <= 0 else 'DÉPASSE'
            card_class = 'card-success' if crit_margin <= 0 else 'card-danger'

            crit_adiv_txt = _fmt_equation_term(crit_adiv)
            crit_aatm_txt = _fmt_equation_term(crit_aatm)
            crit_agr_txt = _fmt_equation_term(crit_agr)
            crit_abar_txt = _fmt_equation_term(crit_abar)
            crit_agr_desc = f"Atténuation due à l’effet de sol (G_eff={crit_g_eff:.2f})"
            crit_abar_desc = "Atténuation due au MDT sur le trajet dominant"
            crit_abar_max_txt = _fmt_equation_term(crit_abar_max)
            crit_abar_mean_txt = _fmt_equation_term(crit_abar_mean)
            crit_abar_ew_txt = _fmt_equation_term(crit_abar_ew)
            try:
                crit_n_turb_i = int(crit_n_turb)
            except Exception:
                crit_n_turb_i = 0

            def _fmt_m_or_na(v):
                try:
                    f = float(v)
                    if f != f:
                        return 'N/A'
                    return f"{f:.2f}"
                except Exception:
                    return 'N/A'

            dominant_height_html = (
                f"<br><b>Hauteurs du trajet dominant :</b> terrain éolienne={_fmt_m_or_na(crit_src_z)} m · "
                f"hub={_fmt_m_or_na(crit_hub_h)} m AGL · hauteur acoustique éolienne={_fmt_m_or_na(crit_src_ac_z)} m · "
                f"terrain récepteur={_fmt_m_or_na(crit_rec_z)} m · h récepteur={_fmt_m_or_na(crit_rec_h)} m AGL · "
                f"hauteur acoustique récepteur={_fmt_m_or_na(crit_rec_ac_z)} m."
            )
            maxabar_height_html = ''
            if float(crit_abar_max or 0.0) > 0.005:
                maxabar_height_html = (
                    f"<br><b>Trajet avec Abar maximal :</b> source={crit_maxab_src} · état={crit_maxab_state or '-'} · "
                    f"obs={_fmt_m_or_na(crit_maxab_obs_h)} m · d1={_fmt_m_or_na(crit_maxab_d1)} m · d2={_fmt_m_or_na(crit_maxab_d2)} m."
                )
            abar_summary_html = ''
            if dem_used and engine == 'iso_aligned':
                abar_summary_html = f"""
                <div class='note'>
                    <b>Lecture correcte d’Abar :</b> la valeur <b>Abar du trajet dominant</b> correspond uniquement à l’éolienne qui contribue le plus au récepteur et à sa bande dominante. Le niveau total du récepteur est obtenu par sommation énergétique de toutes les éoliennes et bandes.
                    <br><b>Abar maximal parmi les éoliennes contributrices :</b> {crit_abar_max_txt} dB · <b>Abar moyen :</b> {crit_abar_mean_txt} dB · <b>Abar pondéré par contribution énergétique :</b> {crit_abar_ew_txt} dB · <b>trajets écrantés :</b> {crit_abar_screen_n}/{crit_n_turb_i if crit_n_turb_i else crit_n_turb}.
                    {dominant_height_html}
                    {maxabar_height_html}
                </div>
                """

            abar_note_html = ''
            if dem_used and engine == 'iso_aligned':
                if abs(float(crit_abar)) < 0.005:
                    reason_map = {
                        'los_clear': 'la ligne de visée entre l’éolienne dominante et ce récepteur est dégagée selon le MDT',
                        'below_threshold': 'un relief a été détecté, mais sous le seuil conservateur d’activation',
                        'no_profile': 'aucun profil MDT valide n’a pu être extrait pour le trajet dominant',
                        'no_dem': 'aucun MDT n’était disponible sur ce trajet',
                    }
                    reason = reason_map.get(crit_abar_state, 'aucun obstacle topographique pertinent n’a été détecté sur le trajet dominant')
                    extra = ''
                    if float(crit_obs_thr) > 0.0:
                        extra = f" Seuil d’activation: {crit_obs_thr:.2f} m."
                    if float(abar_stats.get('max', 0.0) or 0.0) > 0.005:
                        extra += f" D’autres récepteurs présentent bien un écran (Abar max. {float(abar_stats.get('max',0.0)):.2f} dB)."
                    abar_note_html = f"<p style='margin:8px 0 10px 0;color:#495057;'><i>Lecture MDT : Abar=0 au récepteur critique ne signifie pas que le MDT est désactivé ; cela signifie que {reason}.{extra}</i></p>"
                else:
                    abar_note_html = f"<p style='margin:8px 0 10px 0;color:#495057;'><i>Lecture MDT : obstacle dominant estimé {crit_obs_h:.2f} m; d1={crit_obs_d1:.1f} m, d2={crit_obs_d2:.1f} m; état={crit_abar_state or 'actif'}.</i></p>"

            crit_html = f"""
        <div class='{card_class}'>
            <h3>🎯 Récepteur critique (niveau sonore le plus élevé)</h3>
            
            <table style='margin-bottom: 20px;'>
                <tr>
                    <td style='width: 50%; padding-right: 20px;'>
                        <p><b>ID récepteur :</b> {crit_id}</p>
                        <p><b>Niveau total :</b> <span style='font-size:28px; font-weight:bold; color:{'#dc3545' if crit_margin > 0 else '#28a745'};'>{crit_level:.2f} dB(A)</span></p>
                        <p><b>Limite applicable :</b> {crit_limit:.2f} dB(A)</p>
                        <p><b>Marge :</b> {crit_margin:+.2f} dB <span class='{status_badge}'>{status_text}</span></p>
                    </td>
                    <td style='width: 50%;'>
                        <p><b>Modèle dominant :</b> {crit_model}</p>
                        <p><b>Groupe source :</b> {crit_group}</p>
                        <p><b>Éoliennes contributrices dans le rayon :</b> {crit_n_turb}</p>
                        <p><b>Distance :</b> {crit_dist:.1f} m</p>
                    </td>
                </tr>
            </table>
            
            <h4>📊 Décomposition des atténuations</h4>
            <p style='margin: 6px 0 10px 0; color:#495057;'><i>Les valeurs affichées ci-dessous sont les amplitudes d’atténuation utilisées par le modèle. Dans l’équation principale, ces termes sont soustraits au niveau de source.</i></p>
            <table style='margin: 16px 0;'>
                <tr>
                    <th>Terme</th>
                    <th style='text-align: right;'>Valeur [dB]</th>
                    <th>Description</th>
                </tr>
                <tr style='background: #e3f2fd;'>
                    <td><b>LwA source dominante</b></td>
                    <td style='text-align: right;'><b>{crit_lwa:.2f}</b></td>
                    <td>Puissance acoustique de l’éolienne</td>
                </tr>
                <tr>
                    <td>Adiv (divergence)</td>
                    <td style='text-align: right;'>{crit_adiv_txt}</td>
                    <td>Dispersion géométrique</td>
                </tr>
                <tr>
                    <td>Aatm (atmosphérique)</td>
                    <td style='text-align: right;'>{crit_aatm_txt}</td>
                    <td>Absorption dans l’air</td>
                </tr>
                <tr>
                    <td>Agr (sol)</td>
                    <td style='text-align: right;'>{crit_agr_txt}</td>
                    <td>{crit_agr_desc}</td>
                </tr>
                <tr>
                    <td>Abar trajet dominant</td>
                    <td style='text-align: right;'>{crit_abar_txt}</td>
                    <td>{crit_abar_desc}</td>
                </tr>
                <tr>
                    <td>Abar maximal des contributeurs</td>
                    <td style='text-align: right;'>{crit_abar_max_txt}</td>
                    <td>Abar maximal parmi toutes les éoliennes qui contribuent au récepteur</td>
                </tr>
                <tr>
                    <td>Abar pondéré par énergie</td>
                    <td style='text-align: right;'>{crit_abar_ew_txt}</td>
                    <td>Moyenne pondérée par la contribution acoustique de chaque éolienne</td>
                </tr>
                <tr>
                    <td>Trajets écrantés</td>
                    <td style='text-align: right;'>{crit_abar_screen_n}/{crit_n_turb}</td>
                    <td>Nombre d’éoliennes contributrices avec Abar &gt; 0 dB</td>
                </tr>
                <tr style='background: #1e3a5f; color: white; font-weight: bold;'>
                    <td>NIVEAU RÉSULTANT</td>
                    <td style='text-align: right;'>{crit_level:.2f}</td>
                    <td>dB(A)</td>
                </tr>
            </table>
            {abar_note_html}
            {abar_summary_html}
            <p style='margin: 6px 0 10px 0; color:#495057;'><i>Note : le niveau résultant inclut la sommation énergétique multi-source et multi-bande ; ce n’est pas une soustraction directe depuis une seule éolienne.</i></p>
            
            <p style='margin-top: 16px;'>
                <b>Bande dominante :</b> {crit_freq} Hz &nbsp;&nbsp;&nbsp;
                <b>Origine du spectre :</b> {crit_spec_src}
            </p>
        </div>
            """
        else:
            crit_html = "<div class='card'><p>Récepteur critique non disponible.</p></div>"

        rec_types_html = ''.join([f"<li><b>{k}:</b> {v}</li>" for k, v in sorted(receiver_type_counts.items())])
        compliance = self._res.get('receiver_type_compliance') or {}
        compliance_html = ''.join([f"<li><b>{k}:</b> {int((v or {}).get('exceed',0))}/{int((v or {}).get('total',0))} dépassent la limite" + (f" · couverts {int((v or {}).get('covered',0))}" if (v or {}).get('covered') is not None else '') + "</li>" for k, v in sorted(compliance.items())])
        suelo_txt = 'global' if ground_mode != 'landuse' else f"depuis couche ({landuse_layer_name or 'sans nom'})"
        grid_txt = 'non généré'
        if self._res.get('grid_layer') is not None:
            grid_txt = f"oui · résolution demandée {float(grid_diag.get('requested_resolution_m',0.0)):.1f} m · effective {float(grid_diag.get('effective_resolution_m',0.0)):.1f} m"
            if bool(grid_diag.get('auto_adjusted', False)):
                grid_txt += ' · auto-ajustée'
        limit_mode = str(limit_stats.get('mode') or 'global').lower()
        limit_scn = str(limit_stats.get('scenario') or 'custom').lower()
        if limit_mode == 'by_field':
            scn_txt = {'day': 'diurne', 'night': 'nocturne', 'custom': 'personnalisé'}.get(limit_scn, limit_scn or 'personnalisé')
            if abs(float(limit_stats.get('min',45.0)) - float(limit_stats.get('max',45.0))) < 1e-9:
                limit_html = f"<p><b>Limites appliquées :</b> depuis les champs des récepteurs ({scn_txt}) · valeur unique {float(limit_stats.get('min',45.0)):.1f} dB(A)</p>"
            else:
                limit_html = f"<p><b>Limites appliquées :</b> depuis les champs des récepteurs ({scn_txt}) · plage {float(limit_stats.get('min',45.0)):.1f}–{float(limit_stats.get('max',45.0)):.1f} dB(A)</p>"
        else:
            limit_html = f"<p><b>Limite de référence :</b> {float(limit_stats.get('max',45.0)):.1f} dB(A)</p>"

        equations_html = f"<pre style='background:#f6f8fb;border:1px solid #d9e2ef;padding:10px;border-radius:6px;white-space:pre-wrap;'>{equation}</pre>"

        if not crit:
            crit_adiv_txt = crit_aatm_txt = crit_agr_txt = crit_abar_txt = '-'
            crit_agr_desc = 'Effet du sol'
            crit_abar_desc = 'Diffraction topographique'

        param_lines = [
            f"<li><b>Moteur :</b> {engine_label}</li>",
            f"<li><b>Hauteur du récepteur :</b> {rec_h:.1f} m</li>",
            f"<li><b>Rayon maximal :</b> {radius:.0f} m</li>",
            f"<li><b>Mode sol :</b> {suelo_txt}</li>",
        ]
        if ground_mode == 'landuse':
            param_lines.extend([
                f"<li><b>G global de secours:</b> {g:.2f}</li>",
                f"<li><b>G_eff moyen utilisé:</b> {float(g_eff_stats.get('mean', g)):.2f}</li>",
                f"<li><b>G_eff du récepteur critique utilisé:</b> {float(g_eff_stats.get('critical', g)):.2f}</li>",
            ])
        else:
            param_lines.extend([
                f"<li><b>G utilisé:</b> {g:.2f}</li>",
                f"<li><b>G_eff moyen:</b> {float(g_eff_stats.get('mean', g)):.2f}</li>",
                f"<li><b>G_eff du récepteur critique:</b> {float(g_eff_stats.get('critical', g)):.2f}</li>",
            ])
        param_lines.extend([
            f"<li><b>MDT/DSM:</b> {'oui · ' + (dem_layer_name or 'sans nom') if dem_used else 'non'}</li>",
            f"<li><b>Occupation du sol:</b> {'oui · ' + (landuse_layer_name or 'sans nom') if bool(report.get('landuse_used', False)) else 'non'}</li>",
            f"<li><b>Scénario acoustique :</b> {acoustic_txt}</li>",
        ])
        if engine == 'iso_aligned':
            param_lines.extend([
                f"<li><b>Température :</b> {temp_c:.1f} °C</li>",
                f"<li><b>Humidité relative :</b> {hum_pct:.1f} %</li>",
                f"<li><b>Pression :</b> {pressure_kpa:.3f} kPa</li>",
            ])
        else:
            param_lines.append(f"<li><b>α atmosphérique :</b> {alpha:.4f} dB/m</li>")

        term_lines = [
            f"<li><b>Adiv:</b> {'actif' if terms.get('Adiv', True) else 'non'}</li>",
            f"<li><b>Aatm:</b> {'actif' if terms.get('Aatm', True) else 'non'}" + (' (T, HR, P simplifié)' if engine == 'iso_aligned' else ' (α·distance)') + "</li>",
            f"<li><b>Agr/Aground:</b> {'actif' if terms.get('Agr', True) else 'non'}</li>",
            f"<li><b>Abar:</b> {'actif' if terms.get('Abar', False) else 'inactif'}</li>",
            f"<li><b>G effectif depuis l’occupation du sol:</b> {'oui' if terms.get('landuse_g', False) else 'non'}</li>",
        ]

        pressure_warning_html = ''
        if engine == 'iso_aligned' and (pressure_kpa < 85.0 or pressure_kpa > 105.0):
            pressure_warning_html = (
                "<p class='note'><b>Révision recommandée :</b> la pression atmosphérique saisie "
                f"({pressure_kpa:.3f} kPa) est hors de la plage typique utilisée comme référence dans de nombreuses études "
                "préliminaires. Si ce n’est pas une mesure du site, vérifier si elle devrait être proche de 101,325 kPa "
                "ou ajustée à l’altitude.</p>"
            )

        interpretation = (
            "Adiv représente la divergence géométrique. Aatm est calculé par bande et dépend de T, HR et de la pression, avec une formulation simplifiée. "
            "Agr est appliqué comme terme de sol/terrain et Abar comme écran topographique de base lorsqu’un MDT est disponible."
            if engine == 'iso_aligned' else
            "Adiv représente la divergence géométrique, Aatm l’atténuation atmosphérique simplifiée α·distance et Aground une correction simplifiée de l’effet de sol/terrain."
        )

        if engine == 'iso_aligned':
            methodology_flow_html = f"""
            <div class='card card-info'>
                <h3>🧭 Comment le calcul ISO-aligned a été exécuté</h3>
                <p>Cette section explique le flux réel suivi par le plugin afin que le résultat par récepteur soit traçable. Le niveau final de chaque récepteur <b>ne provient pas d’une simple soustraction unique</b>, mais du calcul de toutes les contributions source–récepteur dans le rayon de calcul, puis de leur sommation énergétique.</p>
                <ol>
                    <li><b>Lecture des entrées SIG:</b> les éoliennes/sources acoustiques, les récepteurs, la hauteur du récepteur et le rayon maximal de calcul sont pris en compte (<b>{radius:.0f} m</b>), la couche d’occupation du sol si elle existe et le MDT/DSM s’il est actif.</li>
                    <li><b>État acoustique de chaque groupe source :</b> pour chaque modèle ou groupe d’éoliennes, un <b>LwA opérationnel</b> est obtenu à partir d’une valeur fixe ou d’une courbe <code>LwA(ws)</code>. Dans ce calcul: <b>{acoustic_txt}</b>.</li>
                    <li><b>Conversion en bandes:</b> le moteur ISO-aligned a besoin d’un spectre <code>Lw,b</code> en 8 bandes d’octave. S’il n’existe pas de spectre spécifique, le plugin en reconstruit un à partir d’un gabarit/fallback et l’ajuste pour reproduire le LwA opérationnel.</li>
                    <li><b>Sélection des contributeurs par récepteur:</b> pour chaque récepteur, les éoliennes situées dans le rayon maximal sont recherchées. Les récepteurs sans sources dans ce rayon sont marqués comme <b>hors rayon</b> et ne produisent pas de niveau acoustique utile.</li>
                    <li><b>Calcul par trajet source–récepteur :</b> pour chaque éolienne contributrice, la distance 3D, les cotes acoustiques, <b>G</b> ou <b>G_eff</b> du sol et, si un MDT/DSM est disponible, l’éventuel écran topographique du trajet sont calculés.</li>
                    <li><b>Propagation par bande:</b> dans chaque bande, on applique <code>Lp,b = Lw,b - Adiv - Aatm,b - Agr,b - Abar,b</code>. Adiv dépend de la distance, Aatm,b de la fréquence/de l’atmosphère, Agr,b du sol et Abar,b du MDT s’il existe un obstacle pertinent.</li>
                    <li><b>Sommation par source :</b> les 8 bandes sont pondérées A puis sommées énergétiquement pour obtenir le niveau pondéré A de cette éolienne au récepteur.</li>
                    <li><b>Sommation du récepteur:</b> toutes les éoliennes contributrices sont sommées énergétiquement pour obtenir le <b>niveau total dB(A)</b> du récepteur.</li>
                    <li><b>Comparaison avec les limites:</b> le niveau total est comparé à la limite attribuée au récepteur ou à la limite de référence. La marge, l’état de conformité et le tableau des dépassements en découlent.</li>
                </ol>
                <div class='formula'>LpA,récepteur = 10·log10(Σ_sources 10^(LpA,source/10))</div>
                <p><b>Lecture pratique :</b> le récepteur critique est celui qui présente le niveau total le plus élevé ou la marge la plus défavorable par rapport à la limite. La colonne « source dominante » identifie l’éolienne/le groupe qui contribue le plus, mais le résultat final du récepteur inclut toutes les sources dans le rayon.</p>
            </div>
            <div class='card'>
                <h3>🔎 Ce qui distingue ce mode du mode Screening</h3>
                <p>Le mode ISO-aligned est plus lourd mais plus traçable : il utilise les bandes d’octave, la pondération A finale, l’absorption atmosphérique dépendante de la fréquence, le sol par régions et l’écran topographique <b>Abar</b> lorsqu’un MDT/DSM est disponible. C’est le mode recommandé pour les rapports techniques préliminaires et la revue des récepteurs sensibles.</p>
            </div>
            """
        else:
            methodology_flow_html = f"""
            <div class='card card-info'>
                <h3>🧭 Comment le calcul Screening a été exécuté</h3>
                <p>Cette section explique le flux réel suivi par le plugin en mode rapide. L’objectif est d’obtenir une estimation agile pour les cartes, la comparaison d’alternatives et la détection initiale des récepteurs sensibles.</p>
                <ol>
                    <li><b>Lecture des entrées SIG:</b> les éoliennes/sources acoustiques, les récepteurs, la hauteur du récepteur et le rayon maximal de calcul sont pris en compte (<b>{radius:.0f} m</b>) et la couche d’occupation du sol si elle existe.</li>
                    <li><b>État acoustique de chaque groupe source :</b> chaque modèle ou groupe d’éoliennes utilise un seul <b>LwA opérationnel</b>, défini par une valeur fixe ou par une courbe <code>LwA(ws)</code>. Dans ce calcul: <b>{acoustic_txt}</b>.</li>
                    <li><b>Sélection des contributeurs par récepteur:</b> pour chaque récepteur, les éoliennes situées dans le rayon maximal sont recherchées. Les récepteurs sans sources dans ce rayon sont marqués comme <b>hors rayon</b>.</li>
                    <li><b>Calcul par trajet source–récepteur :</b> pour chaque éolienne contributrice, la distance 3D, la divergence géométrique, une absorption atmosphérique simplifiée <code>α·d</code> et une correction empirique de sol sont calculées.</li>
                    <li><b>Occupation du sol:</b> si une couche de land-use est disponible, le plugin peut calculer un <b>G_eff</b> par trajet ; sinon, il utilise le <b>G global</b> défini par l’utilisateur.</li>
                    <li><b>Propagation simplifiée:</b> <code>Lp = LwA - Adiv - Aatm - Aground</code> est appliqué. Il n’y a ni bandes d’octave ni écran topographique explicite <code>Abar</code>.</li>
                    <li><b>Sommation du récepteur:</b> toutes les éoliennes contributrices sont sommées énergétiquement pour obtenir le <b>niveau total dB(A)</b> du récepteur.</li>
                    <li><b>Comparaison avec les limites:</b> le niveau total est comparé à la limite attribuée au récepteur ou à la limite de référence. La marge, l’état de conformité et le tableau des dépassements en découlent.</li>
                </ol>
                <div class='formula'>LpA,récepteur = 10·log10(Σ_sources 10^(Lp,source/10))</div>
                <p><b>Lecture pratique :</b> ce mode est utile pour le criblage initial. Si un récepteur apparaît proche de la limite ou en dépassement, il est conseillé de le recalculer en mode ISO-aligned et de revoir les spectres, le terrain, l’occupation du sol et les limites appliquées.</p>
            </div>
            <div class='card'>
                <h3>🔎 Ce qui distingue ce mode du mode ISO-aligned</h3>
                <p>Le mode Screening sacrifie le détail pour gagner en vitesse. Il ne propage pas par bandes, n’utilise pas T/HR/P par fréquence, ne calcule pas <b>Abar</b> depuis le MDT et résume l’atmosphère avec un coefficient unique <b>α</b>. Il doit donc être interprété comme une préévaluation rapide, et non comme un rapport acoustique détaillé.</p>
            </div>
            """

        octave_rows = ''.join([
            f"<tr><td>{freq}</td><td style='text-align:right;'>{float(a_w):.1f}</td></tr>"
            for freq, a_w in [(63, -26.2), (125, -16.1), (250, -8.6), (500, -3.2), (1000, 0.0), (2000, 1.2), (4000, 1.0), (8000, -1.1)]
        ])
        atm_rows = ''.join([
            f"<tr><td>{freq}</td><td style='text-align:right;'>{alpha_ref:.4f}</td></tr>"
            for freq, alpha_ref in [(63, 0.0001), (125, 0.0003), (250, 0.0008), (500, 0.0020), (1000, 0.0040), (2000, 0.0095), (4000, 0.0280), (8000, 0.0900)]
        ])
        ground_rows = ''.join([
            "<tr><td>≤ 500 Hz</td><td style='text-align:right;'>A_ground = 1.5 dB</td></tr>",
            "<tr><td>1000 Hz</td><td style='text-align:right;'>1.5·(1 - e^(-h/10))</td></tr>",
            "<tr><td>2000 Hz</td><td style='text-align:right;'>3.0·(1 - e^(-h/10))</td></tr>",
            "<tr><td>4000 Hz</td><td style='text-align:right;'>6.0·(1 - e^(-h/10))</td></tr>",
            "<tr><td>8000 Hz</td><td style='text-align:right;'>12.0·(1 - e^(-h/10))</td></tr>",
        ])

        if engine == 'iso_aligned':
            if dem_used:
                mdt_expl_html = f"""
                <div class='card'>
                    <h3>🗺️ Physique du MDT et de l’écran topographique</h3>
                    <p>Dans le moteur ISO-aligned, le MDT <b>ne modifie pas l’émission de l’éolienne</b> ni l’absorption atmosphérique. Sa fonction est de décrire la <b>géométrie réelle du trajet source–récepteur</b> et d’alimenter le terme d’écran topographique <b>Abar,b</b>.</p>
                    <div class='formula'>Lp,b = Lw,b - Adiv - Aatm,b - Agr,b - Abar,b</div>
                    <h4>Comment le MDT entre dans le calcul</h4>
                    <ol>
                        <li><b>Profil du terrain :</b> le profil source–récepteur est extrait du MDT avec un <b>échantillonnage adaptatif</b>, ajusté à la distance et à la résolution du raster. Le profil est calculé <b>une seule fois</b> par paire source–récepteur et réutilisé sur les 8 bandes afin de réduire le temps de calcul.</li>
                        <li><b>Ligne de visée directe :</b> le profil est comparé à la droite reliant la source acoustique à sa hauteur effective et le récepteur à sa hauteur d’évaluation. Si le terrain reste toujours sous cette droite, il n’y a pas d’obstacle topographique pertinent et <b>Abar,b = 0</b>.</li>
                        <li><b>Détection de l’obstacle dominant :</b> si une colline ou une crête du MDT dépasse au-dessus de la ligne de visée, le modèle considère qu’il existe un écran topographique. La grandeur clé est la hauteur de l’obstacle au-dessus de la ligne de visée:</li>
                    </ol>
                    <div class='formula'>h_obs = z_terrain - z_LOS</div>
                    <p>lorsque <b>h_obs &gt; 0</b>, le relief coupe la vision directe et une atténuation supplémentaire par diffraction peut apparaître.</p>
                    <ol start='4'>
                        <li><b>Activation conservatrice:</b> Abar n’est pas activé pour de petites irrégularités du relief ; un seuil minimal lié à la résolution du MDT est appliqué.</li>
                        <li><b>Géométrie réelle de l’obstacle :</b> le calcul utilise la <b>position réelle</b> de l’obstacle dominant et obtient <b>d1</b> (source → obstacle) et <b>d2</b> (obstacle → récepteur) réels, au lieu de supposer systématiquement un obstacle au point médian.</li>
                        <li><b>Diffraction de type Fresnel:</b> avec cette géométrie, une différence de chemins approximative est estimée et transformée en atténuation dépendante de la fréquence:</li>
                    </ol>
                    <div class='formula'>δ ≈ 0.5·h_obs²·(1/d1 + 1/d2) &nbsp;&nbsp; ; &nbsp;&nbsp; C = (2·f·δ)/c</div>
                    <p>où <b>δ</b> est la différence de chemins approximative, <b>f</b> la fréquence et <b>c</b> la vitesse du son. Le nombre <b>C</b> est ensuite traduit en une atténuation <b>Abar,b</b>, d’autant plus élevée que le relief bloque le trajet. C’est l’approximation implémentée dans le calcul.</p>
                    <p><b>Interprétation physique:</b> en terrain plat ou en l’absence d’intersection avec la ligne de visée, <b>Abar</b> est généralement négligeable. En terrain complexe, le MDT peut introduire plusieurs dB d’atténuation supplémentaire et modifier le récepteur critique.</p>
                    <p><b>Implémentation actuelle :</b> obstacle dominant unique, profil adaptatif avec limites de coût, géométrie réelle de l’obstacle, activation conservatrice et atténuation plafonnée à des valeurs raisonnables.</p>
                    <p><b>MDT utilisé dans ce calcul:</b> {dem_layer_name or 'sans nom'}.</p>
                </div>
                """
            else:
                mdt_expl_html = """
                <div class='card'>
                    <h3>🗺️ Physique du MDT et de l’écran topographique</h3>
                    <p>Dans ce calcul, <b>aucun MDT/DSM n’a été utilisé</b>, donc le terme d’écran topographique est fixé à:</p>
                    <div class='formula'>Abar,b = 0</div>
                    <p>L’évaluation est réalisée sans introduire d’écrans topographiques. La géométrie du trajet est résolue sans profil de terrain et le calcul dépend de Lw,b, Adiv, Aatm,b et Agr,b.</p>
                </div>
                """

            if ground_mode == 'landuse':
                ground_expl_html = f"""
                <div class='card'>
                    <h3>🌱 Physique de l’occupation du sol et calcul de G_eff</h3>
                    <p>Lorsque le mode sol est <b>depuis une couche</b>, le calcul n’utilise pas une seule valeur manuelle pour tout le parc. Pour chaque trajet source–récepteur, un <b>G_eff</b> est calculé depuis la couche d’occupation du sol:</p>
                    <div class='formula'>G_eff = (Σ G_i · L_i) / (Σ L_i)</div>
                    <p>où <b>G_i</b> est la valeur attribuée à chaque polygone intercepté par le trajet et <b>L_i</b> la longueur du trajet à l’intérieur de ce polygone.</p>
                    <ul>
                        <li><b>G = 0</b>: sol dur (urbano/asfalto/roca).</li>
                        <li><b>G = 0.5</b>: terrain mixte.</li>
                        <li><b>G = 1</b>: sol meuble/poreux (agricole, prairie, forestier, végétalisé).</li>
                    </ul>
                    <p><b>Important:</b> le <b>G global</b> affiché dans le rapport est uniquement une valeur de secours. Lorsqu’une couche d’occupation du sol est disponible, le calcul utilise réellement <b>G_eff</b> par trajet. Dans ce calcul, la valeur effective moyenne était <b>{float(g_eff_stats.get('mean', g)):.2f}</b> et celle du récepteur critique <b>{float(g_eff_stats.get('critical', g)):.2f}</b>.</p>
                    <p><b>Couche utilisée:</b> {landuse_layer_name or 'sans nom'}.</p>
                </div>
                """
            else:
                ground_expl_html = f"""
                <div class='card'>
                    <h3>🌱 Physique de l’occupation du sol et calcul de G</h3>
                    <p>Dans ce calcul, l’effet de sol a été calculé avec un <b>G manuel unique</b> pour tout le trajet:</p>
                    <div class='formula'>G = {g:.2f}</div>
                    <p>Cette valeur est appliquée dans le terme de sol du modèle. Aucun G_eff n’a été dérivé depuis une couche d’occupation du sol.</p>
                </div>
                """

            equations_detail_html = f"""
            <div class='card'>
                <h3>📘 Développement physique détaillé du moteur ISO-aligned</h3>
                <p>Ce moteur travaille en <b>8 bandes d’octave</b> (63–8000 Hz). Les bandes ne sont pas un résultat du calcul, mais la <b>grille fréquentielle de la méthode</b>. Pour appliquer la propagation par bandes, le calcul a besoin d’une <b>entrée acoustique par bande</b> de la source <code>Lw,b</code>. Cette entrée peut provenir d’un spectre mesuré/importé ou d’un gabarit/fallback ajusté au niveau global opérationnel.</p>
                <p><b>Scénario opérationnel de ce calcul:</b> {acoustic_txt}.</p>
                <p><b>Équation générale par bande:</b></p>
                <div class='formula'>Lp,b = Lw,b - Adiv - Aatm,b - Agr,b - Abar,b</div>
                <p><b>Sommation finale pondérée A:</b></p>
                <div class='formula'>LpA,total = 10·log10(Σ 10^((Lp,b + A_weight)/10))</div>
                <h4>0. Entrées réellement utilisées dans ce calcul</h4>
                <ul>
                    <li><b>Source acoustique :</b> <code>Lw,b</code> par bandes d’octave. S’il existe un spectre spécifique du groupe source, c’est l’entrée utilisée. Sinon, le plugin utilise une bibliothèque/un gabarit/un fallback et l’ajuste au niveau global opérationnel.</li>
                    <li><b>Niveau opérationnel global:</b> il provient d’un <b>LwA fixe</b> ou d’une <b>courbe acoustique LwA(ws)</b> selon le scénario sélectionné. Ce niveau global ne remplace pas les bandes : il fixe l’état opérationnel et le spectre fournit la répartition fréquentielle.</li>
                    <li><b>Géométrie :</b> coordonnées de source et de récepteur, hauteur du récepteur, hauteur effective de source et distance 3D.</li>
                    <li><b>Atmosphère:</b> température <b>T</b>, humidité relative <b>HR</b> et pression <b>P</b>.</li>
                    <li><b>Sol:</b> un <b>G global manuel</b> ou un <b>G_eff</b> dérivé depuis la couche d’occupation du sol.</li>
                    <li><b>Topographie:</b> MDT/DSM optionnel. Il n’affecte que le calcul de <b>Abar,b</b>.</li>
                </ul>
                <h4>1. Origine de chaque terme de l’équation</h4>
                <table>
                    <tr><th>Terme</th><th>Comment il est obtenu dans ce plugin</th></tr>
                    <tr><td><b>Lw,b</b></td><td>Entrée acoustique par bandes. Elle provient du spectre du groupe source (CSV, bibliothèque, gabarit ou fallback ajusté au niveau global). La courbe acoustique LwA(ws) ou le LwA fixe définit le niveau global opérationnel de l’éolienne, et le spectre par bandes répartit ce niveau entre les 8 bandes.</td></tr>
                    <tr><td><b>Adiv</b></td><td>Calculé à partir de la distance 3D source–récepteur.</td></tr>
                    <tr><td><b>Aatm,b</b></td><td>Calculé par bande avec une table de base d’absorption <code>α_ref(f)</code> et des corrections simplifiées de température, humidité relative et pression. L’implémentation actuelle utilise la formulation exacte du plugin : <code>α = α_ref(f)·corr_T·corr_HR·corr_P</code>.</td></tr>
                    <tr><td><b>Agr,b</b></td><td>Calculé comme effet de sol par régions. Le paramètre de sol utilisé est un <b>G unique par trajet</b> : manuel/global ou <b>G_eff</b> dérivé de la couche d’occupation du sol.</td></tr>
                    <tr><td><b>Abar,b</b></td><td>N’intervient que s’il existe un MDT/DSM et si un écran topographique est détecté. En l’absence de MDT ou d’obstacle pertinent, <b>Abar,b = 0</b>.</td></tr>
                </table>
                <h4>2. Entrée acoustique de la source et bandes</h4>
                <p>Dans ce moteur, le terme <code>Lw,b</code> est une <b>donnée d’entrée par bande</b>. Les <b>bandes d’octave</b> (63–8000 Hz) ne sont pas un résultat ISO ni un tableau calculé par le plugin : ce sont la <b>grille fréquentielle</b> sur laquelle la propagation est résolue.</p>
                <p>Le plugin combine deux éléments:</p>
                <ul>
                    <li><b>Courbe acoustique globale LwA(ws)</b>: fixe le <b>niveau opérationnel global</b> de l’éolienne pour la vitesse de vent ou le cas le plus défavorable sélectionné.</li>
                    <li><b>Spectre par bandes Lw,b</b>: répartit ce niveau global entre les 8 bandes et constitue l’entrée réelle utilisée dans l’équation par bandes.</li>
                </ul>
                <p>Ce spectre peut provenir d’un fichier spécifique du fabricant/de l’utilisateur ou d’un gabarit de référence. Si seule une courbe globale <code>LwA(ws)</code> est disponible, le plugin fixe d’abord le niveau global opérationnel <code>LwA_cible</code>, puis construit un spectre absolu par bandes à partir d’une forme spectrale de référence <code>S_b^ref</code>.</p>
                <p><b>Reconstruction mathématique des bandes lorsqu’il n’existe que LwA(ws):</b></p>
                <div class='formula'>Lw,b = S_b^ref + Δ</div>
                <div class='formula'>Δ = LwA_cible - 10·log10(Σ 10^((S_b^ref + A_weight,b)/10))</div>
                <p>Autrement dit : la courbe acoustique fournit le <b>niveau global opérationnel</b> et le gabarit/la bibliothèque fournit la <b>forme spectrale</b>. Le décalage <b>Δ</b> est calculé de façon à ce que, après pondération A et sommation énergétique des 8 bandes, le spectre reconstruit reproduise exactement le <code>LwA_cible</code> de la courbe importée.</p>
                {spectrum_detail_html}
                <h4>3. Divergence géométrique</h4>
                <div class='formula'>Adiv = 20·log10(d) + 11</div>
                <p>Représente la dispersion géométrique de l’onde sonore avec la distance 3D source–récepteur. Ici, <b>d</b> provient des coordonnées de l’éolienne et du récepteur avec leurs hauteurs d’évaluation.</p>
                <h4>4. Absorption atmosphérique simplifiée</h4>
                <div class='formula'>Aatm,b = α(f, T, HR, P) · d</div>
                <p>L’absorption atmosphérique est calculée par bande à partir d’un coefficient de référence et de trois facteurs correcteurs. La dépendance physique à la température, l’humidité relative et la pression <b>est bien représentée</b>, mais au moyen d’une <b>approximation simplifiée du plugin</b>, et non de la formulation analytique complète de l’ISO 9613-1.</p>
                <div class='formula'>α(f, T, HR, P) = α_ref(f) · corr_T · corr_HR · corr_P</div>
                <div class='formula'>corr_T = 1 + 0.01·(T - 15) &nbsp;&nbsp; ; &nbsp;&nbsp; corr_HR = 1 + 0.003·|HR - 50| &nbsp;&nbsp; ; &nbsp;&nbsp; corr_P = 101.325 / P</div>
                <p><b>Interprétation des corrections:</b> <b>T</b> est introduit en °C par rapport à une référence de 15 °C ; <b>HR</b> est comparée à une humidité optimale de référence de 50 % et la correction augmente lorsque l’on s’en éloigne ; <b>P</b> est introduite en kPa par rapport à une référence de 101,325 kPa avec une correction inverse. Ces facteurs ne modifient que le bloc atmosphérique <b>Aatm,b</b> : ils ne modifient ni l’émission de l’éolienne, ni l’effet de sol, ni le terme MDT/écran.</p>
                <table>
                    <tr><th>Bande [Hz]</th><th style='text-align:right;'>α_ref [dB/m]</th></tr>
                    {atm_rows}
                </table>
                <h4>5. Effet de sol par régions</h4>
                <div class='formula'>Agr,b = As + Am + Ar</div>
                <p>Le terme de sol se décompose en <b>As</b> (région de source), <b>Am</b> (région intermédiaire) et <b>Ar</b> (région du récepteur). Dans cette implémentation, trois paramètres de sol indépendants <code>Gs/Gm/Gr</code> ne sont pas utilisés ; un <b>G unique par trajet</b> est utilisé. Mathématiquement, le plugin applique :</p>
                <div class='formula'>As = G_eff·A_ground(h_s)</div>
                <div class='formula'>Am = G_eff·(1 - G_m)·A_ground(h_medio)</div>
                <div class='formula'>Ar = G_eff·A_ground(h_r)</div>
                <p>où <b>h_s</b> est la hauteur caractéristique de la source, <b>h_r</b> celle du récepteur, <b>h_moy</b> la hauteur moyenne du trajet et <b>G_m≈0</b> dans l’approximation actuelle pour des conditions favorables de propagation. Cette valeur unique de sol peut être :</p>
                <ul>
                    <li><b>G manuel/global</b>, si l’utilisateur fixe une valeur unique.</li>
                    <li><b>G_eff</b>, si une couche d’occupation du sol existe et si une moyenne pondérée par la longueur du trajet est calculée.</li>
                </ul>
                <div class='formula'>G_eff = (Σ G_i · L_i) / (Σ L_i)</div>
                <p><b>Signification physique de G:</b> représente le caractère acoustique du terrain et contrôle l’influence du sol sur la propagation. <b>G≈0</b> indique un sol dur (urbain, asphalte, roche), <b>G≈1</b> un sol meuble/poreux (agricole, prairie, forestier) et les valeurs intermédiaires représentent un terrain mixte.</p>
                <p><b>Ce que signifie « depuis couche » :</b> le plugin intersecte le trajet source–récepteur avec la couche d’occupation du sol, attribue une valeur <b>G_i</b> à chaque polygone intercepté et calcule un <b>G_eff</b> unique pour ce trajet. C’est cette valeur qui entre réellement dans <b>Agr,b</b> ; le <b>G global</b> affiché dans le rapport reste uniquement une valeur de secours.</p>
                <p><b>Convention du rapport :</b> <b>Agr,b</b> est affiché ici comme une <b>amplitude positive d’atténuation</b>. Dans l’équation principale, il est soustrait au niveau de source comme Adiv, Aatm et Abar.</p>
                <table>
                    <tr><th>Bande [Hz]</th><th style='text-align:right;'>Terme base A_ground(h)</th></tr>
                    {ground_rows}
                </table>
                <h4>6. Écran topographique avec MDT</h4>
                <p>Le MDT <b>ne modifie pas l’émission</b> de l’éolienne ni l’absorption atmosphérique. Sa fonction est de décrire la <b>géométrie réelle du trajet</b> et d’alimenter le terme <b>Abar,b</b>.</p>
                <ol>
                    <li><b>Profil du terrain :</b> le profil source–récepteur est extrait du MDT avec un échantillonnage adaptatif.</li>
                    <li><b>Ligne de visée:</b> la droite entre la hauteur effective de source et la hauteur du récepteur est construite. Si le terrain reste toujours en dessous, alors <b>Abar,b = 0</b>.</li>
                    <li><b>Obstacle dominant:</b> si une colline ou une crête dépasse, la hauteur au-dessus de la ligne de visée est calculée :</li>
                </ol>
                <div class='formula'>h_obs = z_terrain - z_LOS</div>
                <p>Lorsque <b>h_obs &gt; 0</b>, le relief coupe la vision directe et une atténuation supplémentaire par diffraction peut apparaître.</p>
                <ol start='4'>
                    <li><b>Géométrie réelle de l’obstacle :</b> le plugin utilise la position réelle de l’obstacle dominant et calcule <b>d1</b> (source → obstacle) et <b>d2</b> (obstacle → récepteur).</li>
                    <li><b>Activation conservatrice:</b> <b>Abar</b> n’est pas activé pour de petites irrégularités du MDT ; un seuil minimal lié à la résolution du raster est exigé.</li>
                    <li><b>Diffraction de type Fresnel:</b> avec cette géométrie, une différence de chemins et un nombre de Fresnel sont estimés :</li>
                </ol>
                <div class='formula'>δ ≈ 0.5·h_obs²·(1/d1 + 1/d2) &nbsp;&nbsp; ; &nbsp;&nbsp; C = (2·f·δ)/c</div>
                <p>Ce nombre est ensuite transformé en une atténuation <b>Abar,b</b> dépendante de la fréquence au moyen de l’approximation actuelle du plugin :</p>
                <div class='formula'>si C ≤ -2 → Abar = 0 &nbsp;&nbsp; ; &nbsp;&nbsp; -2 &lt; C ≤ 0 → Abar = 10·log10(3 + 20·C)</div>
                <div class='formula'>0 &lt; C ≤ 3.5 → Abar = 10·log10(3 + 80·C) &nbsp;&nbsp; ; &nbsp;&nbsp; C &gt; 3.5 → Abar = 10·log10(3 + 280·C)</div>
                <p>Dans l’implémentation actuelle, <b>Abar</b> est également limité à des valeurs raisonnables (plafonnement supérieur) afin d’éviter des suratténuations parasites. En l’absence de MDT ou d’obstacle pertinent, alors <b>Abar,b = 0</b>.</p>
                <h4>7. Pondération A utilisée à la fin</h4>
                <table>
                    <tr><th>Bande [Hz]</th><th style='text-align:right;'>A_weight [dB]</th></tr>
                    {octave_rows}
                </table>
                <p><b>Lecture du récepteur critique:</b> le tableau de la section du récepteur critique affiche des amplitudes d’atténuation pour la traçabilité. Le <b>niveau résultant</b> ne doit pas être interprété comme une soustraction directe depuis une seule éolienne : il est obtenu par sommation énergétique par bandes et par sommation des sources contributrices dans le rayon de calcul.</p>
            </div>
            {ground_expl_html}
            {mdt_expl_html}
            """
        else:
            if ground_mode == 'landuse':
                fast_ground_html = f"""
                <h4>3. Effet de sol simplifié avec occupation du sol</h4>
                <p>Dans le moteur rapide, le terme <b>Aground</b> reste empirique, mais le paramètre de sol peut provenir de la couche d’occupation du sol sous forme de <b>G_eff</b> par trajet :</p>
                <div class='formula'>G_eff = (Σ G_i · L_i) / (Σ L_i)</div>
                <div class='formula'>Aground = min(6, max(0, G_eff · 3·log10(1 + d_xy/100) · 1/(1 + (h_s + h_r)/80)))</div>
                <p>Ce <b>G_eff</b> est ensuite utilisé dans la correction simplifiée du terrain du moteur rapide. La valeur globale <b>G = {g:.2f}</b> reste uniquement une valeur de secours si la couche ne fournit pas d’information valide.</p>
                <p><b>Couche utilisée:</b> {landuse_layer_name or 'sans nom'} · <b>G_eff moyen:</b> {float(g_eff_stats.get('mean', g)):.2f} · <b>G_eff du récepteur critique:</b> {float(g_eff_stats.get('critical', g)):.2f}</p>
                """
            else:
                fast_ground_html = f"""
                <h4>3. Effet de sol simplifié</h4>
                <p>Le terme <b>Aground</b> est une correction empirique du terrain contrôlée par un seul paramètre manuel :</p>
                <div class='formula'>G = {g:.2f}</div>
                <div class='formula'>Aground = min(6, max(0, G · 3·log10(1 + d_xy/100) · 1/(1 + (h_s + h_r)/80)))</div>
                <p>Dans ce calcul, aucun G_eff n’a été dérivé depuis une couche d’occupation du sol. Ici, <b>d_xy</b> est la distance horizontale, <b>h_s</b> la hauteur de source et <b>h_r</b> la hauteur du récepteur.</p>
                """

            fast_mdt_html = """
                <h4>4. MDT / topographie</h4>
                <p>Dans le moteur rapide, le MDT n’introduit pas de terme explicite d’écran topographique. Même si une couche de relief existe dans le projet, ce mode ne calcule pas <b>Abar</b>, n’extrait pas de ligne de visée et n’applique pas de diffraction ; la physique se base donc uniquement sur <b>LwA</b>, <b>Adiv</b>, <b>Aatm = α·d</b> et la correction empirique de terrain <b>Aground</b>.</p>
            """

            equations_detail_html = f"""
            <div class='card'>
                <h3>📘 Développement physique détaillé du moteur rapide</h3>
                <div class='formula'>Lp = LwA - Adiv - Aatm - Aground</div>
                <p>Le moteur rapide travaille avec un seul niveau global <b>LwA</b> par groupe source. Il est conçu pour le criblage, les cartes rapides et les comparaisons rapides, en sacrifiant le détail spectral au profit de la vitesse. Dans ce mode, il n’y a <b>pas de propagation par bandes</b> ni de terme explicite d’écran topographique.</p>
                <p><b>Scénario opérationnel de ce calcul:</b> {acoustic_txt}.</p>
                <h4>0. Entrées réellement utilisées dans ce calcul</h4>
                <ul>
                    <li><b>Source acoustique :</b> un seul niveau global <b>LwA</b> par groupe source.</li>
                    <li><b>Niveau opérationnel global:</b> provient d’un <b>LwA fixe</b> ou d’une <b>courbe acoustique LwA(ws)</b> pour la vitesse ou le cas le plus défavorable sélectionnés.</li>
                    <li><b>Géométrie :</b> coordonnées de source et de récepteur, hauteur du récepteur, hauteur effective de source et distance 3D.</li>
                    <li><b>Atmosphère:</b> dans ce mode, T/HR/P ne sont pas utilisés ; l’absorption est résumée par un coefficient unique <b>α</b>.</li>
                    <li><b>Sol:</b> un <b>G global manuel</b> ou un <b>G_eff</b> dérivé depuis la couche d’occupation du sol.</li>
                    <li><b>Topographie :</b> le MDT n’entre pas comme écran explicite dans ce mode.</li>
                </ul>
                <h4>1. Origine de chaque terme de l’équation</h4>
                <table>
                    <tr><th>Terme</th><th>Comment il est obtenu dans ce plugin</th></tr>
                    <tr><td><b>LwA</b></td><td>Entrée globale de la source. Elle provient d’une valeur fixe par groupe ou d’une courbe acoustique <code>LwA(ws)</code> pour la vitesse/le pire cas sélectionné.</td></tr>
                    <tr><td><b>Adiv</b></td><td>Calculé à partir de la distance 3D source–récepteur.</td></tr>
                    <tr><td><b>Aatm</b></td><td>Calculé avec un coefficient constant unique <code>α</code> multiplié par la distance.</td></tr>
                    <tr><td><b>Aground</b></td><td>Correction empirique de l’effet de sol. Le paramètre de sol peut être un <b>G global manuel</b> ou un <b>G_eff</b> dérivé de la couche d’occupation du sol.</td></tr>
                </table>
                <h4>2. Divergence géométrique</h4>
                <div class='formula'>Adiv = 20·log10(d) + 11</div>
                <p>Représente la dispersion géométrique de l’onde sonore avec la distance 3D source–récepteur.</p>
                <h4>3. Absorption atmosphérique simplifiée</h4>
                <div class='formula'>Aatm = α · d</div>
                <p>Dans ce calcul, <b>α = {alpha:.4f} dB/m</b> a été utilisé. Dans le moteur rapide, l’absorption atmosphérique est résumée par un seul coefficient constant ; <b>T</b>, <b>HR</b> et <b>P</b> <b>n’entrent donc pas explicitement</b> dans le calcul. C’est l’une des simplifications clés par rapport au mode ISO-aligned.</p>
                {fast_ground_html}
                {fast_mdt_html}
                <h4>5. Ce que ce mode ne fait pas</h4>
                <p>Le moteur rapide ne travaille pas par bandes, ne calcule pas <b>Lw,b</b>, n’introduit pas <b>Abar</b> et n’extrait ni ligne de visée ni diffraction depuis le MDT. Il est donc adapté au criblage et aux comparaisons rapides, mais pas à l’analyse spectrale détaillée.</p>
            </div>
            """

        # === CALCULAR TASAS Y FECHA ===
        coverage_rate = (100.0 * n_with / n_receivers) if n_receivers else 0
        exceed_rate = (100.0 * n_exceed / n_with) if n_with else 0
        comply_rate = 100.0 - exceed_rate
        from datetime import datetime
        now = datetime.now()

        # === BANNER DE ALCANCE (lo primero que se lee, antes de cualquier cifra) ===
        if engine == 'iso_aligned':
            scope_what_is = "une évaluation acoustique préliminaire alignée sur la méthodologie ISO 9613-2, destinée à la conception, à la comparaison d’alternatives et au criblage des récepteurs sensibles."
            scope_what_not = "ce n’est pas un rapport acoustique certifié et ne remplace pas une étude réglementaire définitive réalisée avec un logiciel commercial validé."
            scope_simpl_items = [
                "Absorption atmosphérique Aatm via une table de référence avec corrections simplifiées de température, humidité et pression, et non la formulation analytique complète de l’ISO 9613-1.",
                "Sans correction météorologique de long terme Cmet.",
                "Diffraction topographique d’un obstacle dominant unique : sans diffraction latérale ni écrans multiples.",
                "Résolution spectrale en 8 bandes d’octave de 63 à 8000 Hz, pas en tiers d’octave.",
                "Directivité de source Dc supposée égale à 0 dB.",
            ]
        else:
            scope_what_is = "une estimation rapide de criblage pour des cartes agiles et la comparaison d’alternatives d’implantation."
            scope_what_not = "ce n’est ni un calcul spectral détaillé ni un rapport réglementaire ; pour les récepteurs proches de la limite, il convient de recalculer en mode ISO-aligned."
            scope_simpl_items = [
                "Sans propagation par bandes d’octave.",
                "Absorption atmosphérique résumée par un seul coefficient alpha constant.",
                "Sans écran topographique Abar depuis le MDT.",
                "Effet de sol via une correction empirique simplifiée.",
            ]
        scope_reco = "Pour les décisions réglementaires critiques, validez les résultats avec des mesures de terrain ou un logiciel commercial certifié."
        scope_items_html = ''.join(f"<li>{it}</li>" for it in scope_simpl_items)
        scope_banner_html = f"""
        <div style='background:#fff8e1;border:2px solid #f0ad4e;border-left:8px solid #f0ad4e;border-radius:8px;padding:18px 22px;margin:0 0 26px 0;'>
            <h3 style='margin:0 0 10px 0;color:#7a5b00;'>⚠️ Portée de ce rapport — à lire avant d’utiliser les résultats</h3>
            <p style='margin:6px 0;'><b>Ce que c’est :</b> {scope_what_is}</p>
            <p style='margin:6px 0;'><b>Ce que ce n’est pas :</b> {scope_what_not}</p>
            <p style='margin:10px 0 4px 0;'><b>Simplifications appliquées dans ce mode :</b></p>
            <ul style='margin:4px 0 10px 0;'>{scope_items_html}</ul>
            <p style='margin:6px 0 0 0;'><b>Recommandation :</b> {scope_reco}</p>
        </div>
        """

        # === GLOSARIO DE SÍMBOLOS (decodifica fórmulas y tablas en un solo sitio) ===
        glossary_rows = [
            ("LwA", "Niveau de puissance acoustique pondéré A de la source, en dB(A)."),
            ("Lw,b", "Puissance acoustique de la source par bande d’octave, en dB."),
            ("S_b^ref", "Forme spectrale de référence par bande utilisée comme gabarit, en dB."),
            ("A_weight,b", "Pondération A appliquée à chaque bande d’octave, en dB."),
            ("Δ", "Décalage global appliqué au gabarit spectral pour reproduire le LwA cible, en dB."),
            ("LpA", "Niveau de pression acoustique pondéré A résultant au récepteur, en dB(A)."),
            ("Adiv", "Atténuation par divergence géométrique avec la distance, en dB."),
            ("Aatm", "Atténuation due à l’absorption atmosphérique de l’air, en dB."),
            ("Agr", "Atténuation due à l’effet de sol, en dB."),
            ("Abar", "Atténuation due à l’écran topographique, uniquement en mode ISO avec MDT, en dB."),
            ("d", "Distance tridimensionnelle entre source et récepteur, en mètres."),
            ("G / G_eff", "Facteur de sol de 0 (dur) à 1 (meuble) et sa valeur effective par trajet."),
            ("Cmet", "Correction météorologique de long terme, non appliquée dans ce plugin."),
            ("Dc", "Correction de directivité de la source, supposée égale à 0 dB."),
        ]
        glossary_rows_html = ''.join(f"<tr><td><b>{sym}</b></td><td>{desc}</td></tr>" for sym, desc in glossary_rows)
        glossary_html = f"""
        <div class='card card-info'>
            <h3>📖 Glossaire des symboles</h3>
            <p>Définition compacte des symboles qui apparaissent dans les formules et tableaux de ce rapport.</p>
            <table>
                <tr><th>Symbole</th><th>Signification</th></tr>
                {glossary_rows_html}
            </table>
        </div>
        """

        html = f"""
        <style>
            body {{
                font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
                line-height: 1.6;
                color: #212529;
            }}
            h1, h2, h3 {{
                color: #1e3a5f;
                font-weight: 600;
                margin-top: 24px;
                margin-bottom: 12px;
            }}
            h2 {{
                border-left: 4px solid #4a90d9;
                padding-left: 12px;
            }}
            .card {{
                background: #f8f9fa;
                border: 1px solid #e9ecef;
                border-radius: 8px;
                padding: 20px;
                margin: 16px 0;
                box-shadow: 0 2px 4px rgba(0,0,0,0.05);
            }}
            .card-success {{
                border-left: 5px solid #28a745;
            }}
            .card-danger {{
                border-left: 5px solid #dc3545;
            }}
            .card-info {{
                border-left: 5px solid #4a90d9;
            }}
            .metrics-grid {{
                display: grid;
                grid-template-columns: repeat(3, 1fr);
                gap: 16px;
                margin: 20px 0;
            }}
            .metric {{
                background: white;
                padding: 20px;
                border-radius: 8px;
                box-shadow: 0 2px 8px rgba(0,0,0,0.1);
                text-align: center;
                border-top: 4px solid #4a90d9;
            }}
            .metric-value {{
                font-size: 32px;
                font-weight: 700;
                color: #1e3a5f;
                margin: 8px 0;
            }}
            .metric-label {{
                font-size: 14px;
                color: #343a40;
                font-weight: 500;
            }}
            table {{
                width: 100%;
                border-collapse: collapse;
                margin: 16px 0;
                font-size: 13px;
            }}
            th {{
                background: #1e3a5f;
                color: white;
                padding: 12px;
                text-align: left;
                font-weight: 600;
            }}
            td {{
                padding: 10px 12px;
                border-bottom: 1px solid #e9ecef;
            }}
            tr:nth-child(even) {{
                background: #f8f9fa;
            }}
            .badge {{
                display: inline-block;
                padding: 4px 12px;
                border-radius: 12px;
                font-size: 11px;
                font-weight: 700;
                text-transform: uppercase;
            }}
            .badge-success {{
                background: #28a745;
                color: white;
            }}
            .badge-danger {{
                background: #dc3545;
                color: white;
            }}
            .formula {{
                background: #f1f3f5;
                border: 1px solid #dee2e6;
                padding: 16px;
                margin: 12px 0;
                border-radius: 6px;
                font-family: 'Courier New', monospace;
                font-size: 14px;
            }}
            .disclaimer {{
                background: #fff3cd;
                border-left: 5px solid #ffc107;
                padding: 16px;
                margin: 20px 0;
                border-radius: 4px;
            }}
            .note {{
                background: #fff8e1;
                border-left: 4px solid #f0ad4e;
                padding: 10px 12px;
                margin: 10px 0;
                border-radius: 4px;
                color: #5f4300;
            }}
            ol {{
                margin: 12px 0;
                padding-left: 26px;
            }}
            ol li {{
                margin: 8px 0;
            }}
            ul {{
                margin: 12px 0;
                padding-left: 24px;
            }}
            li {{
                margin: 6px 0;
            }}
        </style>
        
        <table width='100%' bgcolor='#1e3a5f' cellpadding='18' cellspacing='0' style='margin-bottom:30px; border-collapse:separate;'><tr><td style='border-bottom:none;'>
            <span style='color:#ffffff; font-size:30px; font-weight:bold;'>📊 RAPPORT TECHNIQUE D’IMPACT ACOUSTIQUE</span><br/>
            <span style='color:#d8e4f5; font-size:16px;'>Évaluation du bruit généré par les éoliennes</span><br/>
            <span style='color:#d8e4f5; font-size:14px;'>📅 {now.strftime('%d/%m/%Y - %H:%M:%S')}</span>
        </td></tr></table>
        
        {scope_banner_html}
        
        <h2>1. RÉSUMÉ EXÉCUTIF</h2>
        
        <table width='100%' cellpadding='14' cellspacing='8' style='margin:16px 0; border-collapse:separate;'><tr>
            <td width='33%' align='center' bgcolor='#ffffff' style='border-bottom:none;'>
                <span style='font-size:12px; color:#5a6b7f; letter-spacing:1px;'>ÉOLIENNES</span><br/>
                <span style='font-size:30px; font-weight:700; color:#1e3a5f;'>{n_sources}</span>
            </td>
            <td width='33%' align='center' bgcolor='#ffffff' style='border-bottom:none;'>
                <span style='font-size:12px; color:#5a6b7f; letter-spacing:1px;'>RÉCEPTEURS ÉVALUÉS</span><br/>
                <span style='font-size:30px; font-weight:700; color:#1e3a5f;'>{n_receivers}</span>
            </td>
            <td width='34%' align='center' bgcolor='#ffffff' style='border-bottom:none;'>
                <span style='font-size:12px; color:#5a6b7f; letter-spacing:1px;'>NIVEAU MAXIMAL (dB(A))</span><br/>
                <span style='font-size:30px; font-weight:700; color:#1e3a5f;'>{max_noise:.1f}</span>
            </td>
        </tr></table>
        
        <table width='100%' cellpadding='0' cellspacing='8' style='border-collapse:separate;'><tr><td width='50%' style='vertical-align:top; border-bottom:none;'>
            <div class='card card-{'success' if coverage_rate > 80 else 'info'}'>
                <h3>📍 Couverture de l’analyse</h3>
                <p><strong>{n_with} récepteurs</strong> dans le rayon<br>
                <strong>{coverage_rate:.1f}%</strong> de couverture<br>
                {n_without} récepteurs hors rayon</p>
            </div>
        </td><td width='50%' style='vertical-align:top; border-bottom:none;'>
            <div class='card card-{'success' if comply_rate > 90 else 'danger' if comply_rate < 50 else 'info'}'>
                <h3>✓ Conformité réglementaire</h3>
                <p><strong>{n_exceed} récepteurs</strong> dépassent les limites<br>
                <strong>{comply_rate:.1f}%</strong> de conformité sur les récepteurs couverts<br>
                Limite : {float(limit_stats.get('min',45)):.1f}–{float(limit_stats.get('max',45)):.1f} dB(A)</p>
            </div>
        </td></tr></table>
        
        <div class='card card-info'>
            <h3>🎯 Méthodologie de calcul</h3>
            <p><b>Moteur utilisé :</b> {engine_label}</p>
            <p><b>Groupes source acoustiques :</b> {n_models} modèle(s) d’éolienne</p>
            <p><b>Méthode :</b> {'Propagation par bandes d’octave selon la méthodologie ISO-aligned' if engine == 'iso_aligned' else 'Calcul acoustique simplifié pour le criblage'}</p>
            <p><b>Carte raster :</b> {grid_txt}</p>
        </div>

        <h2>2. COMMENT LE RÉSULTAT A ÉTÉ GÉNÉRÉ</h2>
        {methodology_flow_html}
        
        <h2>3. RÉCEPTEUR CRITIQUE</h2>
        {crit_html}
        
        <div class='card'>
            <h3>📊 Statistiques des atténuations (récepteurs couverts)</h3>
        <p style='margin: 6px 0 10px 0; color:#495057;'><i>Les amplitudes brutes d’atténuation sont affichées (et non le signe algébrique dans l’équation). Pour Abar, le maximum parmi les éoliennes contributrices de chaque récepteur est utilisé, pas uniquement le trajet dominant.</i></p>
            <table>
                <tr>
                    <th>Terme</th>
                    <th style='text-align: right;'>Moyenne [dB]</th>
                    <th style='text-align: right;'>Maximum [dB]</th>
                </tr>
                <tr>
                    <td><b>Adiv</b> (divergence géométrique)</td>
                    <td style='text-align: right;'>{float(adiv_stats.get('mean',0.0)):.2f}</td>
                    <td style='text-align: right;'>{float(adiv_stats.get('max',0.0)):.2f}</td>
                </tr>
                <tr>
                    <td><b>Aatm</b> (absorption atmosphérique)</td>
                    <td style='text-align: right;'>{float(aatm_stats.get('mean',0.0)):.2f}</td>
                    <td style='text-align: right;'>{float(aatm_stats.get('max',0.0)):.2f}</td>
                </tr>
                <tr>
                    <td><b>Agr/Aground</b> (effet de sol)</td>
                    <td style='text-align: right;'>{float(aground_stats.get('mean',0.0)):.2f}</td>
                    <td style='text-align: right;'>{float(aground_stats.get('max',0.0)):.2f}</td>
                </tr>
                <tr>
                    <td><b>Abar</b> (maximum parmi les contributeurs)</td>
                    <td style='text-align: right;'>{float(abar_stats.get('mean',0.0)):.2f}</td>
                    <td style='text-align: right;'>{float(abar_stats.get('max',0.0)):.2f}</td>
                </tr>
            </table>
        </div>
        
        <h2>4. CONFIGURATION ET PARAMÈTRES</h2>
        
        <div class='card'>
            <h3>⚙️ Équation utilisée</h3>
            <div class='formula'>{equation}</div>
            <p><em>{interpretation}</em></p>
        </div>
        
        <div class='card'>
            <h3>📋 Paramètres du calcul</h3>
            <ul>{''.join(param_lines)}</ul>
            {pressure_warning_html}
            <p><b>Trajets avec G différent du global :</b> {int(ground_diag.get('from_landuse_count',0))} ({float(ground_diag.get('from_landuse_pct',0.0)):.1f}%)</p>
        </div>
        
        <div class='card'>
            <h3>✓ Termes actifs</h3>
            <ul>{''.join(term_lines)}</ul>
        </div>
        
        <h2>5. PHYSIQUE DÉTAILLÉE ET TRAÇABILITÉ DU CALCUL</h2>
        {glossary_html}
        {equations_detail_html}
        
        <h2>6. GROUPES SOURCE ACOUSTIQUES</h2>
        <div class='card'>
            <h3>⚡ LwA effectif par groupe</h3>
            <ul>{''.join(eff_lines) if eff_lines else '<li>Non disponible</li>'}</ul>
        </div>
        
        <h2>7. DISTRIBUTION PAR TYPE DE RÉCEPTEUR</h2>
        <table width='100%' cellpadding='0' cellspacing='8' style='border-collapse:separate;'><tr><td width='50%' style='vertical-align:top; border-bottom:none;'>
            <div class='card'>
                <h3>📍 Récepteurs par catégorie</h3>
                <ul>{rec_types_html if rec_types_html else '<li>Non disponible</li>'}</ul>
            </div>
        </td><td width='50%' style='vertical-align:top; border-bottom:none;'>
            <div class='card'>
                <h3>✓ Conformité par catégorie</h3>
                <ul>{compliance_html if compliance_html else '<li>Non disponible</li>'}</ul>
            </div>
        </td></tr></table>
        
        <div class='disclaimer'>
            <strong>⚠️ Limites et recommandations</strong>
            <p><b>Moteur rapide :</b> Adapté au criblage préliminaire et aux cartes agiles.</p>
            <p><b>Moteur ISO-aligned :</b> Adapté aux études techniques préliminaires, aux comparaisons et à l’itération de conception.</p>
            <p><b>Simplifications connues :</b> Aatm simplifié (tables + corrections) ; Agr et Abar avec approximations de base ; directivité Dc supposée égale à 0 dB ; Cmet/correction météorologique de long terme non appliquée.</p>
            <p><b>Modèles multiples :</b> pris en charge au moyen de couches/groupes source indépendants. Mélanger plusieurs modèles dans une seule couche via attributs n’est pas activé dans cette version expérimentale.</p>
            <p><b>Raster ISO + MDT :</b> utilise la même logique d’écran topographique que les récepteurs ponctuels, mais peut être coûteux sur de grandes cartes.</p>
            <p><b>Recommandation :</b> Pour les études réglementaires critiques, valider avec des mesures ou un logiciel commercial certifié.</p>
        </div>
        """
        lang = str(current_language()).lower()
        if lang.startswith(("es", "en", "fr", "de")):
            # The legacy report template was assembled in French and earlier
            # cleanup passes translated it fragment by fragment. That created
            # hybrid sentences in ES/EN/DE. For the noise report summary we now
            # render the final HTML from native language blocks and the already
            # computed numeric context, so every visible static sentence belongs
            # to one language only.
            html = _render_native_noise_report(lang, dict(locals()))
        elif lang:
            html = translate_html(html)
        self.page_summary.document().setHtml(html)

    def _fill_models(self):
        model_diag = self._res.get("model_diag", {}) or {}
        rows: List[tuple] = []
        for name, d in model_diag.items():
            dia = d.get("diameter")
            hh = d.get("hub_height")
            mode = str(d.get('acoustic_mode') or 'fixed').lower()
            if mode == 'curve' and str(d.get('curve_path') or '').strip():
                note = str(d.get('curve_note') or _tr('Curva acústica activa'))
            else:
                note = _tr('LwA fijo por grupo de fuente acústica')
            rows.append((str(name), int(d.get("count", 0)), float(d.get("lwa", 0.0)), hh, dia, note))
        self.tbl_models.setRowCount(len(rows))
        for r, row in enumerate(rows):
            vals = [
                row[0], str(row[1]), f"{row[2]:.1f}",
                "-" if row[3] is None or (isinstance(row[3], float) and not (row[3] == row[3])) else f"{float(row[3]):.1f}",
                "-" if row[4] is None or (isinstance(row[4], float) and not (row[4] == row[4])) else f"{float(row[4]):.1f}",
                row[5],
            ]
            for c, v in enumerate(vals):
                it = QtWidgets.QTableWidgetItem(v)
                it.setFlags(it.flags() & ~QtCore.Qt.ItemIsEditable)
                self.tbl_models.setItem(r, c, it)
        self.tbl_models.resizeColumnsToContents()

    def _feature_value_last(self, feat, field_name, default=""):
        """Return the last field named ``field_name`` from a QgsFeature.

        Receiver input layers can already contain generic names such as
        ``state`` or ``limit_dba``. QGIS name lookup returns the first match,
        which can silently pick the original receiver attribute instead of the
        computed noise output. The computed fields are appended at the end, so
        use the last matching index for UI/export fallbacks.
        """
        try:
            fields = feat.fields()
            idx = -1
            for i in range(fields.count()):
                if fields.at(i).name() == field_name:
                    idx = i
            if idx >= 0:
                return feat.attribute(idx)
        except Exception:
            pass
        try:
            return feat[field_name]
        except Exception:
            return default

    def _fill_top_receivers(self):
        # Prefer named payload rows. They are created by the engine with stable
        # semantic keys and avoid both duplicate input-field names and raw
        # attribute-order shifts in the QGIS memory layer.
        payload_rows = self._payload_top_receivers()[:15]
        feats = []
        if not payload_rows:
            layer = self._res.get("result_layer")
            if isinstance(layer, QgsVectorLayer):
                try:
                    for f in layer.getFeatures():
                        feats.append(f)
                except Exception:
                    feats = []
            def keyf(f):
                try:
                    return float(self._feature_value_last(f, "noise_dba", 0.0) or 0.0)
                except Exception:
                    return -1e9
            feats = sorted(feats, key=keyf, reverse=True)[:15]
        row_count = len(payload_rows) if payload_rows else len(feats)
        self.tbl_top.setRowCount(row_count)
        iterable = payload_rows if payload_rows else feats
        for r, f in enumerate(iterable):
            if isinstance(f, dict):
                clean_row = self._clean_receiver_row(f)
            else:
                raw = {"fid": f.id()}
                for key in CONSULTANCY_RECEIVER_KEYS:
                    raw[key] = self._feature_value_last(f, key, "")
                clean_row = self._clean_receiver_row(raw)
            for c, header in enumerate(CONSULTANCY_RECEIVER_HEADERS):
                v = str(clean_row.get(header, ""))
                it = QtWidgets.QTableWidgetItem(v)
                it.setFlags(it.flags() & ~QtCore.Qt.ItemIsEditable)
                self.tbl_top.setItem(r, c, it)
        self.tbl_top.resizeColumnsToContents()


    def _fill_mdt_screening(self):
        """Fill a DEM/MDT audit table sorted by criblage, not by noise level."""
        rows = [dict(r) for r in self._payload_receiver_rows() if isinstance(r, dict)]

        def _f(d, key, default=0.0):
            try:
                v = d.get(key, default)
                if v is None or str(v).strip().lower() in ('', 'none', 'nan'):
                    return default
                return float(v)
            except Exception:
                return default

        # Keep covered receivers first.  Sort by active Abar, then by largest
        # detected obstacle, then by acoustic level.  This makes receivers with
        # strong terrain screening visible even if their total sound level is low.
        covered = [r for r in rows if _f(r, 'n_src', 0.0) > 0.0]
        covered.sort(
            key=lambda d: (
                _f(d, 'abar_max_db', 0.0),
                _f(d, 'maxobs_h', 0.0),
                _f(d, 'noise_dba', -1.0e99),
            ),
            reverse=True,
        )
        visible = covered[:30]

        keys = [
            'rec_id', 'noise_dba', 'n_src', 'abar_max_db', 'abar_ew_db',
            'abar_screen_n', 'abar_state', 'abar_db', 'maxab_src',
            'maxab_state', 'maxab_obs_h', 'maxab_thr', 'maxab_d1',
            'maxab_d2', 'maxobs_src', 'maxobs_state', 'maxobs_h',
            'maxobs_thr', 'maxobs_d1', 'maxobs_d2', 'rec_z_m',
            'rec_h_m', 'rec_ac_z_m', 'src_z_m', 'src_ac_z_m',
            'maxab_src_z', 'maxab_src_ac_z',
        ]

        self.tbl_mdt.setRowCount(len(visible))
        for r, row in enumerate(visible):
            for c, k in enumerate(keys):
                val = row.get(k, "")
                if k == "rec_id" and (val is None or str(val).strip() == ""):
                    val = row.get("fid", "")
                if val is None or str(val).strip().lower() in ('none', 'nan'):
                    v = "N/A"
                elif isinstance(val, float):
                    v = f"{val:.2f}"
                else:
                    v = str(val)
                it = QtWidgets.QTableWidgetItem(v)
                it.setFlags(it.flags() & ~QtCore.Qt.ItemIsEditable)
                self.tbl_mdt.setItem(r, c, it)
        self.tbl_mdt.resizeColumnsToContents()


    def _format_receiver_value(self, key: str, val) -> str:
        if key == "rec_id" and (val is None or str(val).strip() == ""):
            return ""
        if key == "exceeds":
            try:
                return _tr("sí") if int(float(val or 0)) == 1 else _tr("no")
            except Exception:
                txt = str(val or "").strip().lower()
                return _tr("sí") if txt in ("true", "yes", "sí", "si", "oui", "1") else _tr("no")
        if val is None:
            return "N/A"
        txt = str(val).strip()
        if txt.lower() in ("", "none", "nan", "n/a"):
            return "N/A"
        try:
            fval = float(txt.replace(",", "."))
        except Exception:
            return txt
        if not (fval == fval):
            return "N/A"
        if key in ("n_src",):
            return str(int(round(fval)))
        if key in ("noise_dba", "limit_dba", "margin_db", "src_lwa", "adiv_db", "aatm_db", "aground_db", "abar_max_db", "ground_g"):
            return f"{fval:.2f}"
        if key in ("near_m", "rec_h_m", "rec_z_m", "rec_ac_z_m"):
            return f"{fval:.1f}"
        return f"{fval:.2f}"


    def _clean_receiver_row(self, row: Dict[str, object]) -> Dict[str, object]:
        out: Dict[str, object] = {}
        for key, label in CONSULTANCY_RECEIVER_COLUMNS:
            val = row.get(key, "") if isinstance(row, dict) else ""
            if key == "rec_id" and (val is None or str(val).strip() == "") and isinstance(row, dict):
                val = row.get("fid", "")
            out[label] = self._format_receiver_value(key, val)
        return out


    def _receiver_rows_for_export(self) -> List[Dict[str, object]]:
        rows = self._res.get('receiver_rows') or []
        if not rows:
            layer = self._res.get('result_layer')
            if isinstance(layer, QgsVectorLayer):
                rows = list(self._iter_layer_dicts(layer))
        if not rows:
            rows = self._payload_top_receivers()
        return [self._clean_receiver_row(r) for r in rows if isinstance(r, dict)]



    def _write_layer_csv(self, layer: QgsVectorLayer, path: str):
        field_names = [f.name() for f in layer.fields()]
        with open(path, 'w', newline='', encoding='utf-8-sig') as fh:
            writer = csv.writer(fh, delimiter=';')
            writer.writerow(['fid'] + field_names)
            for feat in layer.getFeatures():
                row = [feat.id()]
                for name in field_names:
                    try:
                        val = feat[name]
                    except Exception:
                        val = ''
                    row.append(val)
                writer.writerow(row)

    def _iter_layer_dicts(self, layer: QgsVectorLayer):
        field_names = [f.name() for f in layer.fields()]
        for feat in layer.getFeatures():
            row = {"fid": feat.id()}
            for name in field_names:
                try:
                    row[name] = feat[name]
                except Exception:
                    row[name] = ""
            yield row

    def _collect_exceedance_rows(self):
        rows_source = self._res.get('receiver_rows') or []
        layer = self._res.get('result_layer')
        if not rows_source and isinstance(layer, QgsVectorLayer):
            rows_source = list(self._iter_layer_dicts(layer))
        rows = []
        for row in rows_source or []:
            try:
                exceeds = int(float(row.get('exceeds') or 0))
            except Exception:
                exceeds = 0
            if exceeds == 1:
                rows.append(self._clean_receiver_row(row))
        return rows

    def _write_rows_csv(self, rows, path: str):
        rows = list(rows or [])
        headers = []
        for r in rows:
            for k in r.keys():
                if k not in headers:
                    headers.append(k)
        with open(path, 'w', newline='', encoding='utf-8-sig') as fh:
            writer = csv.writer(fh, delimiter=';')
            writer.writerow(headers)
            for row in rows:
                writer.writerow([row.get(h, '') for h in headers])

    def _table_headers(self, table: QtWidgets.QTableWidget) -> List[str]:
        headers: List[str] = []
        for c in range(table.columnCount()):
            item = table.horizontalHeaderItem(c)
            headers.append(item.text() if item is not None else f"col_{c+1}")
        return headers

    def _collect_table_rows(self, table: QtWidgets.QTableWidget) -> List[Dict[str, object]]:
        headers = self._table_headers(table)
        rows: List[Dict[str, object]] = []
        for r in range(table.rowCount()):
            row: Dict[str, object] = {}
            has_value = False
            for c, h in enumerate(headers):
                item = table.item(r, c)
                text = item.text() if item is not None else ""
                if str(text).strip():
                    has_value = True
                row[h] = text
            if has_value:
                rows.append(row)
        return rows

    def _write_table_csv(self, table: QtWidgets.QTableWidget, path: str):
        headers = self._table_headers(table)
        with open(path, 'w', newline='', encoding='utf-8-sig') as fh:
            writer = csv.writer(fh, delimiter=';')
            writer.writerow(headers)
            for row in self._collect_table_rows(table):
                writer.writerow([row.get(h, '') for h in headers])

    def _append_table_sheet(self, wb, title: str, table: QtWidgets.QTableWidget):
        ws = wb.create_sheet(title=title[:31] or 'Hoja')
        headers = self._table_headers(table)
        ws.append(headers)
        rows = self._collect_table_rows(table)
        if not rows:
            ws.append(['sin_datos'])
        else:
            for row in rows:
                ws.append([row.get(h, '') for h in headers])
        try:
            for idx, h in enumerate(headers, start=1):
                width = max(len(str(h)), max((len(str(r.get(h, ''))) for r in rows), default=0))
                ws.column_dimensions[chr(64 + idx) if idx <= 26 else ws.cell(row=1, column=idx).column_letter].width = min(max(width + 2, 10), 45)
        except Exception:
            pass

    def _append_sheet(self, wb, title: str, rows):
        ws = wb.create_sheet(title=title[:31] or 'Hoja')
        rows = list(rows or [])
        headers = []
        for r in rows:
            for k in r.keys():
                if k not in headers:
                    headers.append(k)
        if not headers:
            ws.append(['sin_datos'])
            ws.append([''])
            return
        ws.append(headers)
        for row in rows:
            ws.append([row.get(h, '') for h in headers])
        try:
            for idx, h in enumerate(headers, start=1):
                width = max(len(str(h)), max((len(str(r.get(h, ''))) for r in rows), default=0))
                ws.column_dimensions[chr(64 + idx) if idx <= 26 else ws.cell(row=1, column=idx).column_letter].width = min(max(width + 2, 10), 40)
        except Exception:
            pass

    def _export_summary(self):
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, ('Zusammenfassung exportieren' if str(current_language()).lower().startswith('de') else _tr('Exportar resumen')), os.path.expanduser('~/schall_zusammenfassung.html' if str(current_language()).lower().startswith('de') else '~/ruido_resumen.html'), ('HTML (*.html);;Text (*.txt)' if str(current_language()).lower().startswith('de') else _tr('HTML (*.html);;Texto (*.txt)')))
        if not path:
            return
        try:
            if path.lower().endswith('.txt'):
                with open(path, 'w', encoding='utf-8') as fh:
                    fh.write(self.page_summary.toPlainText())
            else:
                if not path.lower().endswith('.html'):
                    path += '.html'
                with open(path, 'w', encoding='utf-8') as fh:
                    fh.write(self.page_summary.toHtml())
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, ('Zusammenfassung exportieren' if str(current_language()).lower().startswith('de') else _tr('Exportar resumen')), (f'Die Zusammenfassung konnte nicht exportiert werden:\n{e}' if str(current_language()).lower().startswith('de') else _tr('No se pudo exportar el resumen:') + f'\n{e}'))

    def _export_receivers_csv(self):
        rows = self._receiver_rows_for_export()
        if not rows:
            QtWidgets.QMessageBox.information(self, _tr('Exportar receptores'), _tr('No hay filas de receptores para exportar.'))
            return
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, _tr('Exportar receptores CSV'), os.path.expanduser('~/ruido_receptores.csv'), 'CSV (*.csv)')
        if not path:
            return
        try:
            if not path.lower().endswith('.csv'):
                path += '.csv'
            self._write_rows_csv(rows, path)
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, _tr('Exportar receptores'), _tr('No se pudo exportar el CSV:') + f'\n{e}')

    def _write_dict_rows_csv(self, rows, path: str):
        # Deterministic CSV for dictionaries. Keeps debug exports independent
        # from visible table columns and QGIS field ordering.
        keys = []
        for row in rows or []:
            if not isinstance(row, dict):
                continue
            for k in row.keys():
                if k not in keys:
                    keys.append(k)
        with open(path, 'w', newline='', encoding='utf-8-sig') as fh:
            writer = csv.writer(fh, delimiter=';')
            writer.writerow(keys)
            for row in rows or []:
                writer.writerow([row.get(k, '') if isinstance(row, dict) else '' for k in keys])


    def _export_path_diagnostics_csv(self):
        rows = self._res.get('path_diagnostics') or []
        if not rows:
            QtWidgets.QMessageBox.information(self, _tr('Exportar diagnóstico MDT'), _tr('No hay diagnóstico por pares fuente-receptor disponible. Recalcula con el motor ISO-aligned y fuentes dentro del radio. Este CSV permite auditar cada aerogenerador frente a cada receptor.'))
            return
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, _tr('Exportar diagnóstico MDT por pares CSV'), os.path.expanduser('~/ruido_diagnostico_mdt_pares.csv'), 'CSV (*.csv)')
        if not path:
            return
        try:
            if not path.lower().endswith('.csv'):
                path += '.csv'
            self._write_dict_rows_csv(rows, path)
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, _tr('Exportar diagnóstico MDT'), _tr('No se pudo exportar el CSV de diagnóstico MDT:') + f'\n{e}')


    def _export_top_receivers_csv(self):
        if self.tbl_top.rowCount() <= 0:
            QtWidgets.QMessageBox.information(self, _tr('Exportar principales receptores'), _tr('No hay filas de principales receptores para exportar.'))
            return
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, _tr('Exportar top receptores CSV'), os.path.expanduser('~/ruido_top_receptores.csv'), 'CSV (*.csv)')
        if not path:
            return
        try:
            if not path.lower().endswith('.csv'):
                path += '.csv'
            self._write_table_csv(self.tbl_top, path)
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, _tr('Exportar top receptores'), _tr('No se pudo exportar el CSV de top receptores:') + f'\n{e}')

    def _export_mdt_screening_csv(self):
        if self.tbl_mdt.rowCount() <= 0:
            QtWidgets.QMessageBox.information(self, _tr('Exportar screening MDT'), _tr('No hay filas de screening MDT para exportar.'))
            return
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, _tr('Exportar screening MDT CSV'), os.path.expanduser('~/ruido_screening_mdt_receptores.csv'), 'CSV (*.csv)')
        if not path:
            return
        try:
            if not path.lower().endswith('.csv'):
                path += '.csv'
            self._write_table_csv(self.tbl_mdt, path)
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, _tr('Exportar screening MDT'), _tr('No se pudo exportar el CSV de screening MDT:') + f'\n{e}')


    def _export_sources_csv(self):
        layer = self._res.get('sources_layer')
        if not isinstance(layer, QgsVectorLayer):
            QtWidgets.QMessageBox.information(self, _tr('Exportar grupos fuente'), _tr('No hay capa de fuentes para exportar.'))
            return
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, _tr('Exportar grupos fuente CSV'), os.path.expanduser('~/ruido_fuentes.csv'), 'CSV (*.csv)')
        if not path:
            return
        try:
            if not path.lower().endswith('.csv'):
                path += '.csv'
            self._write_layer_csv(layer, path)
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, _tr('Exportar grupos fuente'), _tr('No se pudo exportar el CSV:') + f'\n{e}')

    def _export_exceedances_csv(self):
        rows = self._collect_exceedance_rows()
        if not rows:
            QtWidgets.QMessageBox.information(self, _tr('Exportar excedencias'), _tr('Ningún receptor supera el límite en este cálculo.'))
            return
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, _tr('Exportar excedencias CSV'), os.path.expanduser('~/ruido_excedencias.csv'), 'CSV (*.csv)')
        if not path:
            return
        try:
            if not path.lower().endswith('.csv'):
                path += '.csv'
            self._write_rows_csv(rows, path)
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, _tr('Exportar excedencias'), _tr('No se pudo exportar el CSV:') + f'\n{e}')

    def _export_package_xlsx(self):
        if Workbook is None:
            QtWidgets.QMessageBox.information(self, _tr('Exportar paquete XLSX'), _tr('openpyxl no está disponible en este entorno QGIS. Usa las exportaciones CSV o instala openpyxl.'))
            return
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, _tr('Exportar paquete XLSX'), os.path.expanduser('~/ruido_paquete.xlsx'), 'Excel (*.xlsx)')
        if not path:
            return
        try:
            if not path.lower().endswith('.xlsx'):
                path += '.xlsx'
            wb = Workbook()
            ws0 = wb.active
            ws0.title = _tr('Resumen')[:31]
            plain = self.page_summary.toPlainText().splitlines()
            for line in plain:
                ws0.append([line])
            self._append_table_sheet(wb, _tr('Modelos'), self.tbl_models)
            self._append_sheet(wb, _tr('Receptores'), self._receiver_rows_for_export())
            self._append_sheet(wb, 'Excedencias', self._collect_exceedance_rows())
            self._append_table_sheet(wb, _tr('Capas_creadas'), self.tbl_layers)
            wb.save(path)
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, _tr('Exportar paquete XLSX'), _tr('No se pudo exportar el XLSX:') + f'\n{e}')

    def _fill_layers(self):
        entries = [
            (_tr("Ruido · Receptores"), self._res.get("result_layer") is not None),
            (_tr("Ruido · Fuentes"), self._res.get("sources_layer") is not None),
            (_tr("Ruido · Enlaces dominantes"), self._res.get("links_layer") is not None),
            (_tr("Ruido · Receptores fuera de radio"), self._res.get("uncovered_layer") is not None),
            (_tr("Ruido · Mapa"), self._res.get("grid_layer") is not None),
            (_tr("Ruido · Isófonas"), self._res.get("iso_layer") is not None),
        ]
        self.tbl_layers.setRowCount(len(entries))
        for r, (name, ok) in enumerate(entries):
            for c, v in enumerate([name, _tr("creada") if ok else _tr("no creada")]):
                it = QtWidgets.QTableWidgetItem(v)
                it.setFlags(it.flags() & ~QtCore.Qt.ItemIsEditable)
                self.tbl_layers.setItem(r, c, it)
        self.tbl_layers.resizeColumnsToContents()

    def _infer_limit_stats_from_layer(self) -> dict:
        layer = self._res.get("result_layer")
        default = {"min": 45.0, "max": 45.0, "mode": "global", "scenario": "custom", "unique_count": 1}
        if not isinstance(layer, QgsVectorLayer):
            return default
        vals = []
        mode = None
        scenario = None
        try:
            for f in layer.getFeatures():
                try:
                    v = f["limit_dba"]
                    if v is not None:
                        vals.append(float(v))
                except Exception:
                    pass
                if mode is None:
                    try:
                        mode = str(f["limit_src"] or "").strip().lower() or None
                    except Exception:
                        pass
                if scenario is None:
                    try:
                        scenario = str(f["limit_scn"] or "").strip().lower() or None
                    except Exception:
                        pass
        except Exception:
            return default
        if not vals:
            return default
        return {
            "min": min(vals),
            "max": max(vals),
            "mode": mode or "global",
            "scenario": scenario or "custom",
            "unique_count": len({round(v, 6) for v in vals}),
        }
