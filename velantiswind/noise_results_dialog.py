# -*- coding: utf-8 -*-
from __future__ import annotations

from typing import Dict, List
import csv
import os

from qgis.PyQt import QtCore, QtWidgets, QtGui
from qgis.PyQt.QtGui import QGuiApplication
from .i18n import apply_i18n, install_runtime_i18n_patches, translate_html, tr_text as _tr
from .ui_core.responsive import fit_to_screen, configure_table
from qgis.core import QgsFeatureRequest, QgsVectorLayer

try:
    from .noise_core.noise_common import OCTAVE_BANDS, A_WEIGHTING
except Exception:
    from noise_core.noise_common import OCTAVE_BANDS, A_WEIGHTING

try:
    from openpyxl import Workbook
except Exception:
    Workbook = None


# Client-facing receiver table/export schema.  Detailed MDT/path diagnostics are
# still kept internally in the result payload, but the default dialog and exports
# should stay readable for consultancy workflows.
CONSULTANCY_RECEIVER_COLUMNS = [
    ("rec_id", "id receptor"),
    ("rec_type", "tipo"),
    ("noise_dba", "nivel total dB(A)"),
    ("limit_dba", "límite dB(A)"),
    ("margin_db", "margen límite dB"),
    ("state", "estado"),
    ("exceeds", "supera límite"),
    ("n_src", "nº turbinas"),
    ("near_m", "dist. turbina cercana (m)"),
    ("dom_model", "modelo dominante"),
    ("dom_group", "grupo fuente dom."),
    ("dom_park", "parque dom."),
    ("src_lwa", "LwA fuente dom. dB(A)"),
    ("adiv_db", "Adiv dB"),
    ("aatm_db", "Aatm dB"),
    ("aground_db", "Agr/Aground dB"),
    ("abar_max_db", "Abar máx. dB"),
    ("ground_g", "G suelo"),
    ("ground_md", "modo suelo"),
    ("rec_h_m", "h receptor m"),
    ("rec_z_m", "z terreno receptor m"),
    ("rec_ac_z_m", "z acústica receptor m"),
    ("dom_src_lyr", "capa fuente dominante"),
]

CONSULTANCY_RECEIVER_KEYS = [key for key, _label in CONSULTANCY_RECEIVER_COLUMNS]
CONSULTANCY_RECEIVER_HEADERS = [label for _key, label in CONSULTANCY_RECEIVER_COLUMNS]


class NoiseResultsDialog(QtWidgets.QDialog):
    def __init__(self, parent=None, result: Dict[str, object] | None = None):
        install_runtime_i18n_patches()
        super().__init__(parent)
        self._res = result or {}
        self.setWindowTitle(_tr("Noise · Technical summary"))
        self.setModal(True)
        self._resize_to_screen()
        self._build_ui()
        self._fill()
        apply_i18n(self)

    def _resize_to_screen(self):
        fit_to_screen(self, preferred=(1100, 820), minimum=(680, 460), max_ratio=(0.92, 0.90))

    def _build_ui(self):
        root = QtWidgets.QVBoxLayout(self)
        root.setContentsMargins(12, 12, 12, 12)
        root.setSpacing(8)

        header = QtWidgets.QHBoxLayout()
        title = QtWidgets.QLabel("Noise · Calculation summary")
        title.setStyleSheet("font-size:20px; font-weight:700; color:#103b67;")
        header.addWidget(title, 1)
        header.addStretch(1)
        logo = QtWidgets.QLabel(self)
        logo_path = os.path.join(os.path.dirname(__file__), "assets", "velantiswind_logo.png")
        if os.path.exists(logo_path):
            pix = QtGui.QPixmap(logo_path)
            if not pix.isNull():
                logo.setPixmap(pix.scaled(180, 180, QtCore.Qt.KeepAspectRatio, QtCore.Qt.SmoothTransformation))
                logo.setToolTip("Velantis Wind")
        header.addWidget(logo, 0, QtCore.Qt.AlignRight | QtCore.Qt.AlignTop)
        root.addLayout(header)

        self.tabs = QtWidgets.QTabWidget(self)
        root.addWidget(self.tabs, 1)

        self.page_summary = QtWidgets.QTextBrowser(self)
        self.tabs.addTab(self.page_summary, "Resumen")

        self.tbl_models = QtWidgets.QTableWidget(0, 6, self)
        self.tbl_models.setHorizontalHeaderLabels(["Modelo WT", "Turbinas", "LwA eff.", "HH", "D", "Notas"])
        configure_table(self.tbl_models, stretch_columns=(0, 5))
        self.tbl_models.horizontalHeader().setStretchLastSection(True)
        self.tabs.addTab(self.tbl_models, "Modelos")

        self.tbl_top = QtWidgets.QTableWidget(0, len(CONSULTANCY_RECEIVER_HEADERS), self)
        self.tbl_top.setHorizontalHeaderLabels(CONSULTANCY_RECEIVER_HEADERS)
        self.tbl_top.setToolTip(
            "Tabla resumida para consultoría: resultados acústicos por receptor, "
            "cumplimiento, fuente dominante y atenuaciones principales. "
            "Los diagnósticos internos MDT por pares se conservan en memoria, pero no se muestran por defecto."
        )
        configure_table(self.tbl_top, stretch_columns=(0, 1, 9, 10, 11, 22))
        self.tbl_top.horizontalHeader().setStretchLastSection(True)
        self.tabs.addTab(self.tbl_top, "Top receptores")

        # Internal MDT screening table kept for compatibility with helper methods,
        # but no longer exposed as a default consultancy tab/export.
        self.tbl_mdt = QtWidgets.QTableWidget(0, 27, self)
        self.tbl_mdt.setHorizontalHeaderLabels([
            'id receptor', 'nivel total dB(A)', 'nº turbinas', 'Abar max contrib. dB',
            'Abar ponderada dB', 'turbinas apant.', 'estado MDT dom.', 'Abar dom. dB',
            'ID fuente max Abar', 'estado max Abar', 'obs. max Abar m',
            'umbral max Abar m', 'd1 max Abar m', 'd2 max Abar m',
            'ID fuente max obst.', 'estado max obst.', 'obs. max obst. m',
            'umbral max obst. m', 'd1 max obst. m', 'd2 max obst. m',
            'z terreno receptor m', 'h receptor m', 'z acústica receptor m',
            'z terreno turb. dom. m', 'z acústica turb. dom. m',
            'z terreno turb. max Abar m', 'z acústica turb. max Abar m'
        ])

        self.tbl_layers = QtWidgets.QTableWidget(0, 2, self)
        self.tbl_layers.setHorizontalHeaderLabels(["Capa", "Estado"])
        self.tbl_layers.horizontalHeader().setStretchLastSection(True)
        self.tabs.addTab(self.tbl_layers, "Capas creadas")

        btns = QtWidgets.QHBoxLayout()
        self.btn_export_summary = QtWidgets.QPushButton("Exportar informe…")
        self.btn_export_summary.setToolTip("Guarda el resumen técnico en HTML o TXT.")
        self.btn_export_summary.clicked.connect(self._export_summary)
        self.btn_export_receivers = QtWidgets.QPushButton("Exportar receptores CSV…")
        self.btn_export_receivers.setToolTip("Guarda una tabla limpia con una fila por receptor y las columnas necesarias para consultoría.")
        self.btn_export_receivers.clicked.connect(self._export_receivers_csv)
        self.btn_export_exceed = QtWidgets.QPushButton("Exportar excedencias CSV…")
        self.btn_export_exceed.setToolTip("Guarda solo los receptores que superan su límite acústico.")
        self.btn_export_exceed.clicked.connect(self._export_exceedances_csv)
        self.btn_export_xlsx = QtWidgets.QPushButton("Exportar paquete XLSX…")
        self.btn_export_xlsx.setToolTip("Guarda resumen, modelos, receptores y excedencias en un único libro Excel.")
        self.btn_export_xlsx.clicked.connect(self._export_package_xlsx)
        btns.addWidget(self.btn_export_summary)
        btns.addWidget(self.btn_export_receivers)
        btns.addWidget(self.btn_export_exceed)
        btns.addWidget(self.btn_export_xlsx)
        btns.addStretch(1)
        close_btn = QtWidgets.QPushButton("Cerrar")
        close_btn.setMinimumHeight(34)
        close_btn.clicked.connect(self.accept)
        btns.addWidget(close_btn)
        root.addLayout(btns)

    def _fill(self):
        self._fill_summary()
        self._fill_models()
        self._fill_top_receivers()
        self._fill_mdt_screening()
        self._fill_layers()

    def _payload_top_receivers(self) -> List[Dict[str, object]]:
        rows = self._res.get("top_receivers") or []
        out: List[Dict[str, object]] = []
        for row in rows:
            if isinstance(row, dict):
                out.append(row)
        def _noise(d):
            try:
                return float(d.get("noise_dba") or d.get("total_level_dba") or 0.0)
            except Exception:
                return -1.0e99
        out.sort(key=_noise, reverse=True)
        return out


    def _payload_receiver_rows(self) -> List[Dict[str, object]]:
        rows = self._res.get("receiver_rows") or []
        out: List[Dict[str, object]] = []
        for row in rows:
            if isinstance(row, dict):
                out.append(row)
        if out:
            return out
        # Fallback to visible top rows if the full receiver payload is absent.
        return self._payload_top_receivers()


    def _attenuation_stats_from_payload_rows(self) -> Dict[str, Dict[str, float]]:
        """Compute attenuation statistics from stable named receiver rows.

        The HTML report historically used the precomputed ``*_stats`` entries
        in ``self._res``.  When the calculation is returned by a background
        QgsTask, those entries can remain zero if the QGIS memory layer cannot
        be read at the exact moment the dialog is built, even though
        ``receiver_rows`` and the Top receivers table contain the correct
        values.  This fallback derives the statistics directly from the named
        payload used by the CSV/XLSX exports.
        """
        rows = self._payload_receiver_rows()

        def _f(d: Dict[str, object], *keys: str):
            for key in keys:
                try:
                    v = d.get(key)
                except Exception:
                    v = None
                if v is None:
                    continue
                txt = str(v).strip()
                if txt == '' or txt.lower() in ('none', 'nan', 'n/a'):
                    continue
                try:
                    x = float(txt.replace(',', '.'))
                except Exception:
                    continue
                if x == x:
                    return x
            return None

        def _covered(d: Dict[str, object]) -> bool:
            nsrc = _f(d, 'n_src', 'turbines_in_radius', 'no. turbines')
            if nsrc is not None:
                return nsrc > 0
            covered = _f(d, 'covered')
            if covered is not None:
                return covered > 0
            noise = _f(d, 'noise_dba', 'total_level_dba', 'total level dB(A)')
            return bool(noise is not None and noise > 0)

        vals = {
            'adiv': [],
            'aatm': [],
            'aground': [],
            'abar': [],
        }
        for d in rows:
            if not isinstance(d, dict) or not _covered(d):
                continue
            for name, keys in {
                'adiv': ('adiv_db', 'divergence_loss_db', 'Adiv loss dB', 'pérdida Adiv dB'),
                'aatm': ('aatm_db', 'atmospheric_loss_db', 'Aatm loss dB', 'pérdida Aatm dB'),
                'aground': ('aground_db', 'ground_loss_db', 'Agr/Aground loss dB', 'pérdida Agr/Aground dB'),
                'abar': ('abar_max_db', 'barrier_loss_max_contributors_db', 'Abar max contrib. dB', 'abar_db', 'Abar dom. dB'),
            }.items():
                x = _f(d, *keys)
                if x is not None:
                    vals[name].append(float(x))

        def _stat(seq: List[float]) -> Dict[str, float]:
            if not seq:
                return {'mean': 0.0, 'max': 0.0}
            return {'mean': sum(seq) / float(len(seq)), 'max': max(seq)}

        return {name: _stat(seq) for name, seq in vals.items()}

    def _prefer_payload_stats_if_needed(self, current: Dict[str, object], fallback: Dict[str, float]) -> Dict[str, float]:
        """Use payload-derived stats when the current report stats are empty/zero."""
        try:
            cur_max = float((current or {}).get('max', 0.0) or 0.0)
        except Exception:
            cur_max = 0.0
        try:
            fb_max = float((fallback or {}).get('max', 0.0) or 0.0)
        except Exception:
            fb_max = 0.0
        if fb_max > 0.0 and cur_max <= 0.0:
            return dict(fallback or {})
        return dict(current or {})

    def _infer_critical_receiver_from_layer(self) -> Dict[str, object]:
        """Return the highest-noise receiver as a dict using current layer fields."""
        layer = self._res.get("result_layer")
        payload_rows = self._payload_top_receivers()
        if not isinstance(layer, QgsVectorLayer):
            return dict(payload_rows[0]) if payload_rows else {}
        best_feat = None
        best_level = -1.0e99
        level_keys = ("noise_dba", "total_level_dba", "nivel_total_dba")
        try:
            iterator = layer.getFeatures()
        except Exception:
            return {}
        for feat in iterator:
            level = None
            for key in level_keys:
                try:
                    level = float(feat[key])
                    break
                except Exception:
                    continue
            if level is None:
                continue
            try:
                if level != level:
                    continue
            except Exception:
                continue
            if best_feat is None or level > best_level:
                best_feat = feat
                best_level = level
        if best_feat is None:
            return dict(payload_rows[0]) if payload_rows else {}
        row: Dict[str, object] = {"fid": best_feat.id(), "rec_id": best_feat.id()}
        try:
            for fld in layer.fields():
                name = fld.name()
                try:
                    row[name] = best_feat[name]
                except Exception:
                    pass
        except Exception:
            pass
        if not row.get("rec_id"):
            row["rec_id"] = best_feat.id()
        return row

    def _fill_summary(self):
        n_sources = int(self._res.get("n_sources", 0))
        n_receivers = int(self._res.get("n_receivers", 0))
        n_with = int(self._res.get("n_receivers_with_sources", 0))
        n_without = int(self._res.get("n_uncovered_receivers", max(0, n_receivers - n_with)))
        n_exceed = int(self._res.get("n_receivers_exceeding_limit", 0))
        max_noise = float(self._res.get("max_noise_dba", 0.0))
        model_diag = self._res.get("model_diag", {}) or {}
        n_models = len(model_diag)
        limit_stats = self._res.get('limit_stats') or self._infer_limit_stats_from_layer()
        acoustic = self._res.get('acoustic_scenario', {}) or {}
        crit_raw = self._res.get('critical_receiver') or {}
        crit_layer = self._infer_critical_receiver_from_layer()

        def _has_value(v):
            if v is None:
                return False
            try:
                if isinstance(v, float) and v != v:
                    return False
            except Exception:
                pass
            return str(v).strip() != ''

        # Merge stored critical-receiver metadata with a robust fallback read directly
        # from the result layer. This avoids visual summaries falling back to 0.00
        # when the engine changes field names.
        crit = dict(crit_layer or {})
        for _k, _v in dict(crit_raw or {}).items():
            if _has_value(_v):
                crit[_k] = _v
        payload_att_stats = self._attenuation_stats_from_payload_rows()
        adiv_stats = self._prefer_payload_stats_if_needed(self._res.get('adiv_stats') or {}, payload_att_stats.get('adiv') or {})
        aatm_stats = self._prefer_payload_stats_if_needed(self._res.get('aatm_stats') or {}, payload_att_stats.get('aatm') or {})
        aground_stats = self._prefer_payload_stats_if_needed(self._res.get('aground_stats') or {}, payload_att_stats.get('aground') or {})
        abar_stats = self._prefer_payload_stats_if_needed(self._res.get('abar_stats') or {}, payload_att_stats.get('abar') or {})
        g_eff_stats = self._res.get('g_eff_stats') or {}
        ground_diag = self._res.get('ground_diag') or {}
        receiver_type_counts = self._res.get('receiver_type_counts') or {}
        grid_diag = self._res.get('grid_diag') or {}
        report = self._res.get('report_meta') or {}
        ground_mode = str(report.get('ground_mode') or self._res.get('ground_mode') or 'global')
        landuse_layer_name = str(report.get('landuse_layer_name') or self._res.get('landuse_layer_name') or '')
        dem_layer_name = str(report.get('dem_layer_name') or self._res.get('dem_layer_name') or '')
        dem_used = bool(report.get('dem_used', self._res.get('dem_used', False)))
        engine = str(report.get('engine') or ('iso_aligned' if str(self._res.get('method') or '').startswith('iso_') else 'fast'))
        engine_label = str(report.get('engine_label') or ('ISO-aligned por bandas' if engine == 'iso_aligned' else 'Rápido LwA global'))
        equation = str(report.get('equation') or ('Lp,b = Lw,b - Adiv - Aatm,b - Agr,b - Abar,b' if engine == 'iso_aligned' else 'Lp = LwA - Adiv - Aatm - Aground'))
        alpha = float(report.get('alpha_db_per_m', self._res.get('alpha_db_per_m', 0.0)))
        g = float(report.get('ground_factor_g', self._res.get('ground_factor_g', 0.0)))
        rec_h = float(report.get('receiver_height_m', self._res.get('receiver_height_m', 0.0)))
        radius = float(report.get('max_radius_m', self._res.get('max_radius_m', 0.0)))
        temp_c = float(report.get('temperature_c', 15.0))
        hum_pct = float(report.get('humidity_percent', 70.0))
        pressure_kpa = float(report.get('pressure_kpa', 101.325))
        terms = report.get('active_terms') or {}
        spectrum_rows = report.get('spectrum_sources') or []

        if str(acoustic.get('mode') or 'fixed') == 'curve':
            if bool(acoustic.get('use_curve_worst_case', False)):
                acoustic_txt = 'Curvas acústicas LwA(ws) en peor caso'
            else:
                try:
                    acoustic_txt = f"Curvas acústicas LwA(ws) a {float(acoustic.get('eval_ws_m_s')):.1f} m/s"
                except Exception:
                    acoustic_txt = 'Curvas acústicas LwA(ws)'
        else:
            acoustic_txt = 'LwA fijo por grupo fuente acústico'

        eff_lines = []
        for d in list(acoustic.get('effective_models') or []):
            group_name = str(d.get('name') or 'Grupo')
            park_name = str(d.get('park_name') or '').strip()
            model_name = str(d.get('model_name') or '').strip()
            spec_src = ''
            for sp in spectrum_rows:
                if str(sp.get('group_name') or '') == group_name:
                    spec_src = str(sp.get('spectrum_source') or '')
                    break
            try:
                line = f"<li><b>{group_name}</b>: {float(d.get('lwa_effective')):.2f} dB(A)"
            except Exception:
                line = f"<li><b>{group_name}</b>: sin valor"
            extra = []
            if model_name:
                extra.append(f"modelo {model_name}")
            if park_name:
                extra.append(f"parque {park_name}")
            if str(d.get('curve_note') or '').strip():
                extra.append(str(d.get('curve_note')))
            if spec_src:
                extra.append(f"espectro {spec_src}")
            if extra:
                line += " · " + " · ".join(extra)
            line += "</li>"
            eff_lines.append(line)

        spectrum_detail_blocks = []
        for sp in spectrum_rows:
            group_name = str(sp.get('group_name') or 'Grupo')
            model_name = str(sp.get('model_name') or group_name)
            spec_src = str(sp.get('spectrum_source') or '')
            lw_oct = {int(k): float(v) for k, v in (sp.get('lw_octave') or {}).items()}
            sref = {int(k): float(v) for k, v in (sp.get('spectrum_template_ref') or {}).items()}
            try:
                delta_db = float(sp.get('spectrum_delta_db'))
                if not (delta_db == delta_db):
                    delta_db = None
            except Exception:
                delta_db = None
            rows = []
            for f in OCTAVE_BANDS:
                sref_txt = '-'
                if f in sref:
                    sref_txt = f"{sref[f]:.2f}"
                lw_txt = '-'
                if f in lw_oct:
                    lw_txt = f"{lw_oct[f]:.2f}"
                a_txt = f"{float(A_WEIGHTING.get(f, 0.0)):.1f}"
                rows.append(f"<tr><td>{f}</td><td style='text-align:right;'>{sref_txt}</td><td style='text-align:right;'>{a_txt}</td><td style='text-align:right;'>{lw_txt}</td></tr>")
            delta_line = ''
            if delta_db is not None:
                delta_line = f"<p><b>Δ aplicado:</b> {delta_db:.2f} dB. Este desplazamiento eleva o reduce toda la plantilla espectral para que su suma A-ponderada reproduzca el <code>LwA_objetivo</code> de la curva acústica o del LwA fijo.</p>"
            origin_line = '<p><b>Interpretación:</b> el espectro final <code>Lw,b</code> es el que entra realmente en la ecuación por bandas. Si existe <code>S_b^ref</code>, corresponde a la plantilla de referencia antes del ajuste global <code>Δ</code>.</p>' if sref else '<p><b>Interpretación:</b> en este grupo no se ha usado una plantilla interna visible; el espectro final <code>Lw,b</code> procede directamente del espectro cargado/importado o de una biblioteca externa.</p>'
            spectrum_detail_blocks.append(f"""
                <div class='card'>
                    <h4>2.1 Espectro usado por el grupo fuente: {group_name}</h4>
                    <p><b>Modelo:</b> {model_name} · <b>Origen del espectro:</b> {spec_src or '-'}.</p>
                    <p><b>Qué representa cada columna:</b> <code>S_b^ref</code> es la plantilla espectral de referencia (si existe), <code>A_weight,b</code> la ponderación A de cada banda y <code>Lw,b</code> el nivel final en dB realmente usado por el cálculo.</p>
                    {delta_line}
                    <table>
                        <tr><th>Banda [Hz]</th><th style='text-align:right;'>S_b^ref [dB]</th><th style='text-align:right;'>A_weight,b [dB]</th><th style='text-align:right;'>Lw,b final [dB]</th></tr>
                        {''.join(rows)}
                    </table>
                    {origin_line}
                </div>
            """)
        spectrum_detail_html = ''.join(spectrum_detail_blocks)

        def _fmt_equation_term(value: float) -> str:
            try:
                v = float(value)
            except Exception:
                return '-'
            if v != v:
                return '-'
            if abs(v) < 0.005:
                return '0.00'
            return f"{v:.2f}"

        if crit:
            def _crit_value(*keys, default=None):
                for key in keys:
                    try:
                        val = crit.get(key)
                    except Exception:
                        val = None
                    if val is None:
                        continue
                    try:
                        if isinstance(val, float) and val != val:
                            continue
                    except Exception:
                        pass
                    if str(val).strip() == '':
                        continue
                    return val
                return default

            def _crit_float(*keys, default=0.0):
                val = _crit_value(*keys, default=None)
                if val is None:
                    return float(default)
                try:
                    f = float(val)
                    if f != f:
                        return float(default)
                    return f
                except Exception:
                    return float(default)

            crit_id = _crit_value('rec_id', 'fid', default='-')
            crit_level = _crit_float('nivel_total_dba', 'total_level_dba', 'noise_dba', default=max_noise)
            crit_limit = _crit_float('limite_aplicado_dba', 'limit_dba', default=45.0)
            crit_margin = _crit_float('margen_limite_db', 'limit_margin_db', 'margin_db', default=crit_level - crit_limit)
            crit_model = _crit_value('modelo_dominante', 'dominant_model', 'dom_model', default='-')
            crit_group = _crit_value('grupo_fuente_dominante', 'dominant_source_group', 'dom_group', default='-')
            crit_n_turb = _crit_value('n_turbinas_en_radio', 'turbines_in_radius', 'n_src', default='-')
            crit_lwa = _crit_float('lwa_fuente_dom_dba', 'source_lwa_dba', 'src_lwa', default=0.0)
            crit_dist = _crit_float('dist_fuente_dom_3d_m', 'source_receiver_3d_m', 'dist3d_m', 'near_m', default=0.0)
            crit_adiv = _crit_float('perdida_divergencia_db', 'divergence_loss_db', 'adiv_db', default=0.0)
            crit_aatm = _crit_float('perdida_atmosferica_db', 'atmospheric_loss_db', 'aatm_db', default=0.0)
            crit_agr = _crit_float('perdida_suelo_db', 'ground_loss_db', 'aground_db', default=0.0)
            crit_abar = _crit_float('perdida_barrera_db', 'barrier_loss_db', 'abar_db', default=0.0)
            crit_abar_max = _crit_float('perdida_barrera_max_db', 'barrier_loss_max_contributors_db', 'abar_max_db', default=crit_abar)
            crit_abar_mean = _crit_float('perdida_barrera_media_db', 'barrier_loss_mean_contributors_db', 'abar_mean_db', default=crit_abar)
            crit_abar_ew = _crit_float('perdida_barrera_ponderada_db', 'barrier_loss_energy_weighted_db', 'abar_ew_db', default=crit_abar)
            crit_abar_screen_n = _crit_value('n_fuentes_apantalladas', 'barrier_screened_sources_n', 'abar_screen_n', default=0)
            try:
                crit_abar_screen_n = int(crit_abar_screen_n or 0)
            except Exception:
                crit_abar_screen_n = 0
            crit_g_eff = _crit_float('factor_suelo_g', 'ground_factor_g', 'ground_g', default=float(g_eff_stats.get('critical', g)))
            crit_freq = _crit_value('banda_dominante_hz', 'dominant_band_hz', 'dom_freq', default='-')
            crit_spec_src = _crit_value('origen_espectro', 'spectrum_source', 'spec_src', default='-')
            crit_abar_state = str(_crit_value('mdt_abar_state', 'abar_state', default='') or '').strip()
            crit_obs_h = _crit_float('mdt_obstacle_height_m', 'obs_h_m', default=0.0)
            crit_obs_d1 = _crit_float('mdt_d1_m', 'obs_d1_m', default=0.0)
            crit_obs_d2 = _crit_float('mdt_d2_m', 'obs_d2_m', default=0.0)
            crit_obs_thr = _crit_float('mdt_obstacle_threshold_m', 'obs_thr_m', default=0.0)
            crit_src_z = _crit_float('dominant_source_ground_z_m', 'src_z_m', default=float('nan'))
            crit_hub_h = _crit_float('dominant_source_hub_height_m', 'hub_h_m', default=float('nan'))
            crit_src_ac_z = _crit_float('dominant_source_acoustic_z_m', 'src_ac_z_m', default=float('nan'))
            crit_rec_z = _crit_float('receiver_ground_z_m', 'rec_z_m', default=float('nan'))
            crit_rec_h = _crit_float('receiver_height_agl_m', 'rec_h_m', default=float('nan'))
            crit_rec_ac_z = _crit_float('receiver_acoustic_z_m', 'rec_ac_z_m', default=float('nan'))
            crit_maxab_src = _crit_value('max_abar_source_index', 'maxab_src', default='-')
            crit_maxab_state = str(_crit_value('max_abar_mdt_state', 'maxab_state', default='') or '').strip()
            crit_maxab_obs_h = _crit_float('max_abar_obstacle_height_m', 'maxab_obs_h', default=0.0)
            crit_maxab_d1 = _crit_float('max_abar_source_obstacle_m', 'maxab_d1', default=0.0)
            crit_maxab_d2 = _crit_float('max_abar_obstacle_receiver_m', 'maxab_d2', default=0.0)
            
            status_badge = 'badge-success' if crit_margin <= 0 else 'badge-danger'
            status_text = 'CUMPLE' if crit_margin <= 0 else 'EXCEDE'
            card_class = 'card-success' if crit_margin <= 0 else 'card-danger'

            crit_adiv_txt = _fmt_equation_term(crit_adiv)
            crit_aatm_txt = _fmt_equation_term(crit_aatm)
            crit_agr_txt = _fmt_equation_term(crit_agr)
            crit_abar_txt = _fmt_equation_term(crit_abar)
            crit_agr_desc = f"Atenuación por efecto del terreno (G_eff={crit_g_eff:.2f})"
            crit_abar_desc = "Atenuación por MDT en la trayectoria dominante"
            crit_abar_max_txt = _fmt_equation_term(crit_abar_max)
            crit_abar_mean_txt = _fmt_equation_term(crit_abar_mean)
            crit_abar_ew_txt = _fmt_equation_term(crit_abar_ew)
            try:
                crit_n_turb_i = int(crit_n_turb)
            except Exception:
                crit_n_turb_i = 0

            def _fmt_m_or_na(v):
                try:
                    f = float(v)
                    if f != f:
                        return 'N/A'
                    return f"{f:.2f}"
                except Exception:
                    return 'N/A'

            dominant_height_html = (
                f"<br><b>Alturas trayectoria dominante:</b> terreno turbina={_fmt_m_or_na(crit_src_z)} m · "
                f"hub={_fmt_m_or_na(crit_hub_h)} m AGL · altura acústica turbina={_fmt_m_or_na(crit_src_ac_z)} m · "
                f"terreno receptor={_fmt_m_or_na(crit_rec_z)} m · h receptor={_fmt_m_or_na(crit_rec_h)} m AGL · "
                f"altura acústica receptor={_fmt_m_or_na(crit_rec_ac_z)} m."
            )
            maxabar_height_html = ''
            if float(crit_abar_max or 0.0) > 0.005:
                maxabar_height_html = (
                    f"<br><b>Trayectoria con Abar máximo:</b> fuente={crit_maxab_src} · estado={crit_maxab_state or '-'} · "
                    f"obs={_fmt_m_or_na(crit_maxab_obs_h)} m · d1={_fmt_m_or_na(crit_maxab_d1)} m · d2={_fmt_m_or_na(crit_maxab_d2)} m."
                )
            abar_summary_html = ''
            if dem_used and engine == 'iso_aligned':
                abar_summary_html = f"""
                <div class='note'>
                    <b>Lectura correcta de Abar:</b> el valor <b>Abar trayectoria dominante</b> corresponde solo a la turbina que más contribuye al receptor y a su banda dominante. El nivel total del receptor se obtiene sumando energéticamente todas las turbinas y bandas.
                    <br><b>Abar máximo entre turbinas contribuyentes:</b> {crit_abar_max_txt} dB · <b>Abar medio:</b> {crit_abar_mean_txt} dB · <b>Abar ponderado por contribución energética:</b> {crit_abar_ew_txt} dB · <b>trayectorias apantalladas:</b> {crit_abar_screen_n}/{crit_n_turb_i if crit_n_turb_i else crit_n_turb}.
                    {dominant_height_html}
                    {maxabar_height_html}
                </div>
                """

            abar_note_html = ''
            if dem_used and engine == 'iso_aligned':
                if abs(float(crit_abar)) < 0.005:
                    reason_map = {
                        'los_clear': 'la línea de visión entre la turbina dominante y este receptor queda despejada según el MDT',
                        'below_threshold': 'se detectó relieve, pero por debajo del umbral conservador de activación',
                        'no_profile': 'no se pudo extraer un perfil MDT válido para el trayecto dominante',
                        'no_dem': 'no había MDT disponible en este trayecto',
                    }
                    reason = reason_map.get(crit_abar_state, 'no se detectó obstáculo topográfico relevante en el trayecto dominante')
                    extra = ''
                    if float(crit_obs_thr) > 0.0:
                        extra = f" Umbral de activación: {crit_obs_thr:.2f} m."
                    if float(abar_stats.get('max', 0.0) or 0.0) > 0.005:
                        extra += f" Otros receptores sí presentan apantallamiento (Abar máx. {float(abar_stats.get('max',0.0)):.2f} dB)."
                    abar_note_html = f"<p style='margin:8px 0 10px 0;color:#495057;'><i>Lectura MDT: Abar=0 en el receptor crítico no implica que el MDT esté desactivado; significa que {reason}.{extra}</i></p>"
                else:
                    abar_note_html = f"<p style='margin:8px 0 10px 0;color:#495057;'><i>Lectura MDT: obstáculo dominante estimado {crit_obs_h:.2f} m; d1={crit_obs_d1:.1f} m, d2={crit_obs_d2:.1f} m; estado={crit_abar_state or 'active'}.</i></p>"

            crit_html = f"""
        <div class='{card_class}'>
            <h3>🎯 Receptor crítico (mayor nivel sonoro)</h3>
            
            <table style='margin-bottom: 20px;'>
                <tr>
                    <td style='width: 50%; padding-right: 20px;'>
                        <p><b>ID Receptor:</b> {crit_id}</p>
                        <p><b>Nivel total:</b> <span style='font-size:28px; font-weight:bold; color:{'#dc3545' if crit_margin > 0 else '#28a745'};'>{crit_level:.2f} dB(A)</span></p>
                        <p><b>Límite aplicable:</b> {crit_limit:.2f} dB(A)</p>
                        <p><b>Margen:</b> {crit_margin:+.2f} dB <span class='{status_badge}'>{status_text}</span></p>
                    </td>
                    <td style='width: 50%;'>
                        <p><b>Modelo dominante:</b> {crit_model}</p>
                        <p><b>Grupo fuente:</b> {crit_group}</p>
                        <p><b>Turbinas contribuyentes dentro del radio:</b> {crit_n_turb}</p>
                        <p><b>Distancia:</b> {crit_dist:.1f} m</p>
                    </td>
                </tr>
            </table>
            
            <h4>📊 Desglose de atenuaciones</h4>
            <p style='margin: 6px 0 10px 0; color:#495057;'><i>Los valores mostrados a continuación son las magnitudes de atenuación usadas por el modelo. En la ecuación principal estos términos se restan al nivel de fuente.</i></p>
            <table style='margin: 16px 0;'>
                <tr>
                    <th>Término</th>
                    <th style='text-align: right;'>Valor [dB]</th>
                    <th>Descripción</th>
                </tr>
                <tr style='background: #e3f2fd;'>
                    <td><b>LwA fuente dominante</b></td>
                    <td style='text-align: right;'><b>{crit_lwa:.2f}</b></td>
                    <td>Potencia sonora de la turbina</td>
                </tr>
                <tr>
                    <td>Adiv (divergencia)</td>
                    <td style='text-align: right;'>{crit_adiv_txt}</td>
                    <td>Dispersión geométrica</td>
                </tr>
                <tr>
                    <td>Aatm (atmosférica)</td>
                    <td style='text-align: right;'>{crit_aatm_txt}</td>
                    <td>Absorción en el aire</td>
                </tr>
                <tr>
                    <td>Agr (suelo)</td>
                    <td style='text-align: right;'>{crit_agr_txt}</td>
                    <td>{crit_agr_desc}</td>
                </tr>
                <tr>
                    <td>Abar trayectoria dominante</td>
                    <td style='text-align: right;'>{crit_abar_txt}</td>
                    <td>{crit_abar_desc}</td>
                </tr>
                <tr>
                    <td>Abar máximo contribuyentes</td>
                    <td style='text-align: right;'>{crit_abar_max_txt}</td>
                    <td>Máximo Abar entre todas las turbinas que contribuyen al receptor</td>
                </tr>
                <tr>
                    <td>Abar ponderado por energía</td>
                    <td style='text-align: right;'>{crit_abar_ew_txt}</td>
                    <td>Promedio ponderado por la contribución acústica de cada turbina</td>
                </tr>
                <tr>
                    <td>Trayectorias apantalladas</td>
                    <td style='text-align: right;'>{crit_abar_screen_n}/{crit_n_turb}</td>
                    <td>Número de turbinas contribuyentes con Abar &gt; 0 dB</td>
                </tr>
                <tr style='background: #1e3a5f; color: white; font-weight: bold;'>
                    <td>NIVEL RESULTANTE</td>
                    <td style='text-align: right;'>{crit_level:.2f}</td>
                    <td>dB(A)</td>
                </tr>
            </table>
            {abar_note_html}
            {abar_summary_html}
            <p style='margin: 6px 0 10px 0; color:#495057;'><i>Nota: el nivel resultante incluye la suma energética multi-fuente y multi-banda; no es una resta directa de una única turbina.</i></p>
            
            <p style='margin-top: 16px;'>
                <b>Banda dominante:</b> {crit_freq} Hz &nbsp;&nbsp;&nbsp;
                <b>Origen espectro:</b> {crit_spec_src}
            </p>
        </div>
            """
        else:
            crit_html = "<div class='card'><p>Receptor crítico no disponible.</p></div>"

        rec_types_html = ''.join([f"<li><b>{k}:</b> {v}</li>" for k, v in sorted(receiver_type_counts.items())])
        compliance = self._res.get('receiver_type_compliance') or {}
        compliance_html = ''.join([f"<li><b>{k}:</b> {int((v or {}).get('exceed',0))}/{int((v or {}).get('total',0))} superan el límite" + (f" · cubiertos {int((v or {}).get('covered',0))}" if (v or {}).get('covered') is not None else '') + "</li>" for k, v in sorted(compliance.items())])
        suelo_txt = 'global' if ground_mode != 'landuse' else f"desde capa ({landuse_layer_name or 'sin nombre'})"
        grid_txt = 'no generado'
        if self._res.get('grid_layer') is not None:
            grid_txt = f"sí · resolución pedida {float(grid_diag.get('requested_resolution_m',0.0)):.1f} m · efectiva {float(grid_diag.get('effective_resolution_m',0.0)):.1f} m"
            if bool(grid_diag.get('auto_adjusted', False)):
                grid_txt += ' · autoajustada'
        limit_mode = str(limit_stats.get('mode') or 'global').lower()
        limit_scn = str(limit_stats.get('scenario') or 'custom').lower()
        if limit_mode == 'by_field':
            scn_txt = {'day': 'diurno', 'night': 'nocturno', 'custom': 'personalizado'}.get(limit_scn, limit_scn or 'personalizado')
            if abs(float(limit_stats.get('min',45.0)) - float(limit_stats.get('max',45.0))) < 1e-9:
                limit_html = f"<p><b>Límites aplicados:</b> desde campos de receptor ({scn_txt}) · valor único {float(limit_stats.get('min',45.0)):.1f} dB(A)</p>"
            else:
                limit_html = f"<p><b>Límites aplicados:</b> desde campos de receptor ({scn_txt}) · rango {float(limit_stats.get('min',45.0)):.1f}–{float(limit_stats.get('max',45.0)):.1f} dB(A)</p>"
        else:
            limit_html = f"<p><b>Límite de referencia:</b> {float(limit_stats.get('max',45.0)):.1f} dB(A)</p>"

        equations_html = f"<pre style='background:#f6f8fb;border:1px solid #d9e2ef;padding:10px;border-radius:6px;white-space:pre-wrap;'>{equation}</pre>"

        if not crit:
            crit_adiv_txt = crit_aatm_txt = crit_agr_txt = crit_abar_txt = '-'
            crit_agr_desc = 'Efecto del terreno'
            crit_abar_desc = 'Difracción topográfica'

        param_lines = [
            f"<li><b>Motor:</b> {engine_label}</li>",
            f"<li><b>Altura de receptor:</b> {rec_h:.1f} m</li>",
            f"<li><b>Radio máximo:</b> {radius:.0f} m</li>",
            f"<li><b>Modo suelo:</b> {suelo_txt}</li>",
        ]
        if ground_mode == 'landuse':
            param_lines.extend([
                f"<li><b>G global de respaldo:</b> {g:.2f}</li>",
                f"<li><b>G_eff medio usado:</b> {float(g_eff_stats.get('mean', g)):.2f}</li>",
                f"<li><b>G_eff receptor crítico usado:</b> {float(g_eff_stats.get('critical', g)):.2f}</li>",
            ])
        else:
            param_lines.extend([
                f"<li><b>G usado:</b> {g:.2f}</li>",
                f"<li><b>G_eff medio:</b> {float(g_eff_stats.get('mean', g)):.2f}</li>",
                f"<li><b>G_eff receptor crítico:</b> {float(g_eff_stats.get('critical', g)):.2f}</li>",
            ])
        param_lines.extend([
            f"<li><b>MDT/DSM:</b> {'sí · ' + (dem_layer_name or 'sin nombre') if dem_used else 'no'}</li>",
            f"<li><b>Uso del suelo:</b> {'sí · ' + (landuse_layer_name or 'sin nombre') if bool(report.get('landuse_used', False)) else 'no'}</li>",
            f"<li><b>Escenario acústico:</b> {acoustic_txt}</li>",
        ])
        if engine == 'iso_aligned':
            param_lines.extend([
                f"<li><b>Temperatura:</b> {temp_c:.1f} °C</li>",
                f"<li><b>Humedad relativa:</b> {hum_pct:.1f} %</li>",
                f"<li><b>Presión:</b> {pressure_kpa:.3f} kPa</li>",
            ])
        else:
            param_lines.append(f"<li><b>α atmosférico:</b> {alpha:.4f} dB/m</li>")

        term_lines = [
            f"<li><b>Adiv:</b> {'activo' if terms.get('Adiv', True) else 'no'}</li>",
            f"<li><b>Aatm:</b> {'activo' if terms.get('Aatm', True) else 'no'}" + (' (T, HR, P simplificado)' if engine == 'iso_aligned' else ' (α·distancia)') + "</li>",
            f"<li><b>Agr/Aground:</b> {'activo' if terms.get('Agr', True) else 'no'}</li>",
            f"<li><b>Abar:</b> {'activo' if terms.get('Abar', False) else 'no activo'}</li>",
            f"<li><b>G efectivo desde landuse:</b> {'sí' if terms.get('landuse_g', False) else 'no'}</li>",
        ]

        pressure_warning_html = ''
        if engine == 'iso_aligned' and (pressure_kpa < 85.0 or pressure_kpa > 105.0):
            pressure_warning_html = (
                "<p class='note'><b>Revisión recomendada:</b> la presión atmosférica introducida "
                f"({pressure_kpa:.3f} kPa) está fuera del rango típico usado como referencia en muchos estudios "
                "preliminares. Si no es un dato medido del emplazamiento, revisa si debería estar cerca de 101.325 kPa "
                "o ajustada por altitud.</p>"
            )

        interpretation = (
            "Adiv representa la divergencia geométrica. Aatm se calcula por banda y depende de T, HR y presión, con formulación simplificada. "
            "Agr se aplica como término de suelo/terreno y Abar como apantallamiento topográfico básico cuando hay MDT."
            if engine == 'iso_aligned' else
            "Adiv representa la divergencia geométrica, Aatm la atenuación atmosférica simplificada α·distancia y Aground una corrección simplificada del efecto suelo/terreno."
        )

        if engine == 'iso_aligned':
            methodology_flow_html = f"""
            <div class='card card-info'>
                <h3>🧭 Cómo se ha ejecutado el cálculo ISO-aligned</h3>
                <p>Esta sección explica el flujo real que sigue el plugin para que el resultado por receptor sea trazable. El nivel final de cada receptor <b>no sale de una única resta simple</b>, sino de calcular todas las contribuciones fuente–receptor dentro del radio de cálculo y sumarlas energéticamente.</p>
                <ol>
                    <li><b>Lectura de entradas GIS:</b> se toman las turbinas/fuentes acústicas, los receptores, la altura de receptor, el radio máximo de cálculo (<b>{radius:.0f} m</b>), la capa de uso del suelo si existe y el MDT/DSM si está activo.</li>
                    <li><b>Estado acústico de cada grupo fuente:</b> para cada modelo o grupo de turbinas se obtiene un <b>LwA operativo</b> desde un valor fijo o desde una curva <code>LwA(ws)</code>. En esta corrida: <b>{acoustic_txt}</b>.</li>
                    <li><b>Conversión a bandas:</b> el motor ISO-aligned necesita un espectro <code>Lw,b</code> en 8 bandas de octava. Si no hay espectro específico, el plugin reconstruye uno desde una plantilla/fallback y lo ajusta para reproducir el LwA operativo.</li>
                    <li><b>Selección de contribuyentes por receptor:</b> para cada receptor se buscan las turbinas dentro del radio máximo. Los receptores sin fuentes dentro de ese radio se marcan como <b>fuera de radio</b> y no generan nivel acústico útil.</li>
                    <li><b>Cálculo por trayecto fuente–receptor:</b> para cada turbina contribuyente se calcula distancia 3D, cotas acústicas, <b>G</b> o <b>G_eff</b> del suelo y, si hay MDT/DSM, el posible apantallamiento topográfico del trayecto.</li>
                    <li><b>Propagación por banda:</b> en cada banda se aplica <code>Lp,b = Lw,b - Adiv - Aatm,b - Agr,b - Abar,b</code>. Adiv depende de la distancia, Aatm,b de frecuencia/atmósfera, Agr,b del suelo y Abar,b del MDT si hay obstáculo relevante.</li>
                    <li><b>Suma por fuente:</b> las 8 bandas se ponderan en A y se suman energéticamente para obtener el nivel A-ponderado de esa turbina en el receptor.</li>
                    <li><b>Suma del receptor:</b> todas las turbinas contribuyentes se suman energéticamente para obtener el <b>nivel total dB(A)</b> del receptor.</li>
                    <li><b>Comparación con límites:</b> el nivel total se compara con el límite asignado al receptor o con el límite de referencia. De ahí salen el margen, el estado de cumplimiento y la tabla de excedencias.</li>
                </ol>
                <div class='formula'>LpA,receptor = 10·log10(Σ_fuentes 10^(LpA,fuente/10))</div>
                <p><b>Lectura práctica:</b> el receptor crítico es el de mayor nivel total o el que queda con peor margen frente al límite. La columna “fuente dominante” identifica la turbina/grupo que más contribuye, pero el resultado final del receptor incluye todas las fuentes dentro del radio.</p>
            </div>
            <div class='card'>
                <h3>🔎 Qué diferencia este modo del modo Screening</h3>
                <p>El modo ISO-aligned es más pesado pero más trazable: usa bandas de octava, ponderación A final, absorción atmosférica dependiente de frecuencia, suelo por regiones y apantallamiento topográfico <b>Abar</b> cuando hay MDT/DSM. Es el modo recomendado para informes técnicos preliminares y revisión de receptores sensibles.</p>
            </div>
            """
        else:
            methodology_flow_html = f"""
            <div class='card card-info'>
                <h3>🧭 Cómo se ha ejecutado el cálculo Screening</h3>
                <p>Esta sección explica el flujo real que sigue el plugin en el modo rápido. El objetivo es obtener una estimación ágil para mapas, comparación de alternativas y detección inicial de receptores sensibles.</p>
                <ol>
                    <li><b>Lectura de entradas GIS:</b> se toman las turbinas/fuentes acústicas, los receptores, la altura de receptor, el radio máximo de cálculo (<b>{radius:.0f} m</b>) y la capa de uso del suelo si existe.</li>
                    <li><b>Estado acústico de cada grupo fuente:</b> cada modelo o grupo de turbinas trabaja con un único <b>LwA operativo</b>, definido por un valor fijo o por una curva <code>LwA(ws)</code>. En esta corrida: <b>{acoustic_txt}</b>.</li>
                    <li><b>Selección de contribuyentes por receptor:</b> para cada receptor se buscan las turbinas dentro del radio máximo. Los receptores sin fuentes dentro de ese radio se marcan como <b>fuera de radio</b>.</li>
                    <li><b>Cálculo por trayecto fuente–receptor:</b> para cada turbina contribuyente se calcula la distancia 3D, la divergencia geométrica, una absorción atmosférica simplificada <code>α·d</code> y una corrección empírica de suelo.</li>
                    <li><b>Uso del suelo:</b> si hay capa de land-use, el plugin puede calcular un <b>G_eff</b> por trayecto; si no, usa el <b>G global</b> definido por el usuario.</li>
                    <li><b>Propagación simplificada:</b> se aplica <code>Lp = LwA - Adiv - Aatm - Aground</code>. No hay bandas de octava ni apantallamiento topográfico explícito <code>Abar</code>.</li>
                    <li><b>Suma del receptor:</b> todas las turbinas contribuyentes se suman energéticamente para obtener el <b>nivel total dB(A)</b> del receptor.</li>
                    <li><b>Comparación con límites:</b> el nivel total se compara con el límite asignado al receptor o con el límite de referencia. De ahí salen el margen, el estado de cumplimiento y la tabla de excedencias.</li>
                </ol>
                <div class='formula'>LpA,receptor = 10·log10(Σ_fuentes 10^(Lp,fuente/10))</div>
                <p><b>Lectura práctica:</b> este modo es útil para screening inicial. Si un receptor aparece cerca del límite o en excedencia, conviene recalcularlo con el modo ISO-aligned y revisar espectros, terreno, uso del suelo y límites aplicados.</p>
            </div>
            <div class='card'>
                <h3>🔎 Qué diferencia este modo del modo ISO-aligned</h3>
                <p>El modo Screening sacrifica detalle para ganar velocidad. No propaga por bandas, no usa T/HR/P por frecuencia, no calcula <b>Abar</b> desde MDT y resume la atmósfera con un único coeficiente <b>α</b>. Por eso debe interpretarse como preevaluación rápida, no como informe acústico detallado.</p>
            </div>
            """

        octave_rows = ''.join([
            f"<tr><td>{freq}</td><td style='text-align:right;'>{float(a_w):.1f}</td></tr>"
            for freq, a_w in [(63, -26.2), (125, -16.1), (250, -8.6), (500, -3.2), (1000, 0.0), (2000, 1.2), (4000, 1.0), (8000, -1.1)]
        ])
        atm_rows = ''.join([
            f"<tr><td>{freq}</td><td style='text-align:right;'>{alpha_ref:.4f}</td></tr>"
            for freq, alpha_ref in [(63, 0.0001), (125, 0.0003), (250, 0.0008), (500, 0.0020), (1000, 0.0040), (2000, 0.0095), (4000, 0.0280), (8000, 0.0900)]
        ])
        ground_rows = ''.join([
            "<tr><td>≤ 500 Hz</td><td style='text-align:right;'>A_ground = 1.5 dB</td></tr>",
            "<tr><td>1000 Hz</td><td style='text-align:right;'>1.5·(1 - e^(-h/10))</td></tr>",
            "<tr><td>2000 Hz</td><td style='text-align:right;'>3.0·(1 - e^(-h/10))</td></tr>",
            "<tr><td>4000 Hz</td><td style='text-align:right;'>6.0·(1 - e^(-h/10))</td></tr>",
            "<tr><td>8000 Hz</td><td style='text-align:right;'>12.0·(1 - e^(-h/10))</td></tr>",
        ])

        if engine == 'iso_aligned':
            if dem_used:
                mdt_expl_html = f"""
                <div class='card'>
                    <h3>🗺️ Física del MDT y del apantallamiento topográfico</h3>
                    <p>En el motor ISO-aligned, el MDT <b>no cambia la emisión de la turbina</b> ni la absorción atmosférica. Su función es describir la <b>geometría real del trayecto fuente–receptor</b> y alimentar el término de apantallamiento topográfico <b>Abar,b</b>.</p>
                    <div class='formula'>Lp,b = Lw,b - Adiv - Aatm,b - Agr,b - Abar,b</div>
                    <h4>Cómo entra el MDT en el cálculo</h4>
                    <ol>
                        <li><b>Perfil del terreno:</b> se extrae el perfil fuente–receptor desde el MDT con un <b>muestreo adaptativo</b>, ajustado a la distancia y a la resolución del raster. El perfil se calcula <b>una sola vez</b> por par fuente–receptor y se reutiliza en las 8 bandas para reducir tiempo de cálculo.</li>
                        <li><b>Línea de visión directa:</b> se compara el perfil con la recta que une la fuente acústica a su altura efectiva y el receptor a su altura de evaluación. Si el terreno queda siempre por debajo de esa recta, no hay obstáculo topográfico relevante y <b>Abar,b = 0</b>.</li>
                        <li><b>Detección de obstáculo dominante:</b> si una loma o cresta del MDT sobresale por encima de la línea de visión, el modelo interpreta que existe pantalla topográfica. La magnitud clave es la altura del obstáculo sobre la línea de visión:</li>
                    </ol>
                    <div class='formula'>h_obs = z_terreno - z_LOS</div>
                    <p>cuando <b>h_obs &gt; 0</b>, el relieve corta la visión directa y puede aparecer atenuación adicional por difracción.</p>
                    <ol start='4'>
                        <li><b>Activación conservadora:</b> no se activa Abar por pequeñas irregularidades del relieve; se aplica un umbral mínimo ligado a la resolución del MDT.</li>
                        <li><b>Geometría real del obstáculo:</b> el cálculo usa la <b>posición real</b> del obstáculo dominante y obtiene <b>d1</b> (fuente → obstáculo) y <b>d2</b> (obstáculo → receptor) reales, en lugar de asumir siempre un obstáculo en el punto medio.</li>
                        <li><b>Difracción tipo Fresnel:</b> con esa geometría se estima una diferencia de caminos aproximada y se transforma en una atenuación dependiente de la frecuencia:</li>
                    </ol>
                    <div class='formula'>δ ≈ 0.5·h_obs²·(1/d1 + 1/d2) &nbsp;&nbsp; ; &nbsp;&nbsp; C = (2·f·δ)/c</div>
                    <p>donde <b>δ</b> es la diferencia de caminos aproximada, <b>f</b> la frecuencia y <b>c</b> la velocidad del sonido. El número <b>C</b> se traduce después a una atenuación <b>Abar,b</b> mayor cuanto más bloquea el relieve el trayecto. Esta es la misma aproximación implementada en el cálculo.</p>
                    <p><b>Interpretación física:</b> en terreno plano o cuando no hay intersección con la línea de visión, <b>Abar</b> suele ser despreciable. En terreno complejo, el MDT puede introducir varios dB de atenuación adicional y cambiar el receptor crítico.</p>
                    <p><b>Implementación actual:</b> obstáculo dominante único, perfil adaptativo con límites de coste, geometría real del obstáculo, activación conservadora y atenuación capada a valores razonables.</p>
                    <p><b>MDT usado en esta corrida:</b> {dem_layer_name or 'sin nombre'}.</p>
                </div>
                """
            else:
                mdt_expl_html = """
                <div class='card'>
                    <h3>🗺️ Física del MDT y del apantallamiento topográfico</h3>
                    <p>En esta corrida <b>no se ha usado MDT/DSM</b>, por lo que el término de apantallamiento topográfico se fija en:</p>
                    <div class='formula'>Abar,b = 0</div>
                    <p>La evaluación se realiza sin introducir pantallas topográficas. La geometría del trayecto se resuelve sin perfil del terreno y el cálculo depende de Lw,b, Adiv, Aatm,b y Agr,b.</p>
                </div>
                """

            if ground_mode == 'landuse':
                ground_expl_html = f"""
                <div class='card'>
                    <h3>🌱 Física del uso del suelo y cálculo de G_eff</h3>
                    <p>Cuando el modo suelo es <b>desde capa</b>, el cálculo no usa un único valor manual para todo el parque. Para cada trayecto fuente–receptor se obtiene un <b>G_eff</b> calculado desde la capa de uso del suelo:</p>
                    <div class='formula'>G_eff = (Σ G_i · L_i) / (Σ L_i)</div>
                    <p>donde <b>G_i</b> es el valor asignado a cada polígono interceptado por el trayecto y <b>L_i</b> es la longitud del trayecto dentro de ese polígono.</p>
                    <ul>
                        <li><b>G = 0</b>: suelo duro (urbano/asfalto/roca).</li>
                        <li><b>G = 0.5</b>: terreno mixto.</li>
                        <li><b>G = 1</b>: suelo blando/poroso (agrícola, pradera, forestal, vegetado).</li>
                    </ul>
                    <p><b>Importante:</b> el <b>G global</b> mostrado en el informe es solo un valor de respaldo. Cuando hay capa de uso del suelo, el cálculo usa realmente <b>G_eff</b> por trayecto. En esta corrida, el valor medio efectivo fue <b>{float(g_eff_stats.get('mean', g)):.2f}</b> y el del receptor crítico <b>{float(g_eff_stats.get('critical', g)):.2f}</b>.</p>
                    <p><b>Capa usada:</b> {landuse_layer_name or 'sin nombre'}.</p>
                </div>
                """
            else:
                ground_expl_html = f"""
                <div class='card'>
                    <h3>🌱 Física del uso del suelo y cálculo de G</h3>
                    <p>En esta corrida el efecto suelo se ha calculado con un <b>G único manual</b> para todo el trayecto:</p>
                    <div class='formula'>G = {g:.2f}</div>
                    <p>Ese valor se aplica en el término de suelo del modelo. No se ha derivado un G_eff desde capa de uso del suelo.</p>
                </div>
                """

            equations_detail_html = f"""
            <div class='card'>
                <h3>📘 Desarrollo físico detallado del motor ISO-aligned</h3>
                <p>Este motor trabaja en <b>8 bandas de octava</b> (63–8000 Hz). Las bandas no son un resultado del cálculo, sino la <b>malla frecuencial del método</b>. Para aplicar la propagación por bandas, el cálculo necesita un <b>input acústico por banda</b> de la fuente <code>Lw,b</code>. Ese input puede venir de un espectro medido/importado o de una plantilla/fallback ajustada al nivel global operativo.</p>
                <p><b>Escenario operativo de esta corrida:</b> {acoustic_txt}.</p>
                <p><b>Ecuación general por banda:</b></p>
                <div class='formula'>Lp,b = Lw,b - Adiv - Aatm,b - Agr,b - Abar,b</div>
                <p><b>Suma A-ponderada final:</b></p>
                <div class='formula'>LpA,total = 10·log10(Σ 10^((Lp,b + A_weight)/10))</div>
                <h4>0. Inputs realmente usados en esta corrida</h4>
                <ul>
                    <li><b>Fuente acústica:</b> <code>Lw,b</code> por bandas de octava. Si existe espectro específico del grupo fuente, ese es el input usado. Si no, el plugin usa una biblioteca/plantilla/fallback y la ajusta al nivel global operativo.</li>
                    <li><b>Nivel operativo global:</b> viene de <b>LwA fijo</b> o de una <b>curva acústica LwA(ws)</b> según el escenario seleccionado. Ese nivel global no sustituye a las bandas: fija el estado operativo y el espectro aporta el reparto frecuencial.</li>
                    <li><b>Geometría:</b> coordenadas de fuente y receptor, altura de receptor, altura efectiva de fuente y distancia 3D.</li>
                    <li><b>Atmósfera:</b> temperatura <b>T</b>, humedad relativa <b>HR</b> y presión <b>P</b>.</li>
                    <li><b>Suelo:</b> un <b>G global manual</b> o un <b>G_eff</b> derivado desde la capa de uso del suelo.</li>
                    <li><b>Topografía:</b> MDT/DSM opcional. Solo afecta al cálculo de <b>Abar,b</b>.</li>
                </ul>
                <h4>1. De dónde sale cada término de la ecuación</h4>
                <table>
                    <tr><th>Término</th><th>Cómo se obtiene en este plugin</th></tr>
                    <tr><td><b>Lw,b</b></td><td>Input acústico por bandas. Sale del espectro del grupo fuente (CSV, biblioteca, plantilla o fallback ajustado al nivel global). La curva acústica LwA(ws) o el LwA fijo definen el nivel global operativo de la turbina, y el espectro por bandas reparte ese nivel entre las 8 bandas.</td></tr>
                    <tr><td><b>Adiv</b></td><td>Se calcula a partir de la distancia 3D fuente–receptor.</td></tr>
                    <tr><td><b>Aatm,b</b></td><td>Se calcula por banda con una tabla base de absorción <code>α_ref(f)</code> y correcciones simplificadas por temperatura, humedad relativa y presión. La implementación actual usa la matemática exacta del plugin: <code>α = α_ref(f)·corr_T·corr_HR·corr_P</code>.</td></tr>
                    <tr><td><b>Agr,b</b></td><td>Se calcula como efecto suelo por regiones. El parámetro de suelo usado es un <b>G único por trayecto</b>: manual/global o <b>G_eff</b> derivado desde la capa de uso del suelo.</td></tr>
                    <tr><td><b>Abar,b</b></td><td>Solo entra si hay MDT/DSM y se detecta apantallamiento topográfico. Si no hay MDT o no hay obstáculo relevante, <b>Abar,b = 0</b>.</td></tr>
                </table>
                <h4>2. Input acústico de fuente y bandas</h4>
                <p>En este motor, el término <code>Lw,b</code> es un <b>dato de entrada por banda</b>. Las <b>bandas de octava</b> (63–8000 Hz) no son un resultado de la ISO ni una tabla calculada por el plugin: son la <b>malla frecuencial</b> sobre la que se resuelve la propagación.</p>
                <p>El plugin combina dos piezas:</p>
                <ul>
                    <li><b>Curva acústica global LwA(ws)</b>: fija el <b>nivel operativo global</b> de la turbina para la velocidad de viento o el peor caso seleccionado.</li>
                    <li><b>Espectro por bandas Lw,b</b>: reparte ese nivel global entre las 8 bandas y es el input real usado en la ecuación por bandas.</li>
                </ul>
                <p>Ese espectro puede proceder de un archivo específico del fabricante/usuario o de una plantilla de referencia. Si solo se dispone de una curva global <code>LwA(ws)</code>, el plugin fija primero el nivel global operativo <code>LwA_objetivo</code> y después construye un espectro absoluto por bandas a partir de una forma espectral de referencia <code>S_b^ref</code>.</p>
                <p><b>Reconstrucción matemática de las bandas cuando solo existe LwA(ws):</b></p>
                <div class='formula'>Lw,b = S_b^ref + Δ</div>
                <div class='formula'>Δ = LwA_objetivo - 10·log10(Σ 10^((S_b^ref + A_weight,b)/10))</div>
                <p>Es decir: la curva acústica aporta el <b>nivel global operativo</b> y la plantilla/biblioteca aporta la <b>forma espectral</b>. El desplazamiento <b>Δ</b> se calcula para que, al ponderar en A y sumar energéticamente las 8 bandas, el espectro reconstruido reproduzca exactamente el <code>LwA_objetivo</code> de la curva importada.</p>
                {spectrum_detail_html}
                <h4>3. Divergencia geométrica</h4>
                <div class='formula'>Adiv = 20·log10(d) + 11</div>
                <p>Representa la dispersión geométrica de la onda sonora con la distancia 3D fuente–receptor. Aquí <b>d</b> sale de las coordenadas de turbina y receptor junto con sus alturas de evaluación.</p>
                <h4>4. Absorción atmosférica simplificada</h4>
                <div class='formula'>Aatm,b = α(f, T, HR, P) · d</div>
                <p>La absorción atmosférica se calcula por banda a partir de un coeficiente base de referencia y tres factores correctores. La dependencia física con temperatura, humedad relativa y presión <b>sí se representa</b>, pero mediante una <b>aproximación simplificada del plugin</b>, no mediante la formulación analítica completa de ISO 9613-1.</p>
                <div class='formula'>α(f, T, HR, P) = α_ref(f) · corr_T · corr_HR · corr_P</div>
                <div class='formula'>corr_T = 1 + 0.01·(T - 15) &nbsp;&nbsp; ; &nbsp;&nbsp; corr_HR = 1 + 0.003·|HR - 50| &nbsp;&nbsp; ; &nbsp;&nbsp; corr_P = 101.325 / P</div>
                <p><b>Interpretación de las correcciones:</b> <b>T</b> se introduce en °C respecto a una referencia de 15 °C; <b>HR</b> se compara con una humedad óptima de referencia del 50% y la corrección crece al alejarse de ese valor; y <b>P</b> se introduce en kPa respecto a una referencia de 101.325 kPa con una corrección inversa. Estos factores modifican únicamente el bloque atmosférico <b>Aatm,b</b>: no alteran ni la emisión de la turbina, ni el efecto del suelo, ni el término de MDT/apantallamiento.</p>
                <table>
                    <tr><th>Banda [Hz]</th><th style='text-align:right;'>α_ref [dB/m]</th></tr>
                    {atm_rows}
                </table>
                <h4>5. Efecto suelo por regiones</h4>
                <div class='formula'>Agr,b = As + Am + Ar</div>
                <p>El término de suelo se descompone en <b>As</b> (región de fuente), <b>Am</b> (región media) y <b>Ar</b> (región de receptor). En esta implementación no se usan tres parámetros de suelo independientes <code>Gs/Gm/Gr</code>, sino un <b>único G por trayecto</b>. Matemáticamente, el plugin aplica:</p>
                <div class='formula'>As = G_eff·A_ground(h_s)</div>
                <div class='formula'>Am = G_eff·(1 - G_m)·A_ground(h_medio)</div>
                <div class='formula'>Ar = G_eff·A_ground(h_r)</div>
                <p>donde <b>h_s</b> es la altura característica de la fuente, <b>h_r</b> la del receptor, <b>h_medio</b> la altura media del trayecto y <b>G_m≈0</b> en la aproximación actual para condiciones favorables de propagación. Ese valor único de suelo puede ser:</p>
                <ul>
                    <li><b>G manual/global</b>, si el usuario fija un único valor.</li>
                    <li><b>G_eff</b>, si existe capa de uso del suelo y se calcula una media ponderada por longitud del trayecto.</li>
                </ul>
                <div class='formula'>G_eff = (Σ G_i · L_i) / (Σ L_i)</div>
                <p><b>Significado físico de G:</b> representa el carácter acústico del terreno y controla cómo influye el suelo en la propagación. <b>G≈0</b> indica suelo duro (urbano, asfalto, roca), <b>G≈1</b> suelo blando/poroso (agrícola, pradera, forestal) y valores intermedios representan terreno mixto.</p>
                <p><b>Qué significa “desde capa”:</b> el plugin corta el trayecto fuente–receptor contra la capa de uso del suelo, asigna un valor <b>G_i</b> a cada polígono interceptado y calcula un único <b>G_eff</b> para ese trayecto. Ese es el valor que entra realmente en <b>Agr,b</b>; el <b>G global</b> mostrado en el informe queda solo como respaldo.</p>
                <p><b>Convención del informe:</b> <b>Agr,b</b> se muestra aquí como una <b>magnitud positiva de atenuación</b>. En la ecuación principal se resta al nivel de fuente igual que Adiv, Aatm y Abar.</p>
                <table>
                    <tr><th>Banda [Hz]</th><th style='text-align:right;'>Término base A_ground(h)</th></tr>
                    {ground_rows}
                </table>
                <h4>6. Apantallamiento topográfico con MDT</h4>
                <p>El MDT <b>no cambia la emisión</b> de la turbina ni la absorción atmosférica. Su función es describir la <b>geometría real del trayecto</b> y alimentar el término <b>Abar,b</b>.</p>
                <ol>
                    <li><b>Perfil del terreno:</b> se extrae el perfil fuente–receptor desde el MDT con muestreo adaptativo.</li>
                    <li><b>Línea de visión:</b> se construye la recta entre la altura efectiva de fuente y la altura del receptor. Si el terreno queda siempre por debajo, entonces <b>Abar,b = 0</b>.</li>
                    <li><b>Obstáculo dominante:</b> si una loma o cresta sobresale, se calcula la altura sobre la línea de visión:</li>
                </ol>
                <div class='formula'>h_obs = z_terreno - z_LOS</div>
                <p>Cuando <b>h_obs &gt; 0</b>, el relieve corta la visión directa y puede aparecer atenuación adicional por difracción.</p>
                <ol start='4'>
                    <li><b>Geometría real del obstáculo:</b> el plugin usa la posición real del obstáculo dominante y calcula <b>d1</b> (fuente → obstáculo) y <b>d2</b> (obstáculo → receptor).</li>
                    <li><b>Activación conservadora:</b> no se activa <b>Abar</b> por pequeñas irregularidades del MDT; se exige un umbral mínimo ligado a la resolución del raster.</li>
                    <li><b>Difracción tipo Fresnel:</b> con esa geometría se estima una diferencia de caminos y un número de Fresnel:</li>
                </ol>
                <div class='formula'>δ ≈ 0.5·h_obs²·(1/d1 + 1/d2) &nbsp;&nbsp; ; &nbsp;&nbsp; C = (2·f·δ)/c</div>
                <p>Ese número se transforma después en una atenuación <b>Abar,b</b> dependiente de la frecuencia mediante la aproximación actual del plugin:</p>
                <div class='formula'>si C ≤ -2 → Abar = 0 &nbsp;&nbsp; ; &nbsp;&nbsp; -2 &lt; C ≤ 0 → Abar = 10·log10(3 + 20·C)</div>
                <div class='formula'>0 &lt; C ≤ 3.5 → Abar = 10·log10(3 + 80·C) &nbsp;&nbsp; ; &nbsp;&nbsp; C &gt; 3.5 → Abar = 10·log10(3 + 280·C)</div>
                <p>En la implementación actual, <b>Abar</b> se limita además a valores razonables (capado superior) para evitar sobreatenuaciones espurias. Si no hay MDT o no hay obstáculo relevante, entonces <b>Abar,b = 0</b>.</p>
                <h4>7. Ponderación A usada al final</h4>
                <table>
                    <tr><th>Banda [Hz]</th><th style='text-align:right;'>A_weight [dB]</th></tr>
                    {octave_rows}
                </table>
                <p><b>Lectura del receptor crítico:</b> la tabla del apartado de receptor crítico muestra magnitudes de atenuación para trazabilidad. El <b>nivel resultante</b> no debe interpretarse como una resta directa de una sola turbina: se obtiene mediante suma energética por bandas y suma de las fuentes contribuyentes dentro del radio de cálculo.</p>
            </div>
            {ground_expl_html}
            {mdt_expl_html}
            """
        else:
            if ground_mode == 'landuse':
                fast_ground_html = f"""
                <h4>3. Efecto suelo simplificado con uso del suelo</h4>
                <p>En el motor rápido el término <b>Aground</b> sigue siendo empírico, pero el parámetro de suelo puede venir de la capa de uso del suelo como un <b>G_eff</b> por trayecto:</p>
                <div class='formula'>G_eff = (Σ G_i · L_i) / (Σ L_i)</div>
                <div class='formula'>Aground = min(6, max(0, G_eff · 3·log10(1 + d_xy/100) · 1/(1 + (h_s + h_r)/80)))</div>
                <p>Ese <b>G_eff</b> se usa después dentro de la corrección simplificada de terreno del motor rápido. El valor global <b>G = {g:.2f}</b> queda solo como respaldo si la capa no aporta información válida.</p>
                <p><b>Capa usada:</b> {landuse_layer_name or 'sin nombre'} · <b>G_eff medio:</b> {float(g_eff_stats.get('mean', g)):.2f} · <b>G_eff receptor crítico:</b> {float(g_eff_stats.get('critical', g)):.2f}</p>
                """
            else:
                fast_ground_html = f"""
                <h4>3. Efecto suelo simplificado</h4>
                <p>El término <b>Aground</b> es una corrección empírica del terreno controlada por un único parámetro manual:</p>
                <div class='formula'>G = {g:.2f}</div>
                <div class='formula'>Aground = min(6, max(0, G · 3·log10(1 + d_xy/100) · 1/(1 + (h_s + h_r)/80)))</div>
                <p>En esta corrida no se ha derivado un G_eff desde una capa de uso del suelo. Aquí <b>d_xy</b> es la distancia horizontal, <b>h_s</b> la altura de fuente y <b>h_r</b> la altura del receptor.</p>
                """

            fast_mdt_html = """
                <h4>4. MDT / topografía</h4>
                <p>En el motor rápido el MDT no introduce un término explícito de apantallamiento topográfico. Aunque exista una capa de relieve en el proyecto, este modo no calcula <b>Abar</b>, no extrae línea de visión ni aplica difracción; por tanto la física se basa únicamente en <b>LwA</b>, <b>Adiv</b>, <b>Aatm = α·d</b> y la corrección empírica de terreno <b>Aground</b>.</p>
            """

            equations_detail_html = f"""
            <div class='card'>
                <h3>📘 Desarrollo físico detallado del motor rápido</h3>
                <div class='formula'>Lp = LwA - Adiv - Aatm - Aground</div>
                <p>El motor rápido trabaja con un único nivel global <b>LwA</b> por grupo fuente. Está pensado para screening, mapas ágiles y comparativas rápidas, sacrificando detalle espectral frente a velocidad. En este modo <b>no hay propagación por bandas</b> ni término explícito de apantallamiento topográfico.</p>
                <p><b>Escenario operativo de esta corrida:</b> {acoustic_txt}.</p>
                <h4>0. Inputs realmente usados en esta corrida</h4>
                <ul>
                    <li><b>Fuente acústica:</b> un único nivel global <b>LwA</b> por grupo fuente.</li>
                    <li><b>Nivel operativo global:</b> sale de un <b>LwA fijo</b> o de una <b>curva acústica LwA(ws)</b> para la velocidad o peor caso seleccionados.</li>
                    <li><b>Geometría:</b> coordenadas de fuente y receptor, altura de receptor, altura efectiva de fuente y distancia 3D.</li>
                    <li><b>Atmósfera:</b> en este modo no se usan T/HR/P; la absorción se resume en un único coeficiente <b>α</b>.</li>
                    <li><b>Suelo:</b> un <b>G global manual</b> o un <b>G_eff</b> derivado desde la capa de uso del suelo.</li>
                    <li><b>Topografía:</b> el MDT no entra como apantallamiento explícito en este modo.</li>
                </ul>
                <h4>1. De dónde sale cada término de la ecuación</h4>
                <table>
                    <tr><th>Término</th><th>Cómo se obtiene en este plugin</th></tr>
                    <tr><td><b>LwA</b></td><td>Input global de la fuente. Sale de un valor fijo por grupo o de una curva acústica <code>LwA(ws)</code> para la velocidad/peor caso seleccionados.</td></tr>
                    <tr><td><b>Adiv</b></td><td>Se calcula a partir de la distancia 3D fuente–receptor.</td></tr>
                    <tr><td><b>Aatm</b></td><td>Se calcula con un único coeficiente constante <code>α</code> multiplicado por la distancia.</td></tr>
                    <tr><td><b>Aground</b></td><td>Corrección empírica del efecto suelo. El parámetro de suelo puede ser un <b>G global manual</b> o un <b>G_eff</b> derivado desde la capa de uso del suelo.</td></tr>
                </table>
                <h4>2. Divergencia geométrica</h4>
                <div class='formula'>Adiv = 20·log10(d) + 11</div>
                <p>Representa la dispersión geométrica de la onda sonora con la distancia 3D fuente–receptor.</p>
                <h4>3. Absorción atmosférica simplificada</h4>
                <div class='formula'>Aatm = α · d</div>
                <p>En esta corrida se ha usado <b>α = {alpha:.4f} dB/m</b>. En el motor rápido, la absorción atmosférica se resume en un único coeficiente constante, por lo que <b>T</b>, <b>HR</b> y <b>P</b> <b>no entran explícitamente</b> en el cálculo. Esa es una de las simplificaciones clave frente al modo ISO-aligned.</p>
                {fast_ground_html}
                {fast_mdt_html}
                <h4>5. Qué no hace este modo</h4>
                <p>El motor rápido no trabaja por bandas, no calcula <b>Lw,b</b>, no introduce <b>Abar</b> y no extrae línea de visión ni difracción desde MDT. Por tanto, es apropiado para screening y comparativas rápidas, pero no para análisis espectral detallado.</p>
            </div>
            """

        # === CALCULAR TASAS Y FECHA ===
        coverage_rate = (100.0 * n_with / n_receivers) if n_receivers else 0
        exceed_rate = (100.0 * n_exceed / n_with) if n_with else 0
        comply_rate = 100.0 - exceed_rate
        from datetime import datetime
        now = datetime.now()

        # === BANNER DE ALCANCE (lo primero que se lee, antes de cualquier cifra) ===
        if engine == 'iso_aligned':
            scope_what_is = "una evaluación acústica preliminar alineada con la metodología ISO 9613-2, pensada para diseño, comparación de alternativas y cribado de receptores sensibles."
            scope_what_not = "no es un informe acústico certificado ni sustituye a un estudio regulatorio definitivo realizado con software comercial validado."
            scope_simpl_items = [
                "Absorción atmosférica Aatm mediante tabla de referencia con correcciones simplificadas de temperatura, humedad y presión, no la formulación analítica completa de ISO 9613-1.",
                "Sin corrección meteorológica de largo plazo Cmet.",
                "Difracción topográfica de un único obstáculo dominante: sin difracción lateral ni pantallas múltiples.",
                "Resolución espectral en 8 bandas de octava de 63 a 8000 Hz, no en tercios de octava.",
                "Directividad de fuente Dc asumida 0 dB.",
            ]
        else:
            scope_what_is = "una estimación rápida de cribado (screening) para mapas ágiles y comparación de alternativas de implantación."
            scope_what_not = "no es un cálculo espectral detallado ni un informe regulatorio; para receptores cercanos al límite conviene recalcular en modo ISO-aligned."
            scope_simpl_items = [
                "Sin propagación por bandas de octava.",
                "Absorción atmosférica resumida en un único coeficiente alfa constante.",
                "Sin apantallamiento topográfico Abar desde el MDT.",
                "Efecto suelo mediante corrección empírica simplificada.",
            ]
        scope_reco = "Para decisiones regulatorias críticas, valida los resultados con mediciones de campo o software comercial certificado."
        scope_items_html = ''.join(f"<li>{it}</li>" for it in scope_simpl_items)
        scope_banner_html = f"""
        <div style='background:#fff8e1;border:2px solid #f0ad4e;border-left:8px solid #f0ad4e;border-radius:8px;padding:18px 22px;margin:0 0 26px 0;'>
            <h3 style='margin:0 0 10px 0;color:#7a5b00;'>⚠️ Alcance de este informe — léelo antes de usar los resultados</h3>
            <p style='margin:6px 0;'><b>Qué es:</b> {scope_what_is}</p>
            <p style='margin:6px 0;'><b>Qué no es:</b> {scope_what_not}</p>
            <p style='margin:10px 0 4px 0;'><b>Simplificaciones aplicadas en este modo:</b></p>
            <ul style='margin:4px 0 10px 0;'>{scope_items_html}</ul>
            <p style='margin:6px 0 0 0;'><b>Recomendación:</b> {scope_reco}</p>
        </div>
        """

        # === GLOSARIO DE SÍMBOLOS (decodifica fórmulas y tablas en un solo sitio) ===
        glossary_rows = [
            ("LwA", "Nivel de potencia sonora A-ponderado de la fuente, en dB(A)."),
            ("Lw,b", "Potencia sonora de la fuente por banda de octava, en dB."),
            ("S_b^ref", "Forma espectral de referencia por banda usada como plantilla, en dB."),
            ("A_weight,b", "Ponderación A aplicada a cada banda de octava, en dB."),
            ("Δ", "Desplazamiento global aplicado a la plantilla espectral para reproducir el LwA objetivo, en dB."),
            ("LpA", "Nivel de presión sonora A-ponderado resultante en el receptor, en dB(A)."),
            ("Adiv", "Atenuación por divergencia geométrica con la distancia, en dB."),
            ("Aatm", "Atenuación por absorción atmosférica del aire, en dB."),
            ("Agr", "Atenuación por efecto del suelo, en dB."),
            ("Abar", "Atenuación por apantallamiento topográfico, solo en modo ISO con MDT, en dB."),
            ("d", "Distancia tridimensional entre fuente y receptor, en metros."),
            ("G / G_eff", "Factor de suelo de 0 (duro) a 1 (blando) y su valor efectivo por trayecto."),
            ("Cmet", "Corrección meteorológica de largo plazo, no aplicada en este plugin."),
            ("Dc", "Corrección por directividad de la fuente, asumida 0 dB."),
        ]
        glossary_rows_html = ''.join(f"<tr><td><b>{sym}</b></td><td>{desc}</td></tr>" for sym, desc in glossary_rows)
        glossary_html = f"""
        <div class='card card-info'>
            <h3>📖 Glosario de símbolos</h3>
            <p>Definición compacta de los símbolos que aparecen en las fórmulas y tablas de este informe.</p>
            <table>
                <tr><th>Símbolo</th><th>Significado</th></tr>
                {glossary_rows_html}
            </table>
        </div>
        """

        html = f"""
        <style>
            body {{
                font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
                line-height: 1.6;
                color: #212529;
            }}
            h1, h2, h3 {{
                color: #1e3a5f;
                font-weight: 600;
                margin-top: 24px;
                margin-bottom: 12px;
            }}
            h2 {{
                border-left: 4px solid #4a90d9;
                padding-left: 12px;
            }}
            .card {{
                background: #f8f9fa;
                border: 1px solid #e9ecef;
                border-radius: 8px;
                padding: 20px;
                margin: 16px 0;
                box-shadow: 0 2px 4px rgba(0,0,0,0.05);
            }}
            .card-success {{
                border-left: 5px solid #28a745;
            }}
            .card-danger {{
                border-left: 5px solid #dc3545;
            }}
            .card-info {{
                border-left: 5px solid #4a90d9;
            }}
            .metrics-grid {{
                display: grid;
                grid-template-columns: repeat(3, 1fr);
                gap: 16px;
                margin: 20px 0;
            }}
            .metric {{
                background: white;
                padding: 20px;
                border-radius: 8px;
                box-shadow: 0 2px 8px rgba(0,0,0,0.1);
                text-align: center;
                border-top: 4px solid #4a90d9;
            }}
            .metric-value {{
                font-size: 32px;
                font-weight: 700;
                color: #1e3a5f;
                margin: 8px 0;
            }}
            .metric-label {{
                font-size: 14px;
                color: #343a40;
                font-weight: 500;
            }}
            table {{
                width: 100%;
                border-collapse: collapse;
                margin: 16px 0;
                font-size: 13px;
            }}
            th {{
                background: #1e3a5f;
                color: white;
                padding: 12px;
                text-align: left;
                font-weight: 600;
            }}
            td {{
                padding: 10px 12px;
                border-bottom: 1px solid #e9ecef;
            }}
            tr:nth-child(even) {{
                background: #f8f9fa;
            }}
            .badge {{
                display: inline-block;
                padding: 4px 12px;
                border-radius: 12px;
                font-size: 11px;
                font-weight: 700;
                text-transform: uppercase;
            }}
            .badge-success {{
                background: #28a745;
                color: white;
            }}
            .badge-danger {{
                background: #dc3545;
                color: white;
            }}
            .formula {{
                background: #f1f3f5;
                border: 1px solid #dee2e6;
                padding: 16px;
                margin: 12px 0;
                border-radius: 6px;
                font-family: 'Courier New', monospace;
                font-size: 14px;
            }}
            .disclaimer {{
                background: #fff3cd;
                border-left: 5px solid #ffc107;
                padding: 16px;
                margin: 20px 0;
                border-radius: 4px;
            }}
            .note {{
                background: #fff8e1;
                border-left: 4px solid #f0ad4e;
                padding: 10px 12px;
                margin: 10px 0;
                border-radius: 4px;
                color: #5f4300;
            }}
            ol {{
                margin: 12px 0;
                padding-left: 26px;
            }}
            ol li {{
                margin: 8px 0;
            }}
            ul {{
                margin: 12px 0;
                padding-left: 24px;
            }}
            li {{
                margin: 6px 0;
            }}
        </style>
        
        <div style='background: linear-gradient(135deg, #1e3a5f 0%, #2c5f8d 100%); color: white; padding: 30px; border-radius: 10px; margin-bottom: 30px;'>
            <h1 style='color: white; margin: 0 0 8px 0; font-size: 32px;'>📊 INFORME TÉCNICO DE IMPACTO ACÚSTICO</h1>
            <p style='font-size: 16px; opacity: 0.9; margin: 0;'>Evaluación de ruido generado por aerogeneradores</p>
            <p style='font-size: 14px; opacity: 0.85; margin-top: 12px;'>📅 {now.strftime('%d/%m/%Y - %H:%M:%S')}</p>
        </div>
        
        {scope_banner_html}
        
        <h2>1. RESUMEN EJECUTIVO</h2>
        
        <div class='metrics-grid'>
            <div class='metric'>
                <div class='metric-value'>{n_sources}</div>
                <div class='metric-label'>Aerogeneradores</div>
            </div>
            <div class='metric'>
                <div class='metric-value'>{n_receivers}</div>
                <div class='metric-label'>Receptores Evaluados</div>
            </div>
            <div class='metric'>
                <div class='metric-value'>{max_noise:.1f}</div>
                <div class='metric-label'>Nivel Máximo (dB(A))</div>
            </div>
        </div>
        
        <div style='display: grid; grid-template-columns: 1fr 1fr; gap: 16px;'>
            <div class='card card-{'success' if coverage_rate > 80 else 'info'}'>
                <h3>📍 Cobertura de Análisis</h3>
                <p><strong>{n_with} receptores</strong> dentro del radio<br>
                <strong>{coverage_rate:.1f}%</strong> de cobertura<br>
                {n_without} receptores fuera de radio</p>
            </div>
            
            <div class='card card-{'success' if comply_rate > 90 else 'danger' if comply_rate < 50 else 'info'}'>
                <h3>✓ Cumplimiento Normativo</h3>
                <p><strong>{n_exceed} receptores</strong> superan límites<br>
                <strong>{comply_rate:.1f}%</strong> de cumplimiento sobre receptores cubiertos<br>
                Límite: {float(limit_stats.get('min',45)):.1f}–{float(limit_stats.get('max',45)):.1f} dB(A)</p>
            </div>
        </div>
        
        <div class='card card-info'>
            <h3>🎯 Metodología de Cálculo</h3>
            <p><b>Motor utilizado:</b> {engine_label}</p>
            <p><b>Grupos fuente acústicos:</b> {n_models} modelo(s) de aerogenerador</p>
            <p><b>Método:</b> {'Propagación por bandas de octava según metodología ISO-aligned' if engine == 'iso_aligned' else 'Cálculo acústico simplificado para screening'}</p>
            <p><b>Mapa raster:</b> {grid_txt}</p>
        </div>

        <h2>2. CÓMO SE HA GENERADO EL RESULTADO</h2>
        {methodology_flow_html}
        
        <h2>3. RECEPTOR CRÍTICO</h2>
        {crit_html}
        
        <div class='card'>
            <h3>📊 Estadísticos de Atenuaciones (Receptores Cubiertos)</h3>
        <p style='margin: 6px 0 10px 0; color:#495057;'><i>Se muestran magnitudes brutas de atenuación (no el signo algebraico dentro de la ecuación). Para Abar se usa el máximo entre las turbinas contribuyentes de cada receptor, no solo la trayectoria dominante.</i></p>
            <table>
                <tr>
                    <th>Término</th>
                    <th style='text-align: right;'>Media [dB]</th>
                    <th style='text-align: right;'>Máximo [dB]</th>
                </tr>
                <tr>
                    <td><b>Adiv</b> (divergencia geométrica)</td>
                    <td style='text-align: right;'>{float(adiv_stats.get('mean',0.0)):.2f}</td>
                    <td style='text-align: right;'>{float(adiv_stats.get('max',0.0)):.2f}</td>
                </tr>
                <tr>
                    <td><b>Aatm</b> (absorción atmosférica)</td>
                    <td style='text-align: right;'>{float(aatm_stats.get('mean',0.0)):.2f}</td>
                    <td style='text-align: right;'>{float(aatm_stats.get('max',0.0)):.2f}</td>
                </tr>
                <tr>
                    <td><b>Agr/Aground</b> (efecto suelo)</td>
                    <td style='text-align: right;'>{float(aground_stats.get('mean',0.0)):.2f}</td>
                    <td style='text-align: right;'>{float(aground_stats.get('max',0.0)):.2f}</td>
                </tr>
                <tr>
                    <td><b>Abar</b> (máximo entre contribuyentes)</td>
                    <td style='text-align: right;'>{float(abar_stats.get('mean',0.0)):.2f}</td>
                    <td style='text-align: right;'>{float(abar_stats.get('max',0.0)):.2f}</td>
                </tr>
            </table>
        </div>
        
        <h2>4. CONFIGURACIÓN Y PARÁMETROS</h2>
        
        <div class='card'>
            <h3>⚙️ Ecuación Utilizada</h3>
            <div class='formula'>{equation}</div>
            <p><em>{interpretation}</em></p>
        </div>
        
        <div class='card'>
            <h3>📋 Parámetros del Cálculo</h3>
            <ul>{''.join(param_lines)}</ul>
            {pressure_warning_html}
            <p><b>Trayectos con G distinto del global:</b> {int(ground_diag.get('from_landuse_count',0))} ({float(ground_diag.get('from_landuse_pct',0.0)):.1f}%)</p>
        </div>
        
        <div class='card'>
            <h3>✓ Términos Activos</h3>
            <ul>{''.join(term_lines)}</ul>
        </div>
        
        <h2>5. FÍSICA DETALLADA Y TRAZABILIDAD DEL CÁLCULO</h2>
        {glossary_html}
        {equations_detail_html}
        
        <h2>6. GRUPOS FUENTE ACÚSTICOS</h2>
        <div class='card'>
            <h3>⚡ LwA Efectivo por Grupo</h3>
            <ul>{''.join(eff_lines) if eff_lines else '<li>No disponible</li>'}</ul>
        </div>
        
        <h2>7. DISTRIBUCIÓN POR TIPO DE RECEPTOR</h2>
        <div style='display: grid; grid-template-columns: 1fr 1fr; gap: 16px;'>
            <div class='card'>
                <h3>📍 Receptores por Categoría</h3>
                <ul>{rec_types_html if rec_types_html else '<li>No disponible</li>'}</ul>
            </div>
            <div class='card'>
                <h3>✓ Cumplimiento por Categoría</h3>
                <ul>{compliance_html if compliance_html else '<li>No disponible</li>'}</ul>
            </div>
        </div>
        
        <div class='disclaimer'>
            <strong>⚠️ Limitaciones y recomendaciones</strong>
            <p><b>Motor Rápido:</b> Apropiado para screening preliminar y mapas ágiles.</p>
            <p><b>Motor ISO-aligned:</b> Apropiado para estudios técnicos preliminares, comparativas e iteración de diseño.</p>
            <p><b>Simplificaciones conocidas:</b> Aatm simplificado (tablas + correcciones); Agr y Abar con aproximaciones básicas; directividad Dc asumida 0 dB; Cmet/corrección meteorológica de largo plazo no aplicada.</p>
            <p><b>Modelos múltiples:</b> soportados mediante capas/grupos fuente independientes. Mezclar varios modelos dentro de una sola capa por atributos no está habilitado en esta versión experimental.</p>
            <p><b>Raster ISO + MDT:</b> usa la misma lógica de apantallamiento topográfico que los receptores puntuales, pero puede ser costoso en mapas grandes.</p>
            <p><b>Recomendación:</b> Para estudios regulatorios críticos, validar con mediciones o software comercial certificado.</p>
        </div>
        """
        self.page_summary.setHtml(translate_html(html))

    def _fill_models(self):
        model_diag = self._res.get("model_diag", {}) or {}
        rows: List[tuple] = []
        for name, d in model_diag.items():
            dia = d.get("diameter")
            hh = d.get("hub_height")
            mode = str(d.get('acoustic_mode') or 'fixed').lower()
            if mode == 'curve' and str(d.get('curve_path') or '').strip():
                note = str(d.get('curve_note') or 'Acoustic curve activa')
            else:
                note = 'LwA fijo por grupo fuente acústico'
            rows.append((str(name), int(d.get("count", 0)), float(d.get("lwa", 0.0)), hh, dia, note))
        self.tbl_models.setRowCount(len(rows))
        for r, row in enumerate(rows):
            vals = [
                row[0], str(row[1]), f"{row[2]:.1f}",
                "-" if row[3] is None or (isinstance(row[3], float) and not (row[3] == row[3])) else f"{float(row[3]):.1f}",
                "-" if row[4] is None or (isinstance(row[4], float) and not (row[4] == row[4])) else f"{float(row[4]):.1f}",
                row[5],
            ]
            for c, v in enumerate(vals):
                it = QtWidgets.QTableWidgetItem(v)
                it.setFlags(it.flags() & ~QtCore.Qt.ItemIsEditable)
                self.tbl_models.setItem(r, c, it)
        self.tbl_models.resizeColumnsToContents()

    def _feature_value_last(self, feat, field_name, default=""):
        """Return the last field named ``field_name`` from a QgsFeature.

        Receiver input layers can already contain generic names such as
        ``state`` or ``limit_dba``. QGIS name lookup returns the first match,
        which can silently pick the original receiver attribute instead of the
        computed noise output. The computed fields are appended at the end, so
        use the last matching index for UI/export fallbacks.
        """
        try:
            fields = feat.fields()
            idx = -1
            for i in range(fields.count()):
                if fields.at(i).name() == field_name:
                    idx = i
            if idx >= 0:
                return feat.attribute(idx)
        except Exception:
            pass
        try:
            return feat[field_name]
        except Exception:
            return default

    def _fill_top_receivers(self):
        # Prefer named payload rows. They are created by the engine with stable
        # semantic keys and avoid both duplicate input-field names and raw
        # attribute-order shifts in the QGIS memory layer.
        payload_rows = self._payload_top_receivers()[:15]
        feats = []
        if not payload_rows:
            layer = self._res.get("result_layer")
            if isinstance(layer, QgsVectorLayer):
                try:
                    for f in layer.getFeatures():
                        feats.append(f)
                except Exception:
                    feats = []
            def keyf(f):
                try:
                    return float(self._feature_value_last(f, "noise_dba", 0.0) or 0.0)
                except Exception:
                    return -1e9
            feats = sorted(feats, key=keyf, reverse=True)[:15]
        row_count = len(payload_rows) if payload_rows else len(feats)
        self.tbl_top.setRowCount(row_count)
        iterable = payload_rows if payload_rows else feats
        for r, f in enumerate(iterable):
            if isinstance(f, dict):
                clean_row = self._clean_receiver_row(f)
            else:
                raw = {"fid": f.id()}
                for key in CONSULTANCY_RECEIVER_KEYS:
                    raw[key] = self._feature_value_last(f, key, "")
                clean_row = self._clean_receiver_row(raw)
            for c, header in enumerate(CONSULTANCY_RECEIVER_HEADERS):
                v = str(clean_row.get(header, ""))
                it = QtWidgets.QTableWidgetItem(v)
                it.setFlags(it.flags() & ~QtCore.Qt.ItemIsEditable)
                self.tbl_top.setItem(r, c, it)
        self.tbl_top.resizeColumnsToContents()


    def _fill_mdt_screening(self):
        """Fill a DEM/MDT audit table sorted by screening, not by noise level."""
        rows = [dict(r) for r in self._payload_receiver_rows() if isinstance(r, dict)]

        def _f(d, key, default=0.0):
            try:
                v = d.get(key, default)
                if v is None or str(v).strip().lower() in ('', 'none', 'nan'):
                    return default
                return float(v)
            except Exception:
                return default

        # Keep covered receivers first.  Sort by active Abar, then by largest
        # detected obstacle, then by acoustic level.  This makes receivers with
        # strong terrain screening visible even if their total sound level is low.
        covered = [r for r in rows if _f(r, 'n_src', 0.0) > 0.0]
        covered.sort(
            key=lambda d: (
                _f(d, 'abar_max_db', 0.0),
                _f(d, 'maxobs_h', 0.0),
                _f(d, 'noise_dba', -1.0e99),
            ),
            reverse=True,
        )
        visible = covered[:30]

        keys = [
            'rec_id', 'noise_dba', 'n_src', 'abar_max_db', 'abar_ew_db',
            'abar_screen_n', 'abar_state', 'abar_db', 'maxab_src',
            'maxab_state', 'maxab_obs_h', 'maxab_thr', 'maxab_d1',
            'maxab_d2', 'maxobs_src', 'maxobs_state', 'maxobs_h',
            'maxobs_thr', 'maxobs_d1', 'maxobs_d2', 'rec_z_m',
            'rec_h_m', 'rec_ac_z_m', 'src_z_m', 'src_ac_z_m',
            'maxab_src_z', 'maxab_src_ac_z',
        ]

        self.tbl_mdt.setRowCount(len(visible))
        for r, row in enumerate(visible):
            for c, k in enumerate(keys):
                val = row.get(k, "")
                if k == "rec_id" and (val is None or str(val).strip() == ""):
                    val = row.get("fid", "")
                if val is None or str(val).strip().lower() in ('none', 'nan'):
                    v = "N/A"
                elif isinstance(val, float):
                    v = f"{val:.2f}"
                else:
                    v = str(val)
                it = QtWidgets.QTableWidgetItem(v)
                it.setFlags(it.flags() & ~QtCore.Qt.ItemIsEditable)
                self.tbl_mdt.setItem(r, c, it)
        self.tbl_mdt.resizeColumnsToContents()


    def _format_receiver_value(self, key: str, val) -> str:
        if key == "rec_id" and (val is None or str(val).strip() == ""):
            return ""
        if key == "exceeds":
            try:
                return "sí" if int(float(val or 0)) == 1 else "no"
            except Exception:
                txt = str(val or "").strip().lower()
                return "sí" if txt in ("true", "yes", "sí", "si", "1") else "no"
        if val is None:
            return "N/A"
        txt = str(val).strip()
        if txt.lower() in ("", "none", "nan", "n/a"):
            return "N/A"
        try:
            fval = float(txt.replace(",", "."))
        except Exception:
            return txt
        if not (fval == fval):
            return "N/A"
        if key in ("n_src",):
            return str(int(round(fval)))
        if key in ("noise_dba", "limit_dba", "margin_db", "src_lwa", "adiv_db", "aatm_db", "aground_db", "abar_max_db", "ground_g"):
            return f"{fval:.2f}"
        if key in ("near_m", "rec_h_m", "rec_z_m", "rec_ac_z_m"):
            return f"{fval:.1f}"
        return f"{fval:.2f}"


    def _clean_receiver_row(self, row: Dict[str, object]) -> Dict[str, object]:
        out: Dict[str, object] = {}
        for key, label in CONSULTANCY_RECEIVER_COLUMNS:
            val = row.get(key, "") if isinstance(row, dict) else ""
            if key == "rec_id" and (val is None or str(val).strip() == "") and isinstance(row, dict):
                val = row.get("fid", "")
            out[label] = self._format_receiver_value(key, val)
        return out


    def _receiver_rows_for_export(self) -> List[Dict[str, object]]:
        rows = self._res.get('receiver_rows') or []
        if not rows:
            layer = self._res.get('result_layer')
            if isinstance(layer, QgsVectorLayer):
                rows = list(self._iter_layer_dicts(layer))
        if not rows:
            rows = self._payload_top_receivers()
        return [self._clean_receiver_row(r) for r in rows if isinstance(r, dict)]



    def _write_layer_csv(self, layer: QgsVectorLayer, path: str):
        field_names = [f.name() for f in layer.fields()]
        with open(path, 'w', newline='', encoding='utf-8-sig') as fh:
            writer = csv.writer(fh, delimiter=';')
            writer.writerow(['fid'] + field_names)
            for feat in layer.getFeatures():
                row = [feat.id()]
                for name in field_names:
                    try:
                        val = feat[name]
                    except Exception:
                        val = ''
                    row.append(val)
                writer.writerow(row)

    def _iter_layer_dicts(self, layer: QgsVectorLayer):
        field_names = [f.name() for f in layer.fields()]
        for feat in layer.getFeatures():
            row = {"fid": feat.id()}
            for name in field_names:
                try:
                    row[name] = feat[name]
                except Exception:
                    row[name] = ""
            yield row

    def _collect_exceedance_rows(self):
        rows_source = self._res.get('receiver_rows') or []
        layer = self._res.get('result_layer')
        if not rows_source and isinstance(layer, QgsVectorLayer):
            rows_source = list(self._iter_layer_dicts(layer))
        rows = []
        for row in rows_source or []:
            try:
                exceeds = int(float(row.get('exceeds') or 0))
            except Exception:
                exceeds = 0
            if exceeds == 1:
                rows.append(self._clean_receiver_row(row))
        return rows

    def _write_rows_csv(self, rows, path: str):
        rows = list(rows or [])
        headers = []
        for r in rows:
            for k in r.keys():
                if k not in headers:
                    headers.append(k)
        with open(path, 'w', newline='', encoding='utf-8-sig') as fh:
            writer = csv.writer(fh, delimiter=';')
            writer.writerow(headers)
            for row in rows:
                writer.writerow([row.get(h, '') for h in headers])

    def _table_headers(self, table: QtWidgets.QTableWidget) -> List[str]:
        headers: List[str] = []
        for c in range(table.columnCount()):
            item = table.horizontalHeaderItem(c)
            headers.append(item.text() if item is not None else f"col_{c+1}")
        return headers

    def _collect_table_rows(self, table: QtWidgets.QTableWidget) -> List[Dict[str, object]]:
        headers = self._table_headers(table)
        rows: List[Dict[str, object]] = []
        for r in range(table.rowCount()):
            row: Dict[str, object] = {}
            has_value = False
            for c, h in enumerate(headers):
                item = table.item(r, c)
                text = item.text() if item is not None else ""
                if str(text).strip():
                    has_value = True
                row[h] = text
            if has_value:
                rows.append(row)
        return rows

    def _write_table_csv(self, table: QtWidgets.QTableWidget, path: str):
        headers = self._table_headers(table)
        with open(path, 'w', newline='', encoding='utf-8-sig') as fh:
            writer = csv.writer(fh, delimiter=';')
            writer.writerow(headers)
            for row in self._collect_table_rows(table):
                writer.writerow([row.get(h, '') for h in headers])

    def _append_table_sheet(self, wb, title: str, table: QtWidgets.QTableWidget):
        ws = wb.create_sheet(title=title[:31] or 'Hoja')
        headers = self._table_headers(table)
        ws.append(headers)
        rows = self._collect_table_rows(table)
        if not rows:
            ws.append(['sin_datos'])
        else:
            for row in rows:
                ws.append([row.get(h, '') for h in headers])
        try:
            for idx, h in enumerate(headers, start=1):
                width = max(len(str(h)), max((len(str(r.get(h, ''))) for r in rows), default=0))
                ws.column_dimensions[chr(64 + idx) if idx <= 26 else ws.cell(row=1, column=idx).column_letter].width = min(max(width + 2, 10), 45)
        except Exception:
            pass

    def _append_sheet(self, wb, title: str, rows):
        ws = wb.create_sheet(title=title[:31] or 'Hoja')
        rows = list(rows or [])
        headers = []
        for r in rows:
            for k in r.keys():
                if k not in headers:
                    headers.append(k)
        if not headers:
            ws.append(['sin_datos'])
            ws.append([''])
            return
        ws.append(headers)
        for row in rows:
            ws.append([row.get(h, '') for h in headers])
        try:
            for idx, h in enumerate(headers, start=1):
                width = max(len(str(h)), max((len(str(r.get(h, ''))) for r in rows), default=0))
                ws.column_dimensions[chr(64 + idx) if idx <= 26 else ws.cell(row=1, column=idx).column_letter].width = min(max(width + 2, 10), 40)
        except Exception:
            pass

    def _export_summary(self):
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, 'Exportar resumen', os.path.expanduser('~/ruido_resumen.html'), 'HTML (*.html);;Texto (*.txt)')
        if not path:
            return
        try:
            if path.lower().endswith('.txt'):
                with open(path, 'w', encoding='utf-8') as fh:
                    fh.write(self.page_summary.toPlainText())
            else:
                if not path.lower().endswith('.html'):
                    path += '.html'
                with open(path, 'w', encoding='utf-8') as fh:
                    fh.write(self.page_summary.toHtml())
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, 'Exportar resumen', f'No se pudo exportar el resumen:\n{e}')

    def _export_receivers_csv(self):
        rows = self._receiver_rows_for_export()
        if not rows:
            QtWidgets.QMessageBox.information(self, 'Exportar receptores', 'No hay filas de receptores para exportar.')
            return
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, 'Exportar receptores CSV', os.path.expanduser('~/ruido_receptores.csv'), 'CSV (*.csv)')
        if not path:
            return
        try:
            if not path.lower().endswith('.csv'):
                path += '.csv'
            self._write_rows_csv(rows, path)
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, 'Exportar receptores', f'No se pudo exportar el CSV:\n{e}')

    def _write_dict_rows_csv(self, rows, path: str):
        # Deterministic CSV for dictionaries. Keeps debug exports independent
        # from visible table columns and QGIS field ordering.
        keys = []
        for row in rows or []:
            if not isinstance(row, dict):
                continue
            for k in row.keys():
                if k not in keys:
                    keys.append(k)
        with open(path, 'w', newline='', encoding='utf-8-sig') as fh:
            writer = csv.writer(fh, delimiter=';')
            writer.writerow(keys)
            for row in rows or []:
                writer.writerow([row.get(k, '') if isinstance(row, dict) else '' for k in keys])


    def _export_path_diagnostics_csv(self):
        rows = self._res.get('path_diagnostics') or []
        if not rows:
            QtWidgets.QMessageBox.information(self, 'Exportar diagnóstico MDT', 'No hay diagnóstico por pares fuente-receptor disponible. Recalcula con motor ISO-aligned y fuentes dentro del radio. Este CSV es el que permite auditar cada turbina contra cada receptor.')
            return
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, 'Exportar diagnóstico MDT por pares CSV', os.path.expanduser('~/ruido_mdt_pair_diagnostics.csv'), 'CSV (*.csv)')
        if not path:
            return
        try:
            if not path.lower().endswith('.csv'):
                path += '.csv'
            self._write_dict_rows_csv(rows, path)
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, 'Exportar diagnóstico MDT', f'No se pudo exportar el CSV de diagnóstico MDT:\n{e}')


    def _export_top_receivers_csv(self):
        if self.tbl_top.rowCount() <= 0:
            QtWidgets.QMessageBox.information(self, 'Exportar Top receivers', 'No hay filas de Top receivers para exportar.')
            return
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, 'Exportar Top receivers CSV', os.path.expanduser('~/ruido_top_receivers.csv'), 'CSV (*.csv)')
        if not path:
            return
        try:
            if not path.lower().endswith('.csv'):
                path += '.csv'
            self._write_table_csv(self.tbl_top, path)
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, 'Exportar Top receivers', f'No se pudo exportar el CSV de Top receivers:\n{e}')

    def _export_mdt_screening_csv(self):
        if self.tbl_mdt.rowCount() <= 0:
            QtWidgets.QMessageBox.information(self, 'Exportar screening MDT', 'No hay filas de screening MDT para exportar.')
            return
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, 'Exportar screening MDT CSV', os.path.expanduser('~/ruido_mdt_screening_receivers.csv'), 'CSV (*.csv)')
        if not path:
            return
        try:
            if not path.lower().endswith('.csv'):
                path += '.csv'
            self._write_table_csv(self.tbl_mdt, path)
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, 'Exportar screening MDT', f'No se pudo exportar el CSV de screening MDT:\n{e}')


    def _export_sources_csv(self):
        layer = self._res.get('sources_layer')
        if not isinstance(layer, QgsVectorLayer):
            QtWidgets.QMessageBox.information(self, 'Exportar grupos fuente', 'No hay capa de fuentes para exportar.')
            return
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, 'Exportar grupos fuente CSV', os.path.expanduser('~/ruido_fuentes.csv'), 'CSV (*.csv)')
        if not path:
            return
        try:
            if not path.lower().endswith('.csv'):
                path += '.csv'
            self._write_layer_csv(layer, path)
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, 'Exportar grupos fuente', f'No se pudo exportar el CSV:\n{e}')

    def _export_exceedances_csv(self):
        rows = self._collect_exceedance_rows()
        if not rows:
            QtWidgets.QMessageBox.information(self, 'Exportar excedencias', 'No hay receptores que superen el límite en este cálculo.')
            return
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, 'Exportar excedencias CSV', os.path.expanduser('~/ruido_excedencias.csv'), 'CSV (*.csv)')
        if not path:
            return
        try:
            if not path.lower().endswith('.csv'):
                path += '.csv'
            self._write_rows_csv(rows, path)
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, 'Exportar excedencias', f'No se pudo exportar el CSV:\n{e}')

    def _export_package_xlsx(self):
        if Workbook is None:
            QtWidgets.QMessageBox.information(self, 'Exportar paquete XLSX', 'openpyxl no está disponible en este entorno de QGIS. Usa las exportaciones CSV o instala openpyxl.')
            return
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, 'Exportar paquete XLSX', os.path.expanduser('~/ruido_paquete.xlsx'), 'Excel (*.xlsx)')
        if not path:
            return
        try:
            if not path.lower().endswith('.xlsx'):
                path += '.xlsx'
            wb = Workbook()
            ws0 = wb.active
            ws0.title = 'Resumen'
            plain = self.page_summary.toPlainText().splitlines()
            for line in plain:
                ws0.append([line])
            self._append_table_sheet(wb, 'Modelos', self.tbl_models)
            self._append_sheet(wb, 'Receptores', self._receiver_rows_for_export())
            self._append_sheet(wb, 'Excedencias', self._collect_exceedance_rows())
            self._append_table_sheet(wb, 'Capas_creadas', self.tbl_layers)
            wb.save(path)
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, 'Exportar paquete XLSX', f'No se pudo exportar el XLSX:\n{e}')

    def _fill_layers(self):
        entries = [
            ("Noise · Receivers", self._res.get("result_layer") is not None),
            ("Noise · Sources", self._res.get("sources_layer") is not None),
            ("Noise · Dominant links", self._res.get("links_layer") is not None),
            ("Noise · Receivers outside radius", self._res.get("uncovered_layer") is not None),
            ("Noise · Mapa", self._res.get("grid_layer") is not None),
            ("Noise · Isophones", self._res.get("iso_layer") is not None),
        ]
        self.tbl_layers.setRowCount(len(entries))
        for r, (name, ok) in enumerate(entries):
            for c, v in enumerate([name, "creada" if ok else "no creada"]):
                it = QtWidgets.QTableWidgetItem(v)
                it.setFlags(it.flags() & ~QtCore.Qt.ItemIsEditable)
                self.tbl_layers.setItem(r, c, it)
        self.tbl_layers.resizeColumnsToContents()

    def _infer_limit_stats_from_layer(self) -> dict:
        layer = self._res.get("result_layer")
        default = {"min": 45.0, "max": 45.0, "mode": "global", "scenario": "custom", "unique_count": 1}
        if not isinstance(layer, QgsVectorLayer):
            return default
        vals = []
        mode = None
        scenario = None
        try:
            for f in layer.getFeatures():
                try:
                    v = f["limit_dba"]
                    if v is not None:
                        vals.append(float(v))
                except Exception:
                    pass
                if mode is None:
                    try:
                        mode = str(f["limit_src"] or "").strip().lower() or None
                    except Exception:
                        pass
                if scenario is None:
                    try:
                        scenario = str(f["limit_scn"] or "").strip().lower() or None
                    except Exception:
                        pass
        except Exception:
            return default
        if not vals:
            return default
        return {
            "min": min(vals),
            "max": max(vals),
            "mode": mode or "global",
            "scenario": scenario or "custom",
            "unique_count": len({round(v, 6) for v in vals}),
        }
